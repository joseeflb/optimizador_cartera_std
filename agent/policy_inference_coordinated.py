# ============================================
# agent/policy_inference_coordinated.py
# — Decisión coordinada MICRO+MACRO (estilo analista IB)
# v2.2 (BANK-READY · guardrails fire-sale vs BOOK · reestructura bankable · audit trail)
# ============================================
"""
POC — OPTIMIZADOR DE CARTERAS EN DEFAULT (NPL · Basilea III STD)

Salida exigida (por préstamo):
  - decision_micro
  - decision_macro
  - decision_final
  - final_rationale (explicación coordinada micro↔macro)

Export:
  - Un Excel por estrategia (PRUDENTE / BALANCEADO / DESINVERSION)
  - Cada Excel incluye:
      * Hoja Loan_Decisions
      * Hoja Portfolio_Summary

Cambios clave v2.2 (consistencia con price_simulator.py y restructure_optimizer.py):
  - ✅ Venta: simulate_npl_price usando book_value/coverage_rate si existen.
  - ✅ Guardrail fire-sale: por Price/Book (no por Price/Workout).
  - ✅ RW: coerción robusta (100/150 → 1.0/1.5) y discretización DEFAULT.
  - ✅ Reestructura: usa lógica bankable (optimize_restructure) con gates duros por segmento (PTI/DSCR).
  - ✅ Score económico homogéneo (EUR): workout_value_proxy / sale_price_net / restructure_value_net.
  - ✅ Audit trail: drivers numéricos + reglas disparadas por postura.
"""

from __future__ import annotations

import os
import sys
import time
import math
import logging
import argparse
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.append(ROOT_DIR)

import config as cfg
from optimizer.price_simulator import simulate_npl_price
from optimizer.restructure_optimizer import optimize_restructure  # usa gates bankables + hazard forward al horizonte


# -----------------------------
# Global config handles
# -----------------------------
CFG = cfg.CONFIG
REG = CFG.regulacion

# CAP_RATIO robusto
if hasattr(REG, "required_total_capital_ratio") and callable(REG.required_total_capital_ratio):
    CAP_RATIO = float(REG.required_total_capital_ratio())
else:
    base = float(getattr(REG, "total_capital_min", 0.08))
    buf = getattr(getattr(REG, "buffers", None), "total_buffer", lambda: 0.0)()
    CAP_RATIO = float(base + buf)

HURDLE = float(getattr(REG, "hurdle_rate", 0.12))
COST_FUND_ANNUAL = 0.006  # coherente con env/loan_env.py

REPORTS_DIR = os.path.join(ROOT_DIR, "reports")
LOGS_DIR = os.path.join(ROOT_DIR, "logs")
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOGS_DIR, "policy_inference_coordinated.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("policy_inference_coordinated")


# ============================================================
# Config de inferencia coordinada
# ============================================================
@dataclass
class CoordinatedInferenceConfig:
    portfolio_path: str
    tag: str = "coordinated"
    seed: int = 42

    # Horizonte económico para workout/carry (meses)
    horizon_months: int = 24

    # Si True, genera explicación larga por préstamo
    verbose_explanations: bool = True

    # Export a Excel (además de CSV)
    export_excel: bool = True

    # Multi-postura (genera 3 excels en 1 carpeta)
    multi_posture: bool = True

    # Si se indica, fuerza postura (prudente|balanceado|desinversion)
    risk_posture: Optional[str] = None


def _now_tag() -> str:
    return time.strftime("%Y%m%d_%H%M%S")

def harmonize_portfolio_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    # --- map monthly_* -> legacy names usados por optimizadores
    if "ingreso_mensual" not in df.columns and "monthly_income" in df.columns:
        df["ingreso_mensual"] = df["monthly_income"]
    if "cashflow_operativo_mensual" not in df.columns and "monthly_cfo" in df.columns:
        df["cashflow_operativo_mensual"] = df["monthly_cfo"]
    if "cuota_mensual" not in df.columns and "monthly_payment" in df.columns:
        df["cuota_mensual"] = df["monthly_payment"]

    # --- PTI / DSCR: preferimos pre si existe; si no, derivamos desde monthly_*
    if "PTI" not in df.columns:
        if "PTI_pre" in df.columns:
            df["PTI"] = df["PTI_pre"]
        elif ("monthly_payment" in df.columns) and ("monthly_income" in df.columns):
            df["PTI"] = df["monthly_payment"] / df["monthly_income"].replace(0, np.nan)

    if "DSCR" not in df.columns:
        if "DSCR_pre" in df.columns:
            df["DSCR"] = df["DSCR_pre"]
        elif ("monthly_cfo" in df.columns) and ("monthly_payment" in df.columns):
            df["DSCR"] = df["monthly_cfo"] / df["monthly_payment"].replace(0, np.nan)

    # --- book_value: si no viene, aproximación consistente
    if "book_value" not in df.columns:
        if "coverage_rate" in df.columns and "EAD" in df.columns:
            df["book_value"] = df["EAD"] * (1.0 - df["coverage_rate"].clip(0, 1))
        elif "LGD" in df.columns and "EAD" in df.columns:
            df["book_value"] = df["EAD"] * (1.0 - df["LGD"].clip(0, 1))

    # --- normalización mínima de segment (para gates retail vs corp)
    if "segment" in df.columns:
        df["segment"] = df["segment"].astype(str).str.strip()

    return df

def _load_portfolio_df(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"❌ Cartera no encontrada: {path}")
    ext = os.path.splitext(path.lower())[1]
    df = pd.read_excel(path) if ext in (".xlsx", ".xls") else pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # ✅ Harmonize antes de asegurar loan_id (crea PTI/DSCR/book_value y aliases)
    df = harmonize_portfolio_schema(df)

    if cfg.ID_COL not in df.columns:
        if "id" in df.columns:
            df[cfg.ID_COL] = df["id"].astype(str)
        else:
            df[cfg.ID_COL] = [f"loan_{i}" for i in range(len(df))]

    return df



# ============================================================
# Utilidades
# ============================================================
def _safe_float(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, (int, float, np.floating)):
            return float(x)
        if isinstance(x, str) and x.strip() == "":
            return default
        return float(x)
    except Exception:
        return default


def _is_nan(x: Any) -> bool:
    try:
        return x is None or (isinstance(x, float) and math.isnan(x))
    except Exception:
        return x is None


def _coerce_rw(x: Any) -> float:
    """RW puede venir 1.0/1.5 o 100/150. Devolvemos multiplicador."""
    try:
        v = float(x)
    except Exception:
        return float("nan")
    if v > 10.0:
        v = v / 100.0
    return float(v)


def _segment_enum(row: Dict[str, Any]) -> cfg.Segmento:
    # soporta "segmento_banco"/"segment_raw"/"segment"
    for k in ("segmento_banco", "segment_raw", "segment", "Segment", "SEGMENT"):
        if k in row and row[k] is not None and str(row[k]).strip() != "":
            s = str(row[k]).strip().upper()
            break
    else:
        s = "CORPORATE"
    s = s.replace(" ", "_")
    try:
        return cfg.Segmento[s]
    except Exception:
        return cfg.Segmento.CORPORATE


def _rating_str(row: Dict[str, Any]) -> str:
    r = str(row.get("rating", row.get("Rating", "BBB"))).strip().upper()
    return r if r in getattr(cfg, "VALID_RATINGS", {"AAA", "AA", "A", "BBB", "BB", "B", "CCC"}) else "BBB"


def _is_secured(row: Dict[str, Any]) -> bool:
    # No inventamos: si existe, úsalo. Si no, Mortgage => True, resto => False.
    if "secured" in row and row["secured"] is not None and str(row["secured"]).strip() != "":
        return bool(row["secured"])
    seg = _segment_enum(row)
    return seg == cfg.Segmento.MORTGAGE


def _resolve_rw_default(row: Dict[str, Any]) -> float:
    """RW default discreto y robusto."""
    seg = _segment_enum(row)
    rat = _rating_str(row)
    secured_by_mortgage = (seg == cfg.Segmento.MORTGAGE)
    try:
        rw = CFG.basel_map.resolve_rw(seg, rat, cfg.EstadoCredito.DEFAULT, secured_by_mortgage=secured_by_mortgage)
        rw = _coerce_rw(rw)
        if np.isfinite(rw):
            return float(1.00 if rw < 1.25 else 1.50)
    except Exception:
        pass
    return float(1.00 if seg == cfg.Segmento.MORTGAGE else 1.50)


def _capital_release(row: Dict[str, Any]) -> float:
    ead = _safe_float(row.get("EAD", 0.0))
    rw = _coerce_rw(row.get("RW", np.nan))
    if not np.isfinite(rw):
        rw = _resolve_rw_default(row)
    else:
        rw = 1.00 if rw < 1.25 else 1.50
    rwa = float(ead * rw)
    return float(rwa * CAP_RATIO)


def _capital_carry_savings(cap_release: float, horizon_months: int) -> float:
    years = max(1e-9, horizon_months / 12.0)
    return float(cap_release * HURDLE * years)


def _compute_eva_proxy(row: Dict[str, Any], horizon_months: int) -> float:
    """
    EVA proxy coherente con env/loan_env.py:
      NI = EAD*rate - EL_annual - EAD*fund
      EL_annual = (PD*LGD*EAD)/horizon_years   (PD forward al horizonte)
      RWA = EAD*RW
      EVA = NI - HURDLE*RWA
    """
    ead = _safe_float(row.get("EAD", 0.0))
    if ead <= 0:
        return 0.0

    rate = _safe_float(row.get("rate", row.get("Rate", 0.06)), default=0.06)
    pd_fwd = float(np.clip(_safe_float(row.get("PD", 0.60), default=0.60), 0.0, 0.999))
    lgd = float(np.clip(_safe_float(row.get("LGD", 0.60), default=0.60), 0.0, 1.0))

    rw = _coerce_rw(row.get("RW", np.nan))
    if not np.isfinite(rw):
        rw = _resolve_rw_default(row)
    else:
        rw = 1.00 if rw < 1.25 else 1.50

    horizon_years = max(1e-9, horizon_months / 12.0)
    el_life = float(pd_fwd * lgd * ead)
    el_ann = float(el_life / horizon_years)

    ni = float(ead * rate - el_ann - ead * COST_FUND_ANNUAL)
    rwa = float(ead * rw)
    eva = float(ni - HURDLE * rwa)
    return eva


# ============================================================
# Valoraciones tipo IB para NPL (defendibles)
# ============================================================
def _workout_value_proxy(row: Dict[str, Any], horizon_months: int) -> Tuple[float, Dict[str, float]]:
    """
    Proxy de valor de workout (mantener):
      - Base recovery IFRS: EAD*(1-LGD)
      - Ajustes por DPD (timing/complexity), segmento y colateral
      - Coste legal esperado (config)
      - Descuento temporal a horizon (hurdle)

    Nota: es un proxy en EUR (NPV-like), comparable con precio neto de venta.
    """
    ead = _safe_float(row.get("EAD", 0.0))
    lgd = float(np.clip(_safe_float(row.get("LGD", 0.60), default=0.60), 0.0, 1.0))
    dpd = _safe_float(row.get("DPD", 180.0))
    secured = _is_secured(row)
    seg_enum = _segment_enum(row)

    base_recovery = float(ead * (1.0 - lgd))

    alpha_range = CFG.precio_venta.alpha_by_segment.get(seg_enum, (0.35, 0.55))
    alpha_mean = float(np.mean(alpha_range))

    dpd_a = float(getattr(CFG.precio_venta, "dpd_penalty_a", getattr(CFG.precio_venta, "buyer_discount_rate", 0.10)))
    dpd_mult = 1.0 / (1.0 + dpd_a * math.log1p(max(0.0, dpd)))

    # coherente con price_simulator: unsecured penaliza
    if secured:
        sec_mult = 1.0
    else:
        try:
            legal_mean_tmp = float(np.mean(CFG.precio_venta.coste_legal_estimado_abs))
        except Exception:
            legal_mean_tmp = 5000.0
        sec_mult = float(0.85 - min(legal_mean_tmp, 20000.0) / 200000.0)

    det_mult = max(0.01, alpha_mean * dpd_mult * sec_mult)

    try:
        legal_mean = float(np.mean(CFG.precio_venta.coste_legal_estimado_abs))
    except Exception:
        legal_mean = 2500.0

    gross = float(base_recovery * det_mult)
    net_before_time = float(max(0.0, gross - legal_mean))

    years = max(1e-9, horizon_months / 12.0)
    discount = float(1.0 / ((1.0 + HURDLE) ** years))
    net = float(net_before_time * discount)

    return net, {
        "base_recovery": float(base_recovery),
        "alpha_mean": float(alpha_mean),
        "dpd_mult": float(dpd_mult),
        "sec_mult": float(sec_mult),
        "legal_mean": float(legal_mean),
        "discount": float(discount),
        "gross": float(gross),
        "net": float(net),
    }


def _sale_value(row: Dict[str, Any]) -> Tuple[float, Dict[str, Any]]:
    """
    Precio neto de mercado (simulado) + desglose.
    P&L contable y fire-sale se evalúan vs book_value (si existe).
    """
    ead = _safe_float(row.get("EAD", 0.0))
    lgd = _safe_float(row.get("LGD", 0.60))
    pd_fwd = _safe_float(row.get("PD", 0.60))
    dpd = _safe_float(row.get("DPD", 180.0))
    seg = str(row.get("segment", row.get("segment_raw", "corporate"))).upper()
    secured = _is_secured(row)
    rating = _rating_str(row)

    book_value = row.get("book_value", row.get("BOOK_VALUE", None))
    coverage_rate = row.get("coverage_rate", row.get("coverage", row.get("COVERAGE_RATE", None)))

    pd_prob = _safe_float(row.get("PD", np.nan), default=np.nan)
    book_value = row.get("book_value", None)
    coverage_rate = row.get("coverage_rate", None)

    out = simulate_npl_price(
        ead=ead,
        lgd=lgd,
        pd=None if not np.isfinite(pd_prob) else float(pd_prob),
        dpd=dpd,
        segment=seg,
        secured=secured,
        rating=rating,
        book_value=None if book_value is None else float(book_value),
        coverage_rate=None if coverage_rate is None else float(coverage_rate),
    )

    price_net = float(out.get("precio_optimo", 0.0))
    return price_net, out


def _best_restructure_value(
    row: Dict[str, Any],
    posture: cfg.BankProfile,
    horizon_months: int,
) -> Tuple[float, float, Dict[str, Any]]:
    """
    Devuelve:
      - restructure_value_net (EUR proxy comparable con workout/sale)
      - uplift_vs_workout (EUR)
      - breakdown audit

    Implementación:
      - Usa optimize_restructure (gates PTI/DSCR por segmento, hazard forward, RW perf por cure)
      - Convierte su salida a un valor proxy (workout_value_proxy sobre el estado post)
      - Ajusta por costes + parte de ahorro de capital/carry si cured (postura-dependent)
    """
    # Baseline
    v_workout, _ = _workout_value_proxy(row, horizon_months=horizon_months)

    seg_enum = _segment_enum(row)
    rating = _rating_str(row)

    ead = _safe_float(row.get("EAD", 0.0))
    rate = _safe_float(row.get("rate", 0.06), default=0.06)
    pd_fwd = _safe_float(row.get("PD", 0.60), default=0.60)
    lgd = _safe_float(row.get("LGD", 0.60), default=0.60)

    rw = _coerce_rw(row.get("RW", np.nan))
    if not np.isfinite(rw):
        rw = _resolve_rw_default(row)
    else:
        rw = 1.00 if rw < 1.25 else 1.50

    # Viabilidad inputs (compat con generator)
    monthly_income = row.get("monthly_income", row.get("ingreso_mensual", None))
    monthly_cfo = row.get("monthly_cfo", row.get("cashflow_operativo_mensual", None))

    # Umbrales por postura (bank_strategy)
    strat = CFG.bank_strategy
    esfuerzo_max = float(getattr(strat, "esfuerzo_alto", 0.40))
    dscr_min = float(getattr(strat, "dscr_min", 1.10))

    res = optimize_restructure(
        ead=ead,
        rate=rate,
        pd=pd_fwd,
        lgd=lgd,
        rw=rw,
        segment=seg_enum.name,
        rating=rating,
        ingreso_mensual=None if _is_nan(monthly_income) else float(monthly_income),
        cashflow_operativo_mensual=None if _is_nan(monthly_cfo) else float(monthly_cfo),
        esfuerzo_max=esfuerzo_max,
        dscr_min=dscr_min,
        hurdle=HURDLE,
    )

    if not bool(res.get("ok", False)):
        return v_workout, 0.0, {"feasible": 0, "reason": str(res.get("msg", "no_feasible_restructure")), "raw": res}

    # Estado post para valorar workout tras reestructurar (DPD cae, PD/LGD cambian, EAD cambia)
    row_post = dict(row)
    row_post["EAD"] = float(res.get("EAD_post", ead))
    row_post["LGD"] = float(res.get("LGD_post", lgd))
    row_post["PD"] = float(res.get("PD_post", pd_fwd))
    row_post["DPD"] = 0.0 if bool(res.get("cured", False)) else float(min(_safe_float(row.get("DPD", 180.0)), 60.0))
    # si cured, RW end podría bajar; para el workout_value_proxy no es crítico, pero lo guardamos para auditoría
    row_post["RW"] = float(res.get("RW_post_end", rw))

    v_post, bd_post = _workout_value_proxy(row_post, horizon_months=horizon_months)

    # costes one-off
    admin_cost = float(res.get("admin_cost", 0.0) or 0.0)
    quita_cost = float(res.get("quita_cost", 0.0) or 0.0)
    cost_total = float(admin_cost + quita_cost)

    v_net = float(max(0.0, v_post - cost_total))

    # ahorro de capital (postura-dependent): prudente menos agresivo, desinversión menos relevante
    cap_rel_est = float(res.get("capital_release_est", 0.0) or 0.0)
    cap_sav = _capital_carry_savings(cap_rel_est, horizon_months=horizon_months)

    if posture == cfg.BankProfile.PRUDENTE:
        gamma_cap = 0.35
    elif posture == cfg.BankProfile.DESINVERSION:
        gamma_cap = 0.20
    else:
        gamma_cap = 0.30

    if bool(res.get("cured", False)) and cap_rel_est > 0:
        v_net += float(gamma_cap * cap_sav)

    uplift = float(v_net - v_workout)

    breakdown = {
        "feasible": 1,
        "segment": seg_enum.name,
        "rating": rating,
        "gate_type": res.get("gate_type"),
        "cuota_mensual_post": res.get("cuota_mensual_post"),
        "PTI_post": res.get("PTI_post"),
        "DSCR_post": res.get("DSCR_post"),
        "cured": bool(res.get("cured", False)),
        "value_post_workout_proxy": float(v_post),
        "value_post_breakdown": bd_post,
        "admin_cost": float(admin_cost),
        "quita_cost": float(quita_cost),
        "cost_total": float(cost_total),
        "capital_release_est": float(cap_rel_est),
        "capital_savings": float(cap_sav),
        "gamma_cap": float(gamma_cap),
        "value_net": float(v_net),
        "uplift_vs_workout": float(uplift),
        "raw": res,
    }
    return float(v_net), float(uplift), breakdown


# ============================================================
# Micro decision (préstamo individual)
# ============================================================
def micro_decision(row: Dict[str, Any], posture: cfg.BankProfile, horizon_months: int) -> Dict[str, Any]:
    strat = CFG.bank_strategy

    v_workout, bd_workout = _workout_value_proxy(row, horizon_months=horizon_months)
    v_sale, bd_sale = _sale_value(row)
    fire_sale = bool(bd_sale.get("fire_sale", False))
    if fire_sale and posture != cfg.BankProfile.DESINVERSION:
        sale_guardrail_ok = False
    cap_rel = _capital_release(row)
    cap_sav = _capital_carry_savings(cap_rel, horizon_months=horizon_months)

    v_restruct, uplift_restruct, bd_restruct = _best_restructure_value(row, posture=posture, horizon_months=horizon_months)

    # ---------
    # Guardrails venta (fire-sale vs Book)
    # ---------
    fire_sale = bool(bd_sale.get("fire_sale", False))
    ratio_book = _safe_float(bd_sale.get("price_ratio_book", 0.0))
    thr_book = _safe_float(bd_sale.get("fire_sale_threshold_book", getattr(CFG.precio_venta, "fire_sale_price_ratio_book", 0.85)))

    allow_fire_sale = bool(row.get("allow_fire_sale", False))

    if posture == cfg.BankProfile.PRUDENTE:
        lambda_cap = 0.6
        # En prudente: no fire-sale, salvo excepción explícita
        sale_guardrail_ok = (not fire_sale) or allow_fire_sale
    elif posture == cfg.BankProfile.DESINVERSION:
        lambda_cap = 1.6
        # En desinversión se permite, pero aun así evitamos casos extremos salvo allow_fire_sale
        sale_guardrail_ok = (ratio_book >= 0.70) or allow_fire_sale
    else:
        lambda_cap = 1.0
        # Balanceado: no fire-sale salvo allow_fire_sale
        sale_guardrail_ok = (not fire_sale) or allow_fire_sale

    # Scores homogéneos en EUR
    total_sell = float(v_sale + lambda_cap * cap_sav)
    total_keep = float(v_workout)
    total_restruct = float(v_restruct)

    # Preferencia reestructura por tensión PTI/DSCR (si existen, se usan; si no, se intenta con monthly_*)
    pti = _safe_float(row.get("PTI", np.nan), default=float("nan"))
    dscr = _safe_float(row.get("DSCR", np.nan), default=float("nan"))

    if (not np.isfinite(pti)) or (not np.isfinite(dscr)):
        monthly_payment = _safe_float(row.get("monthly_payment", row.get("cuota_mensual", np.nan)), default=float("nan"))
        monthly_income = _safe_float(row.get("monthly_income", row.get("ingreso_mensual", np.nan)), default=float("nan"))
        monthly_cfo = _safe_float(row.get("monthly_cfo", row.get("cashflow_operativo_mensual", np.nan)), default=float("nan"))

        if np.isfinite(monthly_payment) and monthly_payment > 0 and np.isfinite(monthly_income) and monthly_income > 0:
            pti = monthly_payment / max(1e-9, monthly_income)
        if np.isfinite(monthly_payment) and monthly_payment > 0 and np.isfinite(monthly_cfo) and monthly_cfo > 0:
            dscr = monthly_cfo / max(1e-9, monthly_payment)

    esfuerzo_bajo = float(getattr(strat, "esfuerzo_bajo", 0.35))
    dscr_min = float(getattr(strat, "dscr_min", 1.10))

    prefer_restruct = (np.isfinite(pti) and pti > esfuerzo_bajo) or (np.isfinite(dscr) and dscr > 0 and dscr < dscr_min)

    # Decisión micro:
    # - si reestructura factible y mejora neta, y hay tensión → reestructurar
    # - si venta pasa guardrail y domina → vender
    # - si no, mantener (workout)
    if bd_restruct.get("feasible", 0) == 1 and uplift_restruct > 0 and prefer_restruct:
        decision = "REESTRUCTURAR"
        why = "Viabilidad OK (gates) y uplift neto vs workout con tensión PTI/DSCR."
    else:
        if sale_guardrail_ok and (total_sell > total_keep) and (total_sell >= total_restruct):
            decision = "VENDER"
            why = "Venta domina en valor total tras considerar ahorro de capital/carry; guardrails OK."
        elif (bd_restruct.get("feasible", 0) == 1) and (total_restruct > total_keep) and (total_restruct >= total_sell):
            decision = "REESTRUCTURAR"
            why = "Reestructura domina en valor neto (post-cost) frente a workout/venta."
        else:
            decision = "MANTENER"
            why = "Workout domina o alternativas no justifican (guardrails/viabilidad)."

    # Auditoría venta vs workout (ratios)
    px_ratio_sale_vs_workout = float(v_sale / max(v_workout, 1e-9))

    return {
        "decision_micro": decision,
        "why_micro": why,

        "v_workout": float(v_workout),
        "v_sale": float(v_sale),
        "v_restruct": float(v_restruct),

        "px_ratio_sale_vs_workout": float(px_ratio_sale_vs_workout),

        "cap_release": float(cap_rel),
        "cap_savings": float(cap_sav),

        "total_sell": float(total_sell),
        "total_keep": float(total_keep),
        "uplift_restruct": float(uplift_restruct),
        "total_restruct": float(total_restruct),
        "price_ratio_book": float(bd_sale.get("price_ratio_book", np.nan)),
        "fire_sale": int(1 if fire_sale else 0),
        "fire_sale_threshold_book": float(bd_sale.get("fire_sale_threshold_book", np.nan)),
        "pnl_book": float(bd_sale.get("pnl", np.nan)),
        "book_value_source": str(bd_sale.get("book_value_source", "")),

        "sale_guardrail_ok": int(1 if sale_guardrail_ok else 0),
        "fire_sale": int(1 if fire_sale else 0),
        "price_ratio_book": float(ratio_book),
        "fire_sale_threshold_book": float(thr_book),
        "allow_fire_sale": int(1 if allow_fire_sale else 0),

        "pti": float(pti) if np.isfinite(pti) else float("nan"),
        "dscr": float(dscr) if np.isfinite(dscr) else float("nan"),

        "workout_breakdown": bd_workout,
        "sale_breakdown": bd_sale,
        "restruct_breakdown": bd_restruct,
    }


# ============================================================
# Macro constraints (cupo realista por estrategia)
# ============================================================
def macro_constraints(df: pd.DataFrame, posture: cfg.BankProfile) -> Dict[str, Any]:
    if posture == cfg.BankProfile.PRUDENTE:
        return {"max_sell_share_ead": 0.06, "max_sell_count": 15, "max_restruct_count": 35}
    if posture == cfg.BankProfile.DESINVERSION:
        return {"max_sell_share_ead": 0.18, "max_sell_count": 60, "max_restruct_count": 25}
    return {"max_sell_share_ead": 0.10, "max_sell_count": 25, "max_restruct_count": 30}


def _macro_priority_scores(df_dec: pd.DataFrame, posture: cfg.BankProfile) -> pd.DataFrame:
    """
    Macro ranking (por préstamo) para ejecutar cupos.
    Clave: normalizar por EAD para evitar sesgo por tamaño.
    """
    df_dec = df_dec.copy()

    ead = df_dec["EAD"].clip(lower=1.0)

    eva_pre = df_dec["EVA_pre"].fillna(0.0)
    cap_sav = df_dec["cap_savings"].fillna(0.0)

    # Diferencia venta vs workout (en EUR) como señal
    sale_vs_workout = (df_dec["v_sale"] - df_dec["v_workout"]).fillna(0.0)

    # Ajuste por postura
    if posture == cfg.BankProfile.DESINVERSION:
        w_cap = 1.25
        w_eva = 0.75
    elif posture == cfg.BankProfile.PRUDENTE:
        w_cap = 0.75
        w_eva = 1.00
    else:
        w_cap = 1.00
        w_eva = 0.90

    df_dec["macro_sell_score_n"] = (
        w_cap * (cap_sav / ead) +
        w_eva * (np.maximum(0.0, -eva_pre) / ead) +
        0.50 * (np.maximum(0.0, sale_vs_workout) / ead)
    )

    # Señal macro de reestructura: uplift + tensión PTI/DSCR (si existen)
    uplift = df_dec["uplift_restruct"].fillna(0.0)
    pti = df_dec["PTI"].fillna(0.0)
    dscr = df_dec["DSCR"].fillna(0.0)

    tension = (pti > 0.0).astype(float) * pti + ((dscr > 0) & (dscr < 1.10)).astype(float) * 0.50
    df_dec["macro_restruct_score_n"] = (uplift / ead) + 0.10 * tension

    return df_dec


# ============================================================
# Coordinación: micro + macro + final (por préstamo)
# ============================================================
def coordinated_inference(
    df: pd.DataFrame,
    posture: cfg.BankProfile,
    cfg_inf: CoordinatedInferenceConfig
) -> Tuple[pd.DataFrame, pd.DataFrame]:

    rows = df.to_dict(orient="records")
    out_rows: List[Dict[str, Any]] = []

    # ---------
    # 1) MICRO
    # ---------
    for r in rows:
        loan_id = str(r.get(cfg.ID_COL))

        # RW default si falta
        rw_in = _coerce_rw(r.get("RW", np.nan))
        if not np.isfinite(rw_in):
            r["RW"] = _resolve_rw_default(r)
        else:
            r["RW"] = float(1.00 if rw_in < 1.25 else 1.50)

        micro = micro_decision(r, posture=posture, horizon_months=cfg_inf.horizon_months)

        # EVA_pre: usa columna si existe; si no, calcula proxy
        eva_pre = r.get("EVA", None)
        if eva_pre is None or (isinstance(eva_pre, float) and not np.isfinite(eva_pre)):
            eva_pre = _compute_eva_proxy(r, horizon_months=cfg_inf.horizon_months)

        out = {
            cfg.ID_COL: loan_id,
            "segment": r.get("segment", r.get("segment_raw", r.get("Segment", ""))),
            "rating": r.get("rating", r.get("Rating", "")),
            "EAD": _safe_float(r.get("EAD", 0.0)),
            "PD": _safe_float(r.get("PD", 0.0)),
            "LGD": _safe_float(r.get("LGD", 0.0)),
            "DPD": _safe_float(r.get("DPD", 0.0)),
            "PTI": micro["pti"] if np.isfinite(micro["pti"]) else _safe_float(r.get("PTI", 0.0)),
            "DSCR": micro["dscr"] if np.isfinite(micro["dscr"]) else _safe_float(r.get("DSCR", 0.0)),
            "RW": _safe_float(r.get("RW", 0.0)),
            "EVA_pre": float(eva_pre),
            "RWA_pre": _safe_float(r.get("RWA", _safe_float(r.get("EAD", 0.0)) * _safe_float(r.get("RW", 1.5)))),

            "decision_micro": micro["decision_micro"],
            "why_micro": micro["why_micro"],

            "v_workout": micro["v_workout"],
            "v_sale": micro["v_sale"],
            "v_restruct": micro["v_restruct"],

            "px_ratio_sale_vs_workout": micro["px_ratio_sale_vs_workout"],

            "cap_release": micro["cap_release"],
            "cap_savings": micro["cap_savings"],

            "uplift_restruct": micro["uplift_restruct"],
            "total_sell": micro["total_sell"],
            "total_keep": micro["total_keep"],
            "total_restruct": micro["total_restruct"],

            "sale_guardrail_ok": micro["sale_guardrail_ok"],
            "fire_sale": micro["fire_sale"],
            "price_ratio_book": micro["price_ratio_book"],
            "fire_sale_threshold_book": micro["fire_sale_threshold_book"],
            "allow_fire_sale": micro["allow_fire_sale"],
        }

        if cfg_inf.verbose_explanations:
            exp: List[str] = []
            exp.append(f"[MICRO] {micro['decision_micro']}: {micro['why_micro']}")
            exp.append(
                f"workout≈{micro['v_workout']:,.0f}€, venta≈{micro['v_sale']:,.0f}€ "
                f"(sale/workout={micro['px_ratio_sale_vs_workout']:.2f})."
            )
            exp.append(f"cap_release≈{micro['cap_release']:,.0f}€, cap_savings≈{micro['cap_savings']:,.0f}€ (h={cfg_inf.horizon_months}m).")
            exp.append(
                f"fire-sale(book)={bool(micro['fire_sale'])} "
                f"(Price/Book={micro['price_ratio_book']:.2f} vs thr={micro['fire_sale_threshold_book']:.2f}, allow={bool(micro['allow_fire_sale'])})."
            )
            rb = micro["restruct_breakdown"]
            if rb.get("feasible", 0) == 1:
                exp.append(
                    f"restruct feasible: uplift≈{micro['uplift_restruct']:,.0f}€, "
                    f"PTI_post={rb.get('PTI_post', None)}, DSCR_post={rb.get('DSCR_post', None)}, cured={rb.get('cured', False)}."
                )
            out["micro_explanation"] = " ".join([x for x in exp if x])

        out_rows.append(out)

    df_dec = pd.DataFrame(out_rows)

    # -------------------------
    # 2) MACRO: ranking + cupos
    # -------------------------
    cons = macro_constraints(df_dec, posture)
    max_sell_share_ead = float(cons["max_sell_share_ead"])
    max_sell_count = int(cons["max_sell_count"])
    max_restruct_count = int(cons["max_restruct_count"])

    total_ead = float(df_dec["EAD"].sum()) if not df_dec.empty else 0.0
    sell_budget_ead = float(max_sell_share_ead * total_ead)

    df_scored = _macro_priority_scores(df_dec, posture)

    # SELL pool:
    #  - prudente/balanceado: requiere sale_guardrail_ok==1
    #  - desinversión: permite más, pero seguimos evitando EAD<=0
    if posture == cfg.BankProfile.DESINVERSION:
        sell_pool = df_scored.copy()
    else:
        sell_pool = df_scored[df_scored["sale_guardrail_ok"] == 1].copy()

    sell_pool = sell_pool[sell_pool["EAD"] > 0].copy()
    sell_pool.sort_values("macro_sell_score_n", ascending=False, inplace=True)

    sell_selected: List[str] = []
    acc_ead = 0.0
    for _, rr in sell_pool.iterrows():
        if len(sell_selected) >= max_sell_count:
            break
        ead_i = float(rr["EAD"])
        if acc_ead + ead_i > sell_budget_ead and len(sell_selected) > 0:
            continue
        sell_selected.append(str(rr[cfg.ID_COL]))
        acc_ead += ead_i

    # RESTRUCT pool: uplift positivo y (por construcción) factible (si no, uplift=0)
    restr_pool = df_scored[df_scored["uplift_restruct"] > 0].copy()
    restr_pool.sort_values("macro_restruct_score_n", ascending=False, inplace=True)
    restr_selected = [str(x) for x in restr_pool.head(max_restruct_count)[cfg.ID_COL].tolist()]

    # macro decision por préstamo
    def macro_decision_for_loan(lid: str) -> Tuple[str, str]:
        if lid in sell_selected:
            return "VENDER", "Macro: seleccionado en cupo de venta (ranking por score macro)."
        if lid in restr_selected:
            return "REESTRUCTURAR", "Macro: seleccionado en cupo de reestructura (ranking por uplift/viabilidad)."
        return "MANTENER", "Macro: no seleccionado por cupos; se mantiene en workout."

    df_scored["decision_macro"] = ""
    df_scored["why_macro"] = ""
    for i in range(len(df_scored)):
        lid = str(df_scored.iloc[i][cfg.ID_COL])
        d, w = macro_decision_for_loan(lid)
        df_scored.at[df_scored.index[i], "decision_macro"] = d
        df_scored.at[df_scored.index[i], "why_macro"] = w

    # ---------------------------------
    # 3) DECISIÓN FINAL (coordinación)
    # ---------------------------------
    final_decisions: List[str] = []
    final_rationales: List[str] = []

    tol_rel = 0.05  # 5% del workout como tolerancia de dominancia

    for _, rr in df_scored.iterrows():
        micro_d = str(rr["decision_micro"])
        macro_d = str(rr["decision_macro"])

        v_workout = float(rr["v_workout"])
        total_sell = float(rr["total_sell"])
        total_keep = float(rr["total_keep"])
        total_restruct = float(rr["total_restruct"])
        uplift = float(rr["uplift_restruct"])
        sale_ok = int(rr.get("sale_guardrail_ok", 0)) == 1

        final_d = micro_d
        rationale_parts = [
            f"[MICRO] {micro_d}: {rr.get('why_micro','')}",
            f"[MACRO] {macro_d}: {rr.get('why_macro','')}",
        ]

        if micro_d == macro_d:
            final_d = micro_d
            rationale_parts.append("Coherencia micro↔macro: misma acción en ambos niveles.")
        else:
            # Macro puede diferir por cupo (pacing)
            if macro_d == "MANTENER" and micro_d in ("VENDER", "REESTRUCTURAR"):
                final_d = "MANTENER"
                rationale_parts.append("Coordinación: macro difiere la acción por cupos/ritmo; se mantiene en workout.")
            elif macro_d == "VENDER" and micro_d != "VENDER":
                # Si guardrail no ok (y no desinversión), no vendemos
                if (not sale_ok) and posture != cfg.BankProfile.DESINVERSION:
                    final_d = "MANTENER"
                    rationale_parts.append("Coordinación: venta bloqueada por guardrail (fire-sale vs book / política).")
                else:
                    threshold = v_workout * tol_rel
                    if (total_sell - total_keep) > threshold:
                        final_d = "VENDER"
                        rationale_parts.append(
                            f"Coordinación: macro override por dominancia económica (Δ={total_sell-total_keep:,.0f}€ > {threshold:,.0f}€) y objetivo capital."
                        )
                    else:
                        final_d = "MANTENER"
                        rationale_parts.append("Coordinación: ventaja de venta insuficiente; se mantiene workout.")
            elif macro_d == "REESTRUCTURAR" and micro_d != "REESTRUCTURAR":
                threshold = max(0.0, v_workout * tol_rel)
                if uplift > threshold:
                    final_d = "REESTRUCTURAR"
                    rationale_parts.append(
                        f"Coordinación: macro prioriza cure/quality con uplift suficiente (uplift={uplift:,.0f}€ > {threshold:,.0f}€)."
                    )
                else:
                    final_d = "MANTENER"
                    rationale_parts.append("Coordinación: uplift insuficiente para justificar reestructura; se mantiene.")
            else:
                # Dominancia por score total
                best_alt = max(
                    [("VENDER", total_sell), ("REESTRUCTURAR", total_restruct), ("MANTENER", total_keep)],
                    key=lambda x: x[1]
                )
                final_d = best_alt[0]
                rationale_parts.append(
                    f"Coordinación: decisión por dominancia de valor total "
                    f"(SELL={total_sell:,.0f}€, RESTR={total_restruct:,.0f}€, KEEP={total_keep:,.0f}€) → {final_d}."
                )

        final_decisions.append(final_d)
        final_rationales.append(" ".join([p for p in rationale_parts if p]))

    df_scored["decision_final"] = final_decisions
    df_scored["final_rationale"] = final_rationales

    # Resumen portfolio
    summary = {
        "bank_profile": [posture.value],
        "n_loans": [len(df_scored)],
        "EAD_total": [float(df_scored["EAD"].sum())],
        "sell_budget_share_ead": [max_sell_share_ead],
        "sell_budget_ead": [float(sell_budget_ead)],
        "sell_selected_count": [int(len(sell_selected))],
        "sell_selected_ead": [float(acc_ead)],
        "restruct_selected_count": [int(len(restr_selected))],
        "decision_micro_sell_pct": [float((df_scored["decision_micro"] == "VENDER").mean())],
        "decision_macro_sell_pct": [float((df_scored["decision_macro"] == "VENDER").mean())],
        "decision_final_sell_pct": [float((df_scored["decision_final"] == "VENDER").mean())],
        "decision_final_restruct_pct": [float((df_scored["decision_final"] == "REESTRUCTURAR").mean())],
        "decision_final_keep_pct": [float((df_scored["decision_final"] == "MANTENER").mean())],
        "cap_release_selected_sum": [float(df_scored[df_scored["decision_macro"] == "VENDER"]["cap_release"].sum())],
        "fire_sale_final_count": [int(df_scored[(df_scored["decision_final"] == "VENDER") & (df_scored["fire_sale"] == 1)].shape[0])],
    }
    df_sum = pd.DataFrame(summary)

    logger.info(
        f"[{posture.value}] macro cupos: sell_share={max_sell_share_ead:.2%}, "
        f"sell_count={len(sell_selected)}, sell_ead={acc_ead:,.0f}, restruct_count={len(restr_selected)}"
    )

    return df_scored, df_sum


# ============================================================
# Export Excel (2 hojas) — sin dependencias externas
# ============================================================
def export_excel_two_sheets(df_dec: pd.DataFrame, df_sum: pd.DataFrame, out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Loan_Decisions"
    ws2 = wb.create_sheet("Portfolio_Summary")

    THIN = Side(border_style="thin", color="D9D9D9")
    BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    HEADER_FILL = PatternFill("solid", fgColor="244062")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center")

    # ---- Sheet 1
    for row in dataframe_to_rows(df_dec, index=False, header=True):
        ws1.append(row)

    for c in ws1[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    hdr1 = {ws1.cell(1, j).value: j for j in range(1, ws1.max_column + 1)}

    eur_cols = {
        "EAD", "EVA_pre", "RWA_pre", "v_workout", "v_sale", "v_restruct",
        "cap_release", "cap_savings", "uplift_restruct", "total_sell", "total_keep", "total_restruct"
    }
    pct_cols = {"PD", "LGD", "PTI"}  # RW NO es %, es multiplicador
    ratio_cols = {"px_ratio_sale_vs_workout", "price_ratio_book", "fire_sale_threshold_book"}

    for i in range(2, ws1.max_row + 1):
        for j in range(1, ws1.max_column + 1):
            cell = ws1.cell(i, j)
            cell.border = BORDER
            name = ws1.cell(1, j).value

            if name in eur_cols:
                cell.number_format = '#,##0" €"'
                cell.alignment = RIGHT
            elif name in pct_cols:
                cell.number_format = "0.00%"
                cell.alignment = RIGHT
            elif name in ratio_cols:
                cell.number_format = "0.00"
                cell.alignment = RIGHT
            elif name in ("RW",):
                cell.number_format = "0.00"
                cell.alignment = RIGHT
            elif name in ("DPD", "fire_sale", "sale_guardrail_ok", "allow_fire_sale"):
                cell.number_format = "0"
                cell.alignment = RIGHT
            elif name in ("DSCR",):
                cell.number_format = "0.00"
                cell.alignment = RIGHT
            elif name in ("final_rationale", "micro_explanation"):
                cell.alignment = WRAP
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    # autosize (cap a 55)
    for col in ws1.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[: min(len(col), 500)]:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws1.column_dimensions[col_letter].width = min(max_len + 2, 55)

    ws1.freeze_panes = "A2"

    # ---- Sheet 2
    for row in dataframe_to_rows(df_sum, index=False, header=True):
        ws2.append(row)

    for c in ws2[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    for i in range(2, ws2.max_row + 1):
        for j in range(1, ws2.max_column + 1):
            cell = ws2.cell(i, j)
            cell.border = BORDER
            cell.alignment = RIGHT

    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(out_path)


# ============================================================
# Runner (single o multi-postura)
# ============================================================
def _posture_from_str(s: str) -> cfg.BankProfile:
    s = (s or "").strip().lower()
    if s in ("prudente", "prudencial", "prudent"):
        return cfg.BankProfile.PRUDENTE
    if s in ("desinversion", "desinversión", "selldown", "disposal"):
        return cfg.BankProfile.DESINVERSION
    return cfg.BankProfile.BALANCEADO


def _run_one_posture(df: pd.DataFrame, posture: cfg.BankProfile, cfg_inf: CoordinatedInferenceConfig, out_dir: str) -> Tuple[str, str]:
    df_dec, df_sum = coordinated_inference(df, posture=posture, cfg_inf=cfg_inf)

    decisions_csv = os.path.join(out_dir, f"loan_decisions_{posture.value}.csv")
    summary_csv = os.path.join(out_dir, f"portfolio_summary_{posture.value}.csv")
    df_dec.to_csv(decisions_csv, index=False, encoding="utf-8-sig")
    df_sum.to_csv(summary_csv, index=False, encoding="utf-8-sig")

    excel_path = ""
    if cfg_inf.export_excel:
        excel_path = os.path.join(out_dir, f"decisiones_{posture.value}.xlsx")
        export_excel_two_sheets(df_dec, df_sum, excel_path)

    logger.info(f"✅ [{posture.value}] CSV decisions: {decisions_csv}")
    logger.info(f"✅ [{posture.value}] CSV summary:   {summary_csv}")
    if excel_path:
        logger.info(f"✅ [{posture.value}] Excel:         {excel_path}")

    return excel_path, decisions_csv


def run(cfg_inf: CoordinatedInferenceConfig) -> str:
    ts = _now_tag()
    out_dir = os.path.join(REPORTS_DIR, f"coord_inference_{ts}_{cfg_inf.tag}")
    os.makedirs(out_dir, exist_ok=True)

    df = _load_portfolio_df(cfg_inf.portfolio_path)

    if cfg_inf.multi_posture:
        postures = [cfg.BankProfile.PRUDENTE, cfg.BankProfile.BALANCEADO, cfg.BankProfile.DESINVERSION]
    else:
        if cfg_inf.risk_posture:
            postures = [_posture_from_str(cfg_inf.risk_posture)]
        else:
            postures = [CFG.bank_profile]

    for p in postures:
        _run_one_posture(df, p, cfg_inf, out_dir)

    return out_dir


def parse_args() -> CoordinatedInferenceConfig:
    p = argparse.ArgumentParser(description="Inferencia coordinada MICRO+MACRO (1 decisión final por préstamo).")
    p.add_argument("--portfolio", type=str, required=True, help="Ruta a cartera (xlsx/csv).")
    p.add_argument("--tag", type=str, default="coordinated")
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--horizon-months", type=int, default=24, dest="horizon_months")
    p.add_argument("--no-verbose", action="store_true", help="No incluir explicación micro extendida.")
    p.add_argument("--no-excel", action="store_true", help="No exportar Excel (solo CSV).")
    p.add_argument("--single", action="store_true", help="Solo una postura (usa config o --risk-posture).")
    p.add_argument("--risk-posture", type=str, default=None, help="prudente|balanceado|desinversion (si --single).")

    args = p.parse_args()
    return CoordinatedInferenceConfig(
        portfolio_path=args.portfolio,
        tag=args.tag,
        seed=args.seed,
        horizon_months=args.horizon_months,
        verbose_explanations=(not args.no_verbose),
        export_excel=(not args.no_excel),
        multi_posture=(not args.single),
        risk_posture=args.risk_posture,
    )


if __name__ == "__main__":
    cfg_inf = parse_args()
    out_dir = run(cfg_inf)
    logger.info(f"🏁 Coordinated inference completada. Reporte en: {out_dir}")
