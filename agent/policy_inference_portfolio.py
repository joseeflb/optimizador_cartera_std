# ============================================
# agent/policy_inference_portfolio.py
# — Inferencia PPO a nivel de CARTERA COMPLETA (PortfolioEnv)
# (CORREGIDO v3.7.0 · VecNormalize robusto por shape + decisiones audit-ready)
# ============================================

from __future__ import annotations

import os
import sys
import time
import argparse
import logging
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import dataclasses
import importlib
import numpy as np
import pandas as pd

# -----------------------------------------------------------
# 🔧 Rutas y logging
# -----------------------------------------------------------
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.append(ROOT_DIR)

LOG_DIR = os.path.join(ROOT_DIR, "logs")
MODELS_DIR = os.path.join(ROOT_DIR, "models")
REPORTS_DIR = os.path.join(ROOT_DIR, "reports")
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(MODELS_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOG_DIR, "policy_inference_portfolio.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("policy_inference_portfolio")

# -----------------------------------------------------------
# ⚙️ Config + RL
# -----------------------------------------------------------
import config as cfg

try:
    from stable_baselines3 import PPO
    from stable_baselines3.common.vec_env import DummyVecEnv, VecNormalize
except ImportError as e:
    raise SystemExit("❌ Faltan dependencias RL. Ejecuta install_requirements_smart.py.") from e

CFG = cfg.CONFIG


# -----------------------------------------------------------
# 📊 Exportador específico para cartera final
# -----------------------------------------------------------
def export_styled_excel_portfolio(df: pd.DataFrame, out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera_final"

    THIN = Side(border_style="thin", color="999999")
    HEADER_FILL = PatternFill("solid", fgColor="244062")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

    FILL_ACTIVE = PatternFill("solid", fgColor="E7F5E9")
    FILL_CLOSED = PatternFill("solid", fgColor="F2F2F2")

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER

    hdr = {ws.cell(1, j).value: j for j in range(1, ws.max_column + 1)}
    col_closed = hdr.get("closed", None)
    col_ead = hdr.get("EAD", None)
    col_rwa = hdr.get("RWA", None)
    col_pti = hdr.get("PTI", None)
    col_dscr = hdr.get("DSCR", None)

    for i in range(2, ws.max_row + 1):
        is_closed = False
        if col_closed:
            is_closed = bool(ws.cell(i, col_closed).value)

        for j in range(1, ws.max_column + 1):
            cell = ws.cell(i, j)
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            cell.fill = FILL_CLOSED if is_closed else FILL_ACTIVE

    if col_ead:
        for i in range(2, ws.max_row + 1):
            ws.cell(i, col_ead).number_format = '#,##0" €"'
            ws.cell(i, col_ead).alignment = RIGHT
    if col_rwa:
        for i in range(2, ws.max_row + 1):
            ws.cell(i, col_rwa).number_format = '#,##0" €"'
            ws.cell(i, col_rwa).alignment = RIGHT
    if col_pti:
        for i in range(2, ws.max_row + 1):
            ws.cell(i, col_pti).number_format = "0.00%"
            ws.cell(i, col_pti).alignment = RIGHT
    if col_dscr:
        for i in range(2, ws.max_row + 1):
            ws.cell(i, col_dscr).number_format = "0.00"
            ws.cell(i, col_dscr).alignment = RIGHT

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)

    wb.save(out_path)


# -----------------------------------------------------------
# 🧾 Config de inferencia macro
# -----------------------------------------------------------
@dataclass
class PortfolioInferenceConfig:
    model_path: str
    portfolio_path: str
    device: str = "auto"
    seed: int = 42
    deterministic: bool = True
    tag: str = "portfolio_policy"
    n_steps: int = 1
    top_k: int = 5
    scenario: str = "baseline"  # ✅ baseline | adverse | severe

    save_final_excel: bool = True
    risk_posture: str = "balanceado"

    loan_model_path: Optional[str] = None
    vecnormalize_portfolio_path: Optional[str] = None
    vecnormalize_loan_path: Optional[str] = None

    # ✅ no crear carpetas ni exports intermedios
    persist_outputs: bool = True
    output_dir: Optional[str] = None
    final_output_path: Optional[str] = None


# -----------------------------------------------------------
# 🔧 Utilidades
# -----------------------------------------------------------
_ACTION_MAP: Dict[int, str] = {
    0: "MANTENER todos",
    1: "VENDER top-1 EVA más negativa",
    2: "VENDER top-K EVA más negativa",
    3: "REESTRUCTURAR top-1 EVA más negativa",
    4: "REESTRUCTURAR top-K EVA más negativa",
    5: "VENDER top-1 RORWA más bajo",
    6: "VENDER top-K RORWA más bajo",
    7: "REESTRUCTURAR top-1 PTI más alto",
    8: "REESTRUCTURAR top-K PTI más alto",
    9: "MIX: vender top-1 EVA<0 & reestructurar top-1 PTI alto",
    10: "HEURÍSTICA baseline (regla financiera)",
    11: "NO-OP (mantener)",
}

MIX_FAMILY = {9, 10}
SELL_FAMILY = {1, 2, 5, 6}
RESTR_FAMILY = {3, 4, 7, 8}
KEEP_FAMILY = {0, 11}


def _family(action_id: int) -> str:
    if action_id in MIX_FAMILY:
        return "MIX"
    if action_id in SELL_FAMILY:
        return "SELL"
    if action_id in RESTR_FAMILY:
        return "RESTRUCT"
    return "KEEP"


def _now_tag() -> str:
    return time.strftime("%Y%m%d_%H%M%S")


def _device_auto(device: str) -> str:
    if device != "auto":
        return device
    try:
        import torch  # type: ignore
        return "cuda" if torch.cuda.is_available() else "cpu"
    except Exception:
        return "cpu"

def harmonize_portfolio_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    # --- map monthly_* -> legacy names usados por optimizadores
    if "ingreso_mensual" not in df.columns and "monthly_income" in df.columns:
        df["ingreso_mensual"] = df["monthly_income"]
    if "cashflow_operativo_mensual" not in df.columns and "monthly_cfo" in df.columns:
        df["cashflow_operativo_mensual"] = df["monthly_cfo"]
    if "cuota_mensual" not in df.columns and "monthly_payment" in df.columns:
        df["cuota_mensual"] = df["monthly_payment"]

    # --- PTI / DSCR: preferimos pre si existe; si no, derivamos desde monthly_*
    if "PTI" not in df.columns:
        if "PTI_pre" in df.columns:
            df["PTI"] = df["PTI_pre"]
        elif ("monthly_payment" in df.columns) and ("monthly_income" in df.columns):
            df["PTI"] = df["monthly_payment"] / df["monthly_income"].replace(0, np.nan)

    if "DSCR" not in df.columns:
        if "DSCR_pre" in df.columns:
            df["DSCR"] = df["DSCR_pre"]
        elif ("monthly_cfo" in df.columns) and ("monthly_payment" in df.columns):
            df["DSCR"] = df["monthly_cfo"] / df["monthly_payment"].replace(0, np.nan)

    # --- book_value: si no viene, aproximación consistente
    if "book_value" not in df.columns:
        if "coverage_rate" in df.columns and "EAD" in df.columns:
            df["book_value"] = df["EAD"] * (1.0 - df["coverage_rate"].clip(0, 1))
        elif "LGD" in df.columns and "EAD" in df.columns:
            df["book_value"] = df["EAD"] * (1.0 - df["LGD"].clip(0, 1))

    # --- normalización mínima de segment (para gates retail vs corp)
    if "segment" in df.columns:
        df["segment"] = df["segment"].astype(str).str.strip()

    return df

def _load_portfolio_df(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"❌ Cartera no encontrada: {path}")
    ext = os.path.splitext(path.lower())[1]
    df = pd.read_excel(path) if ext in (".xlsx", ".xls") else pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # ✅ Harmonize: crea/normaliza PTI, DSCR, ingreso_mensual, cuota_mensual, book_value, etc.
    df = harmonize_portfolio_schema(df)

    # (opcional pero recomendado) sanity log rápido
    try:
        miss_pti = int(df["PTI"].isna().sum()) if "PTI" in df.columns else len(df)
        miss_dscr = int(df["DSCR"].isna().sum()) if "DSCR" in df.columns else len(df)
        logger.info(
            f"📥 Cartera armonizada ({len(df):,} préstamos) | missing PTI={miss_pti} | missing DSCR={miss_dscr}"
        )
    except Exception:
        logger.info(f"📥 Cartera armonizada ({len(df):,} préstamos)")

    return df



def _pick_default_macro_model_path() -> Optional[str]:
    cands = [
        os.path.join(MODELS_DIR, "best_model_portfolio.zip"),
        os.path.join(MODELS_DIR, "best_model_macro.zip"),
        os.path.join(MODELS_DIR, "best_model.zip"),  # legacy fallback
    ]
    for p in cands:
        if os.path.exists(p):
            return p
    return None


def _pick_default_micro_model_path() -> Optional[str]:
    cands = [
        os.path.join(MODELS_DIR, "best_model_loan.zip"),
        os.path.join(MODELS_DIR, "best_model_micro.zip"),
        os.path.join(MODELS_DIR, "best_model.zip"),
    ]
    for p in cands:
        if os.path.exists(p):
            return p
    return None


def _pick_default_vn_portfolio_path() -> Optional[str]:
    """
    ✅ Ampliado: aceptamos bundles estándar y legacy, y validamos por SHAPE.
    """
    cands = [
        os.path.join(MODELS_DIR, "vecnormalize_portfolio.pkl"),
        os.path.join(MODELS_DIR, "vecnormalize_macro.pkl"),
        os.path.join(MODELS_DIR, "best_model_vecnormalize.pkl"),
        os.path.join(MODELS_DIR, "vecnormalize_final.pkl"),
    ]
    for p in cands:
        if os.path.exists(p):
            return p
    return None


def _pick_default_vn_loan_path() -> Optional[str]:
    cands = [
        os.path.join(MODELS_DIR, "vecnormalize_loan.pkl"),
        os.path.join(MODELS_DIR, "vecnormalize_micro.pkl"),
        os.path.join(MODELS_DIR, "best_model_vecnormalize.pkl"),
    ]
    for p in cands:
        if os.path.exists(p):
            return p

    legacy = os.path.join(MODELS_DIR, "vecnormalize_final.pkl")
    if os.path.exists(legacy):
        logger.warning(
            "⚠️ Usando VN legacy para LOAN: models/vecnormalize_final.pkl. "
            "Recomendado: renombrar/guardar como vecnormalize_loan.pkl."
        )
        return legacy

    return None


def _vn_name_looks_like(label: str, path: str) -> bool:
    b = os.path.basename(path).lower()
    if label.upper() == "PORTFOLIO":
        # nombre neutral o explícito macro/portfolio es ok
        if ("loan" in b) or ("micro" in b):
            return False
        return True

    # LOAN
    if ("portfolio" in b) or ("macro" in b):
        return False
    return True


def _vn_shape_matches_env(vn: VecNormalize, dummy_env: DummyVecEnv) -> bool:
    try:
        env_shape = getattr(dummy_env.observation_space, "shape", None)
        vn_shape = getattr(getattr(vn, "obs_rms", None), "mean", None)
        vn_shape = getattr(vn_shape, "shape", None)
        if env_shape is None or vn_shape is None:
            return True
        return tuple(env_shape) == tuple(vn_shape)
    except Exception:
        return True


def _load_vecnormalize(vn_path: Optional[str], dummy_env: DummyVecEnv, label: str) -> Optional[VecNormalize]:
    if not vn_path:
        return None
    if not os.path.exists(vn_path):
        logger.warning(f"⚠️ VecNormalize {label} no existe: {vn_path}")
        return None

    if not _vn_name_looks_like(label, vn_path):
        logger.warning(
            f"⚠️ VecNormalize {label} sospechoso por nombre: {vn_path}. "
            f"Se intentará cargar pero se validará shape."
        )
    try:
        vn = VecNormalize.load(vn_path, dummy_env)
        vn.training = False
        vn.norm_reward = False

        if not _vn_shape_matches_env(vn, dummy_env):
            env_shape = getattr(dummy_env.observation_space, "shape", None)
            vn_shape = getattr(getattr(vn, "obs_rms", None), "mean", None)
            vn_shape = getattr(vn_shape, "shape", None)
            logger.warning(
                f"⚠️ VecNormalize {label} INVALIDADO por mismatch de shape: env={env_shape} vs vn={vn_shape} | {vn_path}"
            )
            return None

        logger.info(f"🔄 VecNormalize {label} cargado: {vn_path}")
        return vn
    except Exception as e:
        logger.warning(f"⚠️ VecNormalize {label} incompatible (shape mismatch u otro): {vn_path} | {e}")
        return None


class PolicyAdapter:
    """
    Adaptador para aplicar VecNormalize (si existe) antes de model.predict().
    """

    def __init__(self, model: PPO, vecnorm: Optional[VecNormalize], deterministic_default: bool = True):
        self.model = model
        self.vecnorm = vecnorm
        self.det_default = bool(deterministic_default)

    def predict(self, obs: Any, deterministic: Optional[bool] = None):
        det = self.det_default if deterministic is None else bool(deterministic)

        arr = np.asarray(obs, dtype=np.float32)
        if arr.ndim == 1:
            arr = arr.reshape(1, -1)

        arr = np.nan_to_num(arr, nan=0.0, posinf=0.0, neginf=0.0)

        if self.vecnorm is not None:
            arr = self.vecnorm.normalize_obs(arr)

        return self.model.predict(arr, deterministic=det)


def _load_ppo_model(path: str, device: str, label: str) -> PPO:
    if not os.path.exists(path):
        raise FileNotFoundError(f"❌ Modelo {label} no encontrado: {path}")
    device_final = _device_auto(device)
    logger.info(f"🤖 Cargando PPO {label}: {path} [device={device_final}]")
    return PPO.load(path, device=device_final)


def _select_profile_strategy_reward(posture: str):
    posture = (posture or "balanceado").lower()
    if posture == "prudencial":
        prof = cfg.BankProfile.PRUDENTE
    elif posture == "desinversion":
        prof = cfg.BankProfile.DESINVERSION
    else:
        prof = cfg.BankProfile.BALANCEADO
    strat = cfg.BANK_STRATEGIES[prof]
    reward = cfg.RewardParams.from_strategy(strat)
    return prof, strat, reward


# -----------------------------------------------------------
# 🧠 CAPA FINANCIERA MACRO (Banco L1.5)
# -----------------------------------------------------------
def _macro_thresholds_from_strategy(strat: Any, posture: str) -> Dict[str, float]:
    posture = (posture or "balanceado").lower()

    eva_strong_neg = float(getattr(strat, "eva_strongly_neg", -50_000.0)) * 4.0
    eva_min_impr = float(getattr(strat, "eva_min_improvement", 10_000.0)) * 5.0

    eva_pos_strong = 3.0 * abs(eva_min_impr)
    ambig_band = abs(eva_min_impr)

    base_conc = 0.08
    base_vol = 0.04
    w_conc = float(getattr(strat, "w_concentration", base_conc))
    w_vol = float(getattr(strat, "w_vol", base_vol))

    scale_conc = float(np.clip(base_conc / max(w_conc, 1e-6), 0.6, 1.4))
    scale_vol = float(np.clip(base_vol / max(w_vol, 1e-6), 0.6, 1.6))

    hhi_high = 0.35 * scale_conc
    hhi_extreme = 0.45 * scale_conc
    eva_vol_hi = 150_000.0 * scale_vol

    if posture == "prudencial":
        ambig_band *= 0.7
        hhi_high *= 0.9
        eva_vol_hi *= 0.9
    elif posture == "desinversion":
        ambig_band *= 1.2
        hhi_high *= 1.1
        eva_vol_hi *= 1.1

    return {
        "EVA_STRONGLY_NEG": eva_strong_neg,
        "EVA_POS_STRONG": eva_pos_strong,
        "AMBIG_BAND": ambig_band,
        "HHI_HIGH": hhi_high,
        "HHI_EXTREME": hhi_extreme,
        "EVA_VOL_HIGH": eva_vol_hi,
    }


def _posture_guardrails(posture: str) -> Dict[str, Any]:
    posture = (posture or "balanceado").lower()
    if posture == "prudencial":
        return {"max_steps_sell": 0.50, "allow_sell_in_ambig": False, "allow_rl_family_flip": False}
    if posture == "desinversion":
        return {"max_steps_sell": 1.00, "allow_sell_in_ambig": True, "allow_rl_family_flip": True}
    return {"max_steps_sell": 0.75, "allow_sell_in_ambig": True, "allow_rl_family_flip": False}


def _build_macro_state_from_reset(info_reset: Dict[str, Any]) -> Dict[str, float]:
    pm = info_reset.get("portfolio_metrics", {}) or {}
    return {
        "EVA": float(pm.get("EVA_total", 0.0) or 0.0),
        "RWA": float(pm.get("RWA_total", 0.0) or 0.0),
        "capital_liberado": 0.0,
        "num_loans_active": float(pm.get("num_loans_active", 0) or 0),
        "num_loans_closed": float(pm.get("num_loans_closed", 0) or 0),
        "eva_vol": float(pm.get("eva_volatility", 0.0) or 0.0),
        "hhi_seg": float(pm.get("hhi_segment", 0.0) or 0.0),
        "hhi_rat": float(pm.get("hhi_rating", 0.0) or 0.0),
    }


def _build_macro_state_from_step(pm: Dict[str, Any]) -> Dict[str, float]:
    return {
        "EVA": float(pm.get("EVA_after", pm.get("EVA_before", 0.0)) or 0.0),
        "RWA": float(pm.get("RWA_after", pm.get("RWA_before", 0.0)) or 0.0),
        "capital_liberado": float(pm.get("capital_liberado", 0.0) or 0.0),
        "num_loans_active": float(pm.get("num_loans_active", 0) or 0),
        "num_loans_closed": float(pm.get("num_loans_closed", 0) or 0),
        "eva_vol": float(pm.get("eva_volatility", 0.0) or 0.0),
        "hhi_seg": float(pm.get("hhi_segment", 0.0) or 0.0),
        "hhi_rat": float(pm.get("hhi_rating", 0.0) or 0.0),
    }


def _macro_metrics(macro_state: Dict[str, float], hurdle: float) -> Dict[str, float]:
    eva = float(macro_state.get("EVA", 0.0))
    rwa = float(macro_state.get("RWA", 0.0))
    eva_per_rwa = eva / rwa if rwa > 0 else 0.0
    est_rorwa = eva_per_rwa + float(hurdle)
    return {"eva": eva, "rwa": rwa, "eva_per_rwa": eva_per_rwa, "est_rorwa": est_rorwa}


def _macro_financial_action(
    macro_state: Dict[str, float],
    thresholds: Dict[str, float],
    posture: str,
    hurdle: float,
) -> Tuple[Optional[int], str, str]:
    posture = (posture or "balanceado").lower()
    m = _macro_metrics(macro_state, hurdle=hurdle)
    EVA, RWA = m["eva"], m["rwa"]
    eva_per_rwa = m["eva_per_rwa"]
    est_rorwa = m["est_rorwa"]

    eva_vol = float(macro_state.get("eva_vol", 0.0))
    hhi_seg = float(macro_state.get("hhi_seg", 0.0))
    hhi_rat = float(macro_state.get("hhi_rat", 0.0))
    n_active = int(macro_state.get("num_loans_active", 0))

    EVA_STRONGLY_NEG = float(thresholds["EVA_STRONGLY_NEG"])
    EVA_POS_STRONG = float(thresholds["EVA_POS_STRONG"])
    AMBIG_BAND = float(thresholds["AMBIG_BAND"])
    HHI_HIGH = float(thresholds["HHI_HIGH"])
    EVA_VOL_HIGH = float(thresholds["EVA_VOL_HIGH"])

    rationale: List[str] = []
    rationale.append(
        f"EVA={EVA:,.0f}€, RWA={RWA:,.0f}, EVA/RWA={eva_per_rwa:.4f}, "
        f"RORWA_est≈{est_rorwa:.2%} (hurdle={hurdle:.2%}), "
        f"HHI_seg={hhi_seg:.3f}, HHI_rat={hhi_rat:.3f}, EVA_vol≈{eva_vol:,.0f}, n_active={n_active}."
    )

    if n_active <= 0:
        return 11, "KEEP", " ".join(rationale + ["No hay préstamos activos → NO-OP."])

    if EVA <= EVA_STRONGLY_NEG or (RWA > 0 and est_rorwa < hurdle - 0.004):
        if posture == "prudencial":
            rationale.append("Cartera claramente destructiva → MIX/RESTRUCT antes de vender agresivo.")
            return 9, "MIX", " ".join(rationale)
        rationale.append("Cartera claramente destructiva → baseline (10).")
        return 10, "MIX", " ".join(rationale)

    if EVA >= EVA_POS_STRONG and est_rorwa >= hurdle + 0.003:
        if hhi_seg < HHI_HIGH and hhi_rat < HHI_HIGH and eva_vol < EVA_VOL_HIGH:
            rationale.append("Cartera crea valor + diversificación + volatilidad contenida → KEEP.")
            return 0, "KEEP", " ".join(rationale)
        rationale.append("Cartera buena pero con concentración/volatilidad → MIX táctico.")
        return 9, "MIX", " ".join(rationale)

    if -AMBIG_BAND <= EVA <= AMBIG_BAND:
        rationale.append("EVA en banda ambigua → PPO decide táctica (con guardrails).")
        return None, "AMBIG", " ".join(rationale)

    if EVA < 0:
        if posture == "prudencial":
            rationale.append("EVA negativo moderado → preferencia RESTRUCT (4).")
            return 4, "RESTRUCT", " ".join(rationale)
        rationale.append("EVA negativo moderado → baseline (10).")
        return 10, "MIX", " ".join(rationale)

    rationale.append("EVA positivo moderado → PPO refina (KEEP/MIX).")
    return None, "AMBIG", " ".join(rationale)


# -----------------------------------------------------------
# 🚀 Núcleo (con modo no-persist)
# -----------------------------------------------------------
def _resolve_output_dir(cfg_inf: PortfolioInferenceConfig, ts: str) -> str:
    if cfg_inf.output_dir and str(cfg_inf.output_dir).strip():
        return str(cfg_inf.output_dir)
    return os.path.join(REPORTS_DIR, f"inference_portfolio_{ts}_{cfg_inf.tag}")


def _resolve_final_path(cfg_inf: PortfolioInferenceConfig, out_dir: str, default_name: str) -> str:
    if cfg_inf.final_output_path and str(cfg_inf.final_output_path).strip():
        return str(cfg_inf.final_output_path)
    return os.path.join(out_dir, default_name)


def _run_portfolio_inference_for_posture(
    cfg_inf: PortfolioInferenceConfig,
    out_dir: Optional[str],
    suffix: Optional[str] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, str, str, str]:
    ts = _now_tag()
    posture = (cfg_inf.risk_posture or "balanceado").lower()
    guards = _posture_guardrails(posture)

    prof, strat, _reward = _select_profile_strategy_reward(posture)

    # ✅ Intentar fijar perfil antes de cargar/reload del env
    if hasattr(cfg, "set_bank_profile"):
        try:
            cfg.set_bank_profile(prof)
        except Exception as e:
            logger.warning(f"⚠️ No se pudo fijar bank_profile en config: {e}")

    # ✅ Mitigación: reload del módulo env.portfolio_env para refrescar aliases a nivel de módulo
    # (si en tu PortfolioEnv todavía usa BANK_PROFILE/BANK_STRAT como constantes import-time)
    try:
        import env.portfolio_env as portfolio_env_mod  # type: ignore
        portfolio_env_mod = importlib.reload(portfolio_env_mod)
        PortfolioEnv = portfolio_env_mod.PortfolioEnv
    except Exception as e:
        raise SystemExit(f"❌ No se pudo importar/reload PortfolioEnv: {e}")

    try:
        import env.loan_env as loan_env_mod  # type: ignore
        loan_env_mod = importlib.reload(loan_env_mod)
        LoanEnv = loan_env_mod.LoanEnv
    except Exception as e:
        raise SystemExit(f"❌ No se pudo importar/reload LoanEnv: {e}")

    reg_cfg = cfg.CONFIG.regulacion
    hurdle = float(getattr(reg_cfg, "hurdle_rate", 0.0))
    thresholds = _macro_thresholds_from_strategy(strat, posture=posture)

    logger.info(
        f"[MACRO] posture={posture} | profile={prof.value} | strategy={strat.name} | "
        f"EVA_STRONGLY_NEG={thresholds['EVA_STRONGLY_NEG']:,.0f} | "
        f"EVA_POS_STRONG={thresholds['EVA_POS_STRONG']:,.0f} | AMBIG_BAND={thresholds['AMBIG_BAND']:,.0f} | "
        f"scenario={cfg_inf.scenario} | persist_outputs={bool(cfg_inf.persist_outputs)}"
    )

    df = _load_portfolio_df(cfg_inf.portfolio_path)

    # modelos
    model_macro = _load_ppo_model(cfg_inf.model_path, cfg_inf.device, label="MACRO")

    loan_path = cfg_inf.loan_model_path or _pick_default_micro_model_path()
    model_micro: Optional[PPO] = None
    if loan_path and os.path.exists(loan_path):
        model_micro = _load_ppo_model(loan_path, cfg_inf.device, label="MICRO")
    else:
        logger.warning("⚠️ Modelo micro no encontrado → PortfolioEnv sin re-ranking micro.")

    # VecNormalize (con defaults seguros + validación shape)
    vn_port_path = cfg_inf.vecnormalize_portfolio_path or _pick_default_vn_portfolio_path()
    vn_loan_path = cfg_inf.vecnormalize_loan_path or _pick_default_vn_loan_path()

    dummy_port = DummyVecEnv([lambda: PortfolioEnv(loans_df=df, top_k=cfg_inf.top_k, scenario=cfg_inf.scenario, ppo_micro=None)])
    vn_port = _load_vecnormalize(vn_port_path, dummy_port, label="PORTFOLIO")

    vn_micro = None
    if model_micro is not None:
        dummy_loan = DummyVecEnv([lambda: LoanEnv()])
        vn_micro = _load_vecnormalize(vn_loan_path, dummy_loan, label="LOAN")

    macro_policy = PolicyAdapter(model_macro, vn_port, deterministic_default=cfg_inf.deterministic)
    micro_policy = PolicyAdapter(model_micro, vn_micro, deterministic_default=True) if model_micro is not None else None

    # env real
    env = PortfolioEnv(loans_df=df, top_k=cfg_inf.top_k, scenario=cfg_inf.scenario, ppo_micro=micro_policy)
    obs, info_reset = env.reset(seed=cfg_inf.seed)
    macro_state = _build_macro_state_from_reset(info_reset)

    step_records: List[Dict[str, Any]] = []

    # ✅ Ledger audit-ready por préstamo (decisión final y racional macro del step)
    decision_ledger: Dict[str, Dict[str, Any]] = {}

    done = False
    truncated = False
    sell_steps = 0

    for t in range(cfg_inf.n_steps):
        if done or truncated:
            break

        action_fin, fin_family, rationale_fin = _macro_financial_action(
            macro_state=macro_state,
            thresholds=thresholds,
            posture=posture,
            hurdle=hurdle,
        )

        action_rl_arr, _ = macro_policy.predict(obs, deterministic=cfg_inf.deterministic)
        action_rl = int(np.squeeze(action_rl_arr))
        rl_family = _family(action_rl)

        # -----------------------------
        # Combinación híbrida
        # -----------------------------
        if action_fin is None:
            if (not guards["allow_sell_in_ambig"]) and rl_family == "SELL":
                action_final = 4
                combination_note = "Ambigua: PPO proponía SELL, guardrail prudencial → RESTRUCT top-K."
            else:
                action_final = action_rl
                combination_note = "Ambigua: se sigue PPO (con guardrails)."
        else:
            fin_action = int(action_fin)
            fin_action_family = _family(fin_action)

            if fin_action_family == "MIX":
                if action_rl in MIX_FAMILY:
                    action_final = action_rl
                    combination_note = "Regla financiera=MIX, PPO refina dentro de MIX."
                else:
                    action_final = fin_action
                    combination_note = "Regla financiera=MIX, PPO fuera de MIX → se mantiene MIX."
            else:
                if rl_family == fin_action_family:
                    action_final = action_rl
                    combination_note = "PPO compatible con familia macro → PPO refina táctica."
                else:
                    if guards["allow_rl_family_flip"]:
                        action_final = action_rl
                        combination_note = "Desinversión: se permite cambio de familia por PPO."
                    else:
                        action_final = fin_action
                        combination_note = "PPO contradice familia macro → manda regla financiera."

        # Guardrail: limitar % steps SELL
        if _family(action_final) == "SELL":
            if (sell_steps + 1) / max(cfg_inf.n_steps, 1) > float(guards["max_steps_sell"]):
                action_final = 9 if posture != "prudencial" else 4
                combination_note += " | Guardrail: exceso de SELL → degradado a MIX/RESTRUCT."

        obs, reward, done, truncated, info_step = env.step(action_final)
        pm = info_step.get("portfolio_metrics", {}) or {}
        macro_state = _build_macro_state_from_step(pm)

        action_summary = info_step.get("action_summary", {}) or {}
        sold_ids = action_summary.get("sold_ids", []) or []
        restructured_ids = action_summary.get("restructured_ids", []) or []

        sold_ids_s = [str(x).strip() for x in sold_ids if str(x).strip()]
        restr_ids_s = [str(x).strip() for x in restructured_ids if str(x).strip()]

        # ✅ ledger: SELL manda; RESTRUCT solo si no vendido
        for lid in sold_ids_s:
            prev = decision_ledger.get(lid, {})
            decision_ledger[lid] = {
                "decision": "VENDER",
                "decision_step": int(t),
                "decision_action_id": int(action_final),
                "decision_action_desc": _ACTION_MAP.get(int(action_final), f"FINAL_ACCION_{action_final}"),
                "decision_rationale_macro": str(rationale_fin),
                "decision_note": str(combination_note),
                "decision_posture": posture,
                "decision_prev": prev.get("decision", ""),
            }

        for lid in restr_ids_s:
            if decision_ledger.get(lid, {}).get("decision") == "VENDER":
                continue
            if lid not in decision_ledger:
                decision_ledger[lid] = {
                    "decision": "REESTRUCTURAR",
                    "decision_step": int(t),
                    "decision_action_id": int(action_final),
                    "decision_action_desc": _ACTION_MAP.get(int(action_final), f"FINAL_ACCION_{action_final}"),
                    "decision_rationale_macro": str(rationale_fin),
                    "decision_note": str(combination_note),
                    "decision_posture": posture,
                    "decision_prev": "",
                }

        if _family(action_final) == "SELL":
            sell_steps += 1

        rt = info_step.get("reward_terms", {}) or {}

        step_records.append(
            {
                "step": t,
                "timestamp": ts,
                "scenario": cfg_inf.scenario,
                "risk_posture": posture,
                "bank_profile": prof.value,
                "bank_strategy": strat.name,
                "accion_fin_id": action_fin if action_fin is not None else -1,
                "accion_fin_desc": _ACTION_MAP.get(action_fin, "FIN_N/A") if action_fin is not None else "FIN_None",
                "fin_family": fin_family,
                "accion_rl_id": action_rl,
                "accion_rl_desc": _ACTION_MAP.get(action_rl, f"RL_ACCION_{action_rl}"),
                "rl_family": rl_family,
                "accion_final_id": action_final,
                "accion_final_desc": _ACTION_MAP.get(action_final, f"FINAL_ACCION_{action_final}"),
                "final_family": _family(action_final),
                "rationale_fin": rationale_fin,
                "combination_note": combination_note,
                "reward": float(reward) if reward is not None else np.nan,
                "EVA_before": pm.get("EVA_before"),
                "EVA_after": pm.get("EVA_after"),
                "EVA_gain": pm.get("EVA_gain"),
                "RWA_before": pm.get("RWA_before"),
                "RWA_after": pm.get("RWA_after"),
                "capital_liberado": pm.get("capital_liberado"),
                "risk_before": pm.get("risk_before"),
                "risk_after": pm.get("risk_after"),
                "num_loans_active": pm.get("num_loans_active"),
                "num_loans_closed": pm.get("num_loans_closed"),
                "eva_volatility": pm.get("eva_volatility"),
                "hhi_segment": pm.get("hhi_segment"),
                "hhi_rating": pm.get("hhi_rating"),
                "capital_carry_cost": pm.get("capital_carry_cost"),
                "reward_eva_term": rt.get("eva_term", np.nan),
                "reward_cap_term": rt.get("cap_term", np.nan),
                "reward_risk_term": rt.get("risk_term", np.nan),
                "reward_pti_term": rt.get("pti_term", np.nan),
                "reward_pnl_term": rt.get("pnl_term", np.nan),
                "reward_cure_term": rt.get("cure_term", np.nan),
                "reward_vol_term": rt.get("vol_term", np.nan),
                "reward_conc_term": rt.get("conc_term", np.nan),
                "reward_carry_term": rt.get("carry_term", np.nan),
                "reward_total_check": rt.get("reward_total", np.nan),
                "sold_ids": ";".join(sold_ids_s) if sold_ids_s else "",
                "restructured_ids": ";".join(restr_ids_s) if restr_ids_s else "",
            }
        )

    df_steps = pd.DataFrame(step_records)
    df_final = pd.DataFrame(getattr(env, "portfolio", []))

    # -----------------------------
    # ✅ Enriquecer df_final con “decisión final” audit-ready
    # -----------------------------
    if not df_final.empty:
        if "loan_id" in df_final.columns:
            df_final["loan_id"] = df_final["loan_id"].astype(str)
        else:
            # fallback defensivo
            df_final["loan_id"] = [str(i) for i in range(len(df_final))]

        # defaults: MANTENER
        df_final["decision"] = "MANTENER"
        df_final["decision_step"] = -1
        df_final["decision_posture"] = posture
        df_final["decision_action_id"] = -1
        df_final["decision_action_desc"] = ""
        df_final["decision_rationale_macro"] = ""
        df_final["decision_note"] = ""

        # aplicar ledger
        def _map_field(lid: str, k: str, default: Any):
            return decision_ledger.get(lid, {}).get(k, default)

        df_final["decision"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision", "MANTENER"))
        df_final["decision_step"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_step", -1))
        df_final["decision_posture"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_posture", posture))
        df_final["decision_action_id"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_action_id", -1))
        df_final["decision_action_desc"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_action_desc", ""))
        df_final["decision_rationale_macro"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_rationale_macro", ""))
        df_final["decision_note"] = df_final["loan_id"].map(lambda x: _map_field(x, "decision_note", ""))

        # coherencia adicional por estado/closed (por si no hubo ledger)
        if "closed" in df_final.columns:
            mask_closed = df_final["closed"].astype(bool)
            df_final.loc[mask_closed & (df_final["decision"] == "MANTENER"), "decision"] = "VENDER"

        if "estado" in df_final.columns:
            est = df_final["estado"].astype(str).str.upper()
            df_final.loc[est.str.contains("VEND"), "decision"] = "VENDER"
            # si env etiqueta REESTRUCTURADO/CURADO, reflejar
            df_final.loc[est.str.contains("REESTRUCT"), "decision"] = "REESTRUCTURAR"
            df_final.loc[est.str.contains("CUR"), "decision"] = "REESTRUCTURAR"

        # ordenar columnas “bank-ready” al principio si existen
        lead_cols = [
            "loan_id",
            "decision",
            "decision_step",
            "decision_posture",
            "decision_action_id",
            "decision_action_desc",
            "decision_rationale_macro",
            "decision_note",
        ]
        cols = list(df_final.columns)
        ordered = [c for c in lead_cols if c in cols] + [c for c in cols if c not in lead_cols]
        df_final = df_final[ordered]

    # -----------------------------
    # Persistencia (opcional)
    # -----------------------------
    if (not bool(getattr(cfg_inf, "persist_outputs", True))) or (out_dir is None) or (not str(out_dir).strip()):
        return df_steps, df_final, "", "", ""

    os.makedirs(out_dir, exist_ok=True)

    if suffix is None:
        traj_name = "trajectory_portfolio.csv"
        final_name = "portfolio_final.xlsx" if cfg_inf.save_final_excel else "portfolio_final.csv"
        summary_name = "summary_portfolio.csv"
    else:
        traj_name = f"trajectory_portfolio_{suffix}.csv"
        final_name = f"portfolio_final_{suffix}.xlsx" if cfg_inf.save_final_excel else f"portfolio_final_{suffix}.csv"
        summary_name = f"summary_portfolio_{suffix}.csv"

    traj_path = os.path.join(out_dir, traj_name)
    df_steps.to_csv(traj_path, index=False, encoding="utf-8-sig")
    logger.info(f"📈 Trayectoria macro guardada en {traj_path}")

    final_path = _resolve_final_path(cfg_inf, out_dir, final_name)
    if cfg_inf.save_final_excel:
        export_styled_excel_portfolio(df_final, final_path)
    else:
        df_final.to_csv(final_path, index=False, encoding="utf-8-sig")
    logger.info(f"📊 Cartera final exportada a {final_path}")

    summary_path = os.path.join(out_dir, summary_name)
    if not df_steps.empty:
        last_row = df_steps.iloc[-1]
        df_summary = pd.DataFrame(
            {
                "timestamp": [ts],
                "label": [cfg_inf.tag],
                "scenario": [cfg_inf.scenario],
                "risk_posture": [cfg_inf.risk_posture],
                "bank_profile": [prof.value],
                "bank_strategy": [strat.name],
                "n_steps": [len(df_steps)],
                "EVA_final": [last_row.get("EVA_after", np.nan)],
                "RWA_final": [last_row.get("RWA_after", np.nan)],
                "capital_liberado_total": [float(df_steps["capital_liberado"].fillna(0).sum())],
                "num_loans_active_final": [last_row.get("num_loans_active", np.nan)],
                "num_loans_closed_final": [last_row.get("num_loans_closed", np.nan)],
                "pct_steps_keep": [float((df_steps["final_family"] == "KEEP").mean())],
                "pct_steps_sell": [float((df_steps["final_family"] == "SELL").mean())],
                "pct_steps_restruct": [float((df_steps["final_family"] == "RESTRUCT").mean())],
                "pct_steps_mix": [float((df_steps["final_family"] == "MIX").mean())],
            }
        )
        df_summary.to_csv(summary_path, index=False, encoding="utf-8-sig")
        logger.info(f"💾 Resumen agregado macro guardado en {summary_path}")

    return df_steps, df_final, traj_path, final_path, summary_path


# -----------------------------------------------------------
# API NUEVA: in-memory (NO exports / NO carpetas)
# -----------------------------------------------------------
def run_portfolio_inference_df(cfg_inf: PortfolioInferenceConfig) -> Tuple[pd.DataFrame, pd.DataFrame]:
    cfg_local = dataclasses.replace(cfg_inf, persist_outputs=False)
    df_steps, df_final, _, _, _ = _run_portfolio_inference_for_posture(cfg_local, out_dir=None, suffix=None)
    return df_final, df_steps


# -----------------------------------------------------------
# Legacy: crea carpeta y exporta (solo si persist_outputs=True)
# -----------------------------------------------------------
def run_portfolio_inference(cfg_inf: PortfolioInferenceConfig) -> str:
    if not bool(getattr(cfg_inf, "persist_outputs", True)):
        _ = run_portfolio_inference_df(cfg_inf)
        return ""

    ts = _now_tag()
    out_dir = _resolve_output_dir(cfg_inf, ts)
    _df_steps, _df_final, _traj_path, _final_path, _summary_path = _run_portfolio_inference_for_posture(cfg_inf, out_dir, suffix=None)
    return out_dir


def run_portfolio_inference_multi_posture(cfg_base: PortfolioInferenceConfig) -> List[str]:
    if not bool(getattr(cfg_base, "persist_outputs", True)):
        for posture in ("prudencial", "balanceado", "desinversion"):
            _ = run_portfolio_inference_df(dataclasses.replace(cfg_base, risk_posture=posture))
        return []

    ts = _now_tag()
    out_dir = _resolve_output_dir(cfg_base, ts)
    os.makedirs(out_dir, exist_ok=True)

    outputs: List[str] = []
    escenarios = [
        ("prudencial", "portfolio_prudencial"),
        ("balanceado", "portfolio_balanceado"),
        ("desinversion", "portfolio_desinversion"),
    ]

    for posture, tag in escenarios:
        cfg_inf = dataclasses.replace(cfg_base, risk_posture=posture, tag=tag, persist_outputs=True)
        logger.info(f"[MACRO] Ejecutando inferencia multi-postura (risk_posture={posture})…")
        _, _, _, final_path, _ = _run_portfolio_inference_for_posture(cfg_inf, out_dir, suffix=posture)
        outputs.append(final_path)

    return outputs


# -----------------------------------------------------------
# CLI
# -----------------------------------------------------------
def parse_args() -> PortfolioInferenceConfig:
    p = argparse.ArgumentParser(description="Inferencia PPO a nivel de cartera (PortfolioEnv · Banco L1.5)")
    p.add_argument(
        "--model",
        type=str,
        default=_pick_default_macro_model_path() or os.path.join(MODELS_DIR, "best_model_portfolio.zip"),
    )
    p.add_argument("--portfolio", type=str, default=os.path.join(ROOT_DIR, "data", "portfolio_synth.xlsx"))
    p.add_argument("--device", type=str, choices=["auto", "cpu", "cuda"], default="auto")
    p.add_argument("--seed", type=int, default=42)

    p.add_argument("--deterministic", dest="deterministic", action="store_true", help="Inferencia determinista (default).")
    p.add_argument("--stochastic", dest="deterministic", action="store_false", help="Inferencia estocástica.")
    p.set_defaults(deterministic=True)

    p.add_argument("--tag", type=str, default="portfolio_policy")
    p.add_argument("--n-steps", type=int, default=1, dest="n_steps")
    p.add_argument("--top-k", type=int, default=5, dest="top_k")
    p.add_argument("--no-excel", action="store_true", help="No exportar Excel final, solo CSV.")
    p.add_argument(
        "--risk-posture",
        type=str,
        choices=["prudencial", "balanceado", "desinversion"],
        default="balanceado",
    )
    p.add_argument(
        "--scenario",
        type=str,
        choices=["baseline", "adverse", "severe"],
        default="baseline",
        help="Escenario macro usado dentro de PortfolioEnv.",
    )

    p.add_argument("--loan-model", type=str, default=_pick_default_micro_model_path() or "")

    # Defaults VN seguros:
    p.add_argument("--vn-portfolio", type=str, default=_pick_default_vn_portfolio_path() or "")
    p.add_argument("--vn-loan", type=str, default=_pick_default_vn_loan_path() or "")

    p.add_argument("--no-persist", action="store_true", default=False, help="No exporta ni crea carpetas (solo cálculo).")
    p.add_argument("--out-dir", type=str, default="", help="Carpeta de salida (solo si persist).")
    p.add_argument("--final-path", type=str, default="", help="Path exacto del Excel/CSV final (solo si persist).")

    args = p.parse_args()

    loan_model_path: Optional[str] = args.loan_model if args.loan_model and os.path.exists(args.loan_model) else None
    vn_port: Optional[str] = args.vn_portfolio if args.vn_portfolio and os.path.exists(args.vn_portfolio) else None
    vn_loan: Optional[str] = args.vn_loan if args.vn_loan and os.path.exists(args.vn_loan) else None

    out_dir = args.out_dir.strip() or None
    final_path = args.final_path.strip() or None

    return PortfolioInferenceConfig(
        model_path=args.model,
        portfolio_path=args.portfolio,
        device=args.device,
        seed=args.seed,
        deterministic=bool(args.deterministic),
        tag=args.tag,
        n_steps=args.n_steps,
        top_k=args.top_k,
        scenario=args.scenario,
        save_final_excel=not args.no_excel,
        risk_posture=args.risk_posture,
        loan_model_path=loan_model_path,
        vecnormalize_portfolio_path=vn_port,
        vecnormalize_loan_path=vn_loan,
        persist_outputs=(not bool(args.no_persist)),
        output_dir=out_dir,
        final_output_path=final_path,
    )


if __name__ == "__main__":
    cfg_inf = parse_args()
    out_dir = run_portfolio_inference(cfg_inf)
    if out_dir:
        logger.info(f"🏁 Inferencia macro completada. Reporte en carpeta: {out_dir}")
    else:
        logger.info("🏁 Inferencia macro completada (no-persist).")
