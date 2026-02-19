# -*- coding: utf-8 -*-
# ============================================================
# agent/coordinator_inference.py
# — Inferencia coordinada MICRO+MACRO (decisión única por préstamo)
# (v3.7.0 · macro steering consistente · fire-sale robusto · contrafactuals · schema-stable)
# ============================================================
"""
MEJORAS CLAVE (bank-ready):
- Export estable: enforce_schema() SIEMPRE antes de Excel/CSV.
- Macro asigna SOLO loans “tocados” (sell/restructure). Resto => Macro_Assignment="NO_ASIGNADO".
- Arbitraje explícito: si macro_selected y hay conflicto, MACRO gana salvo guardrails DUROS.
- Fire-sale robusto:
    * Enabled por postura (prudencial/balanceado).
    * Triggered si se cumple cualquiera: simulador flag, price/EAD < threshold,
      pnl/book < umbral, pnl_abs < umbral, price/book < threshold_book (si existe).
- RC10: Guardrail de viabilidad (PTI/DSCR faltantes) bloquea REESTRUCTURAR.
- Contrafactuals mínimos coherentes:
    * capital_release_cf derivable con RWA_pre o EAD*RW y ratio_total (fallback 10.5%).
    * capital_release_realized / pnl_realized coherentes con Accion_final.
- Merge defensivo con portfolio input (rellena EAD/RW/segment si faltan en micro).
"""

from __future__ import annotations

import os
import json
import sys
import argparse
import logging
from dataclasses import dataclass
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import shutil

import numpy as np
import pandas as pd

# -----------------------------------------------------------
# Rutas básicas
# -----------------------------------------------------------
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.append(ROOT_DIR)

DATA_DIR = os.path.join(ROOT_DIR, "data")
MODELS_DIR = os.path.join(ROOT_DIR, "models")
REPORTS_DIR = os.path.join(ROOT_DIR, "reports")
LOGS_DIR = os.path.join(ROOT_DIR, "logs")

for d in (DATA_DIR, MODELS_DIR, REPORTS_DIR, LOGS_DIR):
    os.makedirs(d, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOGS_DIR, "coordinator_inference.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("coordinator_inference")

# -----------------------------------------------------------
# Imports del proyecto
# -----------------------------------------------------------
import config as cfg
from agent.policy_inference import InferenceConfig, run_inference
from agent.policy_inference_portfolio import PortfolioInferenceConfig, run_portfolio_inference
from risk.gates import check_restruct_viability, check_sell_fire_sale
from optimizer.guardrails import check_restructure_constraints, check_sell_constraints

# --- Schema enforcement (columnas estables bank-ready) ---
try:
    from reports.schema import enforce_schema  # type: ignore
except Exception:
    def enforce_schema(df: pd.DataFrame) -> pd.DataFrame:  # fallback no-op
        return df

# -----------------------------------------------------------
# CONSTANTS
# -----------------------------------------------------------
OVERRIDE_LOG_COLS = [
    "loan_id", "level", "from_action", "to_action", "portfolio_context", 
    "posture", "run_id", "macro_action_used", "macro_rationales_short", 
    "pti_actual", "dscr_actual", "pnl"
]



# ===========================================================
# Exporter robusto (evita roturas por path)
# ===========================================================
def _load_exporter():
    try:
        from reports.export_styled_excel import export_styled_excel  # type: ignore
        return export_styled_excel
    except Exception:
        pass
    try:
        from summary.export_styled_excel import export_styled_excel  # type: ignore
        return export_styled_excel
    except Exception:
        pass

    # Fallback mínimo “bank-ready”: OpenPyXL con headers y autofit
    def _fallback_export(df: pd.DataFrame, out_path: str) -> None:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Decisiones"

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        wb.save(out_path)

    logger.warning("[WARN] No se encontró export_styled_excel. Usando fallback OpenPyXL.")
    return _fallback_export


export_styled_excel = _load_exporter()


# ===========================================================
# CONFIG GENERAL DEL COORDINADOR
# ===========================================================
@dataclass
class CoordinatorInferenceConfig:
    model_path_micro: str
    portfolio_path: str

    vecnormalize_path_micro: Optional[str] = None

    model_path_macro: Optional[str] = None
    vecnormalize_path_macro: Optional[str] = None
    vecnormalize_path_loan: Optional[str] = None

    device: str = "auto"
    seed: int = 42
    deterministic: bool = True
    tag: str = "coordinated_policy"

    n_steps: int = 1
    top_k: int = 5

    deliverable_only: bool = True
    export_audit_csv: bool = False
    base_output_dir: Optional[str] = None


# ===========================================================
# Guardrails CIB (defaults + override por BankStrategy)
# ===========================================================
DEFAULT_GUARDRAILS = {
    "prudencial": {
        "CAP_RELEASE_CRITICAL": 250_000.0,
        "MIN_PRICE_TO_EAD": 0.12,
        "MAX_FIRE_SALE_PNL_RATIO_BOOK": -0.65,
        "MAX_FIRE_SALE_PNL_ABS": -50_000.0,
        "MIN_DEVA_RESTRUCT": 10_000.0,
        "DSCR_MIN": 1.10,
        "PTI_MAX": 0.35,
    },
    "balanceado": {
        "CAP_RELEASE_CRITICAL": 150_000.0,
        "MIN_PRICE_TO_EAD": 0.10,
        "MAX_FIRE_SALE_PNL_RATIO_BOOK": -0.75,
        "MAX_FIRE_SALE_PNL_ABS": -75_000.0,
        "MIN_DEVA_RESTRUCT": 5_000.0,
        "DSCR_MIN": 1.05,
        "PTI_MAX": 0.45,
    },
    "desinversion": {
        "CAP_RELEASE_CRITICAL": 0.0,
        "MIN_PRICE_TO_EAD": 0.07,
        "MAX_FIRE_SALE_PNL_RATIO_BOOK": -0.90,
        "MAX_FIRE_SALE_PNL_ABS": -250_000.0,
        "MIN_DEVA_RESTRUCT": 15_000.0,
        "DSCR_MIN": 1.00,
        "PTI_MAX": 0.55,
    },
}


def _posture_to_profile(risk_posture: str):
    rp = (risk_posture or "balanceado").lower().strip()
    if rp == "prudencial":
        return cfg.BankProfile.PRUDENTE
    if rp == "desinversion":
        return cfg.BankProfile.DESINVERSION
    return cfg.BankProfile.BALANCEADO


def _get_guardrails(risk_posture: str) -> Dict[str, Any]:
    rp = (risk_posture or "balanceado").lower().strip()
    g = DEFAULT_GUARDRAILS.get(rp, DEFAULT_GUARDRAILS["balanceado"]).copy()

    prof = _posture_to_profile(rp)
    if hasattr(cfg, "set_bank_profile"):
        try:
            cfg.set_bank_profile(prof)
        except Exception as e:
            logger.warning(f"[WARN] No se pudo fijar bank_profile en config: {e}")

    try:
        strat = cfg.BANK_STRATEGIES.get(prof)
        if strat is not None:
            if hasattr(strat, "dscr_min"):
                g["DSCR_MIN"] = float(getattr(strat, "dscr_min"))
            if hasattr(strat, "eva_min_improvement"):
                g["MIN_DEVA_RESTRUCT"] = float(getattr(strat, "eva_min_improvement"))

            pti_candidate = None
            if hasattr(strat, "pti_max") and getattr(strat, "pti_max") is not None:
                pti_candidate = getattr(strat, "pti_max")
            else:
                if rp == "prudencial" and hasattr(strat, "esfuerzo_bajo"):
                    pti_candidate = getattr(strat, "esfuerzo_bajo")
                elif rp == "desinversion" and hasattr(strat, "esfuerzo_alto"):
                    pti_candidate = getattr(strat, "esfuerzo_alto")
                elif hasattr(strat, "esfuerzo_alto") and hasattr(strat, "esfuerzo_bajo"):
                    pti_candidate = 0.5 * (
                        float(getattr(strat, "esfuerzo_alto")) + float(getattr(strat, "esfuerzo_bajo"))
                    )
            if pti_candidate is not None:
                g["PTI_MAX"] = float(pti_candidate)

    except Exception as e:
        logger.warning(f"[WARN] No se pudo derivar guardrails desde BankStrategy: {e}")

    return g


# ===========================================================
# Helpers
# ===========================================================
def _safe_float(x, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        if isinstance(x, float) and np.isnan(x):
            return default
        return float(x)
    except Exception:
        return default


def _safe_str(x, default: str = "") -> str:
    if x is None:
        return default
    if isinstance(x, float) and np.isnan(x):
        return default
    return str(x)


def _is_missing(x: Any) -> bool:
    return x is None or (isinstance(x, float) and np.isnan(x))


def _fmt_eur(x: float) -> str:
    return f"{x:,.0f}€"


def _pct(x: float) -> str:
    return f"{x*100:.1f}%"


def _get_loan_id_from_row(r: pd.Series) -> str:
    for k in ("loan_id", "Loan_ID", "LOAN_ID", "LoanID", "loanId", "id", "ID"):
        if k in r and pd.notna(r.get(k)):
            return str(r.get(k))
    try:
        if r.name is not None and str(r.name).lower() != "nan":
            return str(r.name)
    except Exception:
        pass
    return ""


def _ensure_loan_id_column(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        if df is None:
            return df
        if "loan_id" not in df.columns:
            df = df.copy()
            df["loan_id"] = ""
        return df

    df = df.copy()

    if "loan_id" in df.columns and df["loan_id"].notna().any():
        df["loan_id"] = df["loan_id"].astype(str)
        return df

    aliases = ["Loan_ID", "LOAN_ID", "LoanID", "loanId", "id", "ID"]
    for a in aliases:
        if a in df.columns and df[a].notna().any():
            df["loan_id"] = df[a].astype(str)
            return df

    try:
        idx = df.index.astype(str)
        if len(idx) == len(df) and pd.Series(idx).notna().any():
            df["loan_id"] = idx
            return df
    except Exception:
        pass

    df["loan_id"] = df.apply(_get_loan_id_from_row, axis=1).astype(str)
    return df


def _normalize_action(x: Any) -> str:
    if x is None:
        return ""
    try:
        if isinstance(x, (int, np.integer)):
            return {0: "MANTENER", 1: "REESTRUCTURAR", 2: "VENDER"}.get(int(x), "")
        if isinstance(x, float) and not np.isnan(x) and abs(x - int(x)) < 1e-9:
            return {0: "MANTENER", 1: "REESTRUCTURAR", 2: "VENDER"}.get(int(x), "")
    except Exception:
        pass

    s = str(x).strip().upper()
    if not s:
        return ""
    if s in ("KEEP", "MANTENER", "HOLD", "MAINTAIN"):
        return "MANTENER"
    if s in ("RESTRUCT", "RESTRUCTURE", "REESTRUCTURAR", "REESTRUCTURA"):
        return "REESTRUCTURAR"
    if s in ("SELL", "VENDER", "VENTA"):
        return "VENDER"
    if "VEND" in s:
        return "VENDER"
    if "REEST" in s:
        return "REESTRUCTURAR"
    if "MANT" in s:
        return "MANTENER"
    if s in ("NO_ASIGNADO", "NO ASIGNADO", "NOT_ASSIGNED", "NOT ASSIGNED"):
        return "NO_ASIGNADO"
    return s


def _pick_default_macro_model() -> Optional[str]:
    for p in (
        os.path.join(MODELS_DIR, "best_model_portfolio.zip"),
        os.path.join(MODELS_DIR, "best_model_macro.zip"),
    ):
        if os.path.exists(p):
            return p
    return None


def _pick_default_micro_vecnorm() -> Optional[str]:
    for p in (
        os.path.join(MODELS_DIR, "vecnormalize_loan.pkl"),
        os.path.join(MODELS_DIR, "vecnormalize_micro.pkl"),
    ):
        if os.path.exists(p):
            return p

    legacy = os.path.join(MODELS_DIR, "vecnormalize_final.pkl")
    if os.path.exists(legacy):
        logger.warning(
            "[WARN] Usando VN legacy para MICRO: models/vecnormalize_final.pkl. "
            "Recomendado: guardar como vecnormalize_loan.pkl."
        )
        return legacy
    return None


def _pick_default_macro_vecnorm() -> Optional[str]:
    for p in (
        os.path.join(MODELS_DIR, "vecnormalize_portfolio.pkl"),
        os.path.join(MODELS_DIR, "vecnormalize_macro.pkl"),
    ):
        if os.path.exists(p):
            return p
    return None


def _sanitize_vecnorm_paths(
    vn_micro: Optional[str],
    vn_macro: Optional[str],
    vn_loan: Optional[str],
) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    def norm(p: Optional[str]) -> Optional[str]:
        if p and os.path.exists(p):
            return p
        return None

    vn_micro = norm(vn_micro)
    vn_macro = norm(vn_macro)
    vn_loan = norm(vn_loan)

    if vn_macro and vn_micro and os.path.abspath(vn_macro) == os.path.abspath(vn_micro):
        logger.warning("[WARN] VN macro == VN micro. Ignorando vn_macro para evitar mismatch loan vs portfolio.")
        vn_macro = None

    if vn_macro:
        b = os.path.basename(vn_macro).lower()
        if "loan" in b or "micro" in b:
            logger.warning("[WARN] vn_macro parece de LoanEnv por nombre. Ignorando vn_macro para evitar mismatch.")
            vn_macro = None

    if vn_loan is None:
        vn_loan = vn_micro

    return vn_micro, vn_macro, vn_loan


def _get_ratio_total_capital() -> float:
    # Ratio total (Tier1+T2) típico ~10.5% si no existe en config
    cands = [
        ("CONFIG", "regulacion", "capital_total_ratio"),
        ("CONFIG", "regulacion", "total_capital_ratio"),
        ("CONFIG", "regulacion", "ratio_total_capital"),
    ]
    for a, b, c in cands:
        try:
            obj = getattr(cfg, a, None)
            obj = getattr(obj, b, None)
            v = getattr(obj, c, None)
            if v is not None:
                return float(v)
        except Exception:
            pass
    return 0.105


def harmonize_portfolio_schema(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # --- map monthly_* -> legacy names usados por optimizadores
    if "ingreso_mensual" not in df.columns and "monthly_income" in df.columns:
        df["ingreso_mensual"] = df["monthly_income"]
    if "cashflow_operativo_mensual" not in df.columns and "monthly_cfo" in df.columns:
        df["cashflow_operativo_mensual"] = df["monthly_cfo"]
    if "cuota_mensual" not in df.columns and "monthly_payment" in df.columns:
        df["cuota_mensual"] = df["monthly_payment"]

    # --- PTI / DSCR: preferimos pre si existe; si no, derivamos desde monthly_*
    if "PTI" not in df.columns:
        if "PTI_pre" in df.columns:
            df["PTI"] = df["PTI_pre"]
        elif ("monthly_payment" in df.columns) and ("monthly_income" in df.columns):
            df["PTI"] = df["monthly_payment"] / df["monthly_income"].replace(0, np.nan)

    if "DSCR" not in df.columns:
        if "DSCR_pre" in df.columns:
            df["DSCR"] = df["DSCR_pre"]
        elif ("monthly_cfo" in df.columns) and ("monthly_payment" in df.columns):
            df["DSCR"] = df["monthly_cfo"] / df["monthly_payment"].replace(0, np.nan)

    # --- book_value: si no viene, aproximación consistente
    if "book_value" not in df.columns:
        if "coverage_rate" in df.columns and "EAD" in df.columns:
            cov = pd.to_numeric(df["coverage_rate"], errors="coerce")
            cov = np.where(cov > 1.0, cov / 100.0, cov)
            cov = pd.Series(cov).clip(0, 1)
            df["book_value"] = pd.to_numeric(df["EAD"], errors="coerce") * (1.0 - cov)

        elif "LGD" in df.columns and "EAD" in df.columns:
            df["book_value"] = df["EAD"] * (1.0 - df["LGD"].clip(0, 1))

    # --- normalización mínima de segment (para gates retail vs corp)
    if "segment" in df.columns:
        df["segment"] = df["segment"].astype(str).str.strip()

    return df


def _load_portfolio_df(portfolio_path: str) -> pd.DataFrame:
    if not os.path.exists(portfolio_path):
        return pd.DataFrame()

    ext = os.path.splitext(portfolio_path)[1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            dfp = pd.read_excel(portfolio_path)
        else:
            dfp = pd.read_csv(portfolio_path)
    except Exception as e:
        logger.warning(f"[WARN] No se pudo leer portfolio para merge defensivo: {e}")
        return pd.DataFrame()

    dfp = harmonize_portfolio_schema(dfp)
    dfp = _ensure_loan_id_column(dfp)
    return dfp


def _fill_from_portfolio(df_final: pd.DataFrame, df_port: pd.DataFrame) -> pd.DataFrame:
    if df_port is None or df_port.empty or df_final is None or df_final.empty:
        return df_final

    df = df_final.copy()
    port = df_port.copy()

    # Normaliza alias comunes en portfolio
    alias_map = {
        "segment": ["segment", "segmento_banco", "segment_raw", "SEGMENT", "Segment"],
        "EAD": ["EAD", "ead", "Ead", "EAD_EUR", "exposure"],
        "RW": ["RW", "rw", "risk_weight", "RW_STD"],
        "PTI_pre": ["PTI_pre", "PTI", "pti"],
        "DSCR_pre": ["DSCR_pre", "DSCR", "dscr"],
    }
    for canon, aliases in alias_map.items():
        if canon in port.columns:
            continue
        for a in aliases:
            if a in port.columns:
                port[canon] = port[a]
                break

    keep_cols = ["loan_id"] + [c for c in ("segment", "EAD", "RW", "PTI_pre", "DSCR_pre") if c in port.columns]
    port = port[keep_cols].drop_duplicates(subset=["loan_id"])

    df = df.merge(port, on="loan_id", how="left", suffixes=("", "_port"))

    def fill_col(c: str) -> None:
        if c not in df.columns:
            if f"{c}_port" in df.columns:
                df[c] = df[f"{c}_port"]
            return
        if f"{c}_port" not in df.columns:
            return
        df[c] = df[c].combine_first(df[f"{c}_port"])

    for c in ("segment", "EAD", "RW", "PTI_pre", "DSCR_pre"):
        fill_col(c)

    drop_ports = [c for c in df.columns if c.endswith("_port")]
    if drop_ports:
        df.drop(columns=drop_ports, inplace=True)

    return df


# ===========================================================
# Helpers CIB (razonamiento y codificación auditable)
# ===========================================================
def _cib_rationale_row(
    r: pd.Series,
    risk_posture: str,
    accion_micro: str,
    macro_action: str,
    accion_final: str,
    macro_rationale: str,
    flags: Dict[str, Any],
) -> Tuple[str, str, str]:
    rp_raw = str(risk_posture or "balanceado").strip().lower()
    rp_raw = (
        rp_raw.replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
    )
    if rp_raw in {"prudente", "prudencial"}:
        rp_l = "prudencial"
    elif rp_raw in {"desinversion", "desinversionn", "desinvertir"}:
        rp_l = "desinversion"
    elif rp_raw in {"balanceado", "balanced"}:
        rp_l = "balanceado"
    else:
        rp_l = "balanceado"

    rp_u = rp_l.upper()

    def _fmt_metric(x, ndigits=2) -> str:
        try:
            if x is None:
                return "NA"
            xf = float(x)
            if pd.isna(xf):
                return "NA"
            return f"{xf:.{ndigits}f}"
        except Exception:
            return "NA"

    eva_pre = _safe_float(r.get("EVA_pre", r.get("EVA_before", r.get("EVA", 0.0))))
    eva_post = _safe_float(r.get("EVA_post", r.get("EVA_after", 0.0)))
    deva = _safe_float(r.get("ΔEVA", r.get("delta_eva", r.get("EVA_gain", 0.0))))
    rorwa_pre = _safe_float(r.get("RORWA_pre", r.get("RORWA", 0.0)))

    cap_rel = _safe_float(r.get("capital_liberado", r.get("Capital_release", 0.0)))
    price = _safe_float(r.get("precio_optimo", r.get("price", 0.0)))
    pnl = _safe_float(r.get("pnl", r.get("PnL", 0.0)))

    pd_ = _safe_float(r.get("PD", 0.0))
    lgd = _safe_float(r.get("LGD", 0.0))
    dpd = _safe_float(r.get("DPD", 0.0))
    secured = _safe_str(r.get("secured", ""))

    pti = r.get("PTI_post", r.get("PTI", None))
    dscr = r.get("DSCR_post", r.get("DSCR", None))

    hurdle = _safe_float(getattr(cfg.CONFIG.regulacion, "hurdle_rate", 0.0))

    governance = "Micro-led"
    if accion_final != accion_micro:
        if accion_final == macro_action:
            governance = "Macro override (portfolio steering) with guardrails"
        else:
            governance = "Guardrail override (risk policy)"

    fire_sale = bool(flags.get("fire_sale", False))
    restruct_feasible = bool(flags.get("restruct_feasible", True))
    micro_meets_hurdle = bool(flags.get("micro_meets_hurdle", False))

    # RC10
    if bool(flags.get("missing_viability_inputs", False)):
        code = "RC10_MISSING_VIABILITY_INPUTS"
        proposed_action = _safe_str(flags.get("proposed_action", ""), "")
        txt = (
            f"[{rp_u}] RESTRUCT blocked: missing viability inputs (PTI/DSCR). "
            f"Proposed={proposed_action or 'REESTRUCTURAR'} (Micro={accion_micro}, Macro={macro_action}) -> Final={accion_final}. "
            f"DSCR={_fmt_metric(dscr)}, PTI={_fmt_metric(pti)}. "
            f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
        )
        governance = "Guardrail override (missing PTI/DSCR)"
        return code, txt, governance

    # SELL bloqueado por fire-sale
    sell_requested = (accion_micro == "VENDER") or (macro_action == "VENDER")
    if sell_requested and accion_final != "VENDER" and fire_sale and rp_l != "desinversion":
        code = "RC02_SELL_BLOCKED_FIRE_SALE"
        txt = (
            f"[{rp_u}] SELL blocked by fire-sale guardrail. "
            f"Requested(Micro={accion_micro}, Macro={macro_action}) -> Final={accion_final}. "
            f"CapRel={_fmt_eur(cap_rel)}, Price={_fmt_eur(price)}, P&L={_fmt_eur(pnl)}. "
            f"Risk PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}, Secured={secured}. "
            f"MacroEvidence={'YES' if _safe_str(macro_rationale).strip() else 'NO'}."
        )
        return code, txt, governance

    # Casos por decisión
    if accion_final == "VENDER":
        if macro_action == "VENDER" and accion_micro != "VENDER":
            code = "RC01_MACRO_SELL_STEERING"
            txt = (
                f"[{rp_u}] SELL to reduce NPE and optimise capital. "
                f"Micro={accion_micro}, Macro={macro_action}. "
                f"CapRel={_fmt_eur(cap_rel)}, Price={_fmt_eur(price)}, P&L={_fmt_eur(pnl)}. "
                f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}, Secured={secured}."
            )
            if fire_sale and rp_l != "desinversion":
                txt += " (Flag: FireSale risk noted.)"
            return code, txt, governance

        code = "RC06_MICRO_SELL_VALUE_NEGATIVE"
        txt = (
            f"[{rp_u}] SELL on weak risk-adjusted economics. "
            f"EVA_pre={_fmt_eur(eva_pre)}, RORWA_pre={_pct(rorwa_pre)} vs hurdle={_pct(hurdle)}. "
            f"CapRel={_fmt_eur(cap_rel)}, Price={_fmt_eur(price)}, P&L={_fmt_eur(pnl)}. "
            f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
        )
        return code, txt, governance

    if accion_final == "REESTRUCTURAR":
        if not restruct_feasible:
            code = "RC09_RESTRUCT_REQUIRES_VIABILITY_CHECK"
            txt = (
                f"[{rp_u}] RESTRUCT selected but viability constraints fail or are incomplete. "
                f"ΔEVA={_fmt_eur(deva)}, EVA_post={_fmt_eur(eva_post)}; "
                f"DSCR={_fmt_metric(dscr)}, PTI={_fmt_metric(pti)}."
            )
            return code, txt, governance

        if macro_action == "REESTRUCTURAR" and accion_micro != "REESTRUCTURAR":
            code = "RC03_MACRO_RESTRUCT_STEERING"
            txt = (
                f"[{rp_u}] RESTRUCT to stabilise and recover value. "
                f"Micro={accion_micro}, Macro={macro_action}. "
                f"ΔEVA={_fmt_eur(deva)}, EVA_post={_fmt_eur(eva_post)}. "
                f"Affordability: DSCR={_fmt_metric(dscr)}, PTI={_fmt_metric(pti)}. "
                f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
            )
            return code, txt, governance

        code = "RC03_MICRO_RESTRUCT_VALUE_UPLIFT"
        txt = (
            f"[{rp_u}] RESTRUCT supported by value uplift. "
            f"ΔEVA={_fmt_eur(deva)}, EVA_post={_fmt_eur(eva_post)}; "
            f"DSCR={_fmt_metric(dscr)}, PTI={_fmt_metric(pti)}. "
            f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
        )
        return code, txt, governance

    # KEEP
    if micro_meets_hurdle:
        code = "RC05_KEEP_MEETS_HURDLE"
        txt = (
            f"[{rp_u}] KEEP: RORWA_pre={_pct(rorwa_pre)} ≥ hurdle={_pct(hurdle)}; "
            f"EVA_pre={_fmt_eur(eva_pre)}. PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
        )
        return code, txt, governance

    code = "RC05_KEEP_ACCEPTABLE_ECONOMICS"
    txt = (
        f"[{rp_u}] KEEP: economics acceptable and alternatives not compelling. "
        f"EVA_pre={_fmt_eur(eva_pre)}, ΔEVA={_fmt_eur(deva)}. "
        f"PD={_pct(pd_)}, LGD={_pct(lgd)}, DPD={dpd:.0f}."
    )
    return code, txt, governance


# ===========================================================
# 1) MICRO
# ===========================================================
def _run_micro_inference(
    model_path: str,
    portfolio_path: str,
    vecnorm_path: Optional[str],
    risk_posture: str,
    tag_suffix: str,
    device: str = "auto",
    seed: int = 42,
    deterministic: bool = True,
) -> Tuple[pd.DataFrame, str]:
    cfg_inf = InferenceConfig(
        model_path=model_path,
        portfolio_path=portfolio_path,
        device=device,
        seed=seed,
        deterministic=deterministic,
        tag=f"coord_micro_{tag_suffix}",
        vecnormalize_path=vecnorm_path,
        risk_posture=risk_posture,
    )

    logger.info(f"[MICRO] Lanzando inferencia (risk_posture={risk_posture})…")
    out_dir_micro = run_inference(cfg_inf)

    xlsx_path = os.path.join(out_dir_micro, "decisiones_explicadas.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"No se encontró decisiones_explicadas.xlsx en {out_dir_micro}")

    df_micro = pd.read_excel(xlsx_path)
    df_micro = _ensure_loan_id_column(df_micro)
    logger.info(f"[MICRO] Decisiones cargadas: {xlsx_path} ({len(df_micro):,} préstamos).")
    return df_micro, out_dir_micro


# ===========================================================
# 2) MACRO
# ===========================================================
def _run_macro_inference(
    model_path_portfolio: str,
    portfolio_path: str,
    risk_posture: str,
    loan_model_path: Optional[str],
    tag_suffix: str,
    n_steps: int,
    top_k: int,
    device: str = "auto",
    seed: int = 42,
    deterministic: bool = True,
    vecnormalize_portfolio_path: Optional[str] = None,
    vecnormalize_loan_path: Optional[str] = None,
) -> Tuple[pd.DataFrame, str]:
    cfg_port = PortfolioInferenceConfig(
        model_path=model_path_portfolio,
        portfolio_path=portfolio_path,
        device=device,
        seed=seed,
        deterministic=deterministic,
        tag=f"coord_portfolio_{tag_suffix}",
        n_steps=n_steps,
        top_k=top_k,
        save_final_excel=True,
        risk_posture=risk_posture,
        loan_model_path=loan_model_path,
        vecnormalize_portfolio_path=vecnormalize_portfolio_path,
        vecnormalize_loan_path=vecnormalize_loan_path,
    )

    logger.info(f"[MACRO] Lanzando inferencia cartera (risk_posture={risk_posture})…")
    out_dir_macro = run_portfolio_inference(cfg_port)

    traj_path = os.path.join(out_dir_macro, "trajectory_portfolio.csv")
    if not os.path.exists(traj_path):
        raise FileNotFoundError(f"No se encontró trajectory_portfolio.csv en {out_dir_macro}")

    df_steps = pd.read_csv(traj_path)
    logger.info(f"[MACRO] Trayectoria macro: {traj_path} ({len(df_steps):,} pasos).")
    return df_steps, out_dir_macro


def _split_ids_cell(v: Any) -> List[str]:
    if v is None:
        return []
    if isinstance(v, float) and np.isnan(v):
        return []
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return []
    parts = [p.strip() for p in s.split(";")]
    return [p for p in parts if p and p.lower() != "nan"]


def _build_macro_actions_per_loan(df_steps: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
    """
    Construye “macro assignment” desde trajectory_portfolio.csv.
    SOLO ids tocados (sell/restructure). El resto se considera NO_ASIGNADO.
    """
    macro_info: Dict[str, Dict[str, Any]] = {}
    if df_steps is None or df_steps.empty:
        return macro_info

    for _, row in df_steps.iterrows():
        step = int(_safe_float(row.get("step", 0), 0))
        rationale = _safe_str(row.get("rationale_fin", ""))
        combo = _safe_str(row.get("combination_note", ""))

        sold_ids = _split_ids_cell(row.get("sold_ids", ""))
        restruct_ids = _split_ids_cell(row.get("restructured_ids", ""))

        for lid in sold_ids:
            info = macro_info.setdefault(
                str(lid),
                {"macro_action": "VENDER", "steps_sold": [], "steps_restructured": [], "rationales": []},
            )
            info["macro_action"] = "VENDER"
            info["steps_sold"].append(step)
            if rationale or combo:
                info["rationales"].append(f"[step {step}] {rationale} | {combo}")

        for lid in restruct_ids:
            info = macro_info.setdefault(
                str(lid),
                {"macro_action": "REESTRUCTURAR", "steps_sold": [], "steps_restructured": [], "rationales": []},
            )
            if info.get("macro_action") != "VENDER":
                info["macro_action"] = "REESTRUCTURAR"
            info["steps_restructured"].append(step)
            if rationale or combo:
                info["rationales"].append(f"[step {step}] {rationale} | {combo}")

    return macro_info

# ===========================================================
# 2.5) KPI PORTFOLIO (Helper Task 4 - Robust)
# ===========================================================
def compute_portfolio_kpis(df: pd.DataFrame, action_col: str = "Accion_final", col_map: Dict[str, str] = None) -> Dict[str, Any]:
    """Calcula métricas agregadas del portafolio con búsqueda robusta de columnas."""
    if df.empty:
        return {}
    
    # 1. Action Column Resolution
    target_action = "" # default placeholder
    if action_col in df.columns:
        target_action = action_col
    else:
        # Fallback search
        for c in ["Accion_micro", "Accion", "action", "Action", "target_action"]:
            if c in df.columns:
                target_action = c
                break
    
    # 2. Robust Column Search
    def _get_col(candidates: List[str]) -> Optional[str]:
        # If we have an explicit mapping, check it first
        if col_map:
             # Just map commonly used keys to the forced column
             # e.g. col_map={"EVA": "EVA_post"}
             # We check if 'candidates' contains any key from col_map
             known_keys = ["EVA_post", "EVA", "RWA", "EAD", "capital"] # broad categories
             # But _get_col is called for specific metrics.
             pass

        for c in candidates:
            if c in df.columns:
                return c
        # Case insensitive check
        cols_lower = {c.lower(): c for c in df.columns}
        for c in candidates:
             if c.lower() in cols_lower:
                 return cols_lower[c.lower()]
        return None

    # Helper: resolve column name using map or fallbacks
    def _resolve(key: str, default_candidates: List[str]) -> Optional[str]:
        if col_map and key in col_map:
            # Respect explicit map even if column is missing (it will fail gracefully in _sum_safe or return None)
            # Actually, check if it exists to be safe, but prioritize map.
            mapped = col_map[key]
            if mapped in df.columns:
                return mapped
            # If mapped column missing, try fallbacks? Or strict failure?
            # Let's try fallbacks if mapped is missing, but log it? 
            # For robust audit, we stick to candidates if mapped fails.
        
        return _get_col(default_candidates)

    col_eva = _resolve("EVA", ["EVA_post", "EVA", "eva", "EVA_pre", "delta_eva", "EVA_gain"])
    col_rwa = _resolve("RWA", ["RWA_post", "RWA", "rwa", "RWA_pre", "RWA_std", "risk_weighted_assets"])
    col_ead = _resolve("EAD", ["EAD", "ead", "EAD_pre", "exposure", "exposure_at_default"])
    col_cap = _resolve("CAP", ["capital_release_realized", "capital_liberado", "capital_release", "release"])

    # 3. Safe Summation
    def _sum_safe(col_name: Optional[str]) -> float:
        if col_name:
            try:
                # Force numeric explicitly
                s = pd.to_numeric(df[col_name], errors='coerce').fillna(0.0)
                return float(s.sum())
            except Exception:
                return 0.0
        return 0.0

    total_eva = _sum_safe(col_eva)
    total_rwa = _sum_safe(col_rwa)
    total_ead = _sum_safe(col_ead)
    total_cap = _sum_safe(col_cap)
    
    # 4. Action Counts
    if target_action and target_action in df.columns:
        counts = df[target_action].value_counts().to_dict()
    else:
        counts = {"UNKNOWN": len(df)}

    # 5. Concentration (HHI)
    hhi_score = 0.0
    col_seg = _get_col(["segment", "segmento", "Segment", "sector"])
    if col_seg:
        shares = df[col_seg].value_counts(normalize=True)
        hhi_score = float((shares ** 2).sum())
    
    # Debug info only if needed
    meta_info = { 
        "cols_used": {
            "EVA": col_eva, "RWA": col_rwa, "EAD": col_ead, "Action": target_action
        }
    }
    
    return {
        "n_loans": int(len(df)),
        "total_eva": total_eva,
        "total_rwa": total_rwa,
        "total_ead": total_ead,
        "total_capital_release": total_cap,
        "action_counts": {str(k): int(v) for k, v in counts.items()},
        "hhi_concentration": hhi_score,
        "_meta": meta_info
    }


# ===========================================================
# 3) COMBINACIÓN MICRO + MACRO
# ===========================================================
def _combine_decisions(
    df_micro: pd.DataFrame,
    macro_actions: Dict[str, Dict[str, Any]],
    risk_posture: str,
    n_steps: int,
    top_k: int,
) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    df = _ensure_loan_id_column(df_micro).copy()

    # Preservar micro
    if "Accion" in df.columns:
        df["Accion_micro"] = df["Accion"]
    elif "action" in df.columns:
        df["Accion_micro"] = df["action"]
    else:
        df["Accion_micro"] = ""

    if "Explanation" in df.columns:
        df["Explanation_micro"] = df["Explanation"]
    else:
        df["Explanation_micro"] = ""

    n_rows = len(df)
    if n_rows == 0:
        logger.warning("[WARN] _combine_decisions(): df_micro vacío -> devolviendo vacío.")
        return df

    def _pad_or_trim(lst, fill):
        if len(lst) == n_rows:
            return lst
        if len(lst) == 0:
            logger.warning(f"[WARN] _combine_decisions(): lista vacía detectada -> rellenando con '{fill}'.")
            return [fill] * n_rows
        if len(lst) < n_rows:
            logger.warning(f"[WARN] _combine_decisions(): len(lista)={len(lst)} < n_rows={n_rows} -> padding con '{fill}'.")
            return lst + [fill] * (n_rows - len(lst))
        logger.warning(f"[WARN] _combine_decisions(): len(lista)={len(lst)} > n_rows={n_rows} -> truncando.")
        return lst[:n_rows]

    def _safe_bool_cell(v) -> bool:
        if v is None:
            return False
        try:
            if pd.isna(v):
                return False
        except Exception:
            pass
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, (int, np.integer)):
            return bool(int(v))
        s = str(v).strip().lower()
        if s in {"true", "t", "1", "yes", "y", "si", "sí"}:
            return True
        if s in {"false", "f", "0", "no", "n"}:
            return False
        return bool(v)

    # Outputs acumulados
    accion_final_list: List[str] = []
    accion_macro_list: List[str] = []
    macro_assignment_list: List[str] = []
    macro_selected_list: List[bool] = []
    convergencia_list: List[str] = []

    # 🆕 Task 4: Structured Override Log In-Memory
    override_log_entries: List[Dict[str, Any]] = []
    override_level_list: List[str] = []
    override_from_list: List[str] = []
    override_to_list: List[str] = []
    
    # 🆕 Task 2: Propagar “macro flags”
    macro_rationales_list: List[str] = []
    macro_steps_sold_list: List[str] = []
    macro_steps_restructured_list: List[str] = []
    
    # 🆕 Task 2 Audit: Explicit Macro Flags
    macro_action_used_list: List[str] = []
    macro_conflict_list: List[bool] = []
    macro_applied_list: List[bool] = []

    explanation_micro_list: List[str] = []
    explanation_macro_list: List[str] = []
    explanation_final_list: List[str] = []

    reason_code_list: List[str] = []
    rationale_cib_list: List[str] = []
    macro_evidence_list: List[str] = []
    governance_list: List[str] = []

    # Métricas persistidas
    fire_sale_list: List[bool] = []
    price_to_ead_list: List[float] = []
    price_ratio_ead_list: List[float] = []
    pnl_ratio_book_list: List[float] = []
    book_value_list: List[float] = []
    pnl_book_list: List[float] = []
    fire_sale_threshold_list: List[float] = []
    fire_sale_enabled_list: List[bool] = []
    fire_sale_triggered_list: List[bool] = []
    fire_sale_triggers_list: List[str] = []
    price_ratio_book_list: List[float] = []
    fire_sale_threshold_book_list: List[float] = []

    constraint_blocked_list: List[bool] = []
    constraint_reason_list: List[str] = []
    constraint_metrics_json_list: List[str] = []

    missing_viability_list: List[bool] = []
    restruct_viable_list: List[bool] = []  # útil para hard-guardrails posteriores

    # ✅ BANK-READY: Parámetros de reestructura (propagar desde micro)
    plazo_optimo_list: List[float] = []
    tasa_nueva_list: List[float] = []
    quita_list: List[float] = []

    g = _get_guardrails(risk_posture)
    hurdle = _safe_float(getattr(cfg.CONFIG.regulacion, "hurdle_rate", 0.0))
    rp_l = (risk_posture or "balanceado").strip().lower()

    for _, r in df.iterrows():
        loan_id = str(r.get("loan_id", _get_loan_id_from_row(r)) or "")
        accion_micro = _normalize_action(r.get("Accion_micro", "")) or "MANTENER"

        macro_info = macro_actions.get(loan_id)
        macro_selected = bool(macro_info is not None)

        # Macro assignment (solo si tocado)
        macro_steps_sold = ""
        macro_steps_restruct = ""
        
        if not macro_selected:
            macro_assignment = "NO_ASIGNADO"
            macro_action_used = accion_micro  # neutral para no inventar steering
            macro_rationale = (
                f"Macro not applied (n_steps={n_steps}, top_k={top_k}) ⇒ "
                f"macro_action defaults to micro action (no portfolio steering)."
            )
        else:
            macro_assignment = _normalize_action(macro_info.get("macro_action", "")) or "MANTENER"
            macro_action_used = macro_assignment
            
            # Rationales
            raw_rationales = macro_info.get("rationales", []) or []
            if raw_rationales:
                macro_rationale = " || ".join(raw_rationales)
                if len(macro_rationale) > 500: # Capar longitud
                    macro_rationale = macro_rationale[:497] + "..."
            else:
                macro_rationale = "Macro selected (no rationale text)."
            
            # Steps
            ss = macro_info.get("steps_sold", [])
            sr = macro_info.get("steps_restructured", [])
            if ss:
                macro_steps_sold = ",".join(map(str, ss))
            if sr:
                macro_steps_restruct = ",".join(map(str, sr))

        # Inputs base
        ead = _safe_float(r.get("EAD", 0.0))
        rw = _safe_float(r.get("RW", np.nan), np.nan)

        price = _safe_float(r.get("precio_optimo", r.get("price", np.nan)), np.nan)
        pnl = _safe_float(r.get("pnl", r.get("PnL", np.nan)), np.nan)
        cap_rel = _safe_float(r.get("capital_liberado", r.get("capital_release", np.nan)), np.nan)
        deva = _safe_float(r.get("ΔEVA", r.get("delta_eva", r.get("EVA_gain", np.nan))), np.nan)
        rorwa_pre = _safe_float(r.get("RORWA_pre", r.get("RORWA", np.nan)), np.nan)

        # Hurdle test
        micro_meets_hurdle = False
        if hurdle > 0 and not np.isnan(rorwa_pre):
            try:
                micro_meets_hurdle = (float(rorwa_pre) >= float(hurdle))
            except Exception:
                micro_meets_hurdle = False

        # Viabilidad inputs (preferimos *_pre)
        pti = r.get("PTI_post", r.get("PTI", None))
        dscr = r.get("DSCR_post", r.get("DSCR", None))

        pti_pre = r.get("PTI_pre", r.get("PTI_post", r.get("PTI", None)))
        dscr_pre = r.get("DSCR_pre", r.get("DSCR_post", r.get("DSCR", None)))

        if isinstance(pti_pre, float) and np.isnan(pti_pre):
            pti_pre = None
        if isinstance(dscr_pre, float) and np.isnan(dscr_pre):
            dscr_pre = None

        # Feasible restructure (numérica)
        dscr_min = float(g["DSCR_MIN"])
        pti_max = g.get("PTI_MAX", None)
        restruct_feasible = True

        dscr_gate_allowed = True
        dscr_post_gate = dscr
        dscr_gate_reason = ""
        gate_income = r.get("monthly_cfo", r.get("cashflow_operativo_mensual", None))
        gate_payment = r.get("cuota_mensual_post", r.get("monthly_payment", r.get("cuota_mensual", None)))
        if gate_income is None or gate_payment is None:
            if dscr is not None and not (isinstance(dscr, float) and np.isnan(dscr)):
                gate_income = dscr
                gate_payment = 1.0

        if gate_income is not None and gate_payment is not None:
            dscr_gate_allowed, dscr_post_gate, dscr_gate_reason = check_restruct_viability(
                current_income=gate_income,
                new_payment=gate_payment,
                dscr_min=dscr_min,
            )
            if np.isfinite(dscr_post_gate):
                dscr = dscr_post_gate
            if not dscr_gate_allowed:
                restruct_feasible &= False

        if pti_max is not None and pti is not None and not (isinstance(pti, float) and np.isnan(pti)):
            try:
                restruct_feasible &= (float(pti) <= float(pti_max))
            except Exception:
                pass

        if not np.isnan(deva):
            restruct_feasible &= (deva >= float(g["MIN_DEVA_RESTRUCT"]))
        else:
            restruct_feasible &= False  # sin ΔEVA, no confirmamos uplift

        # RC10 missing viability (segment-aware)
        seg_hint = _safe_str(r.get("segmento_banco", r.get("segment_raw", r.get("segment", ""))), "").strip().lower()
        needs_pti = any(k in seg_hint for k in ("mortgage", "hipotec", "retail", "minorista", "consumer", "consumo"))
        needs_dscr = not needs_pti

        missing_pti = _is_missing(pti_pre)
        missing_dscr = _is_missing(dscr_pre)

        missing_viability_inputs = (needs_pti and missing_pti) or (needs_dscr and missing_dscr)
        if missing_viability_inputs:
            restruct_feasible = False

        # Fire-sale robusto (preferir simulador micro; fallback mínimo)
        price_ratio_ead = _safe_float(r.get("price_ratio_ead", np.nan), np.nan)
        if np.isnan(price_ratio_ead):
            if (ead > 0) and (not np.isnan(price)):
                price_ratio_ead = float(price / ead)

        book_value = _safe_float(r.get("book_value", np.nan), np.nan)
        pnl_book = _safe_float(r.get("pnl_book", np.nan), np.nan)
        ratio_book = _safe_float(r.get("price_ratio_book", np.nan), np.nan)
        thr_book = _safe_float(r.get("fire_sale_threshold_book", np.nan), np.nan)

        lgd = _safe_float(r.get("LGD", np.nan), np.nan)

        if np.isnan(book_value):
            cov = _safe_float(r.get("coverage_rate", np.nan), np.nan)
            if np.isfinite(cov) and cov > 1.0:
                cov = cov / 100.0
            if np.isfinite(cov) and (ead > 0):
                cov = float(np.clip(cov, 0.0, 1.0))
                book_value = ead * (1.0 - cov)
            elif (ead > 0) and (not np.isnan(lgd)):
                book_value = ead * (1.0 - float(np.clip(lgd, 0.0, 1.0)))

        if np.isnan(pnl_book):
            if (np.isfinite(price) and np.isfinite(book_value) and book_value > 0):
                pnl_book = float(price - book_value)
            else:
                pnl_book = pnl

        if np.isnan(ratio_book) and np.isfinite(price) and np.isfinite(book_value) and book_value > 0:
            ratio_book = float(price / book_value)

        pnl_ratio_book = np.nan
        if np.isfinite(book_value) and book_value > 0 and np.isfinite(pnl_book):
            pnl_ratio_book = float(pnl_book / book_value)

        thr_ead = _safe_float(r.get("fire_sale_threshold", np.nan), np.nan)
        min_price_ratio_ead = float(g["MIN_PRICE_TO_EAD"])
        if np.isfinite(thr_ead):
            min_price_ratio_ead = max(min_price_ratio_ead, float(thr_ead))

        fire_sale_sim = _safe_bool_cell(r.get("fire_sale", r.get("Fire_Sale", False)))

        allow_fire_sale = (rp_l == "desinversion")
        fire_sale_enabled = (rp_l in ("prudencial", "balanceado"))
        fire_sale_triggered = False
        triggers: List[str] = []

        if fire_sale_enabled:
            if fire_sale_sim:
                fire_sale_triggered = True
                triggers.append("SIM_FLAG")

            if np.isfinite(thr_book) and np.isfinite(price) and np.isfinite(book_value) and book_value > 0:
                _, ratio_gate, fire_sale_gate, _ = check_sell_fire_sale(
                    price_neto=price,
                    book_value=book_value,
                    allow_fire_sale=allow_fire_sale,
                    thr_book=float(thr_book),
                )
                if np.isfinite(ratio_gate):
                    ratio_book = float(ratio_gate)
                if fire_sale_gate:
                    fire_sale_triggered = True
                    triggers.append("PX_BOOK")
            else:
                if (not np.isnan(price_ratio_ead)) and (price_ratio_ead < float(min_price_ratio_ead)):
                    fire_sale_triggered = True
                    triggers.append("PX_EAD")

            max_pnl_ratio_book = float(g.get("MAX_FIRE_SALE_PNL_RATIO_BOOK", -0.75))
            if (not np.isnan(pnl_ratio_book)) and (pnl_ratio_book < max_pnl_ratio_book):
                fire_sale_triggered = True
                triggers.append("PNL_BOOK_RATIO")

            max_pnl_abs = float(g.get("MAX_FIRE_SALE_PNL_ABS", -50_000.0))
            if (not np.isnan(pnl_book)) and (pnl_book <= max_pnl_abs):
                fire_sale_triggered = True
                triggers.append("PNL_ABS")

        fire_sale = bool(fire_sale_triggered)
        price_to_ead = (price / ead) if (ead > 0 and not np.isnan(price)) else np.nan

        constraint_blocked = False
        constraint_reason = ""
        constraint_metrics_json = ""

        # -----------------------------------------------------------
        # COORDINACIÓN MICRO vs MACRO (Prioridad configurable)
        # -----------------------------------------------------------
        coord_priority = g.get("COORDINATOR_PRIORITY", "PRUDENCIAL_FIRST")
        risk_map = {"MANTENER": 0, "REESTRUCTURAR": 1, "VENDER": 2}

        # 1. Punto de partida: Micro (Bottom-Up)
        accion_candidate = accion_micro
        proposed_action = accion_micro
        
        override_active = False
        override_level = ""
        override_from_val = ""
        override_to_val = ""

        if macro_selected:
            if accion_micro == macro_action_used:
                # Coinciden: no conflicto
                accion_candidate = accion_micro
            else:
                # Conflicto: Aplicar regla de desempate
                apply_macro = False
                
                if coord_priority == "MACRO_FIRST":
                    apply_macro = True
                
                elif coord_priority == "MICRO_FIRST":
                    apply_macro = False
                
                elif coord_priority == "PRUDENCIAL_FIRST":
                    # Gana la opción de MENOR riesgo (MANTENER < REESTRUCTURAR < VENDER)
                    r_micro = risk_map.get(accion_micro, 0)
                    r_macro = risk_map.get(macro_action_used, 0)
                    if r_macro < r_micro:
                        apply_macro = True
                        override_level = "MACRO_PRUDENTIAL"
                    else:
                        apply_macro = False # Micro es más prudente o igual

                if apply_macro:
                    accion_candidate = macro_action_used
                    override_active = True
                    if not override_level: override_level = "MACRO_STRATEGY"
                    override_from_val = accion_micro
                    override_to_val = macro_action_used
            
            proposed_action = accion_candidate

            # ---------------------------------------
            # Guardrails TÁCTICOS (Conflict resolution safety)
            # ---------------------------------------
            # 1. Fire Sale Block (Prudencial/Balanceado): Macro no puede forzar venta si es Fire Sale
            if accion_candidate == "VENDER" and fire_sale and (rp_l in ("prudencial", "balanceado")):
                # Revertir a Micro si es más segura, o fallback MANTENER
                old_cand = accion_candidate
                accion_candidate = "MANTENER" # Fallback safe
                
                # Si micro era REESTRUCTURAR y es viable, usarla
                if accion_micro == "REESTRUCTURAR" and restruct_feasible:
                    accion_candidate = "REESTRUCTURAR"
                
                if old_cand != accion_candidate:
                    override_active = True # Trigger override if not already (or modify existing)
                    override_level = "GUARDRAIL_FIRE_SALE"
                    override_from_val = old_cand
                    override_to_val = accion_candidate

            # 2. Feasibility Block: Macro no puede forzar REESTRUCTURAR si no es viable
            if accion_candidate == "REESTRUCTURAR" and not restruct_feasible:
                old_cand = accion_candidate
                accion_candidate = "MANTENER"
                
                if old_cand != accion_candidate:
                    override_active = True
                    override_level = "GUARDRAIL_FEASIBILITY"
                    override_from_val = old_cand
                    override_to_val = accion_candidate

        # RC10 (Legacy check): Inputs faltantes
        if accion_candidate == "REESTRUCTURAR" and missing_viability_inputs:
             old_cand = accion_candidate
             accion_candidate = "MANTENER"
             if old_cand != accion_candidate: # If it changed
                 override_active = True
                 override_level = "GUARDRAIL_FEASIBILITY"
                 override_from_val = old_cand
                 override_to_val = accion_candidate
        
        # Fire Sale Check final (por si acaso quedó VENDER)
        blocked_sell_fire_sale = False
        if (accion_candidate == "VENDER") and (rp_l in ("prudencial", "balanceado")) and fire_sale:
            blocked_sell_fire_sale = True
            old_cand = accion_candidate
            accion_candidate = "MANTENER"
            if old_cand != accion_candidate:
                override_active = True
                override_level = "GUARDRAIL_FIRE_SALE" 
                override_from_val = old_cand
                override_to_val = accion_candidate

        accion_final = accion_candidate
        
        # Audit Flags
        macro_conflict = False
        macro_applied_flag = False
        
        if macro_selected:
            if accion_micro != macro_action_used:
                macro_conflict = True
                # Applied if final result matches macro
                if accion_final == macro_action_used:
                    macro_applied_flag = True

        # Registrar Override (Audit Grade)
        if override_active or (override_level and override_level != ""):
             # Shorten rationales
             rat_short = macro_rationale if macro_selected else ""
             if len(rat_short) > 200: 
                 rat_short = rat_short[:197] + "..."
                 
             override_log_entries.append({
                 "loan_id": loan_id,
                 "level": override_level,
                 "from_action": override_from_val,
                 "to_action": override_to_val,
                 "portfolio_context": rp_l,
                 "posture": risk_posture,
                 "run_id": datetime.now().strftime("%Y%m%d_%H%M%S"), # Timestamp as run_id
                 "macro_action_used": macro_action_used if macro_selected else "",
                 "macro_rationales_short": rat_short,
                 "pti_actual": f"{pti_pre:.2f}" if pti_pre is not None else "",
                 "dscr_actual": f"{dscr_pre:.2f}" if dscr_pre is not None else "",
                 "pnl": f"{pnl:.0f}" if np.isfinite(pnl) else "",
                 # "capital_release_net" # Calculated later or approximate
             })
        
        override_level_list.append(override_level)
        override_from_list.append(override_from_val)
        override_to_list.append(override_to_val)
        
        # New lists populate
        macro_action_used_list.append(macro_action_used if macro_selected else "")
        macro_conflict_list.append(macro_conflict)
        macro_applied_list.append(macro_applied_flag)

        if blocked_sell_fire_sale:
            constraint_blocked = True
            constraint_reason = "SELL_FIRE_SALE_BLOCKED"
            constraint_metrics_json = json.dumps(
                {
                    "price_neto": float(price) if np.isfinite(price) else None,
                    "book_value": float(book_value) if np.isfinite(book_value) else None,
                    "price_book_ratio": float(ratio_book) if np.isfinite(ratio_book) else None,
                    "thr_book": float(thr_book) if np.isfinite(thr_book) else None,
                    "allow_fire_sale": bool(allow_fire_sale),
                },
                separators=(",", ":"),
            )
        elif (not dscr_gate_allowed) and dscr_gate_reason == "DSCR_BELOW_MIN":

            if (accion_micro == "REESTRUCTURAR") or (macro_action_used == "REESTRUCTURAR"):
                constraint_blocked = True
                constraint_reason = "RESTRUCT_DSCR_BELOW_MIN"
                constraint_metrics_json = json.dumps(
                    {
                        "dscr_post": float(dscr_post_gate) if np.isfinite(dscr_post_gate) else None,
                        "dscr_min": float(dscr_min),
                        "new_payment": float(gate_payment) if gate_payment is not None else None,
                        "current_income": float(gate_income) if gate_income is not None else None,
                    },
                    separators=(",", ":"),
                )

        if constraint_blocked:
            logger.info(
                f"[CONSTRAINT] loan_id={loan_id} reason={constraint_reason} metrics={constraint_metrics_json}"
            )

        flags = {
            "restruct_feasible": restruct_feasible,
            "fire_sale": fire_sale,
            "micro_meets_hurdle": micro_meets_hurdle,
            "sell_requested": (accion_micro == "VENDER") or (macro_action_used == "VENDER"),
            "blocked_sell_fire_sale": blocked_sell_fire_sale,
            "missing_viability_inputs": bool(missing_viability_inputs),
            "proposed_action": proposed_action,
        }

        # Convergencia (audit)
        if not macro_selected:
            convergencia = "MACRO_NOT_APPLIED"
        else:
            if accion_micro == macro_action_used:
                convergencia = "AGREE_MICRO_MACRO"
            else:
                convergencia = "MACRO_WINS" if accion_final == macro_action_used else "GUARDRAIL_OVERRIDE"

        # Evidence micro (si viene vacío, lo construimos)
        exp_micro_raw = _safe_str(r.get("Explanation_micro", ""))
        steps_micro = _safe_str(r.get("Explain_Steps", ""))
        micro_detail = exp_micro_raw.strip() or steps_micro.strip()

        if (not micro_detail) or (micro_detail == "❓"):
            pead_txt = "NA"
            try:
                if (ead > 0) and (not np.isnan(price_to_ead)):
                    pead_txt = f"{float(price_to_ead):.3f}"
            except Exception:
                pead_txt = "NA"

            micro_detail = (
                f"loan_id={loan_id}, EAD={_fmt_eur(ead)}, RW={rw}, RORWA_pre={_pct(rorwa_pre) if not np.isnan(rorwa_pre) else 'NA'}, "
                f"ΔEVA={_fmt_eur(deva) if not np.isnan(deva) else 'NA'}, "
                f"CapRel={_fmt_eur(cap_rel) if not np.isnan(cap_rel) else 'NA'}, Price/EAD={pead_txt}, "
                f"P&L={_fmt_eur(pnl) if not np.isnan(pnl) else 'NA'}, DSCR={dscr}, PTI={pti}, "
                f"FireSale={fire_sale}, RestructFeasible={restruct_feasible}."
            )

        # Reasoning auditable
        reason_code, rationale_cib, governance = _cib_rationale_row(
            r=r,
            risk_posture=risk_posture,
            accion_micro=accion_micro,
            macro_action=macro_action_used,
            accion_final=accion_final,
            macro_rationale=(macro_rationale if macro_selected else ""),
            flags=flags,
        )

        explanation_micro = f"Micro={accion_micro}. Evidence: {micro_detail}"
        explanation_macro = f"Macro={macro_action_used}. Evidence: {macro_rationale}"
        explanation_final = f"{rationale_cib} | Governance={governance} | Convergencia={convergencia}"

        # Append
        macro_rationales_list.append(macro_rationale)
        macro_steps_sold_list.append(macro_steps_sold)
        macro_steps_restructured_list.append(macro_steps_restruct)

        accion_macro_list.append(macro_action_used)
        macro_assignment_list.append(macro_assignment)
        macro_selected_list.append(bool(macro_selected))
        convergencia_list.append(convergencia)
        accion_final_list.append(accion_final)

        explanation_micro_list.append(explanation_micro)
        explanation_macro_list.append(explanation_macro)
        explanation_final_list.append(explanation_final)

        reason_code_list.append(reason_code)
        rationale_cib_list.append(rationale_cib)
        macro_evidence_list.append(macro_rationale if macro_selected else "")
        governance_list.append(governance)

        fire_sale_list.append(bool(fire_sale))
        fire_sale_threshold_list.append(float(min_price_ratio_ead))
        fire_sale_threshold_book_list.append(float(thr_book) if np.isfinite(thr_book) else np.nan)

        price_to_ead_list.append(float(price_to_ead) if not np.isnan(price_to_ead) else np.nan)
        price_ratio_ead_list.append(float(price_ratio_ead) if not np.isnan(price_ratio_ead) else np.nan)

        pnl_ratio_book_list.append(float(pnl_ratio_book) if not np.isnan(pnl_ratio_book) else np.nan)
        book_value_list.append(float(book_value) if not np.isnan(book_value) else np.nan)
        pnl_book_list.append(float(pnl_book) if not np.isnan(pnl_book) else np.nan)

        fire_sale_enabled_list.append(bool(fire_sale_enabled))
        fire_sale_triggered_list.append(bool(fire_sale_triggered))
        fire_sale_triggers_list.append(";".join(triggers) if triggers else "")

        price_ratio_book_list.append(float(ratio_book) if np.isfinite(ratio_book) else np.nan)
        missing_viability_list.append(bool(missing_viability_inputs))
        restruct_viable_list.append(bool(restruct_feasible))

        constraint_blocked_list.append(bool(constraint_blocked))
        constraint_reason_list.append(str(constraint_reason))
        constraint_metrics_json_list.append(str(constraint_metrics_json))

        # ✅ BANK-READY: Parámetros de reestructura (desde micro)
        plazo_optimo_list.append(_safe_float(r.get("plazo_optimo", np.nan), np.nan))
        tasa_nueva_list.append(_safe_float(r.get("tasa_nueva", np.nan), np.nan))
        quita_list.append(_safe_float(r.get("quita", np.nan), np.nan))

    # saneo final de longitudes
    accion_macro_list = _pad_or_trim(accion_macro_list, "MANTENER")
    macro_assignment_list = _pad_or_trim(macro_assignment_list, "NO_ASIGNADO")
    macro_selected_list = _pad_or_trim(macro_selected_list, False)
    convergencia_list = _pad_or_trim(convergencia_list, "MACRO_NOT_APPLIED")
    accion_final_list = _pad_or_trim(accion_final_list, "MANTENER")

    explanation_micro_list = _pad_or_trim(explanation_micro_list, "Micro=NA")
    explanation_macro_list = _pad_or_trim(explanation_macro_list, "Macro=NA")
    explanation_final_list = _pad_or_trim(explanation_final_list, "Fallback")

    reason_code_list = _pad_or_trim(reason_code_list, "RC00_FALLBACK")
    rationale_cib_list = _pad_or_trim(rationale_cib_list, "Fallback")
    macro_evidence_list = _pad_or_trim(macro_evidence_list, "")
    governance_list = _pad_or_trim(governance_list, "Fallback")

    fire_sale_list = _pad_or_trim(fire_sale_list, False)
    price_to_ead_list = _pad_or_trim(price_to_ead_list, np.nan)
    price_ratio_ead_list = _pad_or_trim(price_ratio_ead_list, np.nan)
    pnl_ratio_book_list = _pad_or_trim(pnl_ratio_book_list, np.nan)
    book_value_list = _pad_or_trim(book_value_list, np.nan)
    pnl_book_list = _pad_or_trim(pnl_book_list, np.nan)
    price_ratio_book_list = _pad_or_trim(price_ratio_book_list, np.nan)
    fire_sale_threshold_book_list = _pad_or_trim(fire_sale_threshold_book_list, np.nan)

    fire_sale_threshold_list = _pad_or_trim(fire_sale_threshold_list, np.nan)
    fire_sale_enabled_list = _pad_or_trim(fire_sale_enabled_list, bool(rp_l in ("prudencial", "balanceado")))
    fire_sale_triggered_list = _pad_or_trim(fire_sale_triggered_list, False)
    fire_sale_triggers_list = _pad_or_trim(fire_sale_triggers_list, "")
    missing_viability_list = _pad_or_trim(missing_viability_list, False)
    restruct_viable_list = _pad_or_trim(restruct_viable_list, False)

    constraint_blocked_list = _pad_or_trim(constraint_blocked_list, False)
    constraint_reason_list = _pad_or_trim(constraint_reason_list, "")
    constraint_metrics_json_list = _pad_or_trim(constraint_metrics_json_list, "")

    # ✅ BANK-READY: Saneo parámetros de reestructura
    plazo_optimo_list = _pad_or_trim(plazo_optimo_list, np.nan)
    tasa_nueva_list = _pad_or_trim(tasa_nueva_list, np.nan)
    quita_list = _pad_or_trim(quita_list, np.nan)

    # 🆕 Task 2: Propagar macro columns (listas)
    macro_rationales_list = _pad_or_trim(macro_rationales_list, "")
    macro_steps_sold_list = _pad_or_trim(macro_steps_sold_list, "")
    macro_steps_restructured_list = _pad_or_trim(macro_steps_restructured_list, "")
    
    # 🆕 Task 2 Audit: Explicit Macro Flags
    macro_action_used_list = _pad_or_trim(macro_action_used_list, "")
    macro_conflict_list = _pad_or_trim(macro_conflict_list, False)
    macro_applied_list = _pad_or_trim(macro_applied_list, False)

    # Columnas finales
    df["Accion_macro"] = accion_macro_list
    df["Macro_Assignment"] = macro_assignment_list  # <- lo que “asignó” macro, solo si tocó
    df["Macro_Selected"] = macro_selected_list
    df["Convergencia_Caso"] = convergencia_list
    
    # 🆕 Task 2 Audit: Persistir en DF con naming consistente
    df["macro_action_used"] = macro_action_used_list
    df["macro_conflict"] = macro_conflict_list
    df["macro_applied"] = macro_applied_list
    df["Macro_Rationales"] = macro_rationales_list  # Standard naming requested
    
    # 🆕 Task 2: Persistir en DF legacy columns kept for backward compatibility if needed
    df["Macro_Rationales"] = macro_rationales_list
    df["Macro_Steps_Sold"] = macro_steps_sold_list
    df["Macro_Steps_Restruct"] = macro_steps_restructured_list

    df["Accion_final"] = accion_final_list
    
    # 🆕 Task 1: Boolean Columns + Defaults (para no tener NaNs)
    df["override_applied"] = [bool(lvl != "") for lvl in override_level_list]
    df["override_level"] = override_level_list
    df["override_from"] = override_from_list
    df["override_to"] = override_to_list
    # Fill defaults en caso de inconsistencia
    df["override_level"] = df["override_level"].fillna("")
    df["override_from"] = df["override_from"].fillna("")
    df["override_to"] = df["override_to"].fillna("")

    df["Accion"] = df["Accion_final"]  # legacy
    # ---- Canonical mirror (audit/KPIs): decision_final SIEMPRE = Accion_final
    df["decision_final"] = df["Accion_final"]

    # Si existen aliases legacy, mantenlos coherentes (no creamos columnas nuevas salvo decision_final)
    for c in ("Decision_Final", "Final_Decision", "accion_final", "Accion_Final", "Decision"):
        if c in df.columns:
            df[c] = df["Accion_final"]

    # Rename internal legacy or ensure consistent
    df["Explanation_micro"] = explanation_micro_list
    df["Explanation_macro"] = explanation_macro_list
    df["Explanation_final"] = explanation_final_list
    df["Explanation"] = df["Explanation_final"]  # legacy

    df["Reason_Code"] = reason_code_list
    df["Rationale_CIB"] = rationale_cib_list
    # df["Macro_Evidence"] = macro_evidence_list # Use Macro_Rationales instead as primary
    df["Macro_Evidence"] = df["Macro_Rationales"] # Alias for backward compatibility
    
    df["Decision_Governance"] = governance_list

    df["Fire_Sale"] = fire_sale_list
    df["Price_to_EAD"] = price_to_ead_list
    df["price_ratio_ead"] = price_ratio_ead_list
    df["pnl_ratio_book"] = pnl_ratio_book_list
    df["book_value"] = book_value_list
    df["pnl_book"] = pnl_book_list
    df["price_ratio_book"] = price_ratio_book_list
    df["fire_sale_threshold_book"] = fire_sale_threshold_book_list
    df["fire_sale_threshold"] = fire_sale_threshold_list
    df["FireSale_Enabled"] = fire_sale_enabled_list
    df["FireSale_Triggered"] = fire_sale_triggered_list
    df["FireSale_Triggers"] = fire_sale_triggers_list
    df["Missing_Viability_Inputs"] = missing_viability_list
    df["restruct_viable"] = restruct_viable_list  # para hard guardrails y auditoría

    df["constraint_blocked"] = constraint_blocked_list
    df["constraint_reason"] = constraint_reason_list
    df["constraint_metrics_json"] = constraint_metrics_json_list

    # ✅ BANK-READY: Parámetros de reestructura
    df["plazo_optimo"] = plazo_optimo_list
    df["tasa_nueva"] = tasa_nueva_list
    df["quita"] = quita_list

    # ✅ BANK-READY: Campos de escalación y comité (inicialización)
    df["case_status"] = "NORMAL"  # NORMAL | HOLD_NO_EXECUTABLE_ACTION | ESCALATED
    df["next_step"] = ""  # WORKOUT_REVIEW | LEGAL_ACTION | PRICING_REVIEW
    df["next_step_reason"] = ""  # motivo específico de escalación
    df["review_due_days"] = np.nan  # días para revisión (ej. 30)
    df["required_data_flags"] = ""  # lista de datos faltantes para acción
    df["override_reason"] = ""  # razón de override si aplica

    # Reorden visual (enforce_schema reordenará canónicas primero)
    preferred = [
        "loan_id", "segment", "EAD", "RW",
        "Accion_micro", "Accion_final", 
        "macro_action_used", "macro_conflict", "macro_applied", "Macro_Rationales",
        "Accion_macro", "Macro_Assignment", "Macro_Selected",
        "Convergencia_Caso",
        "Reason_Code", "Decision_Governance",
        "Missing_Viability_Inputs", "restruct_viable",
        "constraint_blocked", "constraint_reason", "constraint_metrics_json",
        "plazo_optimo", "tasa_nueva", "quita",  # parámetros de reestructura
        "PTI_pre", "PTI_post", "DSCR_pre", "DSCR_post",
        "case_status", "next_step", "next_step_reason", "review_due_days",
        "required_data_flags", "override_reason",
        "Rationale_CIB", "Macro_Evidence",
        "Fire_Sale", "FireSale_Enabled", "FireSale_Triggered", "FireSale_Triggers",
        "Price_to_EAD", "fire_sale_threshold",
        "Explanation_micro", "Explanation_macro", "Explanation_final",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    df = df[cols]

    # ===========================================================
    # Contrafactual vs Realizado (coherencia para summary)
    # ===========================================================
    ratio_total = _get_ratio_total_capital()

    if "capital_release_cf" not in df.columns:
        df["capital_release_cf"] = np.nan

    if "RWA_pre" in df.columns:
        rwa_pre = pd.to_numeric(df["RWA_pre"], errors="coerce")
    else:
        ead_s = pd.to_numeric(df.get("EAD", np.nan), errors="coerce")
        rw_s = pd.to_numeric(df.get("RW", np.nan), errors="coerce")
        rwa_pre = ead_s * rw_s

    df["capital_release_cf"] = pd.to_numeric(df["capital_release_cf"], errors="coerce")
    df.loc[df["capital_release_cf"].isna(), "capital_release_cf"] = (rwa_pre * float(ratio_total))

    # ===========================================================
    # 🆕 EJECUTABILIDAD: RECOMENDACIÓN vs EJECUCIÓN
    # ===========================================================
    # Separación clara entre:
    # - RECOMENDACIÓN: Lo que el modelo/analista recomienda (exported siempre)
    # - EJECUCIÓN: Lo que es realmente ejecutable hoy (gates, mandatos, fallbacks)
    # ===========================================================

    logger.info("[EXECUTABILITY] Aplicando gates de ejecutabilidad (RECOMENDACIÓN vs EJECUCIÓN)...")

    # Cargar BankStrategy con knobs de ejecutabilidad
    from config import BANK_STRATEGIES, BankProfile
    
    posture_map = {
        "prudencial": BankProfile.PRUDENTE,
        "balanceado": BankProfile.BALANCEADO,
        "desinversion": BankProfile.DESINVERSION,
    }
    posture_key = (risk_posture or "balanceado").lower().strip()
    bank_profile = posture_map.get(posture_key, BankProfile.BALANCEADO)
    strategy = BANK_STRATEGIES[bank_profile]
    
    # Knobs ejecutabilidad
    sale_floor_ratio = strategy.sale_floor_ratio
    loss_cap_pct = strategy.loss_cap_pct
    min_acceptance_score = strategy.min_acceptance_score
    max_restructure_share = strategy.max_restructure_share
    # 🆕 PC7 SECOND PASS - Mandatos con tiering + percentiles (NO thresholds absolutos)
    mandate_share_target = strategy.mandate_share_target  # % objetivo (0.25 DESINV)
    mandate_tier1_share = strategy.mandate_tier1_share  # % genuinamente obligatorio (0.05)
    mandate_w_rwa = strategy.mandate_w_rwa  # peso RWA en score
    mandate_w_age = strategy.mandate_w_age  # peso age_npl en score
    mandate_w_recovery = strategy.mandate_w_recovery  # peso (1-recovery) en score
    mandate_loss_tolerance = strategy.mandate_loss_tolerance
    # 🆕 NPL BANK-GRADE - Disciplina económica
    recovery_min_pct = strategy.recovery_min_pct  # recovery mínimo para venta voluntaria
    max_sell_share = strategy.max_sell_share  # cap de ventas VOLUNTARIAS (mandatos exentos)
    
    logger.info(f"[KNOBS] {posture_key.upper()}: sale_floor={sale_floor_ratio:.2f}, "
                f"loss_cap={loss_cap_pct:.2f}, recovery_min={recovery_min_pct:.2%}, "
                f"max_sell={max_sell_share:.0%}, min_accept={min_acceptance_score:.0f}, "
                f"max_restruct={max_restructure_share:.0%}")
    logger.info(f"[MANDATES] target={mandate_share_target:.1%}, tier1={mandate_tier1_share:.1%}, "
                f"weights=(rwa={mandate_w_rwa:.1f}, age={mandate_w_age:.1f}, recovery={mandate_w_recovery:.1f})")

    # -----------------------------------------------------------
    # 1) RECOMENDACIÓN (ya existe en Accion_micro/Accion_macro)
    # -----------------------------------------------------------
    # recommended_action = lo que el análisis técnico recomienda (sin gates)
    # Usamos la decisión "antes de guardrails" que es Accion_final actual
    df["recommended_action"] = df["Accion_final"].astype(str).str.upper().str.strip()
    
    # Métricas de recomendación (ya existen)
    df["recommended_sale_price"] = pd.to_numeric(df.get("precio_optimo", np.nan), errors="coerce")
    df["recommended_sale_recovery"] = pd.to_numeric(df.get("recovery_sale", np.nan), errors="coerce")
    
    # Crear valor_referencia (proxy valor intrínseco) = max(book_value, recovery_restruct * EAD, recovery_sale * EAD * 1.1)
    book_val = pd.to_numeric(df.get("book_value", np.nan), errors="coerce")
    rec_rest = pd.to_numeric(df.get("recovery_restruct", np.nan), errors="coerce")
    rec_sale = pd.to_numeric(df.get("recovery_sale", np.nan), errors="coerce")
    ead_val = pd.to_numeric(df.get("EAD", np.nan), errors="coerce")
    
    val1 = book_val
    val2 = rec_rest * ead_val
    val3 = rec_sale * ead_val * 1.1  # factor esperanza de mejor bid
    
    df["valor_referencia"] = val1.combine(val2, max, fill_value=0).combine(val3, max, fill_value=0)
    df["valor_referencia"] = df["valor_referencia"].fillna(ead_val * 0.40)  # fallback 40% EAD

    # -----------------------------------------------------------
    # 2) MANDATO DE VENTA (PC7 SECOND PASS: tiering + percentiles)
    # -----------------------------------------------------------
    logger.info("[MANDATE] Calculando sale_mandate (tiering + percentile-based scoring)...")
    
    rw_ratio = pd.to_numeric(df.get("RW", np.nan), errors="coerce")
    
    # 🆕 FORZAR RECÁLCULO de recovery_sale como precio_optimo / EAD 
    # (ignorar recovery_sale pre-existente en portfolio que puede ser obsoleto)
    logger.info(f"   [INFO] Calculando recovery_sale = precio_optimo/EAD (NPL actual)")
    precio_opt = pd.to_numeric(df.get("precio_optimo", np.nan), errors="coerce")
    ead_val_mandate = pd.to_numeric(df.get("EAD", np.nan), errors="coerce").replace(0, np.nan)
    rec_sale_val = (precio_opt / ead_val_mandate).replace([np.inf, -np.inf], np.nan).fillna(0)
    df["recovery_sale"] = rec_sale_val  # Sobrescribir con precio actual
    
    # age_npl: priorizar age_npl_m (meses), si no existe usar DPD (días) convertido a meses
    if "age_npl_m" in df.columns:
        age_npl = pd.to_numeric(df.get("age_npl_m", np.nan), errors="coerce")
    elif "DPD" in df.columns:
        dpd_days = pd.to_numeric(df.get("DPD", np.nan), errors="coerce")
        age_npl = (dpd_days / 30.0).fillna(0)  # Convertir días a meses
        logger.info(f"   [INFO] Usando DPD (días) convertido a meses como proxy de age_npl_m")
    else:
        age_npl = pd.Series([0.0] * len(df))  # Default 0 si no hay columna
        logger.warning(f"   [WARNING] Ni age_npl_m ni DPD encontrados - age_npl=0 (mandato por age desactivado)")
    
    # 🆕 PC7 SECOND PASS: score_mandate basado en percentiles (NO thresholds absolutos)
    # score_mandate = w_rwa * RW_normalized + w_age * age_normalized + w_recovery * (1 - recovery_rate)
    # Normalización: min-max a [0,1] para componibilidad
    
    rw_norm = (rw_ratio - rw_ratio.min()) / (rw_ratio.max() - rw_ratio.min() + 1e-9)
    age_norm = (age_npl - age_npl.min()) / (age_npl.max() - age_npl.min() + 1e-9)
    recovery_penalty = (1.0 - rec_sale_val).clip(0, 1)  # ya en [0,1]
    
    score_mandate = (
        mandate_w_rwa * rw_norm + 
        mandate_w_age * age_norm + 
        mandate_w_recovery * recovery_penalty
    )
    df["score_mandate"] = score_mandate.fillna(0)
    
    # TIER 1: Top mandate_tier1_share% (genuinamente obligatorio - worst of worst)
    target_k_tier1 = int(len(df) * mandate_tier1_share)
    n_tier1 = max(1, target_k_tier1) if mandate_tier1_share > 1e-6 else 0
    tier1_threshold = df["score_mandate"].nlargest(n_tier1).min() if n_tier1 > 0 else 999
    
    # TIER 2: Top mandate_share_target% total (incluye TIER1 + capital pressure)
    target_k_total = int(len(df) * mandate_share_target)
    n_total = max(1, target_k_total) if mandate_share_target > 1e-6 else 0
    tier2_threshold = df["score_mandate"].nlargest(n_total).min() if n_total > 0 else 999
    
    df["sale_mandate_tier"] = "NONE"
    df.loc[df["score_mandate"] >= tier1_threshold, "sale_mandate_tier"] = "TIER1_SEVERE"
    df.loc[(df["score_mandate"] >= tier2_threshold) & (df["score_mandate"] < tier1_threshold), "sale_mandate_tier"] = "TIER2_CAPITAL_PRESSURE"
    
    df["sale_mandate"] = df["sale_mandate_tier"] != "NONE"
    
    # sale_mandate_reason (auditable con componentes dominantes)
    def _build_mandate_reason(row):
        if row.get("sale_mandate_tier", "NONE") == "NONE":
            return ""
        tier = row.get("sale_mandate_tier")
        idx = row.name
        score = row.get("score_mandate", 0)
        # Componentes
        rw_val = rw_ratio.iloc[idx] if idx < len(rw_ratio) else 0
        age_val = age_npl.iloc[idx] if idx < len(age_npl) else 0
        rec_val = rec_sale_val.iloc[idx] if idx < len(rec_sale_val) else 0
        
        components = []
        if mandate_w_rwa > 0 and rw_val > 0:
            components.append(f"RWA={rw_val:.2f}")
        if mandate_w_age > 0 and age_val > 0:
            components.append(f"AGE={age_val:.0f}m")
        if mandate_w_recovery > 0:
            components.append(f"RECOVERY={rec_val:.1%}")
        
        comp_str = ",".join(components) if components else "COMPOSITE"
        return f"{tier}(score={score:.2f}|{comp_str})"
    
    df["sale_mandate_reason"] = df.apply(_build_mandate_reason, axis=1)
    
    n_mandate = df["sale_mandate"].sum()
    n_tier1 = (df["sale_mandate_tier"] == "TIER1_SEVERE").sum()
    n_tier2 = (df["sale_mandate_tier"] == "TIER2_CAPITAL_PRESSURE").sum()
    logger.info(f"   [OK] {n_mandate}/{len(df)} préstamos con sale_mandate=True")
    logger.info(f"   [TIER1_SEVERE] {n_tier1} ({n_tier1/len(df)*100:.1f}%)")
    logger.info(f"   [TIER2_CAPITAL_PRESSURE] {n_tier2} ({n_tier2/len(df)*100:.1f}%)")
    logger.info(f"   [SCORE_MANDATE] rango: [{df['score_mandate'].min():.2f}, {df['score_mandate'].max():.2f}]")

    # -----------------------------------------------------------
    # 3) GATE VENTA EJECUTABLE (precio no insultante + loss cap + recovery mínimo)
    # -----------------------------------------------------------
    logger.info("[SALE_GATE] Calculando sale_executable (precio + loss + recovery NPL bank-grade)...")
    
    # Sale price = precio_optimo (simulado por price_simulator)
    sale_price = pd.to_numeric(df.get("precio_optimo", np.nan), errors="coerce")
    
    valor_ref = df["valor_referencia"]
    
    # Gate 1: sale_insulting_flag = precio < sale_floor_ratio * valor_referencia
    df["sale_insulting_flag"] = (sale_price < sale_floor_ratio * valor_ref) | pd.isna(sale_price)
    
    # Gate 2: loss actual = (EAD - sale_price) / EAD
    sale_loss_pct = (ead_val - sale_price) / ead_val.replace(0, np.nan)
    df["sale_loss_pct"] = sale_loss_pct.clip(0, 1).fillna(0)
    
    # trade-off aceptable = loss <= loss_cap (sin mandato) o loss <= mandate_loss_tolerance (con mandato)
    loss_threshold = df["sale_mandate"].apply(lambda x: mandate_loss_tolerance if x else loss_cap_pct)
    df["sale_within_loss_cap"] = df["sale_loss_pct"] <= loss_threshold
    
    # Gate 3 🆕 NPL BANK-GRADE: recovery_rate = sale_price / EAD (disciplina económica)
    recovery_rate = sale_price / ead_val.replace(0, np.nan)
    df["recovery_rate"] = recovery_rate.clip(0, 1).fillna(0)
    df["sale_meets_recovery_min"] = df["recovery_rate"] >= recovery_min_pct
    
    # sale_executable = (NOT insulting) AND (within loss cap) AND (recovery_min O mandate)
    # Rationale: mandatos dominan (venta sí o sí), pero ventas voluntarias deben cumplir recovery mínimo
    df["sale_executable"] = (
        (~df["sale_insulting_flag"]) & 
        df["sale_within_loss_cap"] & 
        (df["sale_meets_recovery_min"] | df["sale_mandate"])  # Mandato exime de recovery_min
    )
    
    # Reason code para venta (auditable con 3 gates: insulting, loss_cap, recovery_min)
    def _build_sale_reason(row):
        if row.get("sale_mandate", False):
            if row.get("sale_executable", False):
                return "CAPITAL_MANDATE_EXECUTABLE"
            else:
                # Priority: insulting > loss_cap > recovery (aunque mandato exime recovery)
                if row.get("sale_insulting_flag"):
                    return "CAPITAL_MANDATE_BLOCKED_INSULTING_PRICE"
                elif not row.get("sale_within_loss_cap"):
                    return "CAPITAL_MANDATE_BLOCKED_LOSS_CAP"
                else:
                    return "CAPITAL_MANDATE_BLOCKED_OTHER"  # fallback (rare)
        else:
            if row.get("sale_executable", False):
                return "VOLUNTARY_SALE_ACCEPTABLE"
            else:
                # Priority: insulting > loss_cap > recovery_min
                if row.get("sale_insulting_flag"):
                    return "SALE_BLOCKED_INSULTING_PRICE"
                elif not row.get("sale_within_loss_cap"):
                    return "SALE_BLOCKED_LOSS_CAP"
                elif not row.get("sale_meets_recovery_min"):
                    return "SALE_BLOCKED_RECOVERY_TOO_LOW"  # 🆕 NPL bank-grade gate
                else:
                    return "SALE_BLOCKED_OTHER"  # fallback
    
    df["sale_reason_code"] = df.apply(_build_sale_reason, axis=1)
    
    n_executable_sales = df["sale_executable"].sum()
    n_insulting = df["sale_insulting_flag"].sum()
    n_recovery_low = (~df["sale_meets_recovery_min"]).sum()
    logger.info(f"   [OK] {n_executable_sales}/{len(df)} ventas ejecutables | "
                f"{n_insulting} insultantes | {n_recovery_low} recovery<{recovery_min_pct:.0%}")

    # -----------------------------------------------------------
    # 4) GATE REESTRUCTURA EJECUTABLE (acceptance_score)
    # -----------------------------------------------------------
    logger.info("[ACCEPTANCE] Calculando acceptance_score (ejecutabilidad reestructura)...")
    
    # Acceptance score basado en agresividad de oferta (0-100)
    # Penaliza: quita alta, plazo largo, tasa nueva alta
    # Premia: mejora DSCR/PTI, coherencia económica
    
    quita_val = pd.to_numeric(df.get("quita", np.nan), errors="coerce").fillna(0)
    plazo_val = pd.to_numeric(df.get("plazo_optimo", np.nan), errors="coerce").fillna(120)
    tasa_val = pd.to_numeric(df.get("tasa_nueva", np.nan), errors="coerce").fillna(0.05)
    dscr_post = pd.to_numeric(df.get("DSCR_post", np.nan), errors="coerce")
    pti_post = pd.to_numeric(df.get("PTI_post", np.nan), errors="coerce")
    
    # Base score = 50
    score = pd.Series(50.0, index=df.index)
    
    # Penalización quita: -30 si quita=30%, 0 si quita=0%
    score -= quita_val * 100  # quita 0.30 → -30 puntos
    
    # Penalización plazo: -20 si plazo=240m, -10 si 120m, 0 si 60m
    score -= (plazo_val - 60) / 180 * 20  # plazo 240 → -20, plazo 120 → -6.7
    
    # Penalización tasa alta: -10 si tasa=10%, 0 si tasa=5%
    score -= (tasa_val - 0.05) / 0.05 * 10  # tasa 10% → -10
    
    # Bonificación DSCR: +20 si DSCR>1.5, +10 si DSCR=1.2, 0 si DSCR=1.0
    if isinstance(dscr_post, pd.Series):
        score += ((dscr_post.fillna(1.0) - 1.0) / 0.5 * 20).clip(0, 25)
    else:
        score += ((dscr_post - 1.0) / 0.5 * 20)

    # Bonificación PTI bajo: +15 si PTI<25%, +7.5 si PTI=30%, 0 si PTI>35%
    if isinstance(pti_post, pd.Series):
        score += ((0.35 - pti_post.fillna(0.35)) / 0.10 * 15).clip(0, 20)
    else:
        score += ((0.35 - pti_post) / 0.10 * 15)
    
    df["acceptance_score"] = score.clip(0, 100)
    
    # restruct_executable = viable AND acceptance >= threshold
    viable_bool = df.get("restruct_viable", pd.Series(False, index=df.index)).apply(_safe_bool_cell)
    df["restruct_executable"] = viable_bool & (df["acceptance_score"] >= min_acceptance_score)
    
    # Reason code reestructura
    def _build_restruct_reason(row):
        if not row.get("restruct_viable", False):
            return "RESTRUCT_NOT_VIABLE_MISSING_DATA_OR_GATES"
        elif row.get("acceptance_score", 0) < min_acceptance_score:
            return f"RESTRUCT_BLOCKED_LOW_ACCEPTANCE({row.get('acceptance_score', 0):.0f}<{min_acceptance_score:.0f})"
        else:
            return "RESTRUCT_EXECUTABLE_CLEAN_TERMS"
    
    df["restructure_reason_code"] = df.apply(_build_restruct_reason, axis=1)
    
    n_executable_restruct = df["restruct_executable"].sum()
    logger.info(f"   [OK] {n_executable_restruct}/{len(df)} reestructuras ejecutables")
    logger.info(f"   [OK] acceptance_score rango: [{df['acceptance_score'].min():.0f}, {df['acceptance_score'].max():.0f}]")

    # -----------------------------------------------------------
    # 5) DECISIÓN FINAL CON FALLBACKS
    # -----------------------------------------------------------
    logger.info("[DECISION] Aplicando lógica de decisión con fallbacks...")
    
    # Guardamos acción original como recomendación
    df["action_before_executability"] = df["Accion_final"].copy()
    
    # Lógica de decisión (orden de precedencia)
    final_actions = []
    execution_status = []  # NORMAL | MANDATE_BLOCKED | RECOMMENDATION_NOT_EXECUTABLE
    
    for idx, row in df.iterrows():
        rec_action = str(row.get("recommended_action", "MANTENER")).upper()
        
        # CASO 1: MANDATO DE VENTA (sí o sí)
        if row.get("sale_mandate", False):
            if row.get("sale_executable", False):
                # Mandato + ejecutable → VENDER
                final_actions.append("VENDER")
                execution_status.append("MANDATE_EXECUTED")
            else:
                # Mandato bloqueado por loss-cap o insulting → MANTENER temporal
                final_actions.append("MANTENER")
                execution_status.append("MANDATE_BLOCKED")
        
        # CASO 2: SIN MANDATO - orden lógico
        else:
            # Primero intentar la recomendación original
            if rec_action == "VENDER":
                if row.get("sale_executable", False):
                    final_actions.append("VENDER")
                    execution_status.append("NORMAL")
                elif row.get("restruct_executable", False):
                    # Fallback: venta no ejecutable → reestructura
                    final_actions.append("REESTRUCTURAR")
                    execution_status.append("FALLBACK_RESTRUCT")
                else:
                    # Ni venta ni restructura → mantener
                    final_actions.append("MANTENER")
                    execution_status.append("NO_EXECUTABLE_ACTION")
            
            elif rec_action == "REESTRUCTURAR":
                if row.get("restruct_executable", False):
                    final_actions.append("REESTRUCTURAR")
                    execution_status.append("NORMAL")
                elif row.get("sale_executable", False):
                    # Fallback: reestructura no ejecutable → venta
                    final_actions.append("VENDER")
                    execution_status.append("FALLBACK_SALE")
                else:
                    # Ni reestructura ni venta → mantener
                    final_actions.append("MANTENER")
                    execution_status.append("NO_EXECUTABLE_ACTION")
            
            else:  # MANTENER
                final_actions.append("MANTENER")
                execution_status.append("NORMAL")
    
    df["final_action_executability"] = final_actions
    df["execution_status"] = execution_status
    
    # Actualizar case_status para mandatos bloqueados
    mandate_blocked = df["execution_status"] == "MANDATE_BLOCKED"
    if mandate_blocked.any():
        df.loc[mandate_blocked, "case_status"] = "MANDATE_BLOCKED"
        df.loc[mandate_blocked, "next_step"] = "MARKET_SALE_ESCALATION"
        df.loc[mandate_blocked, "next_step_reason"] = "CAPITAL_MANDATE_REQUIRES_SALE_BUT_PRICE_INSULTING_OR_LOSS_EXCEEDS_CAP"
        df.loc[mandate_blocked, "review_due_days"] = 14  # revisión urgente
        df.loc[mandate_blocked, "override_reason"] = "TEMPORARY_HOLD_UNTIL_MARKET_IMPROVES_OR_MANDATE_OVERRIDE"
    
    # Actualizar case_status para sin acción ejecutable
    no_executable = df["execution_status"] == "NO_EXECUTABLE_ACTION"
    if no_executable.any():
        df.loc[no_executable, "case_status"] = "HOLD_NO_EXECUTABLE_ACTION"
        df.loc[no_executable, "next_step"] = "WORKOUT_REVIEW"
        df.loc[no_executable, "next_step_reason"] = "NEITHER_SALE_NOR_RESTRUCTURE_EXECUTABLE_TODAY"
        df.loc[no_executable, "review_due_days"] = 30
    
    # Sobrescribir Accion_final con la decisión ejecutable
    df["Accion_final"] = df["final_action_executability"]
    
    # Stats
    action_dist = df["Accion_final"].value_counts()
    status_dist = df["execution_status"].value_counts()
    logger.info(f"[STATS] Distribución final de acciones:")
    for action, count in action_dist.items():
        logger.info(f"   {action}: {count} ({count/len(df)*100:.1f}%)")
    logger.info(f"[STATS] Status de ejecución:")
    for status, count in status_dist.items():
        logger.info(f"   {status}: {count}")

    # -----------------------------------------------------------
    # 6) CAPACIDAD OPERATIVA (max_restructure_share)
    # -----------------------------------------------------------
    logger.info(f"🏭 Aplicando capacidad operativa (max {max_restructure_share:.0%} reestructuras)...")
    
    restruct_mask = df["Accion_final"] == "REESTRUCTURAR"
    n_restruct_current = restruct_mask.sum()
    max_allowed = int(len(df) * max_restructure_share)
    
    if n_restruct_current > max_allowed:
        # Degradar las peores (menor acceptance_score) a MANTENER
        restruct_df = df[restruct_mask].copy()
        restruct_df_sorted = restruct_df.sort_values("acceptance_score", ascending=True)
        n_to_degrade = n_restruct_current - max_allowed
        degrade_indices = restruct_df_sorted.head(n_to_degrade).index
        
        df.loc[degrade_indices, "Accion_final"] = "MANTENER"
        df.loc[degrade_indices, "execution_status"] = "CAPACITY_LIMIT"
        df.loc[degrade_indices, "case_status"] = "HOLD_CAPACITY_LIMIT"
        df.loc[degrade_indices, "next_step"] = "WAIT_CAPACITY_OR_FALLBACK_SALE"
        df.loc[degrade_indices, "next_step_reason"] = f"RESTRUCTURE_CAPACITY_EXCEEDED({n_restruct_current}>{max_allowed})"
        df.loc[degrade_indices, "review_due_days"] = 60
        df.loc[degrade_indices, "override_reason"] = "CAPACITY_LIMIT_RESTRUCT_DOWNGRADED_TO_MAINTAIN"
        
        if "Reason_Code" in df.columns:
            df.loc[degrade_indices, "Reason_Code"] = "RC14_CAPACITY_LIMIT_RESTRUCT"
        
        logger.info(f"   [WARNING] {n_to_degrade} reestructuras degradadas a MANTENER por capacidad")
    else:
        logger.info(f"   [OK] Capacidad OK: {n_restruct_current}/{max_allowed} reestructuras")

    # -----------------------------------------------------------
    # 7) 🆕 PC7 SECOND PASS - CAP DE VENTAS (max_sell_share solo VOLUNTARIAS, mandatos exentos)
    # -----------------------------------------------------------
    if max_sell_share < 1.0:  # Solo si hay cap
        logger.info(f"🆕 Aplicando cap de ventas VOLUNTARIAS (max {max_sell_share:.0%}, mandatos exentos)...")
        
        sell_mask = df["Accion_final"] == "VENDER"
        n_ventas_total = sell_mask.sum()
        n_ventas_mandate = (sell_mask & df["sale_mandate"]).sum()
        n_ventas_voluntary = n_ventas_total - n_ventas_mandate
        
        # 🆕 PC7 SECOND PASS: mandatos NO cuentan contra el cap (dominan)
        max_allowed_voluntary = int(len(df) * max_sell_share)
        
        if n_ventas_voluntary > max_allowed_voluntary and max_allowed_voluntary >= 0:
            # Degradar ventas voluntarias con menor recovery_rate a MANTENER
            voluntary_sell_df = df[sell_mask & (~df["sale_mandate"])].copy()
            voluntary_sorted = voluntary_sell_df.sort_values("recovery_rate", ascending=True)
            n_to_degrade = n_ventas_voluntary - max_allowed_voluntary
            degrade_indices = voluntary_sorted.head(n_to_degrade).index
            
            df.loc[degrade_indices, "Accion_final"] = "MANTENER"
            df.loc[degrade_indices, "execution_status"] = "VOLUME_CAP_LIMIT"
            df.loc[degrade_indices, "case_status"] = "HOLD_VOLUME_CAP"
            df.loc[degrade_indices, "next_step"] = "ALT_DISPOSAL_OR_WAIT"
            df.loc[degrade_indices, "next_step_reason"] = (
                f"SELL_VOLUME_CAP_EXCEEDED(voluntary={n_ventas_voluntary}>{max_allowed_voluntary}, "
                f"mandates={n_ventas_mandate} EXEMPT)"
            )
            df.loc[degrade_indices, "review_due_days"] = 90
            df.loc[degrade_indices, "override_reason"] = "VOLUME_CAP_VOLUNTARY_SELL_DOWNGRADED_TO_MAINTAIN"
            
            if "Reason_Code" in df.columns:
                df.loc[degrade_indices, "Reason_Code"] = "RC15_VOLUME_CAP_VOLUNTARY_SELL"
            
            logger.info(f"   [WARNING] {n_to_degrade} ventas VOLUNTARIAS degradadas a MANTENER por cap")
            logger.info(f"   [MANDATES] {n_ventas_mandate} ventas por MANDATO (exentas de cap, DOMINAN)")
        else:
            logger.info(f"   [OK] Cap OK: voluntary={n_ventas_voluntary}/{max_allowed_voluntary}, "
                        f"mandates={n_ventas_mandate} (exentos)")
        
        # Totales finales
        n_ventas_final = (df["Accion_final"] == "VENDER").sum()
        n_mandate_final = ((df["Accion_final"] == "VENDER") & df["sale_mandate"]).sum()
        logger.info(f"   [FINAL] Ventas totales={n_ventas_final} (mandate={n_mandate_final}, "
                    f"voluntary={n_ventas_final - n_mandate_final})")
    else:
        logger.info(f"   [NO CAP] max_sell_share={max_sell_share:.0%} - Sin límite de ventas")

    # Distribución final
    action_dist_final = df["Accion_final"].value_counts()
    logger.info(f"[STATS] Distribución FINAL de acciones (post-capacidad + post-volume-cap):")
    for action, count in action_dist_final.items():
        logger.info(f"   {action}: {count} ({count/len(df)*100:.1f}%)")

    # ===========================================================
    # HARD GUARDRAILS (Centralized via optimizer/guardrails.py)
    # ===========================================================
    logger.info("🛡️ Aplicando guardrails centralizados (Check Hard Constraints)...")

    def _set_action_all(mask: pd.Series, action: str) -> None:
        """Set de acción en TODAS las columnas espejo que puedan existir."""
        # simple normalization
        a = str(action).upper().strip()
        mirror_cols = (
            "Accion_final", "Accion", "decision_final", "Decision_Final",
            "Final_Decision", "accion_final", "Accion_Final", "Decision"
        )
        for c in mirror_cols:
            if c in df.columns:
                df.loc[mask, c] = a

    # Preparar config dummy con los umbrales de la postura actual
    g_params = _get_guardrails(risk_posture)
    
    class GuardrailConfig:
        def __init__(self, params):
            self.risk_params = type("RiskParams", (), {
                "pti_limit": params.get("PTI_MAX", 0.45),
                "dscr_min": params.get("DSCR_MIN", 1.10)
            })
            self.fire_sale = type("FireSaleParams", (), {
                "threshold_book": params.get("MAX_FIRE_SALE_PNL_RATIO_BOOK", 0.20),
                "allow_fire_sale": risk_posture.lower() != "prudencial" # Solo no prudencial permite fire sale agresivo (simplificado)
            })
    
    cfg_gr = GuardrailConfig(g_params)

    # Iterar y aplicar
    n_blocked_restruct = 0
    n_blocked_sell = 0
    
    for idx, row in df.iterrows():
        action = str(row.get("Accion_final", "")).upper().strip()
        
        # 1. REESTRUCTURAR
        if action == "REESTRUCTURAR":
            # Construir estado
            loan_state = row.to_dict()
            # Mapear nombres si difieren
            loan_state["pti_actual"] = row.get("PTI_post", row.get("PTI", 0.0))
            loan_state["dscr_actual"] = row.get("DSCR_post", row.get("DSCR", 0.0))
            
            ok, reasons, metrics = check_restructure_constraints(loan_state, cfg_gr)
            
            # Persistir metricas de auditoria (siempre, pase o falle)
            df.at[idx, "restructure_ok"] = ok
            df.at[idx, "pti_limit"] = metrics.get("pti_max")
            df.at[idx, "dscr_limit"] = metrics.get("dscr_min")
            df.at[idx, "pti_headroom"] = metrics.get("pti_headroom")
            df.at[idx, "dscr_headroom"] = metrics.get("dscr_headroom")

            if not ok:
                n_blocked_restruct += 1
                df.at[idx, "Accion_final"] = "MANTENER"
                _set_action_all(df.index == idx, "MANTENER")
                
                reason_str = "; ".join(reasons)
                df.at[idx, "guardrail_reasons"] = reason_str
                df.at[idx, "override_applied"] = True
                df.at[idx, "override_from"] = "REESTRUCTURAR"
                df.at[idx, "override_to"] = "MANTENER"
                df.at[idx, "Reason_Code"] = "RC_GUARDRAIL_BLOCK" # Estandarizado
                
                current_gov = str(df.at[idx, "Decision_Governance"]) if pd.notna(df.at[idx, "Decision_Governance"]) else ""
                df.at[idx, "Decision_Governance"] = (current_gov + f"; BLOCK: {reason_str}").strip("; ")

        # 2. VENDER
        elif action == "VENDER":
            loan_state = row.to_dict()
            # Construir pricing_out robusto desde columnas del DF (calculadas previamente)
            # Intentar varios alias comunes
            pr_opt = row.get("precio_optimo", row.get("Price", 0.0))
            pnl_val = row.get("pnl", row.get("PnL", 0.0))
            cap_rel = row.get("capital_liberado", row.get("capital_release", 0.0))
            
            pricing_out = {
                "precio_optimo": pr_opt,
                "price": row.get("precio_bruto", pr_opt),
                "pnl": pnl_val,
                "capital_liberado": cap_rel,
                "coste_tx": row.get("coste_tx", 0.0),
                "fire_sale": row.get("fire_sale", False),
                "fire_sale_reason": row.get("fire_sale_reason", ""),
                "book_value": row.get("book_value", 0.0),
                "rw": row.get("RW", 1.5)
            }
            
            ok, reasons, metrics = check_sell_constraints(loan_state, pricing_out, cfg_gr)
            
            # Persistir metricas audit sell
            df.at[idx, "sell_ok"] = ok
            df.at[idx, "capital_release_net"] = metrics.get("capital_release")
            df.at[idx, "rwa_before"] = metrics.get("rwa_before")
            df.at[idx, "rwa_after"] = metrics.get("rwa_after")
            # Reuse audit_price_book_ratio (legacy name)
            df.at[idx, "audit_price_book_ratio"] = metrics.get("price_book_ratio")

            if not ok:
                n_blocked_sell += 1
                # Degradar a MANTENER 
                
                df.at[idx, "Accion_final"] = "MANTENER"
                _set_action_all(df.index == idx, "MANTENER")
                
                reason_str = "; ".join(reasons)
                df.at[idx, "guardrail_reasons"] = reason_str
                df.at[idx, "override_applied"] = True
                df.at[idx, "override_from"] = "VENDER"
                df.at[idx, "override_to"] = "MANTENER"
                df.at[idx, "Reason_Code"] = "RC_GUARDRAIL_BLOCK"
                
                current_gov = str(df.at[idx, "Decision_Governance"]) if pd.notna(df.at[idx, "Decision_Governance"]) else ""
                df.at[idx, "Decision_Governance"] = (current_gov + f"; BLOCK: {reason_str}").strip("; ")

    logger.info(f"🛡️ Guardrails aplicados: {n_blocked_restruct} reestructuras bloqueadas, {n_blocked_sell} ventas bloqueadas.")


    # ===========================================================
    # (Removed old HARD GUARDRAILS block)
    # ===========================================================




    # ===========================================================
    # Realizados (con Accion_final ya hard-guardrailed)
    # ===========================================================
    df["pnl_realized"] = 0.0
    df["capital_release_realized"] = 0.0

    accion_final_u = df["Accion_final"].astype(str).str.upper()
    sell_mask2 = accion_final_u.eq("VENDER")
    rest_mask2 = accion_final_u.eq("REESTRUCTURAR")

    if "pnl" in df.columns:
        df.loc[sell_mask2, "pnl_realized"] = pd.to_numeric(df.loc[sell_mask2, "pnl"], errors="coerce").fillna(0.0)
    elif "PnL" in df.columns:
        df.loc[sell_mask2, "pnl_realized"] = pd.to_numeric(df.loc[sell_mask2, "PnL"], errors="coerce").fillna(0.0)

    df.loc[sell_mask2, "capital_release_realized"] = pd.to_numeric(
        df.loc[sell_mask2, "capital_release_cf"], errors="coerce"
    ).fillna(0.0)

    if "capital_liberado" in df.columns:
        df.loc[rest_mask2, "capital_release_realized"] = pd.to_numeric(
            df.loc[rest_mask2, "capital_liberado"], errors="coerce"
        ).fillna(0.0)

    # ===========================================================
    # ✅ BANK-READY: Post-procesamiento escalación
    # Identificar casos que fueron bloqueados pero no tienen metadata de escalación
    # ===========================================================
    try:
        # Casos críticos: Sell_Blocked=True + restruct_viable=False + Accion_final=MANTENER
        if all(c in df.columns for c in ["Sell_Blocked", "restruct_viable", "Accion_final", "case_status"]):
            sell_blocked = df["Sell_Blocked"].fillna(False).astype(bool)
            not_viable = ~df["restruct_viable"].fillna(True).astype(bool)
            is_mantener = df["Accion_final"].astype(str).str.upper().eq("MANTENER")
            still_normal = df["case_status"].astype(str).eq("NORMAL")
            
            escalate_mask = sell_blocked & not_viable & is_mantener & still_normal
            
            if escalate_mask.any():
                logger.info(f"[ESCALATION] Aplicando metadata de escalación a {escalate_mask.sum()} casos críticos (Sell_Blocked + restruct_viable=False → MANTENER)")
                
                df.loc[escalate_mask, "case_status"] = "HOLD_NO_EXECUTABLE_ACTION"
                df.loc[escalate_mask, "next_step"] = "WORKOUT_REVIEW"
                df.loc[escalate_mask, "next_step_reason"] = "FIRESALE_BLOCKED_AND_RESTRUCTURE_NOT_VIABLE"
                df.loc[escalate_mask, "review_due_days"] = 30
                
                if "override_reason" in df.columns:
                    current_override = df.loc[escalate_mask, "override_reason"].fillna("").astype(str)
                    new_override = "FIRESALE_BLOCKED_RESTRUCTURE_NOT_VIABLE_FALLBACK_MAINTAIN"
                    # Solo sobrescribir si está vacío
                    df.loc[escalate_mask & (current_override == ""), "override_reason"] = new_override
                
                # Flags de datos requeridos
                if "required_data_flags" in df.columns:
                    def _build_flags_postproc(row):
                        flags = []
                        if pd.isna(row.get("book_value")) or row.get("book_value", 0) == 0:
                            flags.append("book_value_tasacion")
                        if row.get("Missing_Viability_Inputs", False):
                            seg = str(row.get("segment", "")).upper()
                            if "MORTGAGE" in seg or "CONSUMER" in seg:
                                flags.append("monthly_income")
                            else:
                                flags.append("monthly_cfo")
                        if pd.isna(row.get("secured")):
                            flags.append("collateral_info")
                        return "; ".join(flags) if flags else "viability_assessment"
                    
                    df.loc[escalate_mask, "required_data_flags"] = df.loc[escalate_mask].apply(_build_flags_postproc, axis=1)
    except Exception as e:
        logger.warning(f"[WARN] Error en post-procesamiento de escalación: {e}")

    return df, override_log_entries


# ===========================================================
# 4) PIPELINE COORDINADO (UNA POSTURA)
# ===========================================================
def run_coordinator_inference(
    model_micro: str,
    portfolio_path: str,
    vecnorm_micro_path: Optional[str],
    model_macro: Optional[str],
    risk_posture: str,
    n_steps: int,
    top_k: int,
    tag: str,
    base_output_dir: Optional[str] = None,
    device: str = "auto",
    seed: int = 42,
    deterministic: bool = True,
    export_audit_csv: bool = False,
    vecnorm_macro_path: Optional[str] = None,
    vecnorm_loan_path: Optional[str] = None,
) -> Tuple[str, str]:
    if not os.path.exists(model_micro):
        raise FileNotFoundError(f"Modelo MICRO no encontrado: {model_micro}")
    if not os.path.exists(portfolio_path):
        raise FileNotFoundError(f"Cartera no encontrada: {portfolio_path}")

    vecnorm_micro_path, vecnorm_macro_path, vecnorm_loan_path = _sanitize_vecnorm_paths(
        vn_micro=vecnorm_micro_path,
        vn_macro=vecnorm_macro_path,
        vn_loan=vecnorm_loan_path,
    )

    if model_macro is None:
        model_macro = _pick_default_macro_model()

    # 0) Portfolio merge defensivo
    df_port = _load_portfolio_df(portfolio_path)

    # 1) MICRO
    df_micro, _ = _run_micro_inference(
        model_path=model_micro,
        portfolio_path=portfolio_path,
        vecnorm_path=vecnorm_micro_path,
        risk_posture=risk_posture,
        tag_suffix=tag,
        device=device,
        seed=seed,
        deterministic=deterministic,
    )

    # 2) MACRO
    macro_actions: Dict[str, Dict[str, Any]] = {}
    if model_macro is not None and os.path.exists(model_macro):
        df_steps, _ = _run_macro_inference(
            model_path_portfolio=model_macro,
            portfolio_path=portfolio_path,
            risk_posture=risk_posture,
            loan_model_path=model_micro,
            tag_suffix=tag,
            n_steps=n_steps,
            top_k=top_k,
            device=device,
            seed=seed,
            deterministic=deterministic,
            vecnormalize_portfolio_path=vecnorm_macro_path,
            vecnormalize_loan_path=vecnorm_loan_path,
        )
        macro_actions = _build_macro_actions_per_loan(df_steps)
    else:
        logger.warning("[WARN] No se ha encontrado modelo MACRO; se usarán solo decisiones micro.")

    # 3) COMBINAR
    df_final, override_log = _combine_decisions(
        df_micro,
        macro_actions,
        risk_posture=risk_posture,
        n_steps=n_steps,
        top_k=top_k,
    )

    # --- KPI Impact Analysis ---
    logger.info("=== KPI IMPACT ANALYSIS (Micro vs. Coordinated) ===")
    col_act = "Accion" if "Accion" in df_micro.columns else "action"
    if col_act in df_micro.columns:
        logger.info(f"Micro Decisions:\n{df_micro[col_act].value_counts().to_string()}")
    
    col_fin = "Accion" if "Accion" in df_final.columns else "action"
    if col_fin in df_final.columns:
         logger.info(f"Final Decisions:\n{df_final[col_fin].value_counts().to_string()}")
    
    if override_log:
        df_ov = pd.DataFrame(override_log)
        if "override_level" in df_ov.columns:
            logger.info(f"Overrides by Level:\n{df_ov['override_level'].value_counts().to_string()}")
    # ---------------------------

    df_final = _ensure_loan_id_column(df_final)
    df_final = _fill_from_portfolio(df_final, df_port)

    # 3b) schema stable
    df_final = enforce_schema(df_final)

    # 4) EXPORTAR
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    posture_suffix = (risk_posture or "balanceado").lower().strip()

    if base_output_dir:
        out_dir_coord = os.path.join(base_output_dir, f"{ts}_{tag}_{posture_suffix}")
    else:
        out_dir_coord = os.path.join(REPORTS_DIR, f"coordinated_inference_{tag}_{ts}_{posture_suffix}")
    os.makedirs(out_dir_coord, exist_ok=True)
    
    # 🆕 Task 4: KPI Portfolio Before/After (Audit Grade)
    try:
        # 🆕 Task 4: Semantic KPI Structure (Safe for PowerBI/Audit)
        
        # 1. Pre-State (Baseline / Status Quo)
        cols_pre = {"EVA": "EVA", "RWA": "RWA", "EAD": "EAD"}
        if "EVA_pre" in df_micro.columns: cols_pre["EVA"] = "EVA_pre"
        if "RWA_pre" in df_micro.columns: cols_pre["RWA"] = "RWA_pre"
        
        kpis_pre = compute_portfolio_kpis(df_micro, col_map=cols_pre)
        
        # 2. Micro Proposal (Pure Micro Model)
        # Assuming df_micro has 'Accion' or 'action' from the micro model
        action_col_micro = "Accion" if "Accion" in df_micro.columns else "action"
        # Often Micro output already has post-action EVA called 'EVA_post' or just 'EVA' 
        # if the dataframe is the result of applying actions. 
        # But commonly df_micro returned by _run_micro_inference has the *proposed* action
        # and potentially updated metrics. 
        # Let's assume 'EVA_post' exists if actions were simulated, or we rely on the primary 'EVA' column if it was updated.
        # Ideally, we want the theoretical result of the micro model alone.
        cols_micro = {"EVA": "EVA_post", "RWA": "RWA_post", "EAD": "EAD"}
        if "EVA_post" not in df_micro.columns: cols_micro["EVA"] = "EVA" # fallback
        if "RWA_post" not in df_micro.columns: cols_micro["RWA"] = "RWA" # fallback

        kpis_micro = compute_portfolio_kpis(df_micro, action_col=action_col_micro, col_map=cols_micro)

        # 3. Final State (Coordinated)
        kpis_final = compute_portfolio_kpis(
            df_final, 
            action_col="Accion_final",
            col_map={"EVA": "EVA_post", "RWA": "RWA_post", "EAD": "EAD"}
        )
        
        # Calculate Delta (Final - Pre)
        delta_metrics = {
            "total_eva": kpis_final.get("total_eva", 0.0) - kpis_pre.get("total_eva", 0.0),
            "total_rwa": kpis_final.get("total_rwa", 0.0) - kpis_pre.get("total_rwa", 0.0),
            "capital_release": kpis_final.get("total_capital_release", 0.0) # Pre is 0 by definition
        }
        
        portfolio_stats = {
            "metadata": {
                "timestamp": ts,
                "posture": risk_posture,
                "tag": tag,
                "n_loans": len(df_final)
            },
            "pre_state": kpis_pre,
            "micro_proposal": kpis_micro,
            "final_state": kpis_final,
            "delta_impact": delta_metrics
        }
        
        kpi_json = os.path.join(out_dir_coord, f"portfolio_kpis_{posture_suffix}.json")
        with open(kpi_json, "w", encoding="utf-8") as f:
            json.dump(portfolio_stats, f, indent=2)
        logger.info(f"📊 Portfolio KPIs saved (Semantic Format): {kpi_json}")
    except Exception as e:
        logger.warning(f"⚠️ Error computing portfolio KPIs: {e}", exc_info=True)

    # 🆕 Task 1 & 3: Sanitize Audit Columns & Standardize IDs
    if "run_id" not in df_final.columns:
        df_final["run_id"] = f"{tag}_{ts}"
    
    # Audit column sanitization (No NaNs)
    audit_fill_map = {
        "override_applied": False,
        "macro_selected": False,
        "override_from": "NONE",
        "override_to": "NONE",
        "override_level": "NONE",
        "macro_action": "NONE",
        "Convergencia_Caso": "MICRO_DEFAULT",
    }
    for col, fill_val in audit_fill_map.items():
        if col not in df_final.columns:
            # Create column if missing
            df_final[col] = fill_val
        else:
            # Fill existing NaNs/Nones/Empty strings
            # Use inference to ensure object types are handled correctly
            series = df_final[col].fillna(fill_val)
            # Also catch empty strings if they exist
            if pd.api.types.is_object_dtype(series):
                 series = series.replace(r'^\s*$', fill_val, regex=True)
            df_final[col] = series
            
    # Ensure macro_action_used is populated for traceability
    if "macro_action" in df_final.columns and "macro_action_used" not in df_final.columns:
        df_final["macro_action_used"] = df_final["macro_action"]

    excel_final = os.path.join(out_dir_coord, f"decisiones_finales_{posture_suffix}.xlsx")
    export_styled_excel(df_final, excel_final)
    
    # 🆕 Task 3: Override Log Export (ALWAYS) with consistent schema
    csv_overrides = os.path.join(out_dir_coord, f"overrides_log_{posture_suffix}.csv")
    
    if override_log:
        df_ov = pd.DataFrame(override_log)
        # Ensure all standard columns exist
        for col in OVERRIDE_LOG_COLS:
            if col not in df_ov.columns:
                df_ov[col] = None
        # Reorder/Filter columns
        # (We allow extra columns if they exist in log, or strictly filter? 
        # Instructions say "define a unique list... Ensure df = df.reindex". 
        # So we should strictly enforce the list or at least ensure these are present at the start)
        # Let's align with the requested list and keep any extras at the end if we want, 
        # OR just strictly stick to the list. "Standardize" implies strictly stick to the list usually for CSV stability.
        # But if we have extra debug info we might want it.
        # User said: "Defina un listado único... con todas las columnas extendidas... reindex(columns=OVERRIDE_LOG_COLS)"
        # So I will strict reindex.
        df_ov = df_ov.reindex(columns=OVERRIDE_LOG_COLS)
    else:
        df_ov = pd.DataFrame(columns=OVERRIDE_LOG_COLS)
        
    df_ov.to_csv(csv_overrides, index=False, encoding="utf-8-sig")
    logger.info(f"⚡ Override Log saved: {csv_overrides} ({len(df_ov)} triggers)")

    if export_audit_csv:
        csv_final = os.path.join(out_dir_coord, f"decisiones_audit_{posture_suffix}.csv")
        df_final.to_csv(csv_final, index=False, encoding="utf-8-sig")
        logger.info(f"🧾 Audit CSV: {csv_final}")
        
        # (The override log is already saved above with correct schema, no need to overwrite with potentially incorrect schema)


    # Logging convergencia
    try:
        n = len(df_final)
        sel = int(pd.Series(df_final.get("Macro_Selected", False)).sum())
        mw = int((df_final.get("Convergencia_Caso", "") == "MACRO_WINS").sum()) if "Convergencia_Caso" in df_final.columns else 0
        gr = int((df_final.get("Convergencia_Caso", "") == "GUARDRAIL_OVERRIDE").sum()) if "Convergencia_Caso" in df_final.columns else 0
        logger.info(f"[STATS] Macro_Selected: {sel}/{n} ({(sel/max(n,1))*100:.1f}%) | MACRO_WINS={mw} | GUARDRAIL_OVERRIDE={gr}")
    except Exception:
        pass

    logger.info(f"[COMPLETED] Inferencia coordinada ({risk_posture}) completada.")
    logger.info(f"[EXCEL] Excel final: {excel_final}")
    return out_dir_coord, excel_final


# ===========================================================
# 5) PIPELINE MULTI-POSTURA
# ===========================================================
def run_coordinator_inference_multi_posture(cfg_base: CoordinatorInferenceConfig) -> List[str]:
    outputs: List[str] = []
    postures = ["prudencial", "balanceado", "desinversion"]

    runs_root = cfg_base.base_output_dir or os.path.join(REPORTS_DIR, "runs")
    os.makedirs(runs_root, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_base_dir = os.path.join(runs_root, f"{ts}_{cfg_base.tag}_MULTI")
    os.makedirs(run_base_dir, exist_ok=True)

    deliverable_dir = os.path.join(runs_root, f"{ts}_{cfg_base.tag}_DELIVERABLE")
    os.makedirs(deliverable_dir, exist_ok=True)

    logger.info(f"▶ MULTI-POSTURE RUN ROOT: {run_base_dir}")
    logger.info(f"📦 DELIVERABLE ROOT:      {deliverable_dir}")
    logger.info(f"  • deliverable_only:    {cfg_base.deliverable_only}")
    logger.info(f"  • export_audit_csv:    {cfg_base.export_audit_csv}")

    import glob

    for posture in postures:
        logger.info(f"▶ Ejecutando COORDINATOR (risk_posture={posture})…")

        posture_base = os.path.join(run_base_dir, posture)
        os.makedirs(posture_base, exist_ok=True)

        out_dir, excel_path = run_coordinator_inference(
            model_micro=cfg_base.model_path_micro,
            portfolio_path=cfg_base.portfolio_path,
            vecnorm_micro_path=cfg_base.vecnormalize_path_micro,

            model_macro=cfg_base.model_path_macro,
            risk_posture=posture,

            n_steps=int(cfg_base.n_steps),
            top_k=int(cfg_base.top_k),

            tag=f"{cfg_base.tag}_{posture}",
            base_output_dir=posture_base,

            device=cfg_base.device,
            seed=int(cfg_base.seed),
            deterministic=bool(cfg_base.deterministic),

            export_audit_csv=bool(cfg_base.export_audit_csv),

            vecnorm_macro_path=cfg_base.vecnormalize_path_macro,
            vecnorm_loan_path=cfg_base.vecnormalize_path_loan,
        )

        outputs.append(out_dir)

        # Copia Excel al DELIVERABLE
        dest_excel = os.path.join(deliverable_dir, f"decisiones_finales_{posture}.xlsx")
        if excel_path and os.path.exists(excel_path):
            shutil.copy2(excel_path, dest_excel)
            logger.info(f"  ✓ Excel ({posture}) => {dest_excel}")
        else:
            logger.warning(f"  ⚠ No se encontró excel_path para postura={posture}: {excel_path}")

        # Copia AUDIT CSV al DELIVERABLE
        if bool(cfg_base.export_audit_csv):
            if out_dir and os.path.exists(out_dir):
                preferred = glob.glob(os.path.join(out_dir, "**", f"decisiones_audit_{posture}.csv"), recursive=True)
                candidates = preferred or glob.glob(os.path.join(out_dir, "**", "decisiones_audit_*.csv"), recursive=True)

                if candidates:
                    src_audit = candidates[0]
                    dest_audit = os.path.join(deliverable_dir, f"decisiones_audit_{posture}.csv")
                    shutil.copy2(src_audit, dest_audit)
                    logger.info(f"  ✓ Audit CSV ({posture}) => {dest_audit}")
                else:
                    logger.warning(f"  ⚠ No se encontró audit CSV en out_dir para postura={posture}: {out_dir}")
            else:
                logger.warning(f"  ⚠ out_dir inválido para postura={posture}: {out_dir}")

    logger.info("✅ DELIVERABLE generado (3 Excels).")

    if cfg_base.deliverable_only:
        try:
            shutil.rmtree(run_base_dir, ignore_errors=True)
            logger.info(f"🧹 Limpieza deliverable-only: eliminado run_dir={run_base_dir}")
        except Exception as e:
            logger.warning(f"⚠ No se pudo eliminar run_dir={run_base_dir}: {e}")
        return [deliverable_dir]

    return outputs + [deliverable_dir]


# ===========================================================
# 6) CLI
# ===========================================================
def parse_args():
    p = argparse.ArgumentParser(description="Inferencia coordinada MICRO+MACRO (LoanEnv + PortfolioEnv) · Banco L1.5")

    p.add_argument("--model-micro", type=str, required=True, dest="model_micro", help="Ruta a best_model.zip (LoanEnv).")
    p.add_argument("--portfolio", type=str, required=True, help="Ruta a cartera (Excel/CSV).")

    p.add_argument(
        "--vn-micro",
        type=str,
        default=_pick_default_micro_vecnorm(),
        dest="vn_micro",
        help="Ruta VecNormalize MICRO (LoanEnv). Recomendado: models/vecnormalize_loan.pkl",
    )
    p.add_argument(
        "--vecnorm",
        type=str,
        default=None,
        dest="vecnorm",
        help="(DEPRECATED) Alias legacy de --vn-micro (para compatibilidad con .bat antiguos).",
    )

    p.add_argument(
        "--model-macro",
        type=str,
        default=_pick_default_macro_model(),
        dest="model_macro",
        help="Ruta a best_model_portfolio.zip (PortfolioEnv).",
    )
    p.add_argument(
        "--vn-macro",
        type=str,
        default=_pick_default_macro_vecnorm(),
        dest="vn_macro",
        help="Ruta VecNormalize MACRO (PortfolioEnv). Recomendado: models/vecnormalize_portfolio.pkl",
    )
    p.add_argument(
        "--vn-loan",
        type=str,
        default=_pick_default_micro_vecnorm(),
        dest="vn_loan",
        help="Ruta VecNormalize LOAN (para micro re-ranking en macro).",
    )

    p.add_argument("--risk-posture", type=str, choices=["prudencial", "balanceado", "desinversion"], default="balanceado")
    p.add_argument("--n-steps", type=int, default=1, dest="n_steps")
    p.add_argument("--top-k", type=int, default=5, dest="top_k")
    p.add_argument("--tag", type=str, default="run1")
    p.add_argument("--all-postures", action="store_true")

    p.add_argument("--device", type=str, default="auto")
    p.add_argument("--seed", type=int, default=42)

    p.add_argument("--deterministic", dest="deterministic", action="store_true", default=True)
    p.add_argument("--non-deterministic", dest="deterministic", action="store_false")

    p.add_argument("--keep-all", action="store_true", default=False,
                   help="Si se activa, devuelve también carpetas intermedias (no solo DELIVERABLE).")
    p.add_argument("--export-audit-csv", action="store_true", default=False)

    return p.parse_args()


def main():
    args = parse_args()

    # Backward compatibility: --vecnorm (legacy) -> --vn-micro (nuevo)
    if (not getattr(args, "vn_micro", None)) and getattr(args, "vecnorm", None):
        args.vn_micro = args.vecnorm

    deliverable_only = (not bool(args.keep_all))

    if getattr(args, "all_postures", False):
        cfg_base = CoordinatorInferenceConfig(
            model_path_micro=args.model_micro,
            portfolio_path=args.portfolio,
            vecnormalize_path_micro=args.vn_micro if (args.vn_micro and os.path.exists(args.vn_micro)) else None,
            model_path_macro=args.model_macro if (args.model_macro and os.path.exists(args.model_macro)) else None,
            vecnormalize_path_macro=args.vn_macro if (args.vn_macro and os.path.exists(args.vn_macro)) else None,
            vecnormalize_path_loan=args.vn_loan if (args.vn_loan and os.path.exists(args.vn_loan)) else None,
            device=args.device,
            seed=args.seed,
            deterministic=bool(args.deterministic),
            tag=args.tag,
            n_steps=args.n_steps,
            top_k=args.top_k,
            deliverable_only=deliverable_only,
            export_audit_csv=bool(args.export_audit_csv),
        )
        outs = run_coordinator_inference_multi_posture(cfg_base)
        logger.info(f"[COMPLETED] Multi-postura completado. Outputs: {outs}")
    else:
        out_dir, excel_path = run_coordinator_inference(
            model_micro=args.model_micro,
            portfolio_path=args.portfolio,
            vecnorm_micro_path=args.vn_micro if (args.vn_micro and os.path.exists(args.vn_micro)) else None,
            model_macro=args.model_macro if (args.model_macro and os.path.exists(args.model_macro)) else None,
            risk_posture=args.risk_posture,
            n_steps=args.n_steps,
            top_k=args.top_k,
            tag=args.tag,
            base_output_dir=None,
            device=args.device,
            seed=args.seed,
            deterministic=bool(args.deterministic),
            export_audit_csv=bool(args.export_audit_csv),
            vecnorm_macro_path=args.vn_macro if (args.vn_macro and os.path.exists(args.vn_macro)) else None,
            vecnorm_loan_path=args.vn_loan if (args.vn_loan and os.path.exists(args.vn_loan)) else None,
        )
        logger.info(f"[COMPLETED] Inferencia coordinada completada. Carpeta: {out_dir}")
        logger.info(f"[EXCEL] Excel final: {excel_path}")


if __name__ == "__main__":
    main()
