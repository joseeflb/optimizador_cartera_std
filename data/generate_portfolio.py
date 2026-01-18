# -*- coding: utf-8 -*-
"""
===============================================================================
 data/generate_portfolio.py — Generador de cartera sintética NPL (STD · Basilea III)
-------------------------------------------------------------------------------
CORREGIDO (Banco L1.5):
 - 100% préstamos en DEFAULT (DPD >= 90)
 - RW calculado vía BaselSTDMapping.resolve_rw(estado=DEFAULT)
     • DEFAULT + mortgage residencial -> RW = 1.00
     • DEFAULT + resto               -> RW = 1.50
 - PD NO se fuerza a 100%:
     • En NPL, PD se interpreta como riesgo forward (no-curación / pérdida adicional)
 - Campos de compatibilidad inferencia/RL:
     • segment_raw, segment, segmento_banco, segmento_id (numérico)
 - FIXES (según tus sanity checks):
     • segmento_id ahora es numérico (estable) + segmento_banco conserva string
     • RW deja de ser degenerado: se introduce segmento Mortgage (RW=1.00 en default)
     • NI/ EVA dejan de ser “casi siempre negativos”:
         - EL se interpreta como pérdida esperada ACUMULADA a horizonte (PD forward),
           y se anualiza para construir NI proxy (sin salir de NPL/DEFAULT/STD)
 - NUEVO (contable bank-ready):
     • coverage_rate, provisions, book_value (NCA) para poder valorar venta vs book.
 - NUEVO (viabilidad reestructuración, elimina NaNs):
     • maturity_months_remaining, monthly_payment
     • monthly_income (Mortgage/retail-proxy) => PTI_pre
     • monthly_cfo (Corporate/SME/PF)         => DSCR_pre
     • PTI_pre y DSCR_pre NO quedan todos NaN (audit-ready)
===============================================================================
"""
from __future__ import annotations

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import argparse
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import config as cfg


# ---------------------------------------------------------------------
# 🔧 Inicialización
# ---------------------------------------------------------------------
cfg.set_all_seeds(cfg.GLOBAL_SEED)

HURDLE    = float(cfg.CONFIG.regulacion.hurdle_rate)
CAP_RATIO = float(cfg.CONFIG.regulacion.required_total_capital_ratio())

COST_FUND = 0.006  # coste de funding consistente con LoanEnv/PortfolioEnv

# Interpretación consistente con tu POC:
# PD en NPL = PD forward (12–24m). Para construir NI "anual" proxy, anualizamos EL
# usando el horizonte del bloque de sensibilidad (por defecto 24m en config).
HORIZON_YEARS = float(max(1.0, (cfg.CONFIG.sensibilidad_reestructura.horizon_months / 12.0)))


# ---------------------------------------------------------------------
# 🧱 Segmentos sintéticos (legacy -> Enum Segmento)
# ---------------------------------------------------------------------
SEGMENT_WEIGHTS = {
    "Large Corporate": 0.22,
    "Corporate":       0.30,
    "MidCap":          0.18,
    "SME":             0.15,
    "Project Finance": 0.05,
    "Mortgage":        0.10,
}

SEGMENT_TO_ENUM = {
    "Corporate":       cfg.Segmento.CORPORATE,
    "Large Corporate": cfg.Segmento.CORPORATE,
    "MidCap":          cfg.Segmento.SME,
    "SME":             cfg.Segmento.SME,
    "Project Finance": cfg.Segmento.OTHER,
    "Mortgage":        cfg.Segmento.MORTGAGE,
}

# segmento_id debe ser numérico (feature estable para RL)
SEGMENTO_ID_NUM = {
    cfg.Segmento.CORPORATE: 1,
    cfg.Segmento.SME: 2,
    cfg.Segmento.RETAIL: 3,
    cfg.Segmento.MORTGAGE: 4,
    cfg.Segmento.CONSUMER: 5,
    cfg.Segmento.SOVEREIGN: 6,
    cfg.Segmento.BANK: 7,
    cfg.Segmento.LEASING: 8,
    cfg.Segmento.OTHER: 9,
}

# Probabilidad de colateral (proxy) por segmento raw
SECURED_PROB = {
    "Large Corporate": 0.35,
    "Corporate": 0.40,
    "MidCap": 0.45,
    "SME": 0.55,
    "Project Finance": 0.65,
    "Mortgage": 0.95,
}


def simulate_coverage_rate(
    rng_local: np.random.Generator,
    seg_enum: cfg.Segmento,
    secured: int,
    dpd: int,
    lgd: float,
) -> float:
    """
    Coverage / provisiones proxy (Stage 3 / NPL) para poder definir book_value (NCA).
    - Más DPD => más cobertura
    - Más LGD => más cobertura
    - Secured/Mortgage => menos cobertura (por colateral)
    """
    lgd = float(np.clip(lgd, 0.0, 1.0))
    base = 0.35 + 0.40 * lgd  # 0.35..0.75 aprox

    if dpd < 180:
        dpd_add = 0.00
    elif dpd < 360:
        dpd_add = 0.06
    else:
        dpd_add = 0.12

    sec_adj = -0.07 if int(secured) == 1 else 0.03
    seg_adj = -0.05 if seg_enum == cfg.Segmento.MORTGAGE else 0.00

    noise = float(rng_local.normal(0.0, 0.03))
    cov = base + dpd_add + sec_adj + seg_adj + noise

    # clamps conservadores
    return float(np.clip(cov, 0.20, 0.95))


# ---------------------------------------------------------------------
# 🧱 Rating -> base ranges (performing-ish) para luego “NPL-izar”
# ---------------------------------------------------------------------
RATING_DIST = {
    "AAA": 0.02, "AA": 0.05, "A": 0.10,
    "BBB": 0.28, "BB": 0.25, "B": 0.20, "CCC": 0.10
}

# (pd_min, pd_max, lgd_min, lgd_max) base (pre-default)
RATING_MAP = {
    "AAA": (0.0005, 0.005, 0.18, 0.23),
    "AA":  (0.0005, 0.008, 0.20, 0.26),
    "A":   (0.001,  0.015, 0.22, 0.30),
    "BBB": (0.005,  0.025, 0.22, 0.32),
    "BB":  (0.02,   0.06,  0.30, 0.40),
    "B":   (0.06,   0.15,  0.40, 0.55),
    "CCC": (0.15,   0.30,  0.50, 0.65),
}

RATING_NUM = {"AAA": 9, "AA": 8, "A": 7, "BBB": 6, "BB": 5, "B": 4, "CCC": 3}


# ---------------------------------------------------------------------
# 🧱 Productos
# ---------------------------------------------------------------------
PRODUCT_TYPES = [
    "Avales", "Bono", "Cesión de créditos", "Confirming", "Crediglobal",
    "Línea de crédito", "Papel Comercial", "Póliza de crédito",
    "Préstamo", "Préstamo hipotecario", "RCF",
]

PRODUCT_WEIGHTS = {
    "Avales": 0.06, "Bono": 0.08, "Cesión de créditos": 0.06,
    "Confirming": 0.10, "Crediglobal": 0.06,
    "Línea de crédito": 0.15, "Papel Comercial": 0.07,
    "Póliza de crédito": 0.15, "Préstamo": 0.15,
    "Préstamo hipotecario": 0.07, "RCF": 0.05
}

PRODUCT_EAD_MULT = {
    "Avales": 0.6, "Bono": 1.6, "Cesión de créditos": 0.8,
    "Confirming": 0.7, "Crediglobal": 0.7,
    "Línea de crédito": 0.9, "Papel Comercial": 1.3,
    "Póliza de crédito": 0.9, "Préstamo": 1.0,
    "Préstamo hipotecario": 1.2, "RCF": 1.0,
}


# ---------------------------------------------------------------------
# ⚙️ Funciones auxiliares
# ---------------------------------------------------------------------
def simulate_pd_lgd_base(rng_local: np.random.Generator, rating: str) -> tuple[float, float]:
    """Base (pre-default) por rating, luego se ajusta a NPL."""
    if rating not in RATING_MAP:
        rating = "BBB"
    pd_min, pd_max, lgd_min, lgd_max = RATING_MAP[rating]
    return float(rng_local.uniform(pd_min, pd_max)), float(rng_local.uniform(lgd_min, lgd_max))


def sample_dpd_default(rng_local: np.random.Generator) -> int:
    """
    NPL: SIEMPRE DEFAULT -> DPD >= 90
    Distribución típica:
      - 60%:  90-180
      - 30%: 180-360
      - 10%: 360-720
    """
    u = float(rng_local.random())
    if u < 0.60:
        return int(rng_local.integers(90, 180))
    if u < 0.90:
        return int(rng_local.integers(180, 360))
    return int(rng_local.integers(360, 720))


def simulate_rate(segment: str, rating: str, dpd: int) -> float:
    """
    Cupón contractual aproximado (no market yield), con penalización por mora.
    Nota NPL: aquí se usa como proxy de “yield contractual” (no implica cobro real).
    """
    base = {
        "Large Corporate": 0.030,
        "Corporate": 0.035,
        "MidCap": 0.040,
        "SME": 0.045,
        "Project Finance": 0.040,
        "Mortgage": 0.025,
    }
    spread = {
        "AAA": 0.007, "AA": 0.009, "A": 0.012,
        "BBB": 0.018, "BB": 0.030, "B": 0.050, "CCC": 0.080,
    }
    b = float(base.get(segment, 0.040))
    s = float(spread.get(rating, 0.018))
    r = b + s
    if dpd >= 180:
        r += 0.02
    if dpd >= 360:
        r += 0.02
    return float(min(r, 0.22))


def simulate_ead(rng_local: np.random.Generator, segment: str, product: str) -> float:
    params = {
        "Large Corporate": {"median": 20_000_000, "sigma": 0.60},
        "Corporate":       {"median": 5_000_000,  "sigma": 0.75},
        "MidCap":          {"median": 1_500_000,  "sigma": 0.80},
        "SME":             {"median": 400_000,    "sigma": 0.80},
        "Project Finance": {"median": 8_000_000,  "sigma": 0.70},
        "Mortgage":        {"median": 250_000,    "sigma": 0.60},
    }
    p = params.get(segment, {"median": 1_000_000, "sigma": 0.80})
    mu = float(np.log(p["median"]))
    ead_base = float(rng_local.lognormal(mean=mu, sigma=float(p["sigma"])))
    mult = float(PRODUCT_EAD_MULT.get(product, 1.0))
    return float(round(ead_base * mult))


def nplize_pd_lgd(
    pd_base: float,
    lgd_base: float,
    dpd: int,
    rating: str,
    npl: cfg.NPLConventions
) -> tuple[float, float]:
    """
    Ajuste NPL coherente con config.NPLConventions:
      - PD forward NPL acotada a [pd_default_floor, pd_default_cap]
      - LGD NPL acotada a [lgd_default_floor, lgd_default_cap]
      - DPD empuja severidad de PD/LGD (sin romper los clamps)

    Nota: NO es PD=100%. DEFAULT es estado; PD es forward.
    """
    rating_add = {
        "AAA": 0.00, "AA": 0.01, "A": 0.02,
        "BBB": 0.04, "BB": 0.06, "B": 0.08, "CCC": 0.10
    }.get(rating, 0.05)

    # Severidad por DPD: empuja PD y LGD, pero siempre dentro de clamps NPL
    if dpd < 180:
        dpd_add = 0.00
        lgd_mult = 1.00
    elif dpd < 360:
        dpd_add = 0.05
        lgd_mult = 1.03
    else:
        dpd_add = 0.10
        lgd_mult = 1.06

    pd_raw = float(max(pd_base, npl.pd_default_floor + rating_add + dpd_add))
    lgd_raw = float(lgd_base * lgd_mult)

    pd_npl = cfg.clamp_npl_pd(pd_raw, npl)
    lgd_npl = cfg.clamp_npl_lgd(lgd_raw, npl)
    return float(pd_npl), float(lgd_npl)


# ---------------------------------------------------------------------
# 🧮 Viabilidad (PTI/DSCR) - utilidades
# ---------------------------------------------------------------------
def simulate_remaining_months(rng_local: np.random.Generator, seg_raw: str) -> int:
    """
    Tenor remanente (meses). Es proxy para calcular cuota contractual.
    """
    ranges = {
        "Large Corporate": (24, 84),
        "Corporate":       (24, 72),
        "MidCap":          (18, 60),
        "SME":             (12, 48),
        "Project Finance": (60, 180),
        "Mortgage":        (120, 360),
    }
    lo, hi = ranges.get(seg_raw, (24, 60))
    return int(rng_local.integers(lo, hi))


def annuity_payment(principal: float, annual_rate: float, n_months: int) -> float:
    """
    Cuota mensual aproximada (anual_rate nominal, n_months remanentes).
    """
    n = int(max(1, n_months))
    r_m = float(max(0.0, annual_rate) / 12.0)
    if r_m <= 1e-9:
        return float(principal / n)
    denom = 1.0 - (1.0 + r_m) ** (-n)
    if abs(denom) < 1e-9:
        return float(principal / n)
    return float(principal * r_m / denom)


def dpd_bucket(dpd: int) -> int:
    if dpd < 180:
        return 0
    if dpd < 360:
        return 1
    return 2


def rating_bucket(rating: str) -> int:
    """
    0 = investment-ish (AAA..A)
    1 = BBB/BB
    2 = B/CCC
    """
    r = rating.upper()
    if r in ("AAA", "AA", "A"):
        return 0
    if r in ("BBB", "BB"):
        return 1
    return 2


# ---------------------------------------------------------------------
# 🧮 Generador principal (100% DEFAULT + RW 100/150 STD)
# ---------------------------------------------------------------------
def generate_portfolio(n: int = 500, seed: int | None = None) -> pd.DataFrame:
    rng_local = np.random.default_rng(cfg.GLOBAL_SEED if seed is None else seed)

    segments = rng_local.choice(
        list(SEGMENT_WEIGHTS.keys()), size=n, p=list(SEGMENT_WEIGHTS.values())
    )
    ratings = rng_local.choice(
        list(RATING_DIST.keys()), size=n, p=list(RATING_DIST.values())
    )
    productos = rng_local.choice(
        PRODUCT_TYPES, size=n, p=list(PRODUCT_WEIGHTS.values())
    )

    rows: list[dict] = []
    npl_conv = cfg.CONFIG.npl

    for i in range(n):
        seg_raw = str(segments[i])
        rating = str(ratings[i]).upper()
        product = str(productos[i])

        # Coherencia mínima: Mortgage segment -> producto hipotecario
        if seg_raw == "Mortgage":
            product = "Préstamo hipotecario"

        seg_enum = SEGMENT_TO_ENUM.get(seg_raw, cfg.Segmento.OTHER)

        # --- DEFAULT forzado ---
        dpd = int(sample_dpd_default(rng_local))
        dpd = int(cfg.clamp_dpd(float(dpd), npl_conv))
        dpd = max(dpd, 90)
        estado = cfg.EstadoCredito.DEFAULT

        # PD/LGD base y ajuste NPL (clamp centralizado en config)
        pd_base, lgd_base = simulate_pd_lgd_base(rng_local, rating)
        pd_val, lgd_val = nplize_pd_lgd(pd_base, lgd_base, dpd, rating, npl_conv)

        # EAD / rate
        ead = simulate_ead(rng_local, seg_raw, product)
        rate = simulate_rate(seg_raw, rating, dpd)

        # --- RW STD en DEFAULT ---
        rw = cfg.CONFIG.basel_map.resolve_rw(
            segmento=seg_enum,
            rating=rating,
            estado=estado,
            secured_by_mortgage=(seg_enum == cfg.Segmento.MORTGAGE),
        )

        # Validación dura: RW ∈ [1.0, 1.5]
        rw_f = float(rw)
        if not (1.0 <= rw_f <= 1.5):
            raise ValueError(
                f"RW fuera de rango DEFAULT STD: rw={rw_f} seg={seg_enum} rating={rating} loan_id=L{i:06d}"
            )

        # Proxy colateral (independiente de segmento_id)
        secured = int(rng_local.random() < float(SECURED_PROB.get(seg_raw, 0.40)))
        if seg_enum == cfg.Segmento.MORTGAGE:
            secured = 1

        # -----------------------------------------------------------------
        # NUEVO: capa contable bank-ready (coverage/provisiones/book)
        # -----------------------------------------------------------------
        coverage_rate = simulate_coverage_rate(rng_local, seg_enum, secured, dpd, lgd_val)
        provisions = float(ead * coverage_rate)
        book_value = float(max(ead - provisions, 0.0))

        # -----------------------------------------------------------------
        # NUEVO: viabilidad (PTI/DSCR) — evita NaNs masivos en audit
        #   - Mortgage: retail-proxy => PTI_pre calculable
        #   - Resto: corporate-proxy => DSCR_pre calculable
        # -----------------------------------------------------------------
        tenor_m = simulate_remaining_months(rng_local, seg_raw)
        # evita pagos absurdos por rates muy bajos: floor 2%
        rate_eff = float(max(rate, 0.02))
        monthly_payment = float(annuity_payment(float(ead), rate_eff, tenor_m))
        monthly_payment = float(max(monthly_payment, 50.0))

        d_b = dpd_bucket(dpd)
        r_b = rating_bucket(rating)

        # Targets “bank plausible” (sólo para señal/guardrail en sintético)
        # PTI: cuanto peor DPD/rating -> mayor PTI (más estrés)
        pti_target = float(np.clip(0.30 + 0.05 * d_b + 0.03 * r_b + rng_local.normal(0.0, 0.03), 0.18, 0.70))
        # DSCR: cuanto peor DPD/rating -> menor DSCR
        dscr_target = float(np.clip(1.20 - 0.12 * d_b - 0.10 * r_b + rng_local.normal(0.0, 0.10), 0.55, 2.20))

        # mortgage => income definido, DSCR NaN (no aplica)
        if seg_enum == cfg.Segmento.MORTGAGE:
            monthly_income = float(monthly_payment / max(pti_target, 1e-6))
            monthly_cfo = np.nan
            viability_model = "PTI"
            pti_pre = float(monthly_payment / monthly_income) if monthly_income > 0 else np.nan
            dscr_pre = np.nan
        else:
            monthly_income = np.nan
            monthly_cfo = float(monthly_payment * max(dscr_target, 0.01))
            viability_model = "DSCR"
            pti_pre = np.nan
            dscr_pre = float(monthly_cfo / monthly_payment) if monthly_payment > 0 else np.nan

        # -----------------------------------------------------------------
        # Métricas coherentes para RL en NPL:
        #   - EL = EAD * PD * LGD  (pérdida esperada ACUMULADA al horizonte)
        #   - Para NI proxy (anual), anualizamos EL por el horizonte (24m por defecto).
        # -----------------------------------------------------------------
        el_lifetime = float(ead * pd_val * lgd_val)
        el_annual = float(el_lifetime / HORIZON_YEARS)

        ni = float(ead * rate - el_annual - ead * COST_FUND)

        rwa = float(ead * rw_f)
        rorwa = float(ni / rwa) if rwa > 0 else 0.0
        eva = float(rwa * (rorwa - HURDLE))
        rona = float(ni / ead) if ead > 0 else 0.0

        rows.append({
            "loan_id": f"L{i:06d}",

            # Segmentación compat:
            "segment_raw": seg_raw,                  # etiqueta legacy
            "segment": seg_raw,                      # compat legacy
            "segmento_banco": seg_enum.value,        # string enum (audit)
            "segmento_id": float(SEGMENTO_ID_NUM.get(seg_enum, 9)),  # numérico RL

            "producto": product,
            "rating": rating,
            "rating_num": float(RATING_NUM.get(rating, 6)),

            "EAD": float(ead),
            "PD": float(pd_val),
            "LGD": float(lgd_val),

            "RW": float(rw_f),   # 1.00 o 1.50 (DEFAULT STD según tu mapping)
            "RWA": float(rwa),

            "rate": float(rate),
            "DPD": float(dpd),
            "meses_en_default": int(dpd // 30),

            "secured": int(secured),

            # NUEVO contable (para venta vs book)
            "coverage_rate": float(coverage_rate),
            "provisions": float(provisions),
            "book_value": float(book_value),

            # NUEVO viabilidad (para reestructura)
            "maturity_months_remaining": int(tenor_m),
            "monthly_payment": float(monthly_payment),
            "monthly_income": (float(monthly_income) if not (isinstance(monthly_income, float) and np.isnan(monthly_income)) else np.nan),
            "monthly_cfo": (float(monthly_cfo) if not (isinstance(monthly_cfo, float) and np.isnan(monthly_cfo)) else np.nan),
            "viability_model": str(viability_model),
            "PTI_pre": float(pti_pre) if not (isinstance(pti_pre, float) and np.isnan(pti_pre)) else np.nan,
            "DSCR_pre": float(dscr_pre) if not (isinstance(dscr_pre, float) and np.isnan(dscr_pre)) else np.nan,

            # Económicos (audit)
            "EL": float(el_lifetime),        # acumulada a horizonte
            "EL_annual": float(el_annual),   # anualizada para NI proxy
            "NI": float(ni),
            "EVA": float(eva),
            "RORWA": float(rorwa),
            "RONA": float(rona),

            # Estado explícito
            "estado": estado.value,
        })

    df = pd.DataFrame(rows)

    # Sanity checks (duros)
    if (df["DPD"] < 90).any():
        raise AssertionError("Hay préstamos con DPD < 90; no son DEFAULT.")
    if (df["RW"] < 1.0).any() or (df["RW"] > 1.5).any():
        raise AssertionError("Hay RW fuera de [1.0, 1.5]; revisar mapping DEFAULT.")
    if df["segmento_id"].dtype == object:
        raise AssertionError("segmento_id debe ser numérico (float/int) para RL.")
    if (df["coverage_rate"] < 0.0).any() or (df["coverage_rate"] > 1.0).any():
        raise AssertionError("coverage_rate fuera de [0,1].")
    if (df["book_value"] < 0.0).any() or (df["book_value"] > df["EAD"]).any():
        raise AssertionError("book_value fuera de rango [0,EAD].")

    # NUEVO: evitar NaNs masivos (al menos debe haber PTI en Mortgage y DSCR en corporates)
    if df["PTI_pre"].notna().sum() == 0:
        raise AssertionError("PTI_pre es NaN en el 100%: revisar generación (Mortgage/PTI).")
    if df["DSCR_pre"].notna().sum() == 0:
        raise AssertionError("DSCR_pre es NaN en el 100%: revisar generación (Corporate/DSCR).")

    return df


# ---------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------
def export_excel(df: pd.DataFrame, out_path: str):
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    df.to_excel(out_path, index=False, sheet_name="Portfolio")

    wb = load_workbook(out_path)
    ws = wb.active
    ws.title = "Portfolio"

    header_fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    euro = '#,##0.00" €"'
    pct = "0.00%"

    fmt = {
        "EAD": euro, "RWA": euro, "EVA": euro, "NI": euro, "EL": euro, "EL_annual": euro,
        "PD": pct, "LGD": pct, "RW": pct,
        "rate": pct, "RORWA": pct, "RONA": pct,

        # NUEVO contable
        "coverage_rate": pct,
        "book_value": euro,
        "provisions": euro,

        # NUEVO viabilidad
        "monthly_payment": euro,
        "monthly_income": euro,
        "monthly_cfo": euro,
        "PTI_pre": "0.00",
        "DSCR_pre": "0.00",
        "maturity_months_remaining": "0",
    }

    grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_i, col_name in enumerate(df.columns, start=1):
        max_len = len(col_name)
        for row_i in range(2, ws.max_row + 1):
            c = ws.cell(row=row_i, column=col_i)
            if col_name in fmt:
                c.number_format = fmt[col_name]
            c.border = thin
            if row_i % 2 == 0:
                c.fill = grey
            max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 3, 28)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Hoja de resumen
    ws2 = wb.create_sheet("Resumen_Portfolio")
    ws2["A1"] = "Resumen financiero (STD Basilea III) — Cartera 100% DEFAULT"
    ws2["A1"].font = Font(bold=True, size=12, color="244062")

    summary = {
        "Número de préstamos": len(df),
        "EAD total (€)": df["EAD"].sum(),
        "RWA total (€)": df["RWA"].sum(),
        "Capital bloqueado total (€)": df["RWA"].sum() * CAP_RATIO,
        "PD media (%)": df["PD"].mean(),
        "LGD media (%)": df["LGD"].mean(),
        "RORWA medio (%)": df["RORWA"].mean(),
        "EVA medio (€)": df["EVA"].mean(),
        "DPD medio (días)": df["DPD"].mean(),
        "RW min": df["RW"].min(),
        "RW max": df["RW"].max(),
        "Coverage medio (%)": df["coverage_rate"].mean(),
        "Book value total (€)": df["book_value"].sum(),
        "Provisiones total (€)": df["provisions"].sum(),
        "Horizon years (EL annualization)": HORIZON_YEARS,

        # NUEVO viabilidad
        "PTI_pre (Mortgage) media": df["PTI_pre"].mean(skipna=True),
        "DSCR_pre (Corporate) media": df["DSCR_pre"].mean(skipna=True),
        "Cuota mensual media (€)": df["monthly_payment"].mean(),
        "Tenor remanente medio (meses)": df["maturity_months_remaining"].mean(),
    }

    r = 3
    for k, v in summary.items():
        ws2[f"A{r}"] = k
        ws2[f"B{r}"] = float(v) if isinstance(v, (int, float, np.floating)) else v
        if "€" in k:
            ws2[f"B{r}"].number_format = euro
        elif "(%)" in k:
            ws2[f"B{r}"].number_format = pct
        elif "PTI_pre" in k or "DSCR_pre" in k:
            ws2[f"B{r}"].number_format = "0.00"
        else:
            ws2[f"B{r}"].number_format = "0.00"
        r += 1

    wb.save(out_path)
    wb.close()
    print(f"💾 Cartera exportada correctamente → {out_path}")


# ---------------------------------------------------------------------
# ▶️ CLI
# ---------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description="Genera cartera sintética NPL 100% DEFAULT (STD Basilea III · Banco L1.5)"
    )
    ap.add_argument("--n", type=int, default=500)
    ap.add_argument("--seed", type=int, default=cfg.GLOBAL_SEED)
    ap.add_argument("--out", type=str, default=os.path.join("data", "portfolio_synth.xlsx"))
    args = ap.parse_args()

    df = generate_portfolio(n=args.n, seed=args.seed)
    export_excel(df, args.out)

    print(
        f"✅ Cartera NPL (100% DEFAULT) generada "
        f"({len(df)} préstamos) → {args.out} | RW∈[{df['RW'].min():.2f}, {df['RW'].max():.2f}]"
    )


if __name__ == "__main__":
    main()
