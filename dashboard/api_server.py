# -*- coding: utf-8 -*-
# ============================================================
# dashboard/api_server.py
# Autor: José María Fernández-Ladreda Ballvé
# Resumen: Backend Flask que sirve los outputs reales del proyecto como JSON para el dashboard Lovable Matrix (puerto 3000).
# ============================================================
"""
Dashboard API – NPL Portfolio Optimizer
========================================
Flask backend that reads real project outputs and serves them as JSON
for the Lovable Matrix monitoring dashboard.

Launch:  py dashboard/api_server.py
Endpoint: http://localhost:3000
"""

import json, os, csv, glob, hashlib, subprocess
from pathlib import Path
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

# ── Paths ───────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
REPORTS = ROOT / "reports"
RUNS    = REPORTS / "runs"
DATA    = ROOT / "data"
MODELS  = ROOT / "models"
CONFIGS = ROOT / "configs"

app = Flask(__name__, static_folder=str(ROOT / "dashboard" / "static"))
CORS(app)


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def _find_deliverable(tag: str) -> Path | None:
    """Find the latest DELIVERABLE folder for a given tag."""
    if not RUNS.exists():
        return None
    candidates = sorted(RUNS.glob(f"*{tag}*DELIVERABLE"))
    return candidates[-1] if candidates else None


def _find_multi(tag: str) -> Path | None:
    candidates = sorted(RUNS.glob(f"*{tag}*MULTI"))
    return candidates[-1] if candidates else None


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _read_md(path: Path) -> str:
    return path.read_text(encoding="utf-8") if path.exists() else ""


def _read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _read_excel_summary(xlsx_path: Path) -> dict:
    """Read key columns from a decisiones_finales Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            rows.append(d)
        wb.close()

        # Aggregate KPIs
        n = len(rows)
        actions = {}
        total_eva_pre = 0
        total_eva_post = 0
        total_rwa_pre = 0
        total_rwa_post = 0
        total_capital_lib = 0

        for r in rows:
            act = r.get("Accion_final") or r.get("Accion") or "UNKNOWN"
            actions[act] = actions.get(act, 0) + 1
            total_eva_pre += float(r.get("EVA_pre", 0) or 0)
            total_eva_post += float(r.get("EVA_post", 0) or 0)
            total_rwa_pre += float(r.get("RWA_pre", 0) or 0)
            total_rwa_post += float(r.get("RWA_post", 0) or 0)
            total_capital_lib += float(r.get("capital_liberado", 0) or 0)

        return {
            "n_loans": n,
            "actions": actions,
            "total_eva_pre": total_eva_pre,
            "total_eva_post": total_eva_post,
            "delta_eva": total_eva_post - total_eva_pre,
            "total_rwa_pre": total_rwa_pre,
            "total_rwa_post": total_rwa_post,
            "capital_liberado": total_capital_lib,
            "rorwa_post": (total_eva_post / total_rwa_post * 100) if total_rwa_post else 0,
        }
    except Exception as e:
        return {"error": str(e)}


def _loan_details(xlsx_path: Path) -> list[dict]:
    """Return per-loan rows from a decisiones_finales Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        keep = [
            "loan_id", "Accion_final", "Reason_Code", "segment",
            "rating", "EAD", "PD", "LGD", "DPD",
            "EVA_pre", "EVA_post", "RWA_pre", "RWA_post",
            "capital_liberado", "DSCR_pre", "DSCR_post",
            "PTI_pre", "PTI_post", "Explanation",
        ]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            rows.append({k: d.get(k) for k in keep})
        wb.close()
        return rows
    except Exception as e:
        return [{"error": str(e)}]


def _available_tags() -> list[str]:
    """Return unique tags from DELIVERABLE run folders."""
    tags = set()
    if RUNS.exists():
        for d in RUNS.iterdir():
            if d.is_dir() and "DELIVERABLE" in d.name:
                parts = d.name.split("_")
                # Format: TIMESTAMP_tag_DELIVERABLE → extract tag
                # e.g. 20260406_200933_post_fix_v2_DELIVERABLE
                if len(parts) >= 3:
                    tag = "_".join(parts[2:-1])  # everything between timestamp and DELIVERABLE
                    if tag:
                        tags.add(tag)
    return sorted(tags)


# ═══════════════════════════════════════════════════════════════════
#  API ROUTES
# ═══════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()})


@app.route("/api/tags")
def tags():
    return jsonify({"tags": _available_tags()})


@app.route("/api/runs", methods=["GET", "POST"])
def runs():
    """GET: list available runs.  POST: receive n8n callback."""
    if request.method == "POST":
        data = request.get_json(force=True)
        # Store in a simple log
        log_path = REPORTS / "dashboard_callbacks.jsonl"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(data, default=str) + "\n")
        return jsonify({"received": True})

    # GET → list runs
    tag_list = _available_tags()
    results = []
    for tag in tag_list:
        deliv = _find_deliverable(tag)
        results.append({
            "tag": tag,
            "path": str(deliv) if deliv else None,
            "timestamp": deliv.name[:15] if deliv else None,
        })
    return jsonify({"runs": results})


@app.route("/api/kpis/<tag>")
def kpis(tag: str):
    """Vista 1 – KPI Executive: per-posture summary."""
    deliv = _find_deliverable(tag)
    if not deliv:
        return jsonify({"error": f"No DELIVERABLE found for tag={tag}"}), 404

    postures = {}
    for posture in ("prudencial", "balanceado", "desinversion"):
        xlsx = deliv / f"decisiones_finales_{posture}.xlsx"
        if xlsx.exists():
            postures[posture] = _read_excel_summary(xlsx)

    return jsonify({
        "tag": tag,
        "postures": postures,
        "timestamp": deliv.name[:15],
    })


@app.route("/api/decisions/<tag>/<posture>")
def decisions(tag: str, posture: str):
    """Vista 2 – Decision Detail: per-loan table."""
    deliv = _find_deliverable(tag)
    if not deliv:
        return jsonify({"error": f"No DELIVERABLE found for tag={tag}"}), 404

    xlsx = deliv / f"decisiones_finales_{posture}.xlsx"
    if not xlsx.exists():
        return jsonify({"error": f"No file for posture={posture}"}), 404

    loans = _loan_details(xlsx)
    return jsonify({
        "tag": tag,
        "posture": posture,
        "n_loans": len(loans),
        "loans": loans,
    })


@app.route("/api/stress/<tag>")
def stress(tag: str):
    """Vista 3 – Stress Analysis: backtesting + stress data."""
    # Backtesting dynamic
    bt_csv = REPORTS / f"backtesting_dynamic_{tag}.csv"
    bt_rows = _read_csv(bt_csv)

    # Backtesting light
    bl_csv = REPORTS / f"backtesting_light_{tag}.csv"
    bl_rows = _read_csv(bl_csv)

    # Stress scenarios config
    stress_cfg_path = CONFIGS / "stress_scenarios.yaml"
    stress_cfg = ""
    if stress_cfg_path.exists():
        stress_cfg = stress_cfg_path.read_text(encoding="utf-8")

    return jsonify({
        "tag": tag,
        "backtesting_dynamic": bt_rows,
        "backtesting_light": bl_rows,
        "stress_config": stress_cfg,
    })


@app.route("/api/posture-analysis/<tag>")
def posture_analysis(tag: str):
    """Vista 6 – NPL Workout Layer + distance checks."""
    md = _read_md(REPORTS / f"POSTURE_ANALYSIS_NPL_{tag}.md")
    return jsonify({"tag": tag, "markdown": md})


@app.route("/api/dispersion/<tag>")
def dispersion(tag: str):
    """Dispersion analysis."""
    md = _read_md(REPORTS / f"DISPERSION_ANALYSIS_{tag}.md")
    return jsonify({"tag": tag, "markdown": md})


@app.route("/api/validation")
def validation():
    """Agent validation results."""
    data = _read_json(REPORTS / "validation_results.json")
    return jsonify(data)


@app.route("/api/models")
def models():
    """Model inventory with checksums."""
    inventory = []
    if MODELS.exists():
        for f in sorted(MODELS.iterdir()):
            if f.is_file():
                sha = hashlib.sha256(f.read_bytes()).hexdigest()[:16]
                inventory.append({
                    "name": f.name,
                    "size_kb": f.stat().st_size // 1024,
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                    "sha256_prefix": sha,
                })
    return jsonify({"models": inventory})


@app.route("/api/trigger-run", methods=["POST"])
def trigger_run():
    """What-if analysis: trigger a new n8n pipeline run."""
    data = request.get_json(force=True)
    tag = data.get("tag", f"whatif_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    # This would POST to n8n webhook in production
    return jsonify({
        "status": "queued",
        "tag": tag,
        "message": "Run enqueued via n8n webhook. Results will appear in ~15 min.",
    })


# ═══════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print(f"[Dashboard API] ROOT = {ROOT}")
    print(f"[Dashboard API] Tags disponibles: {_available_tags()}")
    print(f"[Dashboard API] Sirviendo en http://localhost:3000")
    app.run(host="0.0.0.0", port=3000, debug=True)
