# -*- coding: utf-8 -*-
# ============================================
# reports/export_financial_decisions.py — Informe financiero detallado
# ============================================
"""
POC — OPTIMIZADOR DE CARTERAS EN DEFAULT (Método Estándar · Basilea III)
Genera un informe financiero individualizado por préstamo tras la inferencia.

Incluye:
- Justificación automática de decisión (mantener / reestructurar / vender)
- Cálculos de impacto económico: ΔEVA, ΔRORWA, ROI y capital liberado
- Exportación profesional a Excel y JSON

NTT Data | 2025
Autor: José María Fernández-Ladreda Ballvé
"""

from __future__ import annotations
import os, json, logging
from typing import Any, Dict
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# -----------------------------------------------------------
# 📣 Logging
# -----------------------------------------------------------
logger = logging.getLogger("export_financial_decisions")
if not logger.handlers:
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s"))
    logger.addHandler(h)
logger.setLevel(logging.INFO)

# -----------------------------------------------------------
# 🎨 Estilos y utilidades
# -----------------------------------------------------------
THIN = Side(border_style="thin", color="999999")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
HDR_FONT = Font(bold=True)
HDR_FILL = PatternFill("solid", fgColor="E6F0FF")

def _safe_div(a, b):
    """División protegida contra 0 y NaN."""
    with np.errstate(divide="ignore", invalid="ignore"):
        res = np.divide(a, b)
        res = np.where(~np.isfinite(res), 0.0, res)
    return res

# -----------------------------------------------------------
# 🧮 Enriquecimiento de decisiones
# -----------------------------------------------------------
def enrich_financial_decisions(df: pd.DataFrame) -> pd.DataFrame:
    """Añade métricas derivadas (ΔEVA, ROI, etc.) y texto justificativo."""
    df = df.copy()

    df["ΔEVA"] = df.get("EVA_post", np.nan) - df.get("EVA_pre", np.nan)
    df["ΔRORWA"] = df.get("RORWA_post", np.nan) - df.get("RORWA_pre", np.nan)
    df["ROI_%"] = _safe_div(df.get("EVA_post", 0), df.get("capital_liberado", 1)) * 100

    justificaciones = []
    for _, r in df.iterrows():
        act = str(r.get("Accion", "")).upper().strip()
        eva_pre = r.get("EVA_pre", 0)
        eva_post = r.get("EVA_post", 0)
        rorwa_pre = r.get("RORWA_pre", 0)
        capital_lib = r.get("capital_liberado", 0)

        if act == "MANTENER":
            msg = (
                f"✅ Se mantiene el préstamo: EVA={eva_pre:,.0f} €, "
                f"RORWA={rorwa_pre:.2%}, superior al hurdle. "
                f"El activo conserva valor económico y estabilidad regulatoria."
            )
        elif act == "REESTRUCTURAR":
            msg = (
                f"🟠 Se reestructura para mejorar EVA a {eva_post:,.0f} €. "
                f"Nuevo plazo={r.get('plazo_optimo','N/D')} meses, "
                f"tasa={r.get('tasa_nueva',r.get('tasa_anual','N/D')):.2%}, "
                f"quita={r.get('quita',0)*100:.1f} %. "
                f"Recalibración ejecutada por optimizador de reestructuración."
            )
        elif act == "VENDER":
            msg = (
                f"🔴 Se vende en mercado secundario (NPL): "
                f"precio óptimo={r.get('precio_optimo',0):,.0f} €, "
                f"liberando {capital_lib:,.0f} € de capital regulatorio "
                f"y mejorando la ratio CET1. "
                f"Estimación simulada por price_simulator.py."
            )
        else:
            msg = "⚪ Acción no identificada o sin datos suficientes."

        justificaciones.append(msg)

    df["Justificación"] = justificaciones
    return df

# -----------------------------------------------------------
# 📊 Exportación a Excel
# -----------------------------------------------------------
def export_financial_excel(df: pd.DataFrame, out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisiones Financieras"

    # Encabezados
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = HDR_FONT
        c.alignment = CENTER
        c.fill = HDR_FILL
        c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    # Datos
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            if isinstance(val, (int, float)):
                c.alignment = RIGHT
            else:
                c.alignment = WRAP
            c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    # Ajustes visuales
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["E"].width = 80

    wb.save(out_path)
    logger.info(f"✅ Excel de decisiones guardado en {out_path}")
    return os.path.abspath(out_path)

# -----------------------------------------------------------
# 💾 Exportación a JSON
# -----------------------------------------------------------
def export_financial_json(df: pd.DataFrame, out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    records = df.to_dict(orient="records")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    logger.info(f"💾 JSON guardado en {out_path}")
    return os.path.abspath(out_path)

# -----------------------------------------------------------
# 🚀 API principal
# -----------------------------------------------------------
def export_financial_decisions(df: pd.DataFrame, out_dir: str) -> Dict[str, Any]:
    """
    Enriquecer y exportar las decisiones financieras de inferencia.
    Retorna las rutas de salida (Excel + JSON).
    """
    os.makedirs(out_dir, exist_ok=True)
    logger.info("📊 Enriqueciendo decisiones con métricas financieras...")
    df_enriched = enrich_financial_decisions(df)

    excel_path = os.path.join(out_dir, "financial_decisions.xlsx")
    json_path = os.path.join(out_dir, "financial_decisions.json")

    export_financial_excel(df_enriched, excel_path)
    export_financial_json(df_enriched, json_path)

    logger.info("✅ Informe financiero detallado exportado correctamente.")
    return {"excel": excel_path, "json": json_path, "records": len(df_enriched)}

# -----------------------------------------------------------
# ▶️ CLI rápido
# -----------------------------------------------------------
if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Exportador de decisiones financieras detalladas")
    p.add_argument("--input", required=True, help="Ruta a decisiones_explicadas.xlsx o .csv")
    p.add_argument("--outdir", default="reports/inference_financial", help="Carpeta de salida")
    args = p.parse_args()

    df_in = pd.read_excel(args.input) if args.input.endswith(".xlsx") else pd.read_csv(args.input)
    out = export_financial_decisions(df_in, args.outdir)
    print(out)
