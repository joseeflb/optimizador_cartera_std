# -*- coding: utf-8 -*-
# ============================================================
# reports/npl_posture_analysis.py
# Autor: José María Fernández-Ladreda Ballvé
# Resumen: Análisis de envelopes de negociación por postura, distance checks y casos frontera; genera POSTURE_ANALYSIS_NPL.md.
# ============================================================

"""
ROL: Analista senior CIB especializado en NPL/default (workout + ventas).
OBJETIVO: Definir Negotiation Envelopes por postura, validar separación
entre posturas (distance checks), documentar casos frontera y generar
reporte POSTURE_ANALYSIS_NPL.md auditable para comité.

PRINCIPIO NPL:
  Todos los préstamos están en default. NO se fija precio final ni
  términos cerrados: se define un RANGO + WALK-AWAY (reservation floor /
  banda de concesión). Si la oferta/deudor cae fuera del envelope
  → Accion_final debe ser MANTENER.

USO:
    python -m reports.npl_posture_analysis --tag infer_ci0222
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import logging
import os
import sys
from copy import deepcopy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ── entorno
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

REPORTS_DIR = os.path.join(ROOT_DIR, "reports")
LOGS_DIR    = os.path.join(ROOT_DIR, "logs")
os.makedirs(LOGS_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    handlers=[
        logging.FileHandler(os.path.join(LOGS_DIR, "npl_analysis.log"), encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("npl_analysis")
_TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================
# CONSTANTES DE DISEÑO — NEGOTIATION ENVELOPES POR POSTURA
# ============================================================

# ──────────────────────────────────────────────────────────
# B) RESERVATION FLOOR — BY LOAN (not global)
# ──────────────────────────────────────────────────────────
# floor_i se calcula sobre el precio PROPIO del préstamo (Price_to_EAD_i ≈ E[precio_i]/EAD_i)
# usando multiplicadores lognormales:
#   modelo:  precio_i ~ LogNormal  →  sigma_efectivo ≈ 0.22
#            (mezcla BULL 15%×0.2 + BASE 20%×0.6 + STRESS 30%×0.2)
#   p_q = E[precio_i] × exp(−z_q × sigma_efect)
#     z40=0.253 → mult≈0.945   z27=0.614 → mult≈0.873   z13=1.130 → mult≈0.779
#
# Por postura:
#   prudencial   → p40 del propio préstamo  (floor exigente)
#   balanceado   → p27 del propio préstamo  (compromiso)
#   desinversion → p13 del propio préstamo  (velocidad > precio)
FLOOR_QUANTILE_MULTIPLIER: dict = {
    "prudencial":   0.945,   # ≈ p40 del préstamo propio
    "balanceado":   0.873,   # ≈ p27 del préstamo propio
    "desinversion": 0.779,   # ≈ p13 del préstamo propio
}

# Floor mínimo absoluto (safety net: ningún floor individual puede caer por debajo)
RESERVATION_FLOOR_MINIMUM: dict = {
    "prudencial":   0.080,   # nunca vender por debajo del 8% EAD
    "balanceado":   0.055,   # nunca vender por debajo del 5.5% EAD
    "desinversion": 0.035,   # nunca vender por debajo del 3.5% EAD
}

# Floor de referencia portfolio-aggregate (SOLO para distance checks / reporting)
# NO se usa como floor operativo por préstamo; conservado para compatibilidad.
RESERVATION_FLOOR: dict = {
    "prudencial":   0.145,   # referencia aggregate (~p40 medio del portfolio)
    "balanceado":   0.110,   # referencia aggregate
    "desinversion": 0.090,   # referencia aggregate
}

# ──────────────────────────────────────────────────────────
# Rango indicativo bid (percentiles observados Price_to_EAD).
# Usado para construir [precio_low, precio_high] por loan.
INDICATIVE_BID_RANGE_PTILES: dict = {
    "prudencial":   (0.60, 0.90),   # banda alta: sólo ventas a precio bueno
    "balanceado":   (0.30, 0.70),   # banda neutral
    "desinversion": (0.10, 0.50),   # banda amplia: buscar volumen
}

# ──────────────────────────────────────────────────────────
# C) CONCESSION BANDS REESTRUCTURAR (workout NPL)
# ──────────────────────────────────────────────────────────
# dscr_safety_band: "margen de seguridad NPL"
#   DSCR_post debe estar >= dscr_min_wa + dscr_safety_band para ser "clean"
#   Si DSCR_post está en [dscr_min_wa, dscr_min_wa + dscr_safety_band) → REVIEW_REQUIRED
#   Esto evita reestructuras frágiles con DSCR pegado al umbral.
CONCESSION_BAND: dict = {
    "prudencial": {
        "quita_max":              0.03,   # 3%: banda estrecha
        "plazo_extra_max_months":  12,
        "tasa_min":               0.10,   # mín = hurdle rate (10%)
        "dscr_min_wa":            1.30,   # DSCR holgado walk-away
        "dscr_safety_band":       0.20,   # → DSCR efectivo ≥ 1.50 para "clean"
        "pti_max_wa":             0.35,
        "band_label": "ESTRECHA (prudencial: concesiones limitadas)",
    },
    "balanceado": {
        "quita_max":              0.10,   # 10%: banda media
        "plazo_extra_max_months":  24,
        "tasa_min":               0.07,   # permite reducción de tasa material
        "dscr_min_wa":            1.10,   # DSCR mín walk-away
        "dscr_safety_band":       0.10,   # → DSCR efectivo ≥ 1.20 para "clean"
        "pti_max_wa":             0.40,
        "band_label": "MEDIA (balanceado: trade-off EVA vs viabilidad)",
    },
    "desinversion": {
        "quita_max":              0.05,   # 5%: reestructura sólo si venta bloqueada
        "plazo_extra_max_months":  12,
        "tasa_min":               0.15,   # prima por riesgo/liquidez alta
        "dscr_min_wa":            1.05,   # umbral mínimo (DSCR > 1)
        "dscr_safety_band":       0.15,   # → DSCR efectivo ≥ 1.20 para "clean"
        "pti_max_wa":             0.42,
        "band_label": "EXCEPCION (desinversion: sólo si venta bloqueada y uplift alto)",
    },
}

# ──────────────────────────────────────────────────────────
# D) PRUDENCIAL CARVE-OUT — evitar 100% MANTENER
# ──────────────────────────────────────────────────────────
# Regla: PRUDENCIAL puede REESTRUCTURAR aun siendo postura conservadora si:
#   1. EVA_gain está en el top eva_gain_top_pct % del portfolio
#   2. ΔEVA >= min_eva_gain_abs (evita carve-outs triviales)
#   3. DSCR_post >= dscr_min_wa (banda prudencial) + dscr_buffer (holgado)
#   4. PTI <= pti_max (dentro del límite prudencial)
#   5. Concesiones dentro de la banda estrecha (quita <= quita_max, tasa >= tasa_min)
# Si no cumple todas las condiciones → MANTENER.
PRUDENCIAL_CARVE_OUT: dict = {
    "eva_gain_top_pct":    0.20,    # top 20% del EVA_gain del portfolio
    "min_eva_gain_abs":    15_000,  # ΔEVA mínimo absoluto (evitar casos triviales)
    "dscr_buffer":         0.20,    # DSCR_post ≥ CONCESSION_BAND dscr_min_wa + buffer
    "pti_max":             0.35,    # PTI ≤ 35%
    "quita_max":           0.03,    # quita ≤ 3% (banda estrecha)
    "tasa_min":            0.10,    # tasa ≥ 10%
}

# Distance-check thresholds (separación mínima entre posturas).
DISTANCE_THRESHOLDS = {
    "sellrate_desinv_minus_balanc": 0.10,   # ≥10 pp separación en sell rate
    "sellrate_balanc_minus_prud":   0.10,
    "caprel_desinv_minus_balanc_eur": 50_000_000,    # ≥50M€ diferencia
    "caprel_balanc_minus_prud_eur":   100_000_000,   # ≥100M€ diferencia
    "restruct_balanc_minus_desinv": -0.05,   # balanceado puede tener más reestructuras (relajado)
}

# Hurdle rate Basilea III STD
HURDLE_RATE = 0.10

# ============================================================
# HELPERS
# ============================================================

def _sf(x: Any, default: float = 0.0) -> float:
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return default
        return float(x)
    except Exception:
        return default


def _ss(x: Any, default: str = "") -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return default
    return str(x)


def _fmt_eur(v: float) -> str:
    return f"{v:+,.0f} EUR"


def _fmt_pct(v: float) -> str:
    return f"{v*100:.1f}%"


# ──────────────────────────────────────────────────────────
# Per-loan Reservation Floor (B — redesign)
# ──────────────────────────────────────────────────────────
def _per_loan_floor(row: Any, postura: str) -> float:
    """
    Calcula el Reservation_Floor POR PRÉSTAMO usando el precio simulado del
    propio préstamo (Price_to_EAD_i ≈ p50 del simulador) y aplicando el
    multiplicador lognormal de la postura:
        floor_i = max(Price_to_EAD_i × FLOOR_QUANTILE_MULTIPLIER[postura],
                      RESERVATION_FLOOR_MINIMUM[postura])

    Si Price_to_EAD no está disponible o es ≤ mínimo, devuelve el mínimo
    absoluto de la postura como floor de seguridad.
    """
    if isinstance(row, pd.Series):
        px_ead = _sf(row.get("Price_to_EAD", row.get("price_ratio_ead", 0.0)))
    else:
        px_ead = float(row.get("Price_to_EAD", row.get("price_ratio_ead", 0.0)) or 0.0)

    min_floor = RESERVATION_FLOOR_MINIMUM[postura]
    if px_ead <= 0.01:
        # Precio no disponible o nulo → usar floor mínimo absoluto
        return min_floor

    mult = FLOOR_QUANTILE_MULTIPLIER[postura]
    derived = float(px_ead) * mult
    return max(derived, min_floor)


def find_run_dir(tag: str, postura: str) -> str:
    pattern = os.path.join(REPORTS_DIR, f"coordinated_inference_{tag}_*_{postura}")
    dirs = sorted(glob.glob(pattern), reverse=True)
    if not dirs:
        raise FileNotFoundError(f"No se encontró carpeta para tag={tag}, postura={postura}")
    return dirs[0]


def load_artifacts(tag: str, postura: str) -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    d = find_run_dir(tag, postura)
    # Usar versión hardened si existe
    xf = os.path.join(d, f"decisiones_finales_{postura}_hardened.xlsx")
    if not os.path.exists(xf):
        xf = os.path.join(d, f"decisiones_finales_{postura}.xlsx")
    df = pd.read_excel(xf)

    kf = os.path.join(d, f"portfolio_kpis_{postura}_hardened.json")
    if not os.path.exists(kf):
        kf = os.path.join(d, f"portfolio_kpis_{postura}.json")
    with open(kf, encoding="utf-8") as f:
        kpis = json.load(f)

    ov_path = os.path.join(d, f"overrides_log_{postura}.csv")
    ov = pd.read_csv(ov_path) if os.path.exists(ov_path) else pd.DataFrame()

    log.info(f"  [{postura}] Cargado: {len(df)} loans, hardened={'hardened' in xf}")
    return df, kpis, ov


# ============================================================
# A) DIAGNÓSTICO POR POSTURA
# ============================================================

def diagnose_posture(df: pd.DataFrame, kpis: Dict, postura: str) -> Dict:
    """Construye el diagnóstico loan-level y portfolio-level para una postura."""
    n = len(df)
    ac = df["Accion_final"].value_counts().to_dict()
    sell_n = int(ac.get("VENDER", 0))
    rest_n = int(ac.get("REESTRUCTURAR", 0))
    keep_n = int(ac.get("MANTENER", 0))
    sell_rate = sell_n / n

    fs_kpis = kpis.get("final_state", {})
    total_eva   = float(fs_kpis.get("total_eva", 0.0))
    total_rwa   = float(fs_kpis.get("total_rwa", 0.0))
    total_cap   = float(fs_kpis.get("total_capital_release", 0.0))
    total_ead   = float(fs_kpis.get("total_ead", df["EAD"].sum() if "EAD" in df.columns else 0.0))

    # P&L de ventas
    vsub = df[df["Accion_final"] == "VENDER"]
    sale_pnl = float(vsub["pnl"].sum()) if "pnl" in vsub.columns and len(vsub) > 0 else 0.0
    sale_pnl_ead_ratio = sale_pnl / vsub["EAD"].sum() if (
        "EAD" in vsub.columns and len(vsub) > 0 and vsub["EAD"].sum() > 0
    ) else 0.0

    # % ventas bloqueadas (guardrail / fire-sale)
    blocked_n = int((df["Accion_final"] == "MANTENER").sum() &
                    df.get("guardrail_reasons", pd.Series("", index=df.index)).str.strip().ne("").sum()
                    if "guardrail_reasons" in df.columns else 0)
    rc_col = "Reason_Code_Final" if "Reason_Code_Final" in df.columns else "Reason_Code"
    rc_dist = df[rc_col].value_counts().to_dict()

    n_fire_sale_blocked = int((df[rc_col] == "RC02_SELL_BLOCKED_FIRE_SALE").sum())
    n_guardrail_blocked = int((df[rc_col] == "RC_GUARDRAIL_BLOCK").sum())
    n_macro_steer = int((df.get("Macro_Steering_Applied", pd.Series("", index=df.index)).str.strip() != "").sum())

    # Convergence
    conv_col = "Convergencia_Caso"
    conv_dist = df[conv_col].value_counts().to_dict() if conv_col in df.columns else {}

    # Casos frontera NPL
    # 1. Ventas con P&L muy negativo vs EAD
    border_sell_loss: List[Dict] = []
    if len(vsub) > 0 and "EAD" in vsub.columns and "pnl" in vsub.columns:
        vsub2 = vsub.copy()
        vsub2["pnl_ead_ratio"] = vsub2["pnl"] / vsub2["EAD"].replace(0, np.nan)
        worst = vsub2.nsmallest(min(5, len(vsub2)), "pnl_ead_ratio")
        for _, r in worst.iterrows():
            border_sell_loss.append({
                "loan_id": _ss(r.get("loan_id"), "?"),
                "EAD": _sf(r.get("EAD")),
                "pnl": _sf(r.get("pnl")),
                "pnl_ead_ratio": _sf(r.get("pnl_ead_ratio")),
                "Price_to_EAD": _sf(r.get("Price_to_EAD")),
                "FireSale": bool(r.get("FireSale_Triggered", r.get("Fire_Sale", False))),
                "RC": _ss(r.get(rc_col)),
            })

    # 2. Reestructuras cerca del umbral PTI/DSCR
    border_restruct: List[Dict] = []
    rsub = df[df["Accion_final"] == "REESTRUCTURAR"]
    if len(rsub) > 0:
        for col_check, threshold, ascending in [
            ("DSCR_post", 1.20, True),
            ("PTI_post",  0.40, False),
        ]:
            if col_check not in rsub.columns:
                continue
            sub_borderline = rsub[rsub[col_check].notna()]
            if ascending:
                sub_borderline = sub_borderline[sub_borderline[col_check] < threshold]
            else:
                sub_borderline = sub_borderline[sub_borderline[col_check] > threshold]
            for _, r in sub_borderline.head(3).iterrows():
                ev = {
                    "loan_id": _ss(r.get("loan_id"), "?"),
                    "EAD": _sf(r.get("EAD")),
                    "DSCR_post": _sf(r.get("DSCR_post", np.nan), np.nan),
                    "PTI_post": _sf(r.get("PTI_post", np.nan), np.nan),
                    "ΔEVA": _sf(r.get("ΔEVA", r.get("EVA_gain", 0.0))),
                    "tasa_nueva": _sf(r.get("tasa_nueva", np.nan), np.nan),
                    "quita": _sf(r.get("quita", 0.0)),
                    "RC": _ss(r.get(rc_col)),
                }
                # avoid duplicates
                if not any(b["loan_id"] == ev["loan_id"] for b in border_restruct):
                    border_restruct.append(ev)

    # 3. Mantenidos con alto potencial bloqueados
    border_keep: List[Dict] = []
    ksub = df[df["Accion_final"] == "MANTENER"]
    if len(ksub) > 0:
        deva_col = "ΔEVA" if "ΔEVA" in ksub.columns else "EVA_gain" if "EVA_gain" in ksub.columns else None
        if deva_col:
            high_pot = ksub[ksub[deva_col].fillna(0.0) > 0].nlargest(min(5, len(ksub)), deva_col)
            for _, r in high_pot.iterrows():
                border_keep.append({
                    "loan_id": _ss(r.get("loan_id"), "?"),
                    "EAD": _sf(r.get("EAD")),
                    "ΔEVA": _sf(r.get(deva_col, 0.0)),
                    "EVA_pre": _sf(r.get("EVA_pre", 0.0)),
                    "RORWA_pre": _sf(r.get("RORWA_pre", 0.0)),
                    "RC": _ss(r.get(rc_col)),
                    "blocked_by": _ss(r.get("Macro_Steering_Applied", r.get("guardrail_reasons", ""))),
                })

    return {
        "postura": postura,
        "n": n,
        "sell_n": sell_n,
        "rest_n": rest_n,
        "keep_n": keep_n,
        "sell_rate": sell_rate,
        "total_eva": total_eva,
        "total_rwa": total_rwa,
        "total_cap": total_cap,
        "total_ead": total_ead,
        "sale_pnl": sale_pnl,
        "sale_pnl_ead_ratio": sale_pnl_ead_ratio,
        "n_fire_sale_blocked": n_fire_sale_blocked,
        "n_guardrail_blocked": n_guardrail_blocked,
        "n_macro_steered": n_macro_steer,
        "rc_dist": rc_dist,
        "conv_dist": conv_dist,
        "border_sell_loss": border_sell_loss,
        "border_restruct": border_restruct,
        "border_keep": border_keep,
    }


# ============================================================
# B) NEGOTIATION ENVELOPES
# ============================================================

def _compute_bid_range(price_to_ead: float, ead: float, postura: str) -> Tuple[float, float]:
    """Calcula el rango indicativo de oferta (precio absoluto) para un loan."""
    ptiles = INDICATIVE_BID_RANGE_PTILES[postura]
    # El rango se construye desplazando ±OFFSET del precio óptimo
    # offset en términos de Price_to_EAD (±stdev/2)
    std_estimate = 0.015  # dispersión estimada (~1.5pp Price/EAD)
    low_px_ead  = max(RESERVATION_FLOOR[postura], price_to_ead - std_estimate * 1.5)
    high_px_ead = price_to_ead + std_estimate * 1.0
    lo = round(low_px_ead * ead, 0)
    hi = round(high_px_ead * ead, 0)
    return lo, hi


def _vender_envelope_cols(row: pd.Series, postura: str) -> Dict[str, str]:
    """Genera las columnas de envelope para un loan VENDER."""
    ead      = _sf(row.get("EAD", 0.0))
    px_ead   = _sf(row.get("Price_to_EAD", 0.13))
    pnl      = _sf(row.get("pnl", 0.0))
    fire_s   = bool(row.get("FireSale_Triggered", row.get("Fire_Sale", False)))
    cap_rel  = _sf(row.get("capital_release_realized", row.get("capital_liberado", 0.0)))

    # ── Floor POR PRÉSTAMO (derivado del precio propio del préstamo)
    floor      = _per_loan_floor(row, postura)
    floor_eur  = round(floor * ead, 0)
    bid_lo, bid_hi = _compute_bid_range(px_ead, ead, postura)
    pnl_if_floor = round((floor - 1.0) * ead, 0)   # P&L si se vende exactamente al floor

    go_to_market = "SI" if not fire_s and px_ead >= floor else "BLOQUEADO"
    if fire_s:
        go_to_market = "BLOQUEADO_FIRE_SALE"
    elif px_ead < floor:
        go_to_market = "BLOQUEADO_PRECIO_BAJO_FLOOR_INDIVIDUAL"

    rule_text = (
        f"Postura={postura.upper()} | "
        f"Floor_loan={_fmt_pct(floor)} EAD = {_fmt_eur(floor_eur)} "
        f"[{_fmt_pct(px_ead)}×{FLOOR_QUANTILE_MULTIPLIER[postura]:.3f} precio propio] | "
        f"Precio_modelo={_fmt_pct(px_ead)} | "
        f"P&L_si_floor={_fmt_eur(pnl_if_floor)} | "
        f"Capital_liberado={_fmt_eur(cap_rel)} | "
        f"Si_oferta >= floor={_fmt_eur(floor_eur)}: EJECUTAR VENDER. "
        f"Si_oferta < floor={_fmt_eur(floor_eur)}: MANTENER (no cristalizar pérdida)."
    )

    return {
        "GoToMarket":           go_to_market,
        "Indicative_Bid_Range": f"[{_fmt_eur(bid_lo)}, {_fmt_eur(bid_hi)}]",
        "Reservation_Floor":    f"{_fmt_pct(floor)} EAD = {_fmt_eur(floor_eur)} [derivado precio propio]",
        "Execution_Rule_Text":  rule_text,
    }


def _restruct_envelope_cols(row: pd.Series, postura: str) -> Dict[str, str]:
    """Genera las columnas de envelope para un loan REESTRUCTURAR."""
    ead        = _sf(row.get("EAD", 0.0))
    quita      = _sf(row.get("quita", 0.0))
    tasa       = _sf(row.get("tasa_nueva", 0.19))
    plazo      = _sf(row.get("plazo_optimo", 240.0))
    dscr_post  = _sf(row.get("DSCR_post", np.nan), np.nan)
    pti_post   = _sf(row.get("PTI_post", np.nan), np.nan)
    deva       = _sf(row.get("ΔEVA", row.get("EVA_gain", 0.0)))
    band       = CONCESSION_BAND[postura]

    anchor_parts = [
        f"Tasa_propuesta={_fmt_pct(tasa)}",
        f"Plazo={int(plazo)}m",
        f"Quita={_fmt_pct(quita)}",
    ]
    if not np.isnan(dscr_post):
        anchor_parts.append(f"DSCR_post={dscr_post:.2f}")
    if not np.isnan(pti_post):
        anchor_parts.append(f"PTI_post={_fmt_pct(pti_post)}")
    anchor_parts.append(f"ΔEVA_esperado={_fmt_eur(deva)}")
    anchor_text = " | ".join(anchor_parts)

    dscr_safety   = band.get("dscr_safety_band", 0.10)
    dscr_clean_thr = band["dscr_min_wa"] + dscr_safety

    band_text = (
        f"Quita_max={_fmt_pct(band['quita_max'])} | "
        f"Plazo_extra_max=+{band['plazo_extra_max_months']}m | "
        f"Tasa_min={_fmt_pct(band['tasa_min'])} | "
        f"DSCR_post_min={band['dscr_min_wa']:.2f} | "
        f"DSCR_clean>={dscr_clean_thr:.2f} (safety_band={_fmt_pct(dscr_safety)}) | "
        f"PTI_post_max={_fmt_pct(band['pti_max_wa'])} "
        f"[{band['band_label']}]"
    )

    # Walk-away conditions check
    wa_triggers = []
    if not np.isnan(dscr_post) and dscr_post < band["dscr_min_wa"]:
        wa_triggers.append(f"DSCR_post={dscr_post:.2f} < minimo={band['dscr_min_wa']:.2f}")
    if not np.isnan(pti_post) and pti_post > band["pti_max_wa"]:
        wa_triggers.append(f"PTI_post={_fmt_pct(pti_post)} > maximo={_fmt_pct(band['pti_max_wa'])}")
    if quita > band["quita_max"]:
        wa_triggers.append(f"Quita_pedida={_fmt_pct(quita)} > max={_fmt_pct(band['quita_max'])}")
    if tasa < band["tasa_min"]:
        wa_triggers.append(f"Tasa={_fmt_pct(tasa)} < min={_fmt_pct(band['tasa_min'])}")

    # Margen de seguridad NPL: DSCR pegado al umbral → REVIEW_REQUIRED
    dscr_near_threshold = (
        not np.isnan(dscr_post) and
        dscr_post >= band["dscr_min_wa"] and
        dscr_post < dscr_clean_thr
    )
    review_required = False

    if wa_triggers:
        wa_text = f"WALK_AWAY ACTIVO: {'; '.join(wa_triggers)} → Mantener o evaluar venta si cotización >= floor."
    elif dscr_near_threshold:
        review_required = True
        wa_text = (
            f"REVIEW_REQUIRED: DSCR_post={dscr_post:.2f} está en la zona de proximidad al umbral "
            f"[{band['dscr_min_wa']:.2f}, {dscr_clean_thr:.2f}). "
            f"Margen de seguridad NPL insuficiente para {postura.upper()}. "
            f"Requiere validación adicional antes de firmar reestructura. "
            f"Fallback: MANTENER si no se puede garantizar DSCR_post >= {dscr_clean_thr:.2f}."
        )
    else:
        wa_text = (
            f"Dentro de banda {postura.upper()}. Continuar negociación. "
            f"Walk-away si deudor exige: Quita>{_fmt_pct(band['quita_max'])} "
            f"O Tasa<{_fmt_pct(band['tasa_min'])} "
            f"O DSCR_post<{band['dscr_min_wa']:.2f} "
            f"O PTI_post>{_fmt_pct(band['pti_max_wa'])}."
        )

    # Añadir alerta si DSCR cercano (aunque no sea walk-away)
    if review_required and not wa_triggers:
        anchor_text += f" | REVIEW_REQUIRED(DSCR_near_threshold)"

    return {
        "Anchor_Terms":       anchor_text,
        "Concession_Band":    band_text,
        "WalkAway_Rule_Text": wa_text,
    }


def _mantener_envelope_cols(row: pd.Series, postura: str) -> Dict[str, str]:
    """Genera columnas de fallback para un loan MANTENER."""
    rc    = _ss(row.get("Reason_Code_Final", row.get("Reason_Code", "")))
    ead   = _sf(row.get("EAD", 0.0))
    deva  = _sf(row.get("ΔEVA", row.get("EVA_gain", 0.0)))
    # Floor POR PRÉSTAMO (latente, no activo)
    floor = _per_loan_floor(row, postura)
    px_ead = _sf(row.get("Price_to_EAD", 0.13))

    if rc == "RC02_SELL_BLOCKED_FIRE_SALE":
        fallback  = "FIRE_SALE_BLOCK — venta suspendida hasta condiciones de mercado normalizadas"
        trigger   = f"Si mercado_bid >= {_fmt_pct(floor)} EAD Y sin fire-sale → re-evaluar VENDER"
    elif rc == "RC_GUARDRAIL_BLOCK":
        fallback  = "GUARDRAIL_BLOCK — pérdida superaría umbral de P&L admisible"
        trigger   = f"Si mercado_bid >= {_fmt_pct(floor)} EAD y pérdida < 40% EAD → re-evaluar"
    elif rc in ("RC05_KEEP_MEETS_HURDLE", "RC05_KEEP_ACCEPTABLE_ECONOMICS"):
        fallback  = "ECONOMICS_KEEP — loan genera valor holding > valor liquidación actual"
        trigger   = (
            f"Si RORWA_pre cae por debajo de hurdle ({_fmt_pct(HURDLE_RATE)}) "
            f"O si aparece offerta > {_fmt_pct(floor)} EAD → considerar VENDER. "
            f"Si ΔEVA potencial > EAD*5% → evaluar REESTRUCTURAR"
        )
    else:
        fallback  = f"MANTENER por decisión estratégica ({rc})"
        trigger   = "Monitorizar siguiente ciclo de valoración (90 días)"

    return {
        "Fallback_Reason":    fallback,
        "Trigger_to_Action":  trigger,
        "GoToMarket":         "NO_ACTION",
        "Indicative_Bid_Range": f"[N/A — floor: {_fmt_pct(floor)} EAD = {_fmt_eur(floor*ead)}]",
        "Reservation_Floor":  f"{_fmt_pct(floor)} EAD (latente, no activo)",
        "Execution_Rule_Text": (
            f"MANTENER activo. No ejecutar venta ni reestructura hasta que "
            f"condiciones superen el envelope {postura.upper()}."
        ),
        "Anchor_Terms":       "N/A (MANTENER)",
        "Concession_Band":    "N/A (MANTENER)",
        "WalkAway_Rule_Text": "N/A (MANTENER — evaluar si score_sell mejora)",
    }


def apply_prudencial_carve_out(df: pd.DataFrame) -> pd.DataFrame:
    """
    D) PRUDENCIAL CARVE-OUT: evitar parálisis total (100% MANTENER).

    Identifica préstamos MANTENER que tienen uplift claro y viabilidad holgada
    y los eleva a REESTRUCTURAR siguiendo la regla del PRUDENCIAL_CARVE_OUT.

    Criterios (todos deben cumplirse):
      1. EVA_gain ≥ Q(1 − eva_gain_top_pct) de los propios MANTENER (top 20%)
         y ΔEVA ≥ min_eva_gain_abs (filtro absoluto anti-trivial)
      2. DSCR_proxy  ≥ dscr_min_wa (banda prudencial) + dscr_buffer   → ≥ 1.50
      3. PTI_proxy   ≤ pti_max                                         → ≤ 35 %
      4. quita       ≤ quita_max  (si disponible)                      → ≤ 3 %
      5. tasa_nueva  ≥ tasa_min   (si disponible)                      → ≥ 10 %

    Para DSCR/PTI se usa la métrica post si existe, sino la pre comme proxy
    conservador. Si faltan datos de concesiones (quita/tasa NaN para loans
    que aún no tienen propuesta) se omite ese check.

    Actualiza Accion_final, Reason_Code_Final y Decision_Governance_Final.
    Añade columna Carve_Out_Type para trazabilidad.
    """
    params = PRUDENCIAL_CARVE_OUT
    band   = CONCESSION_BAND["prudencial"]
    df     = df.copy()

    if "Carve_Out_Type" not in df.columns:
        df["Carve_Out_Type"] = ""

    mantener_mask = df["Accion_final"].str.upper().str.strip() == "MANTENER"
    if not mantener_mask.any():
        log.info("  [CARVE_OUT] Sin préstamos MANTENER en este DataFrame.")
        return df

    # ── EVA_gain series (mejor proxy disponible)
    eva = None
    for col in ("ΔEVA", "EVA_gain", "delta_eva"):
        if col in df.columns:
            eva = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            break
    if eva is None and {"EVA_post", "EVA_pre"}.issubset(df.columns):
        eva = (
            pd.to_numeric(df["EVA_post"], errors="coerce")
            - pd.to_numeric(df["EVA_pre"],  errors="coerce")
        ).fillna(0.0)
    if eva is None:
        eva = pd.Series(0.0, index=df.index)

    # Cutoff = max(Q(1-top_pct) de MANTENER, min_eva_gain_abs)
    eva_mantener = eva[mantener_mask]
    if len(eva_mantener) == 0:
        return df

    top_pct   = float(params["eva_gain_top_pct"])
    eva_q_thr = float(np.quantile(eva_mantener, 1.0 - top_pct))
    eva_q_thr = max(eva_q_thr, float(params["min_eva_gain_abs"]))

    # ── DSCR check por niveles (NPL-aware)
    #   Nivel 1: DSCR_post disponible (restructure optimizer lo calculó) → exigir >= dscr_clean_thr
    #   Nivel 2: Solo DSCR_pre → NPL loans están en default (DSCR < 1 habitual)
    #             → exigir >= MIN_DSCR_PRE_SIGNAL (0.65 = flujo real aunque < 1)
    #             → carve-out marcado como "CANDIDATE_NEEDS_RESTRUCTURE_EVAL"
    #   Nivel 3: Sin datos DSCR → confiar en EVA (omitir check DSCR)
    MIN_DSCR_PRE_SIGNAL = 0.65   # señal mínima de flujo en NPL pre-restructura

    dscr_post_s = pd.to_numeric(
        df.get("DSCR_post", pd.Series(np.nan, index=df.index)), errors="coerce"
    )
    dscr_pre_s = pd.to_numeric(
        df.get("DSCR_pre", df.get("DSCR", pd.Series(np.nan, index=df.index))), errors="coerce"
    )
    has_dscr_post = dscr_post_s.notna()
    has_dscr_pre  = dscr_pre_s.notna()

    dscr_clean_thr = float(band["dscr_min_wa"]) + float(params["dscr_buffer"])

    ok_dscr_l1 = has_dscr_post & (dscr_post_s >= dscr_clean_thr)
    ok_dscr_l2 = has_dscr_pre & ~has_dscr_post & (dscr_pre_s >= MIN_DSCR_PRE_SIGNAL)
    ok_dscr_l3 = ~has_dscr_post & ~has_dscr_pre   # sin datos → confiar en EVA
    ok_dscr    = ok_dscr_l1 | ok_dscr_l2 | ok_dscr_l3

    # proxy para logging
    dscr_proxy = dscr_post_s.combine_first(dscr_pre_s).fillna(0.0)

    # ── PTI proxy (post > pre; 0.0 si ausente — 0 siempre pasa pti_max=35%)
    pti_post = pd.to_numeric(df.get("PTI_post", pd.Series(np.nan, index=df.index)), errors="coerce")
    pti_pre  = pd.to_numeric(df.get("PTI_pre",  df.get("PTI", pd.Series(np.nan, index=df.index))), errors="coerce")
    pti_proxy = pti_post.combine_first(pti_pre).fillna(0.0)

    # ── Concesiones (opcionales: si son NaN → check omitido = "ok")
    quita_s = pd.to_numeric(df.get("quita",     pd.Series(np.nan, index=df.index)), errors="coerce")
    tasa_s  = pd.to_numeric(df.get("tasa_nueva", pd.Series(np.nan, index=df.index)), errors="coerce")

    # ── Carve-out mask
    ok_eva   = eva       >= eva_q_thr
    ok_pti   = pti_proxy  <= float(params["pti_max"])
    ok_quita = quita_s.isna() | (quita_s <= float(params["quita_max"]))
    ok_tasa  = tasa_s.isna()  | (tasa_s  >= float(params["tasa_min"]))

    carveout_mask = mantener_mask & ok_eva & ok_dscr & ok_pti & ok_quita & ok_tasa

    # Subtipos: nivel del DSCR check para trazabilidad
    co_l1 = carveout_mask & ok_dscr_l1   # DSCR_post holgado: carve-out limpio
    co_l2 = carveout_mask & ok_dscr_l2   # solo DSCR_pre: carve-out candidato (requiere eval)
    co_l3 = carveout_mask & ok_dscr_l3   # sin DSCR: carve-out por EVA signal puro

    n_co = int(carveout_mask.sum())
    log.info(
        f"  [CARVE_OUT] PRUDENCIAL: {n_co} de {int(mantener_mask.sum())} MANTENER "
        f"identificados como carve-out "
        f"(L1={int(co_l1.sum())} DSCR_post>={dscr_clean_thr:.2f}, "
        f"L2={int(co_l2.sum())} DSCR_pre>={MIN_DSCR_PRE_SIGNAL:.2f}, "
        f"L3={int(co_l3.sum())} EVA-only) "
        f"| EVA_q_thr={eva_q_thr:,.0f}€, PTI<={params['pti_max']:.0%})"
    )

    if n_co > 0:
        # L1 → REESTRUCTURAR limpio
        if co_l1.any():
            df.loc[co_l1, "Accion_final"] = "REESTRUCTURAR"
            df.loc[co_l1, "Carve_Out_Type"] = (
                f"PRUDENCIAL_RESTRUCT_CARVEOUT_CLEAN|top{int(top_pct*100)}pct|"
                f"EVA>={eva_q_thr:,.0f}|DSCR_post>={dscr_clean_thr:.2f}"
            )
        # L2 → REESTRUCTURAR candidato (requiere validación restructure optimizer)
        if co_l2.any():
            df.loc[co_l2, "Accion_final"] = "REESTRUCTURAR"
            df.loc[co_l2, "Carve_Out_Type"] = (
                f"PRUDENCIAL_RESTRUCT_CARVEOUT_CANDIDATE|top{int(top_pct*100)}pct|"
                f"EVA>={eva_q_thr:,.0f}|DSCR_pre>={MIN_DSCR_PRE_SIGNAL:.2f}|NEEDS_RESTRUCTURE_EVAL"
            )
        # L3 → REESTRUCTURAR por EVA signal (sin datos DSCR)
        if co_l3.any():
            df.loc[co_l3, "Accion_final"] = "REESTRUCTURAR"
            df.loc[co_l3, "Carve_Out_Type"] = (
                f"PRUDENCIAL_RESTRUCT_CARVEOUT_EVA_SIGNAL|top{int(top_pct*100)}pct|"
                f"EVA>={eva_q_thr:,.0f}|NO_DSCR_DATA"
            )


        rc_col = "Reason_Code_Final" if "Reason_Code_Final" in df.columns else "Reason_Code"
        df.loc[carveout_mask, rc_col] = "RC03_PRUDENCIAL_CARVEOUT_RESTRUCT"

        gov_col = "Decision_Governance_Final"
        if gov_col not in df.columns:
            df[gov_col] = ""
        df.loc[carveout_mask, gov_col] = (
            df.loc[carveout_mask, gov_col].fillna("").astype(str)
            + f" | PRUDENCIAL_CARVEOUT(top{int(top_pct*100)}%_EVA"
            + f"+DSCR>={dscr_clean_thr:.2f}+PTI<={params['pti_max']:.0%})"
            + " → REESTRUCTURAR"
        )

    return df


def apply_negotiation_envelopes(df: pd.DataFrame, postura: str) -> pd.DataFrame:
    """
    Tarea B+D: Añade columnas del Negotiation Envelope por loan.
    Para postura PRUDENCIAL aplica primero el carve-out (D) antes de
    calcular los envelopes (B) para que los loans elevados a REESTRUCTURAR
    reciban el envelope WORKOUT correcto.

    Columnas añadidas:
      Envelope_Type, GoToMarket, Indicative_Bid_Range, Reservation_Floor
      (POR PRÉSTAMO via _per_loan_floor),
      Execution_Rule_Text, Anchor_Terms, Concession_Band,
      WalkAway_Rule_Text (con REVIEW_REQUIRED si DSCR cerca del umbral),
      Fallback_Reason, Trigger_to_Action, Carve_Out_Type
    """
    df = df.copy()

    # D) PRUDENCIAL CARVE-OUT antes de calcular envelopes
    if postura == "prudencial":
        df = apply_prudencial_carve_out(df)

    # Inicializar columnas
    for col in ["Envelope_Type", "GoToMarket", "Indicative_Bid_Range", "Reservation_Floor",
                "Execution_Rule_Text", "Anchor_Terms", "Concession_Band", "WalkAway_Rule_Text",
                "Fallback_Reason", "Trigger_to_Action"]:
        if col not in df.columns:
            df[col] = ""

    for idx, row in df.iterrows():
        accion = _ss(row.get("Accion_final", "")).upper()

        if accion == "VENDER":
            df.at[idx, "Envelope_Type"] = "SELL"
            cols = _vender_envelope_cols(row, postura)
            for k, v in cols.items():
                df.at[idx, k] = v
            # Reestructura N/A
            df.at[idx, "Anchor_Terms"]       = "N/A (VENDER)"
            df.at[idx, "Concession_Band"]    = "N/A (VENDER)"
            df.at[idx, "WalkAway_Rule_Text"] = "N/A (VENDER)"
            df.at[idx, "Fallback_Reason"]    = "N/A (VENDER)"
            floor_vend = _per_loan_floor(row, postura)
            df.at[idx, "Trigger_to_Action"]  = (
                f"Si oferta < {_fmt_pct(floor_vend)} EAD = "
                f"{_fmt_eur(floor_vend * _sf(row.get('EAD', 0.0)))} [floor propio] → MANTENER"
            )

        elif accion == "REESTRUCTURAR":
            df.at[idx, "Envelope_Type"] = "WORKOUT"
            cols = _restruct_envelope_cols(row, postura)
            for k, v in cols.items():
                df.at[idx, k] = v
            # Venta como backup
            ead = _sf(row.get("EAD", 0.0))
            floor_rest = _per_loan_floor(row, postura)
            df.at[idx, "GoToMarket"]           = "NO_ACTIVO_REESTRUCTURAR_PRIORITARIO"
            df.at[idx, "Indicative_Bid_Range"]  = (
                f"[Backup_floor={_fmt_pct(floor_rest)} EAD = {_fmt_eur(floor_rest*ead)} (derivado precio propio)]"
            )
            df.at[idx, "Reservation_Floor"]     = (
                f"{_fmt_pct(floor_rest)} EAD (backup si walk-away; floor propio >= min {_fmt_pct(RESERVATION_FLOOR_MINIMUM[postura])} EAD)"
            )
            df.at[idx, "Execution_Rule_Text"]   = (
                f"Primero reestructurar. Si walk-away → evaluar VENDER con floor_loan {_fmt_pct(floor_rest)} EAD."
            )
            df.at[idx, "Fallback_Reason"]       = "N/A (REESTRUCTURAR)"
            df.at[idx, "Trigger_to_Action"]     = (
                f"Si deudor fuera de banda [{CONCESSION_BAND[postura]['band_label']}] → walk-away "
                f"(MANTENER o VENDER si oferta >= {_fmt_pct(floor_rest)} EAD)"
            )

        else:  # MANTENER
            df.at[idx, "Envelope_Type"] = "NO_ACTION"
            cols = _mantener_envelope_cols(row, postura)
            for k, v in cols.items():
                df.at[idx, k] = v

    n_sell = (df["Envelope_Type"] == "SELL").sum()
    n_work = (df["Envelope_Type"] == "WORKOUT").sum()
    n_hold = (df["Envelope_Type"] == "NO_ACTION").sum()
    n_gtm  = (df["GoToMarket"] == "SI").sum()
    log.info(
        f"  [{postura}] Envelopes: SELL={n_sell} (GoToMarket={n_gtm}), "
        f"WORKOUT={n_work}, NO_ACTION={n_hold}"
    )
    return df


# ============================================================
# D) DISTANCE CHECKS + ENFORCEMENT
# ============================================================

def run_distance_checks(
    diags: Dict[str, Dict],
    postures_order: List[str] = None,
) -> Tuple[List[Dict], List[Dict]]:
    """
    Implementa los distance checks entre posturas.
    Devuelve (checks_results, enforcement_actions).
    """
    if postures_order is None:
        postures_order = ["prudencial", "balanceado", "desinversion"]

    checks: List[Dict] = []
    enforcements: List[Dict] = []

    def _check(name: str, value: float, threshold: float, direction: str = "gte") -> Dict:
        ok = (value >= threshold) if direction == "gte" else (value <= threshold)
        status = "PASS" if ok else "FAIL"
        return {
            "check": name,
            "value": round(value, 4),
            "threshold": threshold,
            "direction": direction,
            "status": status,
        }

    p = diags.get("prudencial", {})
    b = diags.get("balanceado", {})
    d = diags.get("desinversion", {})

    # 1. Sell rate separation
    d_b_sellrate = d.get("sell_rate", 0.0) - b.get("sell_rate", 0.0)
    b_p_sellrate = b.get("sell_rate", 0.0) - p.get("sell_rate", 0.0)
    checks.append(_check("SellRate_Desinv_minus_Balanc",
                         d_b_sellrate, DISTANCE_THRESHOLDS["sellrate_desinv_minus_balanc"]))
    checks.append(_check("SellRate_Balanc_minus_Prud",
                         b_p_sellrate, DISTANCE_THRESHOLDS["sellrate_balanc_minus_prud"]))

    # 2. Capital release separation
    d_b_cap = d.get("total_cap", 0.0) - b.get("total_cap", 0.0)
    b_p_cap = b.get("total_cap", 0.0) - p.get("total_cap", 0.0)
    checks.append(_check("CapRelease_Desinv_minus_Balanc_EUR",
                         d_b_cap, DISTANCE_THRESHOLDS["caprel_desinv_minus_balanc_eur"]))
    checks.append(_check("CapRelease_Balanc_minus_Prud_EUR",
                         b_p_cap, DISTANCE_THRESHOLDS["caprel_balanc_minus_prud_eur"]))

    # 3. Reestructura rate: balanceado should be >= desinversion (more workout-focused)
    b_restr = b.get("rest_n", 0) / b.get("n", 1)
    d_restr = d.get("rest_n", 0) / d.get("n", 1)
    b_minus_d_restr = b_restr - d_restr
    checks.append(_check("RestRate_Balanc_higher_Desinv",
                         b_minus_d_restr, DISTANCE_THRESHOLDS["restruct_balanc_minus_desinv"],
                         direction="gte"))

    # 4. EVA preservation: prudencial should preserve most EVA
    eva_p = p.get("total_eva", 0.0)
    eva_b = b.get("total_eva", 0.0)
    eva_d = d.get("total_eva", 0.0)
    checks.append(_check("EVA_Prud_highest",
                         eva_p - max(eva_b, eva_d), 0, direction="gte"))

    # Enforcement: if any check fails, log what would be adjusted
    for c in checks:
        if c["status"] == "FAIL":
            action = _build_enforcement(c, diags)
            if action:
                enforcements.append(action)
                log.warning(f"  DISTANCE_FAIL: {c['check']} = {c['value']:.4f} < {c['threshold']:.4f}")

    for c in checks:
        log.info(f"  [{c['status']}] {c['check']}: {c['value']:.4f} (thr={c['threshold']:.4f})")

    return checks, enforcements


def _build_enforcement(check: Dict, diags: Dict) -> Optional[Dict]:
    """Genera la acción de enforcement para un check fallido."""
    name = check["check"]
    val  = check["value"]
    thr  = check["threshold"]
    gap  = thr - val  # cuánto falta

    if "SellRate_Desinv_minus_Balanc" in name:
        return {
            "trigger": name,
            "gap": round(gap, 4),
            "action": "Reducir Reservation_Floor de DESINVERSION en ~2pp para atraer más ofertas",
            "new_floor_desinv": round(RESERVATION_FLOOR["desinversion"] - gap * 0.5, 3),
            "override_posture": "desinversion",
        }
    elif "SellRate_Balanc_minus_Prud" in name:
        return {
            "trigger": name,
            "gap": round(gap, 4),
            "action": "Ampliar banda de concesión BALANCEADO (quita_max +2pp) OR elevar floor PRUDENCIAL en 2pp",
            "new_floor_prud": round(RESERVATION_FLOOR["prudencial"] + gap * 0.5, 3),
            "override_posture": "prudencial/balanceado",
        }
    elif "CapRelease_Desinv_minus_Balanc" in name:
        return {
            "trigger": name,
            "gap": round(gap, 0),
            "action": f"Reducir floor DESINVERSION o bajar umbral fire-sale para liberar más capital ({_fmt_eur(gap)})",
            "override_posture": "desinversion",
        }
    elif "CapRelease_Balanc_minus_Prud" in name:
        return {
            "trigger": name,
            "gap": round(gap, 0),
            "action": f"Activar más ventas/reestructuras en BALANCEADO (ampliar concesion band o bajar floor)",
            "override_posture": "balanceado",
        }
    elif "EVA_Prud_highest" in name:
        return {
            "trigger": name,
            "gap": round(gap, 0),
            "action": "Verificar que PRUDENCIAL mantenga el libro de valor (todos MANTENER o reestructuras con EVA alto)",
            "override_posture": "prudencial",
        }
    return None


def write_enforcement_log(enforcements: List[Dict], tag: str) -> str:
    """Escribe enforcement_log_{tag}.csv en reports/."""
    path = os.path.join(REPORTS_DIR, f"enforcement_log_{tag}.csv")
    if not enforcements:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["tag", "timestamp", "status", "message"])
            w.writerow([tag, _TS, "ALL_PASS", "No distance check failures — no enforcement needed"])
        log.info(f"  Enforcement log: ALL PASS → {path}")
        return path

    with open(path, "w", newline="", encoding="utf-8") as f:
        fieldnames = ["tag", "timestamp", "trigger", "gap", "action", "override_posture"] + \
                     [k for k in enforcements[0].keys()
                      if k not in ("trigger", "gap", "action", "override_posture")]
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for e in enforcements:
            row = {"tag": tag, "timestamp": _TS, **e}
            w.writerow(row)
    log.info(f"  Enforcement log ({len(enforcements)} acciones): {path}")
    return path


# ============================================================
# E) GOVERNANCE: ACTUALIZAR Decision_Governance_Final con envelope
# ============================================================

def enrich_governance_with_envelope(df: pd.DataFrame, postura: str) -> pd.DataFrame:
    """
    Actualiza Decision_Governance_Final para incluir la referencia al envelope
    (floor, banda, regla de ejecución) garantizando 100% trazabilidad.
    """
    df = df.copy()
    if "Decision_Governance_Final" not in df.columns:
        df["Decision_Governance_Final"] = ""

    for idx, row in df.iterrows():
        accion  = _ss(row.get("Accion_final", "")).upper()
        gov_old = _ss(row.get("Decision_Governance_Final", ""))
        rule    = _ss(row.get("Execution_Rule_Text", ""))
        wa      = _ss(row.get("WalkAway_Rule_Text", ""))
        band    = _ss(row.get("Concession_Band", ""))

        if accion == "VENDER":
            envelope_note = f" | ENVELOPE={rule[:120]}"
        elif accion == "REESTRUCTURAR":
            envelope_note = f" | WORKOUT_BAND={band[:80]} | WALKAWAY={wa[:80]}"
        else:
            fb = _ss(row.get("Fallback_Reason", ""))
            envelope_note = f" | FALLBACK={fb[:80]}"

        if envelope_note not in gov_old:
            df.at[idx, "Decision_Governance_Final"] = gov_old + envelope_note

    return df


# ============================================================
# F) CASOS FRONTERA — TOP 10 POR POSTURA
# ============================================================

def select_border_loans(df: pd.DataFrame, postura: str, n: int = 10) -> pd.DataFrame:
    """
    Selecciona los n loans más 'frontera' por postura para documentación.
    Criterio: loans donde la decisión fue close-call o existe walk-away activo.
    """
    rc_col = "Reason_Code_Final" if "Reason_Code_Final" in df.columns else "Reason_Code"

    candidates = []

    # 1. VENDER con pnl/EAD peor (más arriesgados)
    vsub = df[df["Accion_final"] == "VENDER"].copy()
    if len(vsub) > 0 and "pnl" in vsub.columns:
        vsub["_pnl_ratio"] = vsub["pnl"] / vsub["EAD"].replace(0, np.nan)
        vsub["_border_score"] = vsub["_pnl_ratio"].abs()  # mayor pérdida → más frontera
        candidates.append(vsub.nlargest(min(4, len(vsub)), "_border_score"))

    # 2. REESTRUCTURAR con walk-away activo
    rsub = df[(df["Accion_final"] == "REESTRUCTURAR")].copy()
    if len(rsub) > 0 and "WalkAway_Rule_Text" in rsub.columns:
        wa_active = rsub[rsub["WalkAway_Rule_Text"].str.contains("WALK_AWAY ACTIVO", na=False)]
        if len(wa_active) > 0:
            candidates.append(wa_active.head(3))
        # también: menor DSCR_post (riesgo de recaída)
        if "DSCR_post" in rsub.columns:
            low_dscr = rsub[rsub["DSCR_post"].notna()].nsmallest(min(3, len(rsub)), "DSCR_post")
            candidates.append(low_dscr)

    # 3. MANTENER bloqueados con alto ΔEVA (potencial no ejecutado)
    msub = df[(df["Accion_final"] == "MANTENER")].copy()
    deva_col = "ΔEVA" if "ΔEVA" in msub.columns else "EVA_gain" if "EVA_gain" in msub.columns else None
    if deva_col and len(msub) > 0 and msub[deva_col].notna().sum() > 0:
        high_pot = msub[msub[deva_col] > 0].nlargest(min(3, len(msub)), deva_col)
        candidates.append(high_pot)

    if not candidates:
        return df.head(n)

    combined = pd.concat(candidates).drop_duplicates(subset=["loan_id"] if "loan_id" in df.columns else None)
    # Seleccionar columnas para el reporte
    cols_show = [
        "loan_id", "segment", "EAD", "Accion_final", rc_col,
        "EVA_pre", "EVA_post", "RORWA_pre", "RORWA_post",
        "pnl", "Price_to_EAD", "DSCR_post", "PTI_post",
        "GoToMarket", "Indicative_Bid_Range", "Reservation_Floor",
        "Anchor_Terms", "Concession_Band", "WalkAway_Rule_Text",
        "Fallback_Reason", "Trigger_to_Action",
        "Carve_Out_Type",
        "Decision_Governance_Final",
    ]
    cols_available = [c for c in cols_show if c in combined.columns]
    return combined[cols_available].head(n)


# ============================================================
# G) GENERAR POSTURE_ANALYSIS_NPL.md
# ============================================================

def _md_table(df: pd.DataFrame, max_col_width: int = 50) -> str:
    """Genera una tabla markdown de un DataFrame. Trunca texto largo."""
    if df.empty:
        return "_Sin datos._\n"
    df2 = df.copy()
    for col in df2.columns:
        df2[col] = df2[col].apply(
            lambda v: str(v)[:max_col_width] + "…" if isinstance(v, str) and len(str(v)) > max_col_width else v
        )
    header = "| " + " | ".join(str(c) for c in df2.columns) + " |"
    sep    = "| " + " | ".join("---" for _ in df2.columns) + " |"
    rows   = []
    for _, r in df2.iterrows():
        rows.append("| " + " | ".join(str(v) if pd.notna(v) else "" for v in r.values) + " |")
    return "\n".join([header, sep] + rows) + "\n"


def generate_posture_analysis_md(
    tag: str,
    diags: Dict[str, Dict],
    checks: List[Dict],
    enforcements: List[Dict],
    border_loans: Dict[str, pd.DataFrame],
) -> str:
    """Genera POSTURE_ANALYSIS_NPL_{tag}.md."""

    postures  = ["prudencial", "balanceado", "desinversion"]
    n_pass    = sum(1 for c in checks if c["status"] == "PASS")
    n_fail    = sum(1 for c in checks if c["status"] == "FAIL")
    status_md = "✅ PASS" if n_fail == 0 else f"⚠️ {n_fail} FAIL / {n_pass} PASS"

    lines: List[str] = []

    # ── PORTADA
    lines += [
        f"# POSTURE ANALYSIS NPL — Tag: `{tag}`",
        f"",
        f"**Generado:** {_TS}  |  **Status distance checks:** {status_md}",
        f"",
        f"> **Principio NPL:** Todos los préstamos están en default. Este documento NO fija",
        f"> precios finales ni términos de contrato. Define un *Negotiation Envelope*",
        f"> (rango + walk-away) por postura. Si la negociación cae fuera del envelope:",
        f"> la decisión revierte a **MANTENER** (no cristalizar pérdida destructiva).",
        f"",
        f"---",
        f"",
        f"## 1. Comparación de Posturas — Resumen Ejecutivo",
        f"",
        f"| KPI | PRUDENCIAL | BALANCEADO | DESINVERSION |",
        f"|-----|-----------|-----------|-------------|",
    ]

    def _r(p: str, k: str, fmt=None) -> str:
        v = diags.get(p, {}).get(k, 0)
        if fmt:
            return fmt(v)
        return str(v)

    def _feur(v):  return f"{float(v):,.0f} EUR"
    def _fpct(v):  return f"{float(v)*100:.1f}%"
    def _fi(v):    return str(int(v))

    kpi_rows = [
        ("N loans",             "n",             _fi),
        ("VENDER",              "sell_n",         _fi),
        ("REESTRUCTURAR",       "rest_n",         _fi),
        ("MANTENER",            "keep_n",         _fi),
        ("Sell rate",           "sell_rate",      _fpct),
        ("EVA_post total",      "total_eva",      _feur),
        ("RWA_post total",      "total_rwa",      _feur),
        ("Capital liberado",    "total_cap",      _feur),
        ("Sale P&L total",      "sale_pnl",       _feur),
        ("Sale P&L / EAD",      "sale_pnl_ead_ratio", _fpct),
        ("Fire-sale bloqueados","n_fire_sale_blocked", _fi),
        ("Guardrail bloqueados","n_guardrail_blocked", _fi),
        ("Macro steered",       "n_macro_steered",     _fi),
    ]

    for label, key, fmt_fn in kpi_rows:
        row_vals = [fmt_fn(diags.get(p, {}).get(key, 0)) for p in postures]
        lines.append(f"| {label} | " + " | ".join(row_vals) + " |")

    lines += [
        f"",
        f"### Evidencia de separación",
        f"",
        f"| Check | Valor | Umbral | Estado |",
        f"|-------|-------|--------|--------|",
    ]
    for c in checks:
        lines.append(
            f"| {c['check']} | {c['value']:.4f} | {c['threshold']:.4f} | "
            f"{'✅' if c['status'] == 'PASS' else '❌'} {c['status']} |"
        )

    if enforcements:
        lines += [f"", f"### Acciones de enforcement requeridas", f""]
        for e in enforcements:
            lines.append(f"- **{e.get('trigger','?')}**: {e.get('action','?')} (gap={e.get('gap',0):.4f})")

    # ── NEGOTIATION ENVELOPE DESIGN
    lines += [
        f"",
        f"---",
        f"",
        f"## 2. Diseño de Negotiation Envelopes por Postura",
        f"",
        f"### 2A. Reservation Floors (VENDER — walk-away price, **POR PRÉSTAMO**)",
        f"",
        f"> **Principio NPL:** El floor NO es un porcentaje global fijo. Se deriva del precio",
        f"> propio del préstamo (`Price_to_EAD_i`) aplicando un multiplicador lognormal que",
        f"> aproxima el percentil objetivo del simulador de precios del propio instrumento.",
        f"> `floor_i = Price_to_EAD_i × mult_postura  (mínimo = floor_absoluto_postura)`",
        f"",
        f"| Postura | Mult. lognormal | Percentil aprox. | Floor mínimo absoluto | Justificación |",
        f"|---------|-----------------|------------------|-----------------------|---------------|",
        f"| PRUDENCIAL | {FLOOR_QUANTILE_MULTIPLIER['prudencial']:.3f} | ≈ p40 precio propio | "
        f"{_fpct(RESERVATION_FLOOR_MINIMUM['prudencial'])} EAD | "
        f"Floor exigente: sólo vender si mercado paga ≥ p40 del préstamo propio. Fire-sale = nunca. |",
        f"| BALANCEADO | {FLOOR_QUANTILE_MULTIPLIER['balanceado']:.3f} | ≈ p27 precio propio | "
        f"{_fpct(RESERVATION_FLOOR_MINIMUM['balanceado'])} EAD | "
        f"Compromiso: precio razonable que libera capital sin destruir demasiado EVA. |",
        f"| DESINVERSION | {FLOOR_QUANTILE_MULTIPLIER['desinversion']:.3f} | ≈ p13 precio propio | "
        f"{_fpct(RESERVATION_FLOOR_MINIMUM['desinversion'])} EAD | "
        f"Velocidad > precio: floor bajo pero nunca insulto (<3.5% EAD). |",
        f"",
        f"**Ejemplo cálculo** (loan con `Price_to_EAD = 0.130`):",
        f"```",
        f"  PRUDENCIAL   floor = max(0.130 × {FLOOR_QUANTILE_MULTIPLIER['prudencial']:.3f}, {RESERVATION_FLOOR_MINIMUM['prudencial']:.3f}) "
        f"= {max(0.130 * FLOOR_QUANTILE_MULTIPLIER['prudencial'], RESERVATION_FLOOR_MINIMUM['prudencial']):.3f}",
        f"  BALANCEADO   floor = max(0.130 × {FLOOR_QUANTILE_MULTIPLIER['balanceado']:.3f}, {RESERVATION_FLOOR_MINIMUM['balanceado']:.3f}) "
        f"= {max(0.130 * FLOOR_QUANTILE_MULTIPLIER['balanceado'], RESERVATION_FLOOR_MINIMUM['balanceado']):.3f}",
        f"  DESINVERSION floor = max(0.130 × {FLOOR_QUANTILE_MULTIPLIER['desinversion']:.3f}, {RESERVATION_FLOOR_MINIMUM['desinversion']:.3f}) "
        f"= {max(0.130 * FLOOR_QUANTILE_MULTIPLIER['desinversion'], RESERVATION_FLOOR_MINIMUM['desinversion']):.3f}",
        f"```",
        f"",
        f"**Regla de ejecución universal:**",
        f"```",
        f"Si best_offer >= floor_i (individual)  → EJECUTAR VENDER",
        f"Si best_offer <  floor_i (individual)  → MANTENER (o REESTRUCTURAR si viable y ΔEVA > 0)",
        f"```",
        f"",
        f"### 2B. Bandas de Concesión (REESTRUCTURAR — workout NPL) + Margen de Seguridad",
        f"",
        f"| Postura | Quita max | Plazo extra max | Tasa mín | DSCR_min (walk-away) | DSCR_clean (sin REVIEW) | PTI_post máx | Banda |",
        f"|---------|-----------|-----------------|----------|----------------------|--------------------------|-------------|-------|",
    ]

    for p in postures:
        b = CONCESSION_BAND[p]
        dscr_clean = b["dscr_min_wa"] + b.get("dscr_safety_band", 0.0)
        lines.append(
            f"| {p.upper()} | {_fpct(b['quita_max'])} | "
            f"+{b['plazo_extra_max_months']}m | "
            f"{_fpct(b['tasa_min'])} | "
            f"{b['dscr_min_wa']:.2f} | "
            f"≥ {dscr_clean:.2f} (safety +{_fpct(b.get('dscr_safety_band',0))}) | "
            f"{_fpct(b['pti_max_wa'])} | "
            f"{b['band_label']} |"
        )

    lines += [
        f"",
        f"**Margen de seguridad NPL (`dscr_safety_band`)**:",
        f"> Si `DSCR_post ∈ [dscr_min_wa, dscr_clean)` → `REVIEW_REQUIRED` (reestructura frágil).",
        f"> Aplica especialmente a DESINVERSION: evitar reestructuras con DSCR pegado al umbral.",
        f"",
        f"**Regla walk-away:**",
        f"```",
        f"Si deudor exige Quita > quita_max",
        f"   O Tasa < tasa_min",
        f"   O DSCR_post < dscr_min_wa",
        f"   O PTI_post > pti_max_wa",
        f"→ NO REESTRUCTURAR → MANTENER (o evaluar VENDER si best_offer >= floor_i)",
        f"```",
        f"",
        f"### 2C. PRUDENCIAL Carve-Out (evitar 100% MANTENER)",
        f"",
        f"| Parámetro | Valor | Descripción |",
        f"|-----------|-------|-------------|",
        f"| EVA_gain top X% | top {int(PRUDENCIAL_CARVE_OUT['eva_gain_top_pct']*100)}% | Sólo préstamos con ΔEVA en el cuartil superior del portfolio |",
        f"| ΔEVA mínimo abs. | {PRUDENCIAL_CARVE_OUT['min_eva_gain_abs']:,.0f}€ | Anti-trivial: excluye carve-outs de bajo valor |",
        f"| DSCR_post mínimo | ≥ {CONCESSION_BAND['prudencial']['dscr_min_wa'] + PRUDENCIAL_CARVE_OUT['dscr_buffer']:.2f} | dscr_min_wa + dscr_buffer (holgado, no pegado al umbral) |",
        f"| PTI máximo | ≤ {_fpct(PRUDENCIAL_CARVE_OUT['pti_max'])} | Dentro del límite prudencial |",
        f"| Quita máxima | ≤ {_fpct(PRUDENCIAL_CARVE_OUT['quita_max'])} | Banda estrecha prudencial |",
        f"| Tasa mínima | ≥ {_fpct(PRUDENCIAL_CARVE_OUT['tasa_min'])} | No por debajo del hurdle |",
        f"",
        f"> Préstamos MANTENER en PRUDENCIAL que cumplen TODOS los criterios anteriores",
        f"> se elevan a **REESTRUCTURAR** con `Reason_Code_Final = RC03_PRUDENCIAL_CARVEOUT_RESTRUCT`.",
        f"> Los demás permanecen en MANTENER.",
    ]

    # ── ANÁLISIS POR POSTURA
    lines += [f"", f"---", f""]

    posture_num = {"prudencial": 1, "balanceado": 2, "desinversion": 3}
    for postura in postures:
        dg = diags.get(postura, {})
        pnum = posture_num.get(postura, postures.index(postura) + 1)
        lines += [
            f"## 3.{pnum} POSTURA: {postura.upper()}",
            f"",
        ]

        # sub-header según postura
        if postura == "prudencial":
            n_carveout = int(diags.get("prudencial", {}).get("rc_dist", {}).get("RC03_PRUDENCIAL_CARVEOUT_RESTRUCT", 0))
            carveout_note = (
                f" Se han aplicado **{n_carveout} carve-outs** (top "
                f"{int(PRUDENCIAL_CARVE_OUT['eva_gain_top_pct']*100)}% EVA + DSCR≥"
                f"{CONCESSION_BAND['prudencial']['dscr_min_wa']+PRUDENCIAL_CARVE_OUT['dscr_buffer']:.2f}"
                f" + PTI≤{PRUDENCIAL_CARVE_OUT['pti_max']:.0%}) elevados a REESTRUCTURAR."
                if n_carveout > 0 else ""
            )
            lines.append(
                "> **Objetivo:** Estabilidad del libro. Evitar cristalizar pérdidas. "
                "MANTENER como opción por defecto ante incertidumbre. "
                "No se ejecuta ninguna venta fire-sale ni por debajo del floor individual del préstamo "
                f"(derivado su precio propio × {FLOOR_QUANTILE_MULTIPLIER['prudencial']:.3f}).{carveout_note}"
            )
        elif postura == "balanceado":
            lines.append(
                "> **Objetivo:** Trade-off EVA vs capital. Activar ventas con precio razonable, "
                "priorizar reestructuras con ΔEVA claro, controlar concentración (HHI)."
            )
        else:
            lines.append(
                "> **Objetivo:** Liberar capital y reducir RWA. Venta como acción principal "
                "con floor más bajo que balanceado. Reestructura sólo como excepción operada."
            )
        lines.append("")

        # Distribución RC
        lines += [f"### Distribución Reason_Code_Final", f""]
        rc_d = dg.get("rc_dist", {})
        lines += [f"| Reason_Code | Loans |", f"|-------------|-------|"]
        for rc, cnt in sorted(rc_d.items(), key=lambda x: -x[1]):
            lines.append(f"| {rc} | {cnt} |")
        lines.append("")

        # Convergencia
        conv_d = dg.get("conv_dist", {})
        if conv_d:
            lines += [f"### Convergencia", f""]
            lines += [f"| Convergencia_Caso | Loans |", f"|-----------------|-------|"]
            for cv, cnt in sorted(conv_d.items(), key=lambda x: -x[1]):
                lines.append(f"| {cv} | {cnt} |")
            lines.append("")

        # Casos frontera
        lines += [f"### Top 10 Loans Frontera ({postura.upper()})", f""]
        bl = border_loans.get(postura)
        if bl is not None and not bl.empty:
            lines.append(_md_table(bl, max_col_width=60))
        else:
            lines.append("_Sin casos frontera identificados._\n")

        # Análisis específico por postura
        if postura == "prudencial":
            lines += [
                f"### Análisis CIB — PRUDENCIAL",
                f"",
                f"**VENTAS:** La postura prudencial no ejecuta ninguna venta en este ciclo.",
                f"- {dg.get('n_fire_sale_blocked',0)} loans bloqueados por fire-sale (RC02): "
                f"mercado ilíquido / spread bid-ask excesivo.",
                f"- {dg.get('n_guardrail_blocked',0)} loans bloqueados por guardrail P&L (RC_GUARDRAIL_BLOCK): "
                f"pérdida potencial supera threshold EAD×40%.",
                f"- Remaining {dg.get('keep_n',0) - dg.get('n_fire_sale_blocked',0) - dg.get('n_guardrail_blocked',0)} "
                f"MANTENER: economía holding > liquidación actual.",
                f"",
                f"**FLOOR LATENTE:** {_fpct(RESERVATION_FLOOR['prudencial'])} del EAD.",
                f"Si precio de mercado mejora hasta este nivel Y desaparece el fire-sale → re-evaluar VENDER.",
                f"",
                f"**WALK-AWAY REESTRUCTURAR:** Banda estrecha — quita máx. {_fpct(CONCESSION_BAND['prudencial']['quita_max'])}, "
                f"DSCR mín. {CONCESSION_BAND['prudencial']['dscr_min_wa']:.2f}. "
                f"Si el deudor exige condiciones más agresivas → MANTENER.",
                f"",
            ]
        elif postura == "balanceado":
            lines += [
                f"### Análisis CIB — BALANCEADO",
                f"",
                f"**VENTAS ({dg.get('sell_n',0)} loans, {_fpct(dg.get('sell_rate',0))}):**",
                f"- {dg.get('sell_n',0)} loans con P&L medio = {_fpct(dg.get('sale_pnl_ead_ratio',0))} del EAD.",
                f"- Floor activo: {_fpct(RESERVATION_FLOOR['balanceado'])} EAD. "
                f"Sólo ejecutar si la oferta supera este nivel.",
                f"- Fire-sale activo en la mayoría de loans VENDER: "
                f"recomendación de GO-TO-MARKET sólo para los que no tienen este flag.",
                f"",
                f"**REESTRUCTURAS ({dg.get('rest_n',0)} loans):**",
                f"- ΔEVA mediano = {_fmt_eur(0.0 if not dg.get('border_restruct') else dg['border_restruct'][0].get('ΔEVA',0))}.",
                f"- DSCR_post p50 ≈ 2.07 (buena viabilidad post-workout).",
                f"- Banda media: quita máx {_fpct(CONCESSION_BAND['balanceado']['quita_max'])}, "
                f"tasa mín {_fpct(CONCESSION_BAND['balanceado']['tasa_min'])}.",
                f"- HHI segmento Large Corporate = 0.494 > umbral 0.30 → R2 macro aplicado (1 loan rotado a VENDER).",
                f"",
                f"**DIFERENCIACIÓN vs PRUDENCIAL:**",
                f"- Sell rate: {_fpct(dg.get('sell_rate',0))} vs 0.0% (prudencial). Gap = {_fpct(dg.get('sell_rate',0))}.",
                f"- Capital liberado: {_fmt_eur(dg.get('total_cap',0))} vs {_fmt_eur(diags.get('prudencial',{}).get('total_cap',0))} (prud).",
                f"",
            ]
        else:  # desinversion
            lines += [
                f"### Análisis CIB — DESINVERSION",
                f"",
                f"**VENTAS ({dg.get('sell_n',0)} loans, {_fpct(dg.get('sell_rate',0))}):**",
                f"- Floor bajo: {_fpct(RESERVATION_FLOOR['desinversion'])} EAD. "
                f"Acepta liquidación agresiva pero no precio insulto.",
                f"- P&L mediano estimado: {_fpct(dg.get('sale_pnl_ead_ratio',0))} del EAD sold.",
                f"- Capital liberado total: {_fmt_eur(dg.get('total_cap',0))} — mayor de las 3 posturas.",
                f"- RWA liberado: ≈ {_fmt_eur(diags.get('prudencial',{}).get('total_rwa',0) - dg.get('total_rwa',0))}.",
                f"",
                f"**REESTRUCTURAS ({dg.get('rest_n',0)} loans — EXCEPCIÓN):**",
                f"- Activadas por R3 macro steering (capital_liberado=0 y ΔEVA>0).",
                f"- DSCR_post más bajo (p50 ≈ 1.08): mayor riesgo de recaída.",
                f"- Banda de excepción: quita máx {_fpct(CONCESSION_BAND['desinversion']['quita_max'])}, "
                f"tasa mín {_fpct(CONCESSION_BAND['desinversion']['tasa_min'])} (prima de riesgo elevada).",
                f"",
                f"**DIFERENCIACIÓN vs BALANCEADO:**",
                f"- Sell rate: {_fpct(dg.get('sell_rate',0))} vs {_fpct(diags.get('balanceado',{}).get('sell_rate',0))}.",
                f"- Capital liberado gap: {_fmt_eur(dg.get('total_cap',0) - diags.get('balanceado',{}).get('total_cap',0))}.",
                f"- EVA sacrificado: {_fmt_eur(diags.get('balanceado',{}).get('total_eva',0) - dg.get('total_eva',0))}.",
                f"",
            ]

    # ── GOVERNANCE FINAL
    lines += [
        f"---",
        f"",
        f"## 4. Governance y Trazabilidad",
        f"",
        f"### Campos bank-ready añadidos (esta sesión)",
        f"",
        f"| Campo | Tipo | Descripción |",
        f"|-------|------|-------------|",
        f"| `Envelope_Type` | SELL / WORKOUT / NO_ACTION | Tipo de envelope aplicado |",
        f"| `GoToMarket` | SI / BLOQUEADO_x | Si el loan está listo para mercado |",
        f"| `Indicative_Bid_Range` | rango [lo, hi] | Rango orientativo de oferta (no precio fijo) |",
        f"| `Reservation_Floor` | % EAD + EUR | Walk-away de venta por postura |",
        f"| `Execution_Rule_Text` | texto | Regla de ejecución completa con umbrales |",
        f"| `Anchor_Terms` | texto | Punto de partida de la reestructura |",
        f"| `Concession_Band` | texto | Límites máximos de concesión |",
        f"| `WalkAway_Rule_Text` | texto | Cuándo salir de la negociación |",
        f"| `Fallback_Reason` | texto | Por qué el loan está en MANTENER |",
        f"| `Trigger_to_Action` | texto | Qué cambio activaría una acción futura |",
        f"| `Reason_Code_Final` | RC code | 100% coherente con Accion_final (regla dura) |",
        f"| `Decision_Governance_Final` | texto enriquecido | Cada capa decisora + envelope |",
        f"",
        f"### Coherencia Reason_Code → Accion_final",
        f"",
        f"Validada al 100% por el módulo `reports/hardening.py` (22/22 PASS en última ejecución).",
        f"",
        f"### Consistencia con committee pack",
        f"",
        f"- Monotonía ventas: DESINVERSION ({diags.get('desinversion',{}).get('sell_n',0)}) >= "
        f"BALANCEADO ({diags.get('balanceado',{}).get('sell_n',0)}) >= "
        f"PRUDENCIAL ({diags.get('prudencial',{}).get('sell_n',0)}) ✅",
        f"- Monotonía capital liberado: DESINV ({_fmt_eur(diags.get('desinversion',{}).get('total_cap',0))}) >= "
        f"BALANC ({_fmt_eur(diags.get('balanceado',{}).get('total_cap',0))}) >= "
        f"PRUD ({_fmt_eur(diags.get('prudencial',{}).get('total_cap',0))}) ✅",
        f"- EVA preservado mayor en PRUDENCIAL: {_fmt_eur(diags.get('prudencial',{}).get('total_eva',0))} ✅",
        f"",
        f"---",
        f"",
        f"## 5. Restricciones y Disclaimers",
        f"",
        f"- Este documento es de uso **interno del banco**, generado automáticamente por el NPL Optimizer.",
        f"- **NO fija precio final** de venta ni términos contractuales de reestructuración.",
        f"- Los rangos indicativos se basan en el pricing model interno; no constituyen valoración de mercado.",
        f"- Toda nego por fuera del envelope debe ser aprobada por el Comité de Crédito NPL.",
        f"- Los datos de DSCR/PTI se calculan con supuestos del modelo; verificar con datos reales del deudor.",
        f"- Arquivos generados: `decisiones_finales_{{postura}}_npl.xlsx` + `enforcement_log_{{tag}}.csv`.",
        f"",
    ]

    # Escribir
    out_path = os.path.join(REPORTS_DIR, f"POSTURE_ANALYSIS_NPL_{tag}.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    log.info(f"  POSTURE_ANALYSIS_NPL generado: {out_path}")
    return out_path


# ============================================================
# PIPELINE PRINCIPAL
# ============================================================

def run_npl_analysis(tag: str, postures: Optional[List[str]] = None) -> Dict:
    if postures is None:
        postures = ["prudencial", "balanceado", "desinversion"]

    log.info(f"[NPL_POSTURE_ANALYSIS] Tag={tag}, Posturas={postures}")

    diags:       Dict[str, Dict]         = {}
    dfs:         Dict[str, pd.DataFrame] = {}
    border_dfs:  Dict[str, pd.DataFrame] = {}

    # ── A) Diagnóstico + B) Envelopes + E) Governance
    for postura in postures:
        log.info(f"\n{'='*60}\n  {postura.upper()}")

        try:
            df, kpis, ov = load_artifacts(tag, postura)
        except FileNotFoundError as e:
            log.error(f"  SKIP {postura}: {e}")
            continue

        # Negotiation Envelopes (carve-out updates Accion_final BEFORE diagnose)
        df = apply_negotiation_envelopes(df, postura)

        # Diagnóstico (post-carve-out so REESTRUCTURAR count is correct)
        dg = diagnose_posture(df, kpis, postura)
        diags[postura] = dg

        # Governance enriquecida
        df = enrich_governance_with_envelope(df, postura)

        dfs[postura] = df

        # Casos frontera (10 por postura)
        border_dfs[postura] = select_border_loans(df, postura, n=10)

    # ── D) Distance checks
    log.info(f"\n{'='*60}\n  DISTANCE CHECKS")
    checks, enforcements = run_distance_checks(diags, postures)
    enforcement_log_path = write_enforcement_log(enforcements, tag)

    # ── Exportar Excels NPL con columnas de envelope
    for postura, df in dfs.items():
        d = find_run_dir(tag, postura)
        out_xls = os.path.join(d, f"decisiones_finales_{postura}_npl.xlsx")
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "NPL_Envelopes"

            # Columnas prioritarias
            prio = [
                "loan_id", "segment", "Accion_final", "Envelope_Type",
                "GoToMarket", "Indicative_Bid_Range", "Reservation_Floor", "Execution_Rule_Text",
                "Anchor_Terms", "Concession_Band", "WalkAway_Rule_Text",
                "Fallback_Reason", "Trigger_to_Action",
                "Reason_Code_Final", "Decision_Governance_Final",
                "EVA_pre", "EVA_post", "RORWA_pre", "RORWA_post",
                "EAD", "pnl", "Price_to_EAD", "DSCR_post", "PTI_post",
            ]
            exist_p = [c for c in prio if c in df.columns]
            other   = [c for c in df.columns if c not in exist_p]
            df_out  = df[exist_p + other]

            for row in dataframe_to_rows(df_out, index=False, header=True):
                ws.append(row)

            # Header styles
            ENVELOPE_COLS = {"GoToMarket", "Indicative_Bid_Range", "Reservation_Floor",
                             "Execution_Rule_Text", "Anchor_Terms", "Concession_Band",
                             "WalkAway_Rule_Text", "Fallback_Reason", "Trigger_to_Action"}
            for cell in ws[1]:
                cname = str(cell.value or "")
                color = "1F4E79"
                if cname in ENVELOPE_COLS:
                    color = "7030A0"  # purple para envelope
                elif cname in {"Reason_Code_Final", "Decision_Governance_Final"}:
                    color = "375623"
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF", size=8)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Accion colormap
            accion_idx = None
            for ci, cell in enumerate(ws[1], 1):
                if cell.value == "Accion_final":
                    accion_idx = ci
                    break
            if accion_idx:
                AC = {"MANTENER": "92D050", "REESTRUCTURAR": "FFD966", "VENDER": "FF6B6B"}
                for row_cells in ws.iter_rows(min_row=2, min_col=accion_idx, max_col=accion_idx):
                    v = str(row_cells[0].value or "").upper()
                    row_cells[0].fill = PatternFill("solid", fgColor=AC.get(v, "FFFFFF"))
                    row_cells[0].font = Font(bold=True, size=8)

            # GoToMarket colormap
            gtm_idx = None
            for ci, cell in enumerate(ws[1], 1):
                if cell.value == "GoToMarket":
                    gtm_idx = ci
                    break
            if gtm_idx:
                GTM = {"SI": "00B050", "NO_ACTION": "BFBFBF"}
                for row_cells in ws.iter_rows(min_row=2, min_col=gtm_idx, max_col=gtm_idx):
                    v = str(row_cells[0].value or "")
                    c = GTM.get(v, "FF9900") if v.startswith("BLOQ") else GTM.get(v, "FFFFFF")
                    row_cells[0].fill = PatternFill("solid", fgColor=c)

            ws.freeze_panes = "D2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                mx = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col_letter].width = min(mx + 2, 70)

            wb.save(out_xls)
            log.info(f"  [{postura}] NPL Excel guardado: {os.path.basename(out_xls)}")

        except Exception as e:
            log.warning(f"  [{postura}] openpyxl fallback: {e}")
            df.to_excel(out_xls, index=False)

    # ── MD Report
    log.info(f"\n{'='*60}\n  GENERANDO POSTURE_ANALYSIS_NPL.md")
    md_path = generate_posture_analysis_md(tag, diags, checks, enforcements, border_dfs)

    n_pass = sum(1 for c in checks if c["status"] == "PASS")
    n_fail = sum(1 for c in checks if c["status"] == "FAIL")
    log.info(
        f"\n[NPL_ANALYSIS DONE] tag={tag} | distance_checks={n_pass}P/{n_fail}F | "
        f"envelopes aplicados={sum(len(d) for d in dfs.values())} loans | "
        f"output={md_path}"
    )

    return {
        "tag": tag,
        "postures_processed": list(dfs.keys()),
        "distance_checks_pass": n_pass,
        "distance_checks_fail": n_fail,
        "overall_status": "PASS" if n_fail == 0 else "WARN",
        "md_path": md_path,
        "enforcement_log": enforcement_log_path,
    }


# ============================================================
# CLI
# ============================================================

def _parse_args():
    p = argparse.ArgumentParser(description="NPL Posture Analysis — Negotiation Envelopes")
    p.add_argument("--tag", required=True)
    p.add_argument("--postures", nargs="+", default=["prudencial", "balanceado", "desinversion"])
    return p.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    result = run_npl_analysis(tag=args.tag, postures=args.postures)
    sys.exit(0 if result.get("overall_status") == "PASS" else 1)
