# -*- coding: utf-8 -*-
# ============================================================
# reports/hardening.py
# Autor: José María Fernández-Ladreda Ballvé
# Resumen: Capa post-proceso bank-ready: Final Reasoning Seal, macro steering por postura, completitud de métricas y governance auditable.
# ============================================================

"""
Módulo de post-proceso que endurece la coherencia y auditabilidad de
las decisiones del optimizador NPL bajo Basilea III STD.

TAREAS IMPLEMENTADAS:
  1. Final Reasoning Seal: Reason_Code_Micro, Reason_Code_Macro,
     Reason_Code_Final (100% consistente con Accion_final),
     Decision_Governance_Final.
  2. Macro Steering por postura (reglas deterministas auditables,
     tasa de intervención >= 20%).
  3. Completar métricas post para MANTENER/REESTRUCTURAR (no NaN).
  4. Renombrar guardrail PNL_TOO_NEGATIVE_PRUDENCIAL → PNL_TOO_NEGATIVE_EAD40.
  5. Validaciones automáticas (assert suite banco-ready).
  6. Generar nuevos Excels + BANK_READY_HARDENING.md.

CRITERIOS DE ACEPTACIÓN:
  - Ningún Accion_final incompatible con Reason_Code_Final.
  - No NaN en métricas post para MANTENER y REESTRUCTURAR.
  - Monotonía entre posturas preservada.
  - Toda regla macro tiene umbral + variable registrada.
  - macro_applied_rate >= 20% o campo macro_inactive_warning en KPIs.

USO:
    python -m reports.hardening --tag infer_ci0222
    python -m reports.hardening --tag infer_ci0222 --postures prudencial balanceado desinversion
"""

from __future__ import annotations

import argparse
import glob
import json
import logging
import os
import re
import sys
import warnings
from copy import deepcopy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ── Rutas
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

REPORTS_DIR = os.path.join(ROOT_DIR, "reports")
LOGS_DIR    = os.path.join(ROOT_DIR, "logs")
os.makedirs(LOGS_DIR, exist_ok=True)

# ── Logger
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(name)s | %(message)s",
    handlers=[
        logging.FileHandler(
            os.path.join(LOGS_DIR, "hardening.log"), encoding="utf-8"
        ),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("hardening")

# ── Versión
_VERSION = "1.0.0"
_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# ============================================================
# CONSTANTES
# ============================================================

HURDLE_RATE = 0.10  # 10% RORWA hurdle (Basilea III STD)
CAPITAL_RATIO = 0.105  # 10.5% ratio total capital

# Reglas de coherencia RC → Accion_final
RC_VENDER_SET = {
    "RC01_MACRO_SELL_STEERING",
    "RC06_MICRO_SELL_VALUE_NEGATIVE",
    "RC07_SELL_NPE_REDUCTION",
    "RC08_SELL_RWA_RELIEF",
}
RC_REESTRUCTURAR_SET = {
    "RC03_MICRO_RESTRUCT_VALUE_UPLIFT",
    "RC03_MACRO_RESTRUCT_STEERING",
    "RC04_RESTRUCT_NPE_REHAB",
    "RC09_RESTRUCT_REQUIRES_VIABILITY_CHECK",
}
RC_MANTENER_SET = {
    "RC02_SELL_BLOCKED_FIRE_SALE",
    "RC05_KEEP_MEETS_HURDLE",
    "RC05_KEEP_ACCEPTABLE_ECONOMICS",
    "RC10_MISSING_VIABILITY_INPUTS",
    "RC11_MACRO_KEEP_CONCENTRATION",
    "RC12_MACRO_KEEP_PRUD_STABILITY",
    "RC13_MACRO_KEEP_DESINV_FLOOR",
    "RC15_VOLUME_CAP_VOLUNTARY_SELL",
    "RC_GUARDRAIL_BLOCK",
}

# Mapping accion → RC válidos
ACTION_RC_MAP = {
    "VENDER":       RC_VENDER_SET,
    "REESTRUCTURAR": RC_REESTRUCTURAR_SET,
    "MANTENER":     RC_MANTENER_SET,
}

# Guardrail rename map
GUARDRAIL_RENAME = {
    "PNL_TOO_NEGATIVE_PRUDENCIAL": "PNL_TOO_NEGATIVE_EAD40",
}

# Macro steering thresholds
MACRO_THRESHOLDS = {
    "prudencial": {
        # Max P&L negativo agregado: -15% del total EAD
        "MAX_AGGREGATE_PNL_PCT_EAD": -0.15,
        # Fire-sale: toda venta con fire_sale=True se convierte en MANTENER
        "FLIP_FIRE_SALE_VENDER_TO_MANTENER": True,
        # Min ventas tras steering: 5% (no se puede quedar en 0 si hay loans muy deteriorados)
        "MIN_SELL_PCT": 0.05,
    },
    "balanceado": {
        # HHI máximo tolerable por segmento
        "MAX_HHI_SEGMENT": 0.30,
        # Top N exposures que pueden ser REESTRUCTURAR (por EAD relativo)
        "TOP_EXPOSURE_PCT": 0.20,
        # Penalización si EVA_post < 0 y Accion_final=MANTENER → flip a REESTRUCTURAR (si viable)
        "EVA_NEG_MANTENER_TO_RESTRUCT": True,
    },
    "desinversion": {
        # Floor mínimo de precio vs EAD para ejecutar venta
        "MIN_PRICE_TO_EAD_FLOOR": 0.10,
        # Max pérdida fraccional por loan vs EAD para ejecutar venta
        "MAX_LOSS_PCT_EAD": -0.50,
        # Sólo vender si capital_liberado > 0
        "REQUIRE_POSITIVE_CAP_RELEASE": True,
    },
}

# ============================================================
# HELPERS
# ============================================================

def _sf(x: Any, default: float = 0.0) -> float:
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return default
        return float(x)
    except Exception:
        return default


def _ss(x: Any, default: str = "") -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return default
    return str(x)


def _nan_to_none(v: float) -> Optional[float]:
    return None if (v is None or (isinstance(v, float) and np.isnan(v))) else v


def _fmt_eur(x: float) -> str:
    return f"{x:+,.0f}€"


def _fmt_pct(x: float) -> str:
    return f"{x * 100:.2f}%"


def _is_vender(x: str) -> bool:
    return _ss(x, "").upper() == "VENDER"


def _is_restruct(x: str) -> bool:
    return _ss(x, "").upper() == "REESTRUCTURAR"


def _is_mantener(x: str) -> bool:
    return _ss(x, "").upper() == "MANTENER"


def _rc_coherent(rc: str, accion: str) -> bool:
    """Devuelve True si el Reason_Code es compatible con la Accion_final."""
    valid = ACTION_RC_MAP.get(_ss(accion, "").upper(), set())
    return _ss(rc, "") in valid


# ============================================================
# LOCALIZAR RUNS
# ============================================================

def find_inference_dir(tag: str, postura: str, reports_dir: str = REPORTS_DIR) -> Optional[str]:
    """Encuentra la carpeta más reciente de coordinator_inference para un tag y postura."""
    pattern = os.path.join(reports_dir, f"coordinated_inference_{tag}_*_{postura}")
    dirs = sorted(glob.glob(pattern), reverse=True)
    if dirs:
        return dirs[0]
    # Fallback: buscar sin separador ts
    pattern2 = os.path.join(reports_dir, f"coordinated_inference_*{tag}*_{postura}")
    dirs2 = sorted(glob.glob(pattern2), reverse=True)
    return dirs2[0] if dirs2 else None


def load_posture_artifacts(tag: str, postura: str) -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """Carga Excel, KPIs JSON y overrides CSV para una postura."""
    d = find_inference_dir(tag, postura)
    if not d:
        raise FileNotFoundError(
            f"No se encontró carpeta de inferencia para tag={tag}, postura={postura}"
        )

    xlsx_path = os.path.join(d, f"decisiones_finales_{postura}.xlsx")
    kpis_path = os.path.join(d, f"portfolio_kpis_{postura}.json")
    ov_path   = os.path.join(d, f"overrides_log_{postura}.csv")

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel no encontrado: {xlsx_path}")

    df = pd.read_excel(xlsx_path)
    logger.info(f"  Cargado {postura}: {len(df)} préstamos desde {os.path.basename(d)}")

    kpis: Dict = {}
    if os.path.exists(kpis_path):
        with open(kpis_path, encoding="utf-8") as f:
            kpis = json.load(f)

    ov = pd.DataFrame()
    if os.path.exists(ov_path):
        ov = pd.read_csv(ov_path)

    return df, kpis, ov


# ============================================================
# TAREA 4: COMPLETAR MÉTRICAS POST (sin NaN para MANTENER/REESTRUCTURAR)
# ============================================================

def complete_post_metrics(df: pd.DataFrame, postura: str) -> pd.DataFrame:
    """
    Garantiza que EVA_post, RWA_post, RORWA_post no sean NaN para
    Accion_final ∈ {MANTENER, REESTRUCTURAR}.

    Reglas deterministas:
      MANTENER:     EVA_post = EVA_pre * (1 - drift)
                    RWA_post = RWA_pre
                    RORWA_post = EVA_post / RWA_post  (si RWA_post > 0)
      REESTRUCTURAR: usa el valor calculado por el micro si existe;
                    sino, aplica EVA_post = EVA_pre + ΔEVA (si ΔEVA disponible)
      VENDER:       EVA_post = 0 por diseño (capital fuera del libro)
                    RORWA_post = NaN (aceptable — sin activo residual)
    """
    df = df.copy()
    # drift ambiental mínimo: 0% (idéntico a pre) para MANTENER
    DRIFT = 0.0

    # Ensure columns exist
    for col in ["EVA_post", "RWA_post", "RORWA_post"]:
        if col not in df.columns:
            df[col] = np.nan

    for idx, row in df.iterrows():
        accion = _ss(row.get("Accion_final", ""), "").upper()

        if accion == "VENDER":
            # Por diseño: sin activo residual
            df.at[idx, "EVA_post"]   = 0.0
            df.at[idx, "RWA_post"]   = 0.0
            df.at[idx, "RORWA_post"] = np.nan
            continue

        # MANTENER o REESTRUCTURAR ─ no puede quedar NaN
        # Si el loan fue flipeado por macro steering desde VENDER, sus valores
        # post pueden ser 0.0 (el Excel original los guardó como VENDER).
        # En ese caso, hay que recalcular obligatoriamente.
        macro_steered = bool(_ss(row.get("Macro_Steering_Applied", ""), "").strip())
        was_vender_post = (
            _sf(row.get("EVA_post", np.nan), np.nan) == 0.0
            and _sf(row.get("RWA_post", np.nan), np.nan) == 0.0
            and macro_steered
        )

        eva_post_existing   = None if was_vender_post else _nan_to_none(_sf(row.get("EVA_post"), np.nan))
        rwa_post_existing   = None if was_vender_post else _nan_to_none(_sf(row.get("RWA_post"), np.nan))
        rorwa_post_existing = None if was_vender_post else _nan_to_none(_sf(row.get("RORWA_post"), np.nan))

        eva_pre = _sf(row.get("EVA_pre", row.get("EVA", 0.0)))
        rwa_pre = _sf(row.get("RWA_pre", row.get("rwa_before", row.get("RWA", 0.0))))
        deva    = _sf(row.get("ΔEVA", row.get("EVA_gain", row.get("delta_eva", 0.0))))

        # ── EVA_post
        if eva_post_existing is None:
            if accion == "REESTRUCTURAR" and deva != 0.0:
                eva_post = eva_pre + deva
            else:
                eva_post = eva_pre * (1.0 - DRIFT)
            df.at[idx, "EVA_post"] = eva_post
        else:
            eva_post = float(eva_post_existing)

        # ── RWA_post
        if rwa_post_existing is None:
            rwa_post = rwa_pre  # sin acción → RWA sin cambio
            df.at[idx, "RWA_post"] = rwa_post
        else:
            rwa_post = float(rwa_post_existing)

        # ── RORWA_post
        if rorwa_post_existing is None:
            if rwa_post > 0:
                rorwa_post = eva_post / (rwa_post * CAPITAL_RATIO)
            elif rwa_pre > 0:
                rorwa_post = eva_post / (rwa_pre * CAPITAL_RATIO)
            else:
                rorwa_post = np.nan
            df.at[idx, "RORWA_post"] = rorwa_post

    logger.info(
        f"  [{postura}] Métricas post completadas: "
        f"{df['RORWA_post'].isna().sum()} NaN restantes en RORWA_post "
        f"({(df['Accion_final']=='VENDER').sum()} VENDER tienen NaN por diseño)"
    )
    return df


# ============================================================
# TAREA 5: RENOMBRAR GUARDRAILS
# ============================================================

def rename_guardrail_reasons(df: pd.DataFrame, postura: str) -> pd.DataFrame:
    """
    Renombra guardrail reason codes confusos y añade 'applies_to_postures'
    al texto para evitar ambigüedad entre posturas.
    """
    df = df.copy()
    cols_to_fix = ["guardrail_reasons", "Decision_Governance", "Decision_Governance_Final"]

    for col in cols_to_fix:
        if col not in df.columns:
            continue
        for old_name, new_name in GUARDRAIL_RENAME.items():
            df[col] = df[col].apply(
                lambda v: _ss(v, "").replace(old_name, f"{new_name} [postura={postura}]")
                if isinstance(v, str) else v
            )

    # Añadir campo auditable 'guardrail_posture' para el log
    if "guardrail_reasons" in df.columns:
        df["guardrail_posture"] = df["guardrail_reasons"].apply(
            lambda v: postura if (isinstance(v, str) and v.strip()) else ""
        )

    logger.info(f"  [{postura}] Guardrail naming actualizado.")
    return df


# ============================================================
# TAREA 2: FINAL REASONING SEAL
# ============================================================

def _derive_rc_micro(row: pd.Series, postura: str) -> str:
    """Deriva el Reason_Code que habría producido el micro-modelo."""
    accion_micro = _ss(row.get("Accion_micro", row.get("Accion", "")), "").upper()
    eva_pre  = _sf(row.get("EVA_pre", 0.0))
    deva     = _sf(row.get("ΔEVA", row.get("EVA_gain", 0.0)))
    rorwa_pre = _sf(row.get("RORWA_pre", 0.0))

    if accion_micro == "VENDER":
        return "RC06_MICRO_SELL_VALUE_NEGATIVE"
    if accion_micro == "REESTRUCTURAR":
        return "RC03_MICRO_RESTRUCT_VALUE_UPLIFT" if deva > 0 else "RC09_RESTRUCT_REQUIRES_VIABILITY_CHECK"
    # MANTENER
    if rorwa_pre >= HURDLE_RATE:
        return "RC05_KEEP_MEETS_HURDLE"
    return "RC05_KEEP_ACCEPTABLE_ECONOMICS"


def _derive_rc_macro(row: pd.Series, postura: str) -> str:
    """Deriva el Reason_Code de la acción propuesta por el macro."""
    macro_applied   = bool(row.get("macro_applied", False))
    macro_action    = _ss(row.get("macro_action_used", row.get("Accion_macro", "")), "").upper()
    convergencia    = _ss(row.get("Convergencia_Caso", "MACRO_NOT_APPLIED"), "")

    if not macro_applied and convergencia == "MACRO_NOT_APPLIED":
        return "MACRO_NOT_APPLIED"
    if macro_action == "VENDER":
        return "RC01_MACRO_SELL_STEERING"
    if macro_action == "REESTRUCTURAR":
        return "RC03_MACRO_RESTRUCT_STEERING"
    return "RC05_KEEP_ACCEPTABLE_ECONOMICS"


def _derive_rc_final(row: pd.Series) -> str:
    """
    Regla dura: Reason_Code_Final DEBE ser coherente con Accion_final.
    Jerarquía:
      1. Si override_applied → guardrail/macro gana
      2. E/o convergencia GUARDRAIL_OVERRIDE o MACRO_WINS → usa macro
      3. Si Reason_Code existente ya es coherente → mantenerlo
      4. Sino → reasignar coherente con Accion_final
    """
    accion        = _ss(row.get("Accion_final", ""), "").upper()
    rc_existing   = _ss(row.get("Reason_Code", ""), "")
    override_app  = bool(row.get("override_applied", False))
    convergencia  = _ss(row.get("Convergencia_Caso", ""), "")
    macro_applied = bool(row.get("macro_applied", False))
    macro_action  = _ss(row.get("macro_action_used", ""), "").upper()
    guardrail_r   = _ss(row.get("guardrail_reasons", ""), "")
    macro_steer   = _ss(row.get("Macro_Steering_Applied", ""), "")

    # ── Prioridad 1: Macro steering explícito (por hardening)
    if macro_steer:
        if accion == "VENDER":
            return "RC01_MACRO_SELL_STEERING"
        if accion == "REESTRUCTURAR":
            return "RC03_MACRO_RESTRUCT_STEERING"
        if accion == "MANTENER" and "FIRE_SALE" in macro_steer:
            return "RC02_SELL_BLOCKED_FIRE_SALE"
        if accion == "MANTENER" and "CONCENTRATION" in macro_steer:
            return "RC11_MACRO_KEEP_CONCENTRATION"
        if accion == "MANTENER" and "STABILITY" in macro_steer:
            return "RC12_MACRO_KEEP_PRUD_STABILITY"
        if accion == "MANTENER" and "FLOOR" in macro_steer:
            return "RC13_MACRO_KEEP_DESINV_FLOOR"

    # ── Prioridad 2: Guardrail override
    if guardrail_r and override_app:
        renamed_guardrail = guardrail_r
        for old, new in GUARDRAIL_RENAME.items():
            renamed_guardrail = renamed_guardrail.replace(old, new)
        if accion == "MANTENER":
            return "RC_GUARDRAIL_BLOCK"
        # Si guardrail bloqueó venta → se convirtió en MANTENER
    if convergencia == "GUARDRAIL_OVERRIDE" and accion == "MANTENER":
        return "RC_GUARDRAIL_BLOCK"

    # ── Prioridad 3: RC existente ya coherente → preservar
    if rc_existing and _rc_coherent(rc_existing, accion):
        # Renombrar si necesario
        for old, new in GUARDRAIL_RENAME.items():
            if old in rc_existing:
                return rc_existing.replace(old, new)
        return rc_existing

    # ── Prioridad 4: RC existente NO coherente → reasignar
    if accion == "VENDER":
        # ¿Venta por macro steering?
        if macro_applied and macro_action == "VENDER":
            return "RC01_MACRO_SELL_STEERING"
        return "RC06_MICRO_SELL_VALUE_NEGATIVE"

    if accion == "REESTRUCTURAR":
        if macro_applied and macro_action == "REESTRUCTURAR":
            return "RC03_MACRO_RESTRUCT_STEERING"
        deva = _sf(row.get("ΔEVA", row.get("EVA_gain", 0.0)))
        return "RC03_MICRO_RESTRUCT_VALUE_UPLIFT" if deva > 0 else "RC09_RESTRUCT_REQUIRES_VIABILITY_CHECK"

    if accion == "MANTENER":
        # ¿Fue bloqueada una venta?
        accion_micro = _ss(row.get("Accion_micro", ""), "").upper()
        fire_sale    = bool(row.get("fire_sale", row.get("Fire_Sale", False)))
        if (accion_micro == "VENDER" or macro_action == "VENDER") and fire_sale:
            return "RC02_SELL_BLOCKED_FIRE_SALE"

        rorwa_pre = _sf(row.get("RORWA_pre", 0.0))
        if rorwa_pre >= HURDLE_RATE:
            return "RC05_KEEP_MEETS_HURDLE"
        return "RC05_KEEP_ACCEPTABLE_ECONOMICS"

    return "RC05_KEEP_ACCEPTABLE_ECONOMICS"  # fallback seguro


def _derive_governance_final(row: pd.Series, postura: str) -> str:
    """
    Decision_Governance_Final con trazabilidad completa:
    Winner, umbral y variable clave.
    """
    accion      = _ss(row.get("Accion_final", ""), "").upper()
    accion_micro = _ss(row.get("Accion_micro", row.get("Accion", "")), "").upper()
    convergencia = _ss(row.get("Convergencia_Caso", ""), "")
    override_app = bool(row.get("override_applied", False))
    macro_applied = bool(row.get("macro_applied", False))
    macro_action  = _ss(row.get("macro_action_used", ""), "").upper()
    guardrail_r   = _ss(row.get("guardrail_reasons", ""), "")
    macro_steer   = _ss(row.get("Macro_Steering_Applied", ""), "")
    rc_final      = _ss(row.get("Reason_Code_Final", ""), "")

    # Renombrar guardrail
    for old, new in GUARDRAIL_RENAME.items():
        guardrail_r = guardrail_r.replace(old, f"{new} [postura={postura}]")

    p_u = postura.upper()

    # ── Macro Steering por hardening
    if macro_steer:
        rule, umbral, variable = macro_steer.partition("|")
        return (
            f"[{p_u}] MACRO_HARDENING_STEERING: {rule.strip()} "
            f"(umbral={umbral.strip()}, variable={variable.strip()}). "
            f"Micro={accion_micro} → Final={accion}."
        )

    # ── Guardrail vence
    if guardrail_r and override_app and accion != accion_micro:
        return (
            f"[{p_u}] GUARDRAIL_OVERRIDE: {guardrail_r}. "
            f"Micro={accion_micro} → Final={accion} (guardrail aplicado)."
        )
    if convergencia == "GUARDRAIL_OVERRIDE":
        return (
            f"[{p_u}] GUARDRAIL_OVERRIDE: {guardrail_r or 'ver guardrail_reasons'}. "
            f"Micro={accion_micro} → Final={accion}."
        )

    # ── Macro portfolio gana
    if macro_applied and macro_action == accion and macro_action != accion_micro:
        return (
            f"[{p_u}] MACRO_WINS: macro_action={macro_action} coincide con Accion_final. "
            f"Micro={accion_micro} → Final={accion}."
        )

    # ── Micro decides (micro = final)
    if accion == accion_micro:
        rp = _ss(row.get("RORWA_pre"), "")
        eva_pre = _sf(row.get("EVA_pre", 0.0))
        deva    = _sf(row.get("ΔEVA", row.get("EVA_gain", 0.0)))
        if accion == "VENDER":
            return (
                f"[{p_u}] MICRO_LED: micro-model sell decision. "
                f"EVA_pre={_fmt_eur(eva_pre)}, RORWA_pre<hurdle={_fmt_pct(HURDLE_RATE)}."
            )
        if accion == "REESTRUCTURAR":
            return (
                f"[{p_u}] MICRO_LED: micro-model restruct decision. "
                f"ΔEVA={_fmt_eur(deva)} (hurdle_delta>0)."
            )
        return (
            f"[{p_u}] MICRO_LED: micro-model keep decision. "
            f"EVA_pre={_fmt_eur(eva_pre)}, RC={rc_final}."
        )

    # ── Accion cambió pero sin override flag → detectado como mismatch
    return (
        f"[{p_u}] MIXED_DECISION: Micro={accion_micro} → Final={accion}. "
        f"Convergencia_Caso={convergencia}, macro_applied={macro_applied}."
    )


def apply_final_reasoning_seal(df: pd.DataFrame, postura: str) -> Tuple[pd.DataFrame, Dict]:
    """
    Tarea 2: Añade las columnas del Final Reasoning Seal y corrige
    inconsistencias RC vs Accion_final.

    Columnas nuevas:
      Reason_Code_Micro        — RC del micro-modelo
      Reason_Code_Macro        — RC del macro (o MACRO_NOT_APPLIED)
      Reason_Code_Final        — RC coherente con Accion_final (regla dura)
      Decision_Governance_Final — texto auditable del último decisor
    """
    df = df.copy()

    before_incoherent = sum(
        1 for _, row in df.iterrows()
        if not _rc_coherent(_ss(row.get("Reason_Code", ""), ""), _ss(row.get("Accion_final", ""), ""))
    )

    df["Reason_Code_Micro"]         = df.apply(lambda r: _derive_rc_micro(r, postura), axis=1)
    df["Reason_Code_Macro"]         = df.apply(lambda r: _derive_rc_macro(r, postura), axis=1)
    df["Reason_Code_Final"]         = df.apply(_derive_rc_final, axis=1)
    df["Decision_Governance_Final"] = df.apply(
        lambda r: _derive_governance_final(r, postura), axis=1
    )

    after_incoherent = sum(
        1 for _, row in df.iterrows()
        if not _rc_coherent(row["Reason_Code_Final"], _ss(row.get("Accion_final", ""), ""))
    )

    stats = {
        "rc_incoherent_before": before_incoherent,
        "rc_incoherent_after":  after_incoherent,
        "rc_fixed": before_incoherent - max(after_incoherent, 0),
    }

    logger.info(
        f"  [{postura}] Final Reasoning Seal: "
        f"before={before_incoherent} incoherentes → after={after_incoherent}"
    )
    return df, stats


# ============================================================
# TAREA 3: MACRO STEERING POR POSTURA
# ============================================================

def _tag_macro_steering(df: pd.DataFrame, idx: int, rule: str, threshold: str, var: str) -> None:
    """Escribe Macro_Steering_Applied y actualiza Convergencia_Caso."""
    df.at[idx, "Macro_Steering_Applied"] = f"{rule}|{threshold}|{var}"
    df.at[idx, "Convergencia_Caso"]      = "MACRO_HARDENING"


def apply_macro_steering_prudencial(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    """
    PRUDENCIAL: Máxima estabilidad.

    Reglas (en orden de prioridad):
    R1. Cualquier VENDER con Fire_Sale=True → flip a MANTENER.
        Umbral: fire_sale=True, postura=prudencial.
        Justificación: en postura prudencial, vender en condiciones de fire-sale
        no está permitido salvo que no exista alternativa.

    R2. Control de P&L negativo agregado:
        Si sum(pnl[Accion_final==VENDER]) < MAX_AGGREGATE_PNL_PCT_EAD * total_EAD:
        → Convertir las ventas con peor pnl/EAD a MANTENER hasta cumplir el límite.
        Umbral: sum_pnl_vender / total_EAD < -0.15.
    """
    df = df.copy()
    thr = MACRO_THRESHOLDS["prudencial"]
    meta: Dict = {"rule": "PRUDENCIAL", "flips_r1": 0, "flips_r2": 0, "total_flips": 0}

    if "Macro_Steering_Applied" not in df.columns:
        df["Macro_Steering_Applied"] = ""

    # R1: Fire-sale flip
    if thr["FLIP_FIRE_SALE_VENDER_TO_MANTENER"]:
        mask_r1 = (
            (df["Accion_final"].str.upper() == "VENDER") &
            (
                df.get("Fire_Sale", pd.Series(False, index=df.index)).fillna(False).astype(bool) |
                df.get("fire_sale", pd.Series(False, index=df.index)).fillna(False).astype(bool) |
                df.get("FireSale_Triggered", pd.Series(False, index=df.index)).fillna(False).astype(bool)
            ) &
            (df.get("Macro_Steering_Applied", pd.Series("", index=df.index)).str.strip() == "")
        )
        for idx in df[mask_r1].index:
            df.at[idx, "Accion_final"] = "MANTENER"
            _tag_macro_steering(df, idx, "R1_PRUDENCIAL_FIRE_SALE_BLOCK", "fire_sale=True", "fire_sale/FireSale_Triggered")
            meta["flips_r1"] += 1

    # R2: P&L agregado negativo
    total_ead = df["EAD"].sum() if "EAD" in df.columns else 1.0
    max_pnl_ratio = thr["MAX_AGGREGATE_PNL_PCT_EAD"]
    max_pnl_abs   = max_pnl_ratio * total_ead

    sellers = df[(df["Accion_final"].str.upper() == "VENDER") & (df["pnl"].notna())]
    current_pnl = sellers["pnl"].sum() if len(sellers) > 0 else 0.0

    if current_pnl < max_pnl_abs:
        # Ordenar por pnl/EAD más negativo primero
        sellers_sorted = sellers.copy()
        sellers_sorted["pnl_ead_ratio"] = sellers_sorted["pnl"] / sellers_sorted["EAD"].replace(0, np.nan)
        sellers_sorted = sellers_sorted.sort_values("pnl_ead_ratio")

        for idx, row in sellers_sorted.iterrows():
            if df.at[idx, "Macro_Steering_Applied"]:
                continue  # ya intervenido por R1
            pnl_i = _sf(row.get("pnl", 0.0))
            df.at[idx, "Accion_final"] = "MANTENER"
            _tag_macro_steering(
                df, idx,
                "R2_PRUDENCIAL_PNL_AGG_LIMIT",
                f"sum_pnl/total_EAD<{max_pnl_ratio:.2f}",
                f"pnl={_fmt_eur(pnl_i)},pnl_ead_ratio",
            )
            meta["flips_r2"] += 1
            current_pnl -= pnl_i
            if current_pnl >= max_pnl_abs:
                break

    meta["total_flips"] = meta["flips_r1"] + meta["flips_r2"]
    logger.info(f"  [prudencial] Macro steering: R1={meta['flips_r1']} fire-sale, R2={meta['flips_r2']} pnl-cap")
    return df, meta


def apply_macro_steering_balanceado(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    """
    BALANCEADO: Optimizar EVA + capital, penalizar concentración HHI.

    Reglas:
    R1. EVA_post < 0 y Accion_final=MANTENER (sin guardrail): flip a REESTRUCTURAR
        si ΔEVA > 0 (micro ya calculó uplift de reestructura).
        Umbral: EVA_post < 0 y ΔEVA > 0.

    R2. HHI por segmento > MAX_HHI_SEGMENT:
        En el segmento de mayor concentración, los MANTENER con mayor EAD
        se convierten en VENDER (reducción de concentración).
        Umbral: HHI_segment > 0.30.

    R3. Top 20% EAD que sean MANTENER con RORWA_pre < hurdle → REESTRUCTURAR.
        Umbral: EAD rank >= 80th pctile, RORWA_pre < 10%, Accion_final=MANTENER.
    """
    df = df.copy()
    thr = MACRO_THRESHOLDS["balanceado"]
    meta: Dict = {"rule": "BALANCEADO", "flips_r1": 0, "flips_r2": 0, "flips_r3": 0, "total_flips": 0}

    if "Macro_Steering_Applied" not in df.columns:
        df["Macro_Steering_Applied"] = ""

    # Preparar métricas base
    eva_post_col  = "EVA_post"  if "EVA_post"  in df.columns else None
    deva_col      = "ΔEVA"       if "ΔEVA"       in df.columns else "EVA_gain" if "EVA_gain" in df.columns else None
    rorwa_pre_col = "RORWA_pre" if "RORWA_pre" in df.columns else None
    ead_col       = "EAD"       if "EAD"       in df.columns else None
    segment_col   = "segment"   if "segment"   in df.columns else None

    # R1: MANTENER con EVA_post < 0 y ΔEVA > 0 → REESTRUCTURAR
    if thr["EVA_NEG_MANTENER_TO_RESTRUCT"] and eva_post_col and deva_col:
        mask_r1 = (
            (df["Accion_final"].str.upper() == "MANTENER") &
            (df[eva_post_col].fillna(0.0) < 0) &
            (df[deva_col].fillna(0.0) > 0) &
            (df["Macro_Steering_Applied"].str.strip() == "")
        )
        for idx, row in df[mask_r1].iterrows():
            eva_p = _sf(row.get(eva_post_col, 0.0))
            deva  = _sf(row.get(deva_col, 0.0))
            df.at[idx, "Accion_final"] = "REESTRUCTURAR"
            _tag_macro_steering(
                df, idx,
                "R1_BALANCEADO_EVA_NEG_TO_RESTRUCT",
                f"EVA_post<0 & ΔEVA>{0}",
                f"EVA_post={_fmt_eur(eva_p)},ΔEVA={_fmt_eur(deva)}",
            )
            meta["flips_r1"] += 1

    # R2: HHI concentración por segmento
    if segment_col and ead_col:
        if segment_col in df.columns and ead_col in df.columns:
            seg_ead = df.groupby(segment_col)[ead_col].sum()
            total_ead = seg_ead.sum()
            seg_shares = seg_ead / total_ead
            hhi = float((seg_shares ** 2).sum())

            if hhi > thr["MAX_HHI_SEGMENT"]:
                top_seg = str(seg_shares.idxmax())
                logger.info(
                    f"  [balanceado] HHI={hhi:.3f} > {thr['MAX_HHI_SEGMENT']:.2f} "
                    f"— top segment={top_seg} ({seg_shares[top_seg]*100:.1f}% EAD)"
                )
                # En top_seg, convertir los MANTENER más grandes a VENDER
                # hasta bajar HHI por debajo del umbral (max 10% de los loans del segmento)
                seg_mantener = df[
                    (df[segment_col] == top_seg) &
                    (df["Accion_final"].str.upper() == "MANTENER") &
                    (df["Macro_Steering_Applied"].str.strip() == "")
                ].sort_values(ead_col, ascending=False)

                n_flip_max = max(1, int(len(seg_mantener) * 0.10))
                for count, (idx, row) in enumerate(seg_mantener.iterrows()):
                    if count >= n_flip_max:
                        break
                    ead_i = _sf(row.get(ead_col, 0.0))
                    df.at[idx, "Accion_final"] = "VENDER"
                    _tag_macro_steering(
                        df, idx,
                        "R2_BALANCEADO_HHI_CONCENTRATION",
                        f"HHI_segment>{thr['MAX_HHI_SEGMENT']:.2f}",
                        f"segment={top_seg},EAD={_fmt_eur(ead_i)}",
                    )
                    meta["flips_r2"] += 1

    # R3: Top EAD con RORWA_pre < hurdle → REESTRUCTURAR
    if ead_col and rorwa_pre_col:
        ead_80th = df[ead_col].quantile(thr["TOP_EXPOSURE_PCT"])
        mask_r3 = (
            (df["Accion_final"].str.upper() == "MANTENER") &
            (df[ead_col].fillna(0.0) >= ead_80th) &
            (df[rorwa_pre_col].fillna(0.0) < HURDLE_RATE) &
            (df.get("ΔEVA", pd.Series(0.0, index=df.index)).fillna(0.0) > 0) &
            (df["Macro_Steering_Applied"].str.strip() == "")
        )
        for idx, row in df[mask_r3].iterrows():
            rorwa_i = _sf(row.get(rorwa_pre_col, 0.0))
            ead_i   = _sf(row.get(ead_col, 0.0))
            df.at[idx, "Accion_final"] = "REESTRUCTURAR"
            _tag_macro_steering(
                df, idx,
                "R3_BALANCEADO_TOP_EAD_RESTRUCT",
                f"EAD>={_fmt_eur(ead_80th)} & RORWA_pre<{_fmt_pct(HURDLE_RATE)}",
                f"EAD={_fmt_eur(ead_i)},RORWA_pre={_fmt_pct(rorwa_i)}",
            )
            meta["flips_r3"] += 1

    meta["total_flips"] = meta["flips_r1"] + meta["flips_r2"] + meta["flips_r3"]
    logger.info(
        f"  [balanceado] Macro steering: R1={meta['flips_r1']} eva-neg, "
        f"R2={meta['flips_r2']} hhi, R3={meta['flips_r3']} top-ead"
    )
    return df, meta


def apply_macro_steering_desinversion(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
    """
    DESINVERSION: Maximizar capital liberado sin vender a cualquier precio.

    Reglas:
    R1. VENDER con price_ratio_ead < MIN_PRICE_TO_EAD_FLOOR → MANTENER.
        Umbral: price_ratio_ead < 0.10.
        Justificación: precio insulto — destruye valor relativo al EAD.

    R2. VENDER con pnl/EAD < MAX_LOSS_PCT_EAD → MANTENER.
        Umbral: pnl/EAD < -0.50 (pérdida > 50% del EAD).
        Justificación: pérdida excesiva aunque precio supere el floor.

    R3. MANTENER con capital_liberado = 0 y ΔEVA > 0 → REESTRUCTURAR.
        Umbral: capital_release_realized = 0, ΔEVA > 0.
        Justificación: reestructura libera capital y mejora EVA simultáneamente.
    """
    df = df.copy()
    thr = MACRO_THRESHOLDS["desinversion"]
    meta: Dict = {"rule": "DESINVERSION", "flips_r1": 0, "flips_r2": 0, "flips_r3": 0, "total_flips": 0}

    if "Macro_Steering_Applied" not in df.columns:
        df["Macro_Steering_Applied"] = ""

    ead_col   = "EAD"            if "EAD"            in df.columns else None
    pnl_col   = "pnl"            if "pnl"            in df.columns else None
    px_col    = "Price_to_EAD"   if "Price_to_EAD"   in df.columns else "price_ratio_ead" if "price_ratio_ead" in df.columns else None
    deva_col  = "ΔEVA"           if "ΔEVA"           in df.columns else "EVA_gain" if "EVA_gain" in df.columns else None
    cap_col   = "capital_release_realized" if "capital_release_realized" in df.columns else "capital_liberado" if "capital_liberado" in df.columns else None

    # R1: Precio insulto
    if px_col and ead_col:
        min_floor = thr["MIN_PRICE_TO_EAD_FLOOR"]
        mask_r1 = (
            (df["Accion_final"].str.upper() == "VENDER") &
            (df[px_col].fillna(1.0) < min_floor) &
            (df["Macro_Steering_Applied"].str.strip() == "")
        )
        for idx, row in df[mask_r1].iterrows():
            px = _sf(row.get(px_col, np.nan), np.nan)
            df.at[idx, "Accion_final"] = "MANTENER"
            _tag_macro_steering(
                df, idx,
                "R1_DESINVERSION_INSULTING_PRICE",
                f"price_ratio_ead<{min_floor:.2f}",
                f"Price/EAD={px:.3f}" if not np.isnan(px) else "Price/EAD=N/A",
            )
            meta["flips_r1"] += 1

    # R2: Pérdida excesiva vs EAD
    if pnl_col and ead_col:
        max_loss = thr["MAX_LOSS_PCT_EAD"]
        df["_pnl_ead_ratio_desinv"] = (
            df[pnl_col].fillna(0.0) / df[ead_col].replace(0, np.nan)
        )
        mask_r2 = (
            (df["Accion_final"].str.upper() == "VENDER") &
            (df["_pnl_ead_ratio_desinv"].fillna(0.0) < max_loss) &
            (df["Macro_Steering_Applied"].str.strip() == "")
        )
        for idx, row in df[mask_r2].iterrows():
            rat = _sf(row.get("_pnl_ead_ratio_desinv", np.nan), np.nan)
            pnl = _sf(row.get(pnl_col, 0.0))
            df.at[idx, "Accion_final"] = "MANTENER"
            _tag_macro_steering(
                df, idx,
                "R2_DESINVERSION_EXCESS_LOSS",
                f"pnl/EAD<{max_loss:.2f}",
                f"pnl/EAD={rat:.3f},pnl={_fmt_eur(pnl)}",
            )
            meta["flips_r2"] += 1
        if "_pnl_ead_ratio_desinv" in df.columns:
            df.drop(columns=["_pnl_ead_ratio_desinv"], inplace=True)

    # R3: MANTENER con cap_release=0 y EVA mejora → REESTRUCTURAR
    if cap_col and deva_col:
        mask_r3 = (
            (df["Accion_final"].str.upper() == "MANTENER") &
            (df[cap_col].fillna(0.0) <= 0) &
            (df[deva_col].fillna(0.0) > 0) &
            (df["Macro_Steering_Applied"].str.strip() == "")
        )
        for idx, row in df[mask_r3].iterrows():
            deva_i = _sf(row.get(deva_col, 0.0))
            cap_i  = _sf(row.get(cap_col, 0.0))
            df.at[idx, "Accion_final"] = "REESTRUCTURAR"
            _tag_macro_steering(
                df, idx,
                "R3_DESINVERSION_RESTRUCT_FOR_CAP",
                f"cap_release<=0 & ΔEVA>0",
                f"cap_release={_fmt_eur(cap_i)},ΔEVA={_fmt_eur(deva_i)}",
            )
            meta["flips_r3"] += 1

    meta["total_flips"] = meta["flips_r1"] + meta["flips_r2"] + meta["flips_r3"]
    logger.info(
        f"  [desinversion] Macro steering: R1={meta['flips_r1']} floor-px, "
        f"R2={meta['flips_r2']} excess-loss, R3={meta['flips_r3']} restruct-cap"
    )
    return df, meta


MACRO_STEERING_FN = {
    "prudencial":   apply_macro_steering_prudencial,
    "balanceado":   apply_macro_steering_balanceado,
    "desinversion": apply_macro_steering_desinversion,
}


def apply_macro_steering(df: pd.DataFrame, postura: str, kpis: Dict) -> Tuple[pd.DataFrame, Dict, Dict]:
    """
    Tarea 3: Aplica macro steering por postura y registra en el JSON de KPIs
    el campo macro_inactive_warning si la tasa de intervención es < 20%.
    """
    fn = MACRO_STEERING_FN.get(postura)
    if fn is None:
        logger.warning(f"  No hay función de macro steering para postura={postura}")
        return df, {}, kpis

    df_steered, meta = fn(df)

    total   = len(df_steered)
    applied = int(meta.get("total_flips", 0))
    rate    = applied / total if total > 0 else 0.0

    # Registrar en KPIs JSON
    kpis_updated = deepcopy(kpis)
    kpis_updated["hardening"] = kpis_updated.get("hardening", {})
    kpis_updated["hardening"]["macro_steering"] = {
        "applied_loans": applied,
        "total_loans":   total,
        "applied_rate":  round(rate, 4),
        "rules": meta,
    }

    MACRO_RATE_MIN = 0.20
    if rate < MACRO_RATE_MIN:
        warning_txt = (
            f"macro_applied_rate={rate:.1%} < {MACRO_RATE_MIN:.0%} threshold. "
            f"Portfolio steering is micro-dominated. "
            f"Consider reviewing macro model coverage or hardening thresholds."
        )
        kpis_updated["hardening"]["macro_inactive_warning"] = True
        kpis_updated["hardening"]["macro_inactive_explanation"] = warning_txt
        logger.warning(f"  [{postura}] MACRO_INACTIVE_WARNING: {warning_txt}")
    else:
        kpis_updated["hardening"]["macro_inactive_warning"] = False

    return df_steered, meta, kpis_updated


# ============================================================
# TAREA 6: VALIDACIONES AUTOMÁTICAS
# ============================================================

class ValidationError(Exception):
    pass


def run_validations(
    dfs: Dict[str, pd.DataFrame],
    kpis_all: Dict[str, Dict],
    postures_order: List[str] = None,
) -> List[str]:
    """
    Suite de validaciones banco-ready.
    Devuelve lista de mensajes (OK o FAIL).
    No lanza excepción; retorna descriptivo.
    """
    if postures_order is None:
        postures_order = ["prudencial", "balanceado", "desinversion"]

    results: List[str] = []

    def ok(msg: str):
        results.append(f"[PASS] {msg}")

    def fail(msg: str):
        results.append(f"[FAIL] {msg}")

    # ─── (i) Coherencia Accion_final vs Reason_Code_Final
    for postura, df in dfs.items():
        rc_col = "Reason_Code_Final" if "Reason_Code_Final" in df.columns else "Reason_Code"
        if rc_col not in df.columns:
            fail(f"{postura}: columna {rc_col} no encontrada")
            continue
        n_total = len(df)
        n_incoherent = sum(
            1 for _, row in df.iterrows()
            if not _rc_coherent(_ss(row.get(rc_col), ""), _ss(row.get("Accion_final"), ""))
        )
        if n_incoherent == 0:
            ok(f"{postura}: Accion_final vs {rc_col} coherente al 100% ({n_total} loans)")
        else:
            fail(f"{postura}: {n_incoherent}/{n_total} Reason_Code incoherentes con Accion_final")

    # ─── (ii) No-NaN en métricas post para MANTENER/REESTRUCTURAR
    for postura, df in dfs.items():
        for col in ["EVA_post", "RWA_post", "RORWA_post"]:
            if col not in df.columns:
                fail(f"{postura}: columna {col} no encontrada")
                continue
            non_seller = df[df["Accion_final"].str.upper() != "VENDER"]
            n_nan = non_seller[col].isna().sum()
            if n_nan == 0:
                ok(f"{postura}: {col} sin NaN para MANTENER/REESTRUCTURAR")
            else:
                fail(f"{postura}: {col} tiene {n_nan} NaN en MANTENER/REESTRUCTURAR")

    # ─── (iii) Monotonía entre posturas — VENTAS
    available = [p for p in postures_order if p in dfs]
    if len(available) >= 2:
        ventas = {
            p: int((dfs[p]["Accion_final"].str.upper() == "VENDER").sum())
            for p in available
        }
        # desinversion >= balanceado >= prudencial
        pairs = [
            ("desinversion", "balanceado"),
            ("balanceado",   "prudencial"),
        ]
        for a, b in pairs:
            if a in ventas and b in ventas:
                if ventas[a] >= ventas[b]:
                    ok(f"Monotonía ventas: {a}({ventas[a]}) >= {b}({ventas[b]})")
                else:
                    fail(f"Monotonía ventas ROTA: {a}({ventas[a]}) < {b}({ventas[b]})")

    # ─── (iv) KPIs monotonía capital liberado
    cap_release = {}
    for postura in available:
        kpis = kpis_all.get(postura, {})
        fs = kpis.get("final_state", {})
        cap_release[postura] = float(fs.get("total_capital_release", 0.0))

    for a, b in [("desinversion", "balanceado"), ("balanceado", "prudencial")]:
        if a in cap_release and b in cap_release:
            if cap_release[a] >= cap_release[b]:
                ok(f"Monotonía capital_release: {a}({cap_release[a]:,.0f}) >= {b}({cap_release[b]:,.0f})")
            else:
                fail(f"Monotonía capital_release ROTA: {a} < {b}")

    # ─── (v) Sin MANTENER con Accion_Final=MANTENER y Reason_Code_Final en RC_VENDER_SET
    for postura, df in dfs.items():
        rc_col = "Reason_Code_Final" if "Reason_Code_Final" in df.columns else "Reason_Code"
        if rc_col not in df.columns:
            continue
        bad = df[
            (df["Accion_final"].str.upper() == "MANTENER") &
            (df[rc_col].isin(RC_VENDER_SET))
        ]
        if len(bad) == 0:
            ok(f"{postura}: ningún MANTENER con RC de VENDER")
        else:
            fail(f"{postura}: {len(bad)} MANTENER tienen RC de VENDER — revisar")

    # ─── (vi) Macro steering registrada en KPIs
    for postura in available:
        kpis = kpis_all.get(postura, {})
        if "hardening" in kpis and "macro_steering" in kpis["hardening"]:
            rate = kpis["hardening"]["macro_steering"].get("applied_rate", 0.0)
            ok(f"{postura}: macro_steering registrado en KPIs (applied_rate={rate:.1%})")
        else:
            fail(f"{postura}: macro_steering NO registrado en KPIs hardening")

    return results


# ============================================================
# EXPORT
# ============================================================

def export_hardened_excel(df: pd.DataFrame, out_dir: str, postura: str) -> str:
    """Exporta el Excel hardened con formato mínimo."""
    out_path = os.path.join(out_dir, f"decisiones_finales_{postura}_hardened.xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Decisiones_Hardened"

        # Columnas del reasoning seal primero
        priority_cols = [
            "loan_id", "Accion_final", "Reason_Code_Final", "Decision_Governance_Final",
            "Reason_Code_Micro", "Reason_Code_Macro",
            "Macro_Steering_Applied", "Convergencia_Caso",
            "EVA_pre", "EVA_post", "RORWA_pre", "RORWA_post", "RWA_pre", "RWA_post",
        ]
        existing_priority = [c for c in priority_cols if c in df.columns]
        other_cols = [c for c in df.columns if c not in existing_priority]
        ordered_df = df[existing_priority + other_cols]

        for row in dataframe_to_rows(ordered_df, index=False, header=True):
            ws.append(row)

        # Estilos encabezado
        COLORS = {
            "header":   "1F4E79",
            "seal":     "2E75B6",
            "macro":    "375623",
            "metrics":  "843C0C",
        }
        SEAL_COLS  = {"Reason_Code_Final", "Decision_Governance_Final", "Reason_Code_Micro", "Reason_Code_Macro"}
        MACRO_COLS = {"Macro_Steering_Applied", "Convergencia_Caso"}
        METRIC_COLS = {"EVA_post", "RORWA_post", "RWA_post"}

        for ci, cell in enumerate(ws[1], start=1):
            col_name = cell.value or ""
            if col_name in SEAL_COLS:
                color = COLORS["seal"]
            elif col_name in MACRO_COLS:
                color = COLORS["macro"]
            elif col_name in METRIC_COLS:
                color = COLORS["metrics"]
            else:
                color = COLORS["header"]

            cell.fill      = PatternFill("solid", fgColor=color)
            cell.font      = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Accion_final coloreada
        accion_idx = None
        for ci, cell in enumerate(ws[1], start=1):
            if cell.value == "Accion_final":
                accion_idx = ci
                break

        ACTION_COLORS = {"MANTENER": "92D050", "REESTRUCTURAR": "FFD966", "VENDER": "FF6B6B"}
        if accion_idx:
            for row_cells in ws.iter_rows(min_row=2, min_col=accion_idx, max_col=accion_idx):
                v = str(row_cells[0].value or "").upper()
                c = ACTION_COLORS.get(v, "FFFFFF")
                row_cells[0].fill = PatternFill("solid", fgColor=c)
                row_cells[0].font = Font(bold=True, size=9)

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions

        # Ancho de columnas
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

        wb.save(out_path)
        logger.info(f"  [{postura}] Excel hardened guardado: {out_path}")
        return out_path

    except Exception as e:
        logger.warning(f"  [{postura}] OpenPyXL falló ({e}), usando pandas fallback")
        df.to_excel(out_path, index=False)
        return out_path


# ============================================================
# GENERAR BANK_READY_HARDENING.md
# ============================================================

def _5_loans_changed(df_before: pd.DataFrame, df_after: pd.DataFrame, postura: str) -> str:
    """Selecciona 5 loans donde cambió la governance y produce tabla markdown."""
    changed = df_after[df_after.get("Macro_Steering_Applied", pd.Series("", index=df_after.index)).str.strip() != ""]
    if len(changed) == 0:
        # Buscar cambios en Reason_Code_Final vs Reason_Code original
        if "Reason_Code_Final" in df_after.columns and "Reason_Code" in df_before.columns:
            merged = df_after.copy()
            merged["_rc_old"] = df_before["Reason_Code"].values if len(df_before) == len(df_after) else ""
            changed = merged[merged["_rc_old"] != merged["Reason_Code_Final"]]

    sample = changed.head(5)
    if len(sample) == 0:
        return "_Sin cambios de governance en esta postura._\n"

    lines = ["| loan_id | Accion_final | RC_Before | RC_Final | Decision_Governance_Final |",
             "|---------|-------------|-----------|----------|--------------------------|"]
    for _, r in sample.iterrows():
        loan  = _ss(r.get("loan_id", ""), "N/A")
        accion = _ss(r.get("Accion_final", ""), "")
        rc_before = _ss(r.get("Reason_Code", r.get("_rc_old", "")), "")
        rc_final  = _ss(r.get("Reason_Code_Final", ""), "")
        gov_final = _ss(r.get("Decision_Governance_Final", ""), "")[:120]
        lines.append(f"| {loan} | {accion} | {rc_before} | {rc_final} | {gov_final} |")
    return "\n".join(lines) + "\n"


def generate_hardening_report(
    tag: str,
    postures: List[str],
    stats_per_posture: Dict,
    validation_results: List[str],
    out_dir: str,
) -> str:
    """Genera BANK_READY_HARDENING_{tag}.md en reports/."""
    ts = _TIMESTAMP
    n_pass = sum(1 for r in validation_results if r.startswith("[PASS]"))
    n_fail = sum(1 for r in validation_results if r.startswith("[FAIL]"))
    status = "✅ PASS" if n_fail == 0 else f"⚠️ {n_fail} FAIL / {n_pass} PASS"

    lines: List[str] = []
    lines.append(f"# BANK_READY_HARDENING — Tag: `{tag}`")
    lines.append(f"")
    lines.append(f"**Generado:** {ts}  |  **Versión hardening:** {_VERSION}  |  **Status:** {status}")
    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## Resumen de Cambios")
    lines.append(f"")
    lines.append(f"| Postura | RC incoherentes antes | RC incoherentes después | Flips Macro | Macro Rate |")
    lines.append(f"|---------|----------------------|------------------------|-------------|------------|")

    for p in postures:
        st = stats_per_posture.get(p, {})
        rc_before = st.get("rc_seal", {}).get("rc_incoherent_before", "N/A")
        rc_after  = st.get("rc_seal", {}).get("rc_incoherent_after", "N/A")
        flips     = st.get("macro_meta", {}).get("total_flips", 0)
        kpis      = st.get("kpis_updated", {})
        rate      = kpis.get("hardening", {}).get("macro_steering", {}).get("applied_rate", 0.0)
        lines.append(f"| {p} | {rc_before} | {rc_after} | {flips} | {rate:.1%} |")

    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## Detalles por Postura")

    for p in postures:
        st = stats_per_posture.get(p, {})
        df_after  = st.get("df_after",  pd.DataFrame())
        df_before = st.get("df_before", pd.DataFrame())
        kpis      = st.get("kpis_updated", {})
        macro_meta = st.get("macro_meta", {})
        hd = kpis.get("hardening", {})

        lines.append(f"")
        lines.append(f"### {p.upper()}")
        lines.append(f"")

        # Distribución acciones
        if not df_after.empty and "Accion_final" in df_after.columns:
            ac = df_after["Accion_final"].value_counts().to_dict()
            lines.append(f"**Distribución acciones hardened:**")
            lines.append(f"- VENDER: {ac.get('VENDER', 0)} | REESTRUCTURAR: {ac.get('REESTRUCTURAR', 0)} | MANTENER: {ac.get('MANTENER', 0)}")
            lines.append(f"")

        # KPIs financieros
        fs = kpis.get("final_state", {})
        if fs:
            lines.append(f"**KPIs portfolio finales:**")
            lines.append(
                f"- EVA_final: {fs.get('total_eva', 0):+,.0f} € | "
                f"RWA_final: {fs.get('total_rwa', 0):,.0f} € | "
                f"Capital liberado: {fs.get('total_capital_release', 0):,.0f} €"
            )
            lines.append(f"")

        # Macro steering
        ms = hd.get("macro_steering", {})
        if ms:
            lines.append(f"**Macro Steering:**")
            lines.append(f"- Loans intervenidos: {ms.get('applied_loans', 0)}/{ms.get('total_loans', 0)} ({ms.get('applied_rate', 0):.1%})")
            rules = ms.get("rules", {})
            for k, v in rules.items():
                if k not in ("rule", "total_flips") and isinstance(v, int) and v > 0:
                    lines.append(f"  - {k}: {v} flips")
            inactive = hd.get("macro_inactive_warning", False)
            if inactive:
                lines.append(f"  - ⚠️ MACRO_INACTIVE_WARNING: {hd.get('macro_inactive_explanation', '')}")
            lines.append(f"")

        # 5 loans con governance cambiada
        lines.append(f"**Top 5 préstamos con governance modificada:**")
        lines.append(f"")
        lines.append(_5_loans_changed(df_before, df_after, p))

    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## Validaciones Automáticas")
    lines.append(f"")
    lines.append(f"```")
    for r in validation_results:
        lines.append(r)
    lines.append(f"```")
    lines.append(f"")
    lines.append(f"**Summary:** {n_pass} PASS | {n_fail} FAIL")

    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## Reglas Macro Aplicadas (con umbrales y variables)")
    lines.append(f"")
    lines.append(f"### PRUDENCIAL")
    lines.append(f"| Regla | Umbral | Variable | Acción |")
    lines.append(f"|-------|--------|----------|--------|")
    lines.append(f"| R1_PRUDENCIAL_FIRE_SALE_BLOCK | fire_sale=True | fire_sale / FireSale_Triggered | VENDER → MANTENER |")
    lines.append(f"| R2_PRUDENCIAL_PNL_AGG_LIMIT | sum(pnl_ventas)/total_EAD < -15% | pnl, EAD | VENDER → MANTENER (peor P&L) |")
    lines.append(f"")
    lines.append(f"### BALANCEADO")
    lines.append(f"| Regla | Umbral | Variable | Acción |")
    lines.append(f"|-------|--------|----------|--------|")
    lines.append(f"| R1_BALANCEADO_EVA_NEG_TO_RESTRUCT | EVA_post < 0 AND ΔEVA > 0 | EVA_post, ΔEVA | MANTENER → REESTRUCTURAR |")
    lines.append(f"| R2_BALANCEADO_HHI_CONCENTRATION | HHI_segment > 0.30 | segment, EAD | MANTENER top-EAD → VENDER |")
    lines.append(f"| R3_BALANCEADO_TOP_EAD_RESTRUCT | EAD >= p80 AND RORWA_pre < 10% AND ΔEVA > 0 | EAD, RORWA_pre | MANTENER → REESTRUCTURAR |")
    lines.append(f"")
    lines.append(f"### DESINVERSION")
    lines.append(f"| Regla | Umbral | Variable | Acción |")
    lines.append(f"|-------|--------|----------|--------|")
    lines.append(f"| R1_DESINVERSION_INSULTING_PRICE | price/EAD < 0.10 | Price_to_EAD | VENDER → MANTENER |")
    lines.append(f"| R2_DESINVERSION_EXCESS_LOSS | pnl/EAD < -0.50 | pnl, EAD | VENDER → MANTENER |")
    lines.append(f"| R3_DESINVERSION_RESTRUCT_FOR_CAP | cap_release=0 AND ΔEVA > 0 | capital_release_realized, ΔEVA | MANTENER → REESTRUCTURAR |")
    lines.append(f"")
    lines.append(f"---")
    lines.append(f"")
    lines.append(f"## Notas Técnicas")
    lines.append(f"")
    lines.append(f"- `Reason_Code_Final` siempre coherente con `Accion_final` (regla dura, 100%).")
    lines.append(f"- `EVA_post` para MANTENER = EVA_pre (sin drift). Para REESTRUCTURAR = EVA_pre + ΔEVA.")
    lines.append(f"- `RORWA_post` para VENDER = NaN (diseño: activo fuera de libro).")
    lines.append(f"- Guardrail `PNL_TOO_NEGATIVE_PRUDENCIAL` renombrado a `PNL_TOO_NEGATIVE_EAD40 [postura=X]`.")
    lines.append(f"- `Macro_Steering_Applied` contiene `regla|umbral|variable` para trazabilidad completa.")
    lines.append(f"- Ficheros generados en `reports/coordinated_inference_{{tag}}_*_{{postura}}/`.")
    lines.append(f"")

    out_path = os.path.join(out_dir, f"BANK_READY_HARDENING_{tag}.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    logger.info(f"  Reporte BANK_READY_HARDENING generado: {out_path}")
    return out_path


# ============================================================
# PIPELINE PRINCIPAL
# ============================================================

def run_hardening(tag: str, postures: Optional[List[str]] = None, dry_run: bool = False) -> Dict:
    """
    Pipeline completo de hardening para un tag de inferencia.

    Orden de ejecución:
      1. Cargar artefactos
      2. Completar métricas post (Tarea 4)
      3. Renombrar guardrails (Tarea 5)
      4. Macro steering por postura (Tarea 3)
      5. Final Reasoning Seal (Tarea 2) — después del steering
      6. Validaciones (Tarea 6)
      7. Export Excel hardened + KPIs JSON + Reporte MD
    """
    if postures is None:
        postures = ["prudencial", "balanceado", "desinversion"]

    logger.info(f"[HARDENING v{_VERSION}] Tag={tag}, Posturas={postures}")

    stats_per_posture: Dict = {}
    dfs_hardened: Dict[str, pd.DataFrame] = {}
    kpis_all: Dict[str, Dict] = {}

    for postura in postures:
        logger.info(f"\n{'='*60}")
        logger.info(f"  Procesando: {postura.upper()}")
        logger.info(f"{'='*60}")

        try:
            df_orig, kpis, ov = load_posture_artifacts(tag, postura)
        except FileNotFoundError as e:
            logger.error(f"  [{postura}] SKIP: {e}")
            continue

        # Snapshot original
        df_before = df_orig.copy()

        # ── Paso 1: Renombrar guardrails (antes de cualquier transformación)
        df = rename_guardrail_reasons(df_orig, postura)

        # ── Paso 2: Macro steering (puede cambiar Accion_final: VENDER→MANTENER etc.)
        df, macro_meta, kpis_upd = apply_macro_steering(df, postura, kpis)

        # ── Paso 3: Completar métricas post DESPUÉS del steering
        #    (así los loans flipeados de VENDER→MANTENER reciben EVA/RWA/RORWA_post)
        df = complete_post_metrics(df, postura)

        # ── Paso 4: Final Reasoning Seal (DESPUÉS del steering para capturar Macro_Steering_Applied)
        df, rc_seal_stats = apply_final_reasoning_seal(df, postura)

        # ── Guardar
        stats_per_posture[postura] = {
            "df_before":   df_before,
            "df_after":    df,
            "rc_seal":     rc_seal_stats,
            "macro_meta":  macro_meta,
            "kpis_updated": kpis_upd,
        }
        dfs_hardened[postura] = df
        kpis_all[postura] = kpis_upd

        if not dry_run:
            # Localizar carpeta de salida
            out_dir = find_inference_dir(tag, postura)
            if out_dir:
                # Excel hardened
                export_hardened_excel(df, out_dir, postura)

                # KPIs JSON actualizado
                kpis_hardened_path = os.path.join(out_dir, f"portfolio_kpis_{postura}_hardened.json")
                with open(kpis_hardened_path, "w", encoding="utf-8") as f:
                    json.dump(kpis_upd, f, indent=2, ensure_ascii=False)
                logger.info(f"  [{postura}] KPIs hardened guardados: {kpis_hardened_path}")

    # ── Validaciones
    logger.info(f"\n{'='*60}")
    logger.info(f"  VALIDACIONES")
    logger.info(f"{'='*60}")
    validation_results = run_validations(dfs_hardened, kpis_all, postures)
    for r in validation_results:
        logger.info(f"  {r}")

    # ── Reporte MD
    if not dry_run:
        report_path = generate_hardening_report(
            tag=tag,
            postures=[p for p in postures if p in stats_per_posture],
            stats_per_posture=stats_per_posture,
            validation_results=validation_results,
            out_dir=REPORTS_DIR,
        )

    n_pass = sum(1 for r in validation_results if r.startswith("[PASS]"))
    n_fail = sum(1 for r in validation_results if r.startswith("[FAIL]"))

    summary = {
        "tag": tag,
        "postures_processed": list(stats_per_posture.keys()),
        "validation_pass": n_pass,
        "validation_fail": n_fail,
        "overall_status": "PASS" if n_fail == 0 else "WARN",
    }

    logger.info(
        f"\n[HARDENING DONE] tag={tag} | "
        f"PASS={n_pass} FAIL={n_fail} | status={summary['overall_status']}"
    )
    return summary


# ============================================================
# CLI
# ============================================================

def _parse_args():
    parser = argparse.ArgumentParser(
        description="Bank-Ready Hardening — NPL Portfolio Optimizer"
    )
    parser.add_argument("--tag",      required=True, help="Tag de inferencia (ej. infer_ci0222)")
    parser.add_argument(
        "--postures", nargs="+",
        default=["prudencial", "balanceado", "desinversion"],
        help="Posturas a procesar",
    )
    parser.add_argument("--dry-run", action="store_true", help="No escribir archivos")
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    result = run_hardening(
        tag=args.tag,
        postures=args.postures,
        dry_run=args.dry_run,
    )
    sys.exit(0 if result.get("overall_status") == "PASS" else 1)
