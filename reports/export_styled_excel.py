# -*- coding: utf-8 -*-
# ============================================
# export_styled_excel.py — v7.5 (CIB Executive Notes · COORD-aware)
# ============================================
"""
Excel corporativo con:
• Comentarios hover estilo ejecutivo CIB (muy breves, sin scroll)
• Sin columna Effective_RONA_pct
• Hojas: decisiones, resumen_segmento, resumen_global, parámetros

CORRECCIONES v7.5:
- Compatible con salida del COORDINATOR:
    loan_id, Accion_micro, Accion_macro, Macro_Selected, Accion_final,
    Convergencia_Caso, Explanation_micro, Explanation_macro, Explanation_final, Explanation
- "Accion" se deriva (si falta) desde Accion_final
- Agrupaciones usan Accion_final si existe (evita mezclar con legacy Accion)
- Comentario hover usa Accion_final + governance/razón si están
- Borde y header: no rompe por header en filas no-1 (resumen_segmento)
"""

from __future__ import annotations

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import config as cfg

# ============================================
# 🎨 ESTILOS
# ============================================
THIN = Side(border_style="thin", color="999999")
HDR_FILL = PatternFill("solid", fgColor="DCE6F1")
HDR_FONT = Font(bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

COLOR_KEEP = PatternFill("solid", fgColor="C8E6C9")
COLOR_REST = PatternFill("solid", fgColor="FFE082")
COLOR_SELL = PatternFill("solid", fgColor="FFCDD2")

NUM_EURO = '#,##0.00" €"'
NUM_PCT  = "0.0000%"
NUM_INT  = "0"

ACCION_COLOR = {
    "MANTENER": COLOR_KEEP,
    "REESTRUCTURAR": COLOR_REST,
    "VENDER": COLOR_SELL,
}

# ============================================
# 📌 UTILIDADES
# ============================================
def _autofit(ws):
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except Exception:
            continue
        max_len = 0
        for c in col:
            try:
                max_len = max(max_len, len(str(c.value)) if c.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def _style_header_row(ws, header_row: int = 1):
    for c in ws[header_row]:
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _borders(ws, start_row: int = 2):
    for row in ws.iter_rows(min_row=start_row):
        for c in row:
            c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _format_numeric(ws):
    hdr = {ws.cell(1, j).value: j for j in range(1, ws.max_column + 1)}

    def fmt(cols, fmt_code, align=RIGHT):
        for name in cols:
            if name not in hdr:
                continue
            j = hdr[name]
            for i in range(2, ws.max_row + 1):
                c = ws.cell(i, j)
                c.number_format = fmt_code
                c.alignment = align

    euros = ["EVA_pre","EVA_post","ΔEVA","EAD","RWA_pre","RWA_post","capital_liberado","precio_optimo","pnl"]
    pct   = ["PD","LGD","RORWA_pre","RORWA_post","tasa_nueva","quita"]
    ints  = ["DPD","plazo_optimo","rating_num","segmento_id"]

    fmt(euros, NUM_EURO)
    fmt(pct, NUM_PCT)
    fmt(ints, NUM_INT, CENTER)


def _safe_float(x, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        return default


def _safe_str(x, default: str = "N/D") -> str:
    try:
        if x is None:
            return default
        s = str(x).strip()
        return s if s else default
    except Exception:
        return default


def _get_action_for_row(r: dict) -> str:
    # Prefer coordinator fields
    a = r.get("Accion_final", None)
    if a is None or str(a).strip() == "":
        a = r.get("Accion", None)
    if a is None or str(a).strip() == "":
        a = r.get("Accion_micro", None)
    return str(a).upper().strip()


def _get_segment_col(df: pd.DataFrame) -> str:
    # dataset puede usar "segment", "segmento_banco", etc.
    for c in ["segment", "segmento_banco", "segmento", "seg", "segment_id"]:
        if c in df.columns:
            return c
    return "segment"  # se creará si no existe


# ============================================
# 🧠 NOTA EJECUTIVA (muy breve, sin scroll)
# ============================================
def _build_comment_short(r: dict) -> str:
    act = _get_action_for_row(r)

    eva0 = _safe_float(r.get("EVA_pre", 0))
    eva1 = _safe_float(r.get("EVA_post", 0))
    deva = _safe_float(r.get("ΔEVA", r.get("delta_eva", 0)))
    rwa0 = _safe_float(r.get("RWA_pre", 0))

    pd_  = _safe_float(r.get("PD", 0)) * 100
    lgd  = _safe_float(r.get("LGD", 0)) * 100
    rorwa = _safe_float(r.get("RORWA_pre", 0)) * 100

    try:
        hurdle = float(cfg.CONFIG.regulacion.hurdle_rate) * 100
    except Exception:
        hurdle = 0.0

    pti   = _safe_str(r.get("PTI_post", r.get("PTI", "N/D")))
    dscr  = _safe_str(r.get("DSCR_post", r.get("DSCR", "N/D")))
    price = _safe_float(r.get("precio_optimo", r.get("price", 0)))
    pnl   = _safe_float(r.get("pnl", r.get("PnL", 0)))
    cap   = _safe_float(r.get("capital_liberado", r.get("capital_release", 0)))
    ead   = _safe_float(r.get("EAD", 0))
    rwa_ead = (rwa0 / ead * 100) if ead > 0 else 0.0

    # Coordinator info (si existe)
    micro = _safe_str(r.get("Accion_micro", "N/D"), "N/D")
    macro = _safe_str(r.get("Accion_macro", "N/D"), "N/D")
    macro_sel = r.get("Macro_Selected", None)
    conv = _safe_str(r.get("Convergencia_Caso", "N/D"), "N/D")
    gov = _safe_str(r.get("Decision_Governance", "N/D"), "N/D")
    code = _safe_str(r.get("Reason_Code", "N/D"), "N/D")

    # Cabecera compacta
    header = (
        f"Final: {act}\n"
        f"Micro/Macro: {micro}/{macro} | Macro_Selected: {macro_sel} | {conv}\n"
        f"Reason: {code} | Gov: {gov}\n"
    )

    # ----------- Plantilla compacta por decisión -----------
    if act == "VENDER":
        body = (
            f"EVA_pre {eva0:,.0f}€, RORWA_pre {rorwa:.2f}% < hurdle {hurdle:.2f}%.\n"
            f"Riesgo: PD {pd_:.1f}%, LGD {lgd:.1f}%, RWA/EAD {rwa_ead:.1f}%.\n"
            f"Viabilidad: ΔEVA {deva:,.0f}€, PTI/DSCR {pti}/{dscr}.\n"
            f"Venta: precio {price:,.0f}€, P&L {pnl:,.0f}€, libera {cap:,.0f}€."
        )
        return header + body

    if act == "REESTRUCTURAR":
        body = (
            f"EVA_pre {eva0:,.0f}€, RORWA_pre {rorwa:.2f}% < hurdle {hurdle:.2f}%.\n"
            f"Uplift: EVA_post {eva1:,.0f}€, ΔEVA {deva:,.0f}€.\n"
            f"Asequibilidad: PTI {pti}, DSCR {dscr}.\n"
            f"Riesgo: PD {pd_:.1f}%, LGD {lgd:.1f}%."
        )
        return header + body

    if act == "MANTENER":
        body = (
            f"EVA_pre {eva0:,.0f}€, RORWA_pre {rorwa:.2f}% ≥ hurdle {hurdle:.2f}%.\n"
            f"Riesgo: PD {pd_:.1f}%, LGD {lgd:.1f}%, RWA/EAD {rwa_ead:.1f}%.\n"
            f"Alternativas: ΔEVA {deva:,.0f}€."
        )
        return header + body

    return header + "Decisión no disponible."


# ============================================
# 📄 HOJA: decisiones
# ============================================
def _write_decisiones(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("decisiones")

    df = df.copy()

    # ---- Quitar Effective_RONA_pct ----
    if "Effective_RONA_pct" in df.columns:
        df = df.drop(columns=["Effective_RONA_pct"])

    # ---- Coherencia Acción (COORD-aware) ----
    if "Accion_final" in df.columns:
        # "Accion" legacy debe reflejar final si existe
        df["Accion"] = df["Accion_final"]
    elif "Accion" not in df.columns:
        df["Accion"] = df.get("Accion_micro", "❓")

    # ---- Quitar Explain_Steps si existe (ruido) ----
    if "Explain_Steps" in df.columns:
        df = df.drop(columns=["Explain_Steps"])

    # ---- Asegurar Explanation (legacy) ----
    if "Explanation" not in df.columns:
        # si viene Explanation_final del coordinator, úsalo
        if "Explanation_final" in df.columns:
            df["Explanation"] = df["Explanation_final"]
        else:
            df["Explanation"] = "❓"

    # ---- Orden preferente (incluye coordinator columns) ----
    order = [
        "loan_id",
        "Accion_micro", "Accion_macro", "Macro_Selected", "Accion_final", "Convergencia_Caso",
        "Reason_Code", "Decision_Governance",
        "Rationale_CIB", "Macro_Evidence",
        "Explanation_micro", "Explanation_macro", "Explanation_final",
        "Accion", "Explanation",
        "EVA_pre","EVA_post","ΔEVA",
        "RORWA_pre","RORWA_post",
        "RWA_pre","RWA_post",
        "capital_liberado","precio_optimo","pnl",
        "plazo_optimo","tasa_nueva","quita",
        "segment","segmento_banco","rating","rating_num","EAD","PD","LGD","rate","DPD","secured",
    ]

    cols = [c for c in order if c in df.columns] + [c for c in df.columns if c not in order]
    df2 = df[cols].copy()

    for row in dataframe_to_rows(df2, index=False, header=True):
        ws.append(row)

    _style_header_row(ws, header_row=1)
    _borders(ws, start_row=2)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _format_numeric(ws)

    # ---- Colorear por acción final ----
    if "Accion" in cols:
        act_col = cols.index("Accion") + 1
        for i in range(2, ws.max_row + 1):
            act = ws.cell(i, act_col).value
            fill = ACCION_COLOR.get(str(act).upper())
            if fill:
                for j in range(1, ws.max_column + 1):
                    ws.cell(i, j).fill = fill

    # ---- Comentarios hover ejecutivos ----
    # Preferimos pegar el comment en Explanation_final, si existe; si no, en Explanation.
    exp_target = "Explanation_final" if "Explanation_final" in cols else "Explanation"
    exp_col = cols.index(exp_target) + 1 if exp_target in cols else (cols.index("Explanation") + 1)

    # mapa rápido loan_id -> dict (evita búsqueda O(N^2))
    if "loan_id" in df2.columns:
        by_id = {str(k): v for k, v in df2.set_index("loan_id").to_dict(orient="index").items()}
    else:
        by_id = {}

    for i in range(2, ws.max_row + 1):
        loan_id = ws.cell(i, 1).value
        if loan_id is None:
            continue
        rdict = by_id.get(str(loan_id))
        if not rdict:
            continue

        txt = _build_comment_short(rdict)
        cmt = Comment(txt, "Analyst CIB")
        cmt.width = 350
        cmt.height = 160  # pequeño para no requerir scroll
        ws.cell(i, exp_col).comment = cmt

    _autofit(ws)


# ============================================
# 📄 HOJA: resumen_segmento
# ============================================
def _write_resumen_segmento(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("resumen_segmento")
    df = df.copy()

    seg_col = _get_segment_col(df)
    if seg_col not in df.columns:
        df[seg_col] = "unknown"

    # Acción a usar en agregación (final preferente)
    action_col = "Accion_final" if "Accion_final" in df.columns else ("Accion" if "Accion" in df.columns else None)
    if action_col is None:
        df["Accion"] = "❓"
        action_col = "Accion"

    # KPIs (robusto si faltan columnas)
    for c in ["EVA_post", "ΔEVA", "capital_liberado", "loan_id"]:
        if c not in df.columns:
            df[c] = 0 if c != "loan_id" else ""
    cap_col = "capital_release_realized" if "capital_release_realized" in df.columns else "capital_liberado"
    agg = df.groupby([seg_col, action_col]).agg(
        EVA=("EVA_post", "sum"),
        ΔEVA=("ΔEVA", "sum"),
        Capital=(cap_col, "sum"),
        Count=("loan_id", "count"),
    ).reset_index()

    # Escribimos tabla directamente con header en fila 1 (para no romper header style)
    for row in dataframe_to_rows(agg, index=False, header=True):
        ws.append(row)

    _style_header_row(ws, header_row=1)
    _borders(ws, start_row=2)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autofit(ws)


# ============================================
# 📄 HOJA: resumen_global
# ============================================
def _write_resumen_global(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("resumen_global")
    df = df.copy()

    # Robustez columnas
    for c in ["EVA_post", "ΔEVA", "capital_liberado"]:
        if c not in df.columns:
            df[c] = 0.0

    action_col = "Accion_final" if "Accion_final" in df.columns else ("Accion" if "Accion" in df.columns else None)
    if action_col is None:
        df["Accion"] = "❓"
        action_col = "Accion"

    ws["A1"] = "KPIs Globales del Portfolio"
    ws["A1"].font = Font(bold=True, size=14)

    eva_total = float(df["EVA_post"].sum())
    deva_total = float(df["ΔEVA"].sum())
    cap_col = "capital_release_realized" if "capital_release_realized" in df.columns else "capital_liberado"
    cap_total = float(pd.to_numeric(df.get(cap_col, 0.0), errors="coerce").fillna(0.0).sum())
    counts = df[action_col].value_counts()

    n = len(df) if len(df) > 0 else 1

    rows = [
        ("EVA total", eva_total, NUM_EURO),
        ("ΔEVA total", deva_total, NUM_EURO),
        ("Capital liberado total", cap_total, NUM_EURO),
        ("% Mantener", counts.get("MANTENER", 0)/n, NUM_PCT),
        ("% Reestructurar", counts.get("REESTRUCTURAR", 0)/n, NUM_PCT),
        ("% Vender", counts.get("VENDER", 0)/n, NUM_PCT),
    ]

    r = 3
    for k, v, fmt in rows:
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        ws[f"B{r}"].number_format = fmt
        r += 1

    _borders(ws, start_row=3)
    _autofit(ws)


# ============================================
# 📄 HOJA: parámetros
# ============================================
def _write_params(wb: Workbook):
    ws = wb.create_sheet("parametros")
    ws.append(["Parámetro", "Valor"])
    _style_header_row(ws, header_row=1)

    # defensivo por si alguna rama falta
    try:
        cap_ratio = cfg.CONFIG.regulacion.required_total_capital_ratio()
    except Exception:
        cap_ratio = None

    try:
        buffers = cfg.CONFIG.regulacion.buffers.total_buffer()
    except Exception:
        buffers = None

    items = [
        ("Hurdle rate", getattr(cfg.CONFIG.regulacion, "hurdle_rate", None)),
        ("Capital ratio total", cap_ratio),
        ("Buffers CET1", buffers),
        ("Reward.w_eva", getattr(cfg.CONFIG.reward, "w_eva", None)),
        ("Reward.w_capital", getattr(cfg.CONFIG.reward, "w_capital", None)),
        ("Reward.w_stab", getattr(cfg.CONFIG.reward, "w_stab", None)),
        ("Timesteps PPO", getattr(cfg.CONFIG.ppo, "total_timesteps", None)),
    ]

    for k, v in items:
        ws.append([k, v])

    _borders(ws, start_row=2)
    _autofit(ws)


# ============================================
# MAIN
# ============================================
def export_styled_excel(df: pd.DataFrame, out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    _write_decisiones(wb, df)
    _write_resumen_segmento(wb, df)
    _write_resumen_global(wb, df)
    _write_params(wb)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(out_path)
    return out_path
