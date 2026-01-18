# -*- coding: utf-8 -*-
# ============================================
# reports/export_styled_excel_summary.py
# — Styler RUN-LEVEL para summary_consolidated (no requiere loan_id)
# ============================================

from __future__ import annotations
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule


def export_styled_excel_summary(df: pd.DataFrame, out_path: str, sheet_name: str = "summary") -> str:
    if df is None or df.empty:
        raise ValueError("No se puede exportar summary: DataFrame vacío.")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto width
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for c in col:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    # Formatos típicos
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    money_like = {"evamean", "delta_eva_mean", "capital_liberado_mean"}
    uplift_cols = {c for c in col_map if isinstance(c, str) and c.lower().startswith("uplift_") and c.endswith("_%")}

    for name, col_idx in col_map.items():
        if name in money_like:
            for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cells:
                    c.number_format = '#,##0.00'
        if name in uplift_cols:
            for cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cells:
                    c.number_format = '0.00"%"'

    # Condicional para uplifts
    for name in uplift_cols:
        col_idx = col_map[name]
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        rng = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    wb.save(out_path)
    return out_path
